from __future__ import annotations

import argparse
import csv
import json
import statistics
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from fc_slot_family import is_cath_lab_slot_service
from fc_payer_basis_resolution import (
    AUTO_BASIS,
    PAYER_BASIS_OPTIONS,
    load_resolution_rows,
    selection_lookup_formula,
    supported_basis_options_from_resolution_rows,
)
from professional_fee_review_workbook import (
    PF_PAYOR_ORDER,
    build_pf_summary_lookup,
    get_pf_summary_row,
    load_csv_rows as load_pf_csv_rows,
    load_json as load_pf_json,
    write_professional_fees_review_sheet,
)
from workbook_postprocess import apply_default_calc_settings, recalculate_workbook_with_soffice


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = REPO_ROOT / "output" / "general_medical_management_fc"
DEFAULT_OUTPUT = REPO_ROOT / "output" / "fc_estimate_builder_general_medical_management_cash_tr1.xlsx"
DEFAULT_PHARMACY_TEMPLATE = DEFAULT_OUTPUT_DIR / "01_clean_chemo_pharmacy_template_cash.csv"
DEFAULT_PHARMACY_PER_PATIENT = DEFAULT_OUTPUT_DIR / "02_per_patient_pharmacy_bucket_totals_cash.csv"
DEFAULT_IP_PHARMACY_PER_DAY = DEFAULT_OUTPUT_DIR / "05_ip_bucket_los_normalized_percentiles_cash.csv"
DEFAULT_SERVICES_TEMPLATE = DEFAULT_OUTPUT_DIR / "10_clean_chemo_services_template_cash.csv"
DEFAULT_CLEANED_SERVICES = DEFAULT_OUTPUT_DIR / "11_clean_chemo_services_template_for_fc_cash.csv"
DEFAULT_DEFAULT_SERVICES = DEFAULT_OUTPUT_DIR / "12_default_included_services_cash.csv"
DEFAULT_OPTIONAL_SERVICES = DEFAULT_OUTPUT_DIR / "13_optional_service_add_ons_cash.csv"
DEFAULT_SERVICE_LINE_COUNT = DEFAULT_OUTPUT_DIR / "14_service_line_count_metrics_cash.json"
DEFAULT_ROOM_METRICS = DEFAULT_OUTPUT_DIR / "16_los_icu_ward_room_metrics_cash.json"
DEFAULT_RATE_CSV = DEFAULT_OUTPUT_DIR / "tr1_cash_rates_general_medical_management_full_codes.csv"
DEFAULT_CATH_LAB_METRICS = DEFAULT_OUTPUT_DIR / "17_cath_lab_metrics_cash.json"
DEFAULT_PAYER_BASIS_RESOLUTION = DEFAULT_OUTPUT_DIR / "30_payer_basis_resolution_summary.csv"
DEFAULT_IP_ACTUALS_CSV = DEFAULT_OUTPUT_DIR / "18a_all_cash_patients_fc_bucket_rollup_reconciled.csv"

SHEET_BUILDER = "Builder"
SHEET_SUMMARY = "Estimate Summary"
SHEET_ADVANCED = "Advanced Controls"
SHEET_SERVICE_ADDONS = "Service Add-Ons"
SHEET_GROUPED_ADJUSTMENTS = "Grouped Adjustments"
SHEET_GROUPING_REVIEW = "Grouping Review"
SHEET_BREAKDOWN = "Estimate Breakdown"
SHEET_DETAIL = "Line Item Detail"
SHEET_PHARMACY_TEMPLATE = "Pharmacy Template"
SHEET_SERVICE_TEMPLATE = "Service Template"
SHEET_PHARMACY_METRICS = "Pharmacy Metrics"
SHEET_IP_ACTUALS = "IP FC Actuals"
SHEET_PF_REVIEW = "Professional Fees Review"
SHEET_REFERENCE = "Reference"

THIN_GREY = Side(style="thin", color="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="EAF4EA")
RESULT_FILL = PatternFill("solid", fgColor="FFF2CC")
REFERENCE_FILL = PatternFill("solid", fgColor="F4F6F8")

MODE_LOW = "Low"
MODE_TYPICAL = "Typical"
MODE_HIGH = "High"
INCLUDE = "Include"
EXCLUDE = "Exclude"
GROUP_PRESENCE_AUTO = 90.0
GROUP_PRESENCE_OPTIONAL = 75.0

DRIVER_AUTO = "Auto"
DRIVER_P25 = "P25"
DRIVER_P50 = "P50"
DRIVER_P75 = "P75"
DRIVER_MANUAL = "Manual"

ROOM_GENERAL = "General"
ROOM_TWIN = "Twin"
ROOM_SINGLE = "Single"
ROOM_ICU = "ICU"

SUMMARY_BUCKETS = [
    "Room Charges",
    "Investigations",
    "Doctors / Professionals",
    "Bedside Services",
    "Other Services",
    "Pharmacy",
    "Optional Add-Ons",
]

DRUG_CLASS = "Drugs / Medicines / IVs / Nutrition Products"
SUPPLY_CLASS = "Treatment Supplies"
PHARMACY_SHORTLIST_MAX_ROWS = 10
PHARMACY_SHORTLIST_TARGET_CUM_SHARE = 0.80

LOGIC_CODES = {
    "ROM0001",
    "ROM0024",
    "ROM0036",
    "ROM5009",
    "ROM5189",
    "ROM0093",
    "HSP5013",
    "ICC0001",
    "ICC0002",
    "EME0019",
}

CORE_OVERRIDE_CODES = {
    "PAT0042",
    "BIO0003",
    "PAT0041",
    "EME0020",
}

BED_CODES = {
    ROOM_GENERAL: "ROM0001",
    ROOM_TWIN: "ROM0024",
    ROOM_SINGLE: "ROM0036",
    ROOM_ICU: "ROM5009",
}


@dataclass
class RateRow:
    item_name: str
    general: float | None
    twin: float | None
    single: float | None
    icu: float | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the General Medical Management cash/TR1 FC estimate workbook.")
    parser.add_argument("--pharmacy-template", type=Path, default=DEFAULT_PHARMACY_TEMPLATE)
    parser.add_argument("--pharmacy-per-patient", type=Path, default=DEFAULT_PHARMACY_PER_PATIENT)
    parser.add_argument("--ip-pharmacy-per-day", type=Path, default=DEFAULT_IP_PHARMACY_PER_DAY)
    parser.add_argument("--services-template", type=Path, default=DEFAULT_SERVICES_TEMPLATE)
    parser.add_argument("--cleaned-services", type=Path, default=DEFAULT_CLEANED_SERVICES)
    parser.add_argument("--default-services", type=Path, default=DEFAULT_DEFAULT_SERVICES)
    parser.add_argument("--optional-services", type=Path, default=DEFAULT_OPTIONAL_SERVICES)
    parser.add_argument("--service-line-count", type=Path, default=DEFAULT_SERVICE_LINE_COUNT)
    parser.add_argument("--room-metrics", type=Path, default=DEFAULT_ROOM_METRICS)
    parser.add_argument("--rate-csv", type=Path, default=DEFAULT_RATE_CSV)
    parser.add_argument("--cath-lab-metrics", type=Path, default=DEFAULT_CATH_LAB_METRICS)
    parser.add_argument("--payer-basis-resolution-csv", type=Path, default=DEFAULT_PAYER_BASIS_RESOLUTION)
    parser.add_argument("--ip-actuals-csv", type=Path, default=DEFAULT_IP_ACTUALS_CSV)
    parser.add_argument("--pf-payor-summary-csv", type=Path)
    parser.add_argument("--pf-shape-review-json", type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_code(value: Any) -> str:
    return normalize_text(value).replace(" ", "").upper()


def maybe_float(value: Any) -> float | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def as_float(value: Any) -> float:
    number = maybe_float(value)
    return 0.0 if number is None else float(number)


def inclusive_quartiles(values: list[float]) -> tuple[float, float, float]:
    cleaned = sorted(float(value) for value in values)
    if not cleaned:
        return 0.0, 0.0, 0.0
    if len(cleaned) == 1:
        return cleaned[0], cleaned[0], cleaned[0]

    def percentile(p: float) -> float:
        index = (len(cleaned) - 1) * p
        lower = int(index)
        upper = min(lower + 1, len(cleaned) - 1)
        fraction = index - lower
        return cleaned[lower] + (cleaned[upper] - cleaned[lower]) * fraction

    return percentile(0.25), percentile(0.50), percentile(0.75)


def summarize_numeric_series(values: list[float]) -> dict[str, float]:
    cleaned = [float(value) for value in values]
    if not cleaned:
        return {"min": 0.0, "max": 0.0, "average": 0.0, "p25": 0.0, "p50": 0.0, "p75": 0.0}
    p25, p50, p75 = inclusive_quartiles(cleaned)
    return {
        "min": min(cleaned),
        "max": max(cleaned),
        "average": float(statistics.mean(cleaned)),
        "p25": p25,
        "p50": p50,
        "p75": p75,
    }


def load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def load_cath_lab_metrics(path: Path) -> tuple[float, float, float]:
    payload = load_json(path)
    metrics = payload.get("metrics") or {}
    return (
        as_float(metrics.get("p25")),
        as_float(metrics.get("p50")),
        as_float(metrics.get("p75")),
    )


def filter_out_cath_lab_slot_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [
        row
        for row in rows
        if not is_cath_lab_slot_service(
            code=row.get("item_code"),
            grouping=row.get("grouping"),
            service_name=row.get("item_name"),
        )
    ]


def load_rate_lookup(path: Path) -> dict[str, RateRow]:
    lookup: dict[str, RateRow] = {}
    for row in load_csv_rows(path):
        code = normalize_code(row.get("item_code"))
        if not code:
            continue
        lookup[code] = RateRow(
            item_name=normalize_text(row.get("item_name")),
            general=maybe_float(row.get("general")),
            twin=maybe_float(row.get("twin")),
            single=maybe_float(row.get("single")),
            icu=maybe_float(row.get("icu")),
        )
    return lookup


def style_cell(cell, *, fill=None, bold=False, font_color="000000", align="center", wrap=False):
    if fill is not None:
        cell.fill = fill
    cell.font = Font(bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)


def style_row(ws, row_number: int, start_col: int, end_col: int, *, fill=None, bold=False, font_color="000000", align="center", wrap=False):
    for col in range(start_col, end_col + 1):
        style_cell(ws.cell(row=row_number, column=col), fill=fill, bold=bold, font_color=font_color, align=align, wrap=wrap)


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def room_rate_cell(code: str, room: str) -> str:
    room_col = {
        ROOM_GENERAL: "B",
        ROOM_TWIN: "C",
        ROOM_SINGLE: "D",
        ROOM_ICU: "E",
    }[room]
    return (
        f'=IFERROR(0+INDEX(\'{SHEET_REFERENCE}\'!${room_col}:${room_col},'
        f'MATCH("{code}",\'{SHEET_REFERENCE}\'!$A:$A,0)),0)'
    )


def choose_amount_formula(selected_room_ref: str, selected_mode_ref: str, row_num: int) -> str:
    return (
        f'=IF({selected_room_ref}="{ROOM_GENERAL}",IF({selected_mode_ref}="{MODE_LOW}",P{row_num},IF({selected_mode_ref}="{MODE_HIGH}",R{row_num},Q{row_num})),'
        f'IF({selected_room_ref}="{ROOM_TWIN}",IF({selected_mode_ref}="{MODE_LOW}",S{row_num},IF({selected_mode_ref}="{MODE_HIGH}",U{row_num},T{row_num})),'
        f'IF({selected_room_ref}="{ROOM_SINGLE}",IF({selected_mode_ref}="{MODE_LOW}",V{row_num},IF({selected_mode_ref}="{MODE_HIGH}",X{row_num},W{row_num})),'
        f'IF({selected_mode_ref}="{MODE_LOW}",Y{row_num},IF({selected_mode_ref}="{MODE_HIGH}",AA{row_num},Z{row_num})))))'
    )


def normalize_summary_bucket(raw_bucket: str) -> str:
    bucket = normalize_text(raw_bucket)
    if bucket == "Room Charges":
        return "Room Charges"
    if bucket == "Investigations":
        return "Investigations"
    if bucket in {"Doctors & Professionals - General - Needed", "Anesthetist - General - Needed"}:
        return "Doctors / Professionals"
    if bucket == "Bedside Services":
        return "Bedside Services"
    return "Other Services"


def dedupe_service_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    deduped: dict[str, dict[str, str]] = {}
    for row in rows:
        code = normalize_code(row.get("item_code"))
        key = code or normalize_text(row.get("canonical_item_key")) or normalize_text(row.get("item_name"))
        if not key:
            continue
        current = deduped.get(key)
        if current is None:
            deduped[key] = row
            continue
        current_case_count = as_float(current.get("case_count"))
        new_case_count = as_float(row.get("case_count"))
        current_rate_count = as_float(current.get("rate_cash_count"))
        new_rate_count = as_float(row.get("rate_cash_count"))
        if (new_case_count, new_rate_count, normalize_text(row.get("item_name"))) > (
            current_case_count,
            current_rate_count,
            normalize_text(current.get("item_name")),
        ):
            deduped[key] = row
    return sorted(
        deduped.values(),
        key=lambda row: (
            -as_float(row.get("case_count")),
            -as_float(row.get("amount_cash_typical")),
            normalize_text(row.get("item_name")),
        ),
    )


def build_core_service_rows(default_rows: list[dict[str, str]], cleaned_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    by_code = {normalize_code(row.get("item_code")): row for row in dedupe_service_rows(cleaned_rows)}
    selected: dict[str, dict[str, str]] = {}
    for row in dedupe_service_rows(default_rows):
        code = normalize_code(row.get("item_code"))
        if code and code not in LOGIC_CODES and not is_cath_lab_slot_service(code=code, grouping=row.get("grouping"), service_name=row.get("item_name")):
            selected[code] = by_code.get(code, row)
    for code in CORE_OVERRIDE_CODES:
        if code in LOGIC_CODES:
            continue
        row = by_code.get(code)
        if row:
            selected[code] = row
    return sorted(selected.values(), key=lambda row: (normalize_summary_bucket(row.get("fc_estimate_bucket", "")), normalize_text(row.get("item_name"))))


def build_optional_service_rows(optional_rows: list[dict[str, str]], rate_lookup: dict[str, RateRow]) -> list[dict[str, Any]]:
    selected: list[dict[str, Any]] = []
    excluded_codes = LOGIC_CODES | set(CORE_OVERRIDE_CODES) | {normalize_code(row.get("item_code")) for row in dedupe_service_rows(optional_rows) if normalize_code(row.get("item_code")) in LOGIC_CODES}
    for row in dedupe_service_rows(optional_rows):
        code = normalize_code(row.get("item_code"))
        if not code or code in excluded_codes:
            continue
        if is_cath_lab_slot_service(code=code, grouping=row.get("grouping"), service_name=row.get("item_name")):
            continue
        rate_row = rate_lookup.get(code)
        if not rate_row:
            continue
        selected.append(
            {
                "item_code": code,
                "item_name": normalize_text(row.get("item_name")),
                "fc_bucket": normalize_text(row.get("fc_estimate_bucket")),
                "grouping": normalize_text(row.get("grouping")),
                "presence_rate": as_float(row.get("case_presence_rate")),
                "qty_p25": as_float(row.get("quantity_p25")) or 1.0,
                "qty_p50": as_float(row.get("quantity_p50")) or 1.0,
                "qty_p75": as_float(row.get("quantity_p75")) or 1.0,
                "rate_general": rate_row.general,
                "rate_twin": rate_row.twin,
                "rate_single": rate_row.single,
                "rate_icu": rate_row.icu,
            }
        )
    return selected


def preferred_general_rate(rate_row: RateRow | None) -> float:
    if not rate_row:
        return 0.0
    for value in (rate_row.general, rate_row.twin, rate_row.single, rate_row.icu):
        if value is not None:
            return float(value)
    return 0.0


def classify_group_residual_band(group_presence_rate: float, residual_p50: float) -> str:
    if residual_p50 <= 0.01:
        return ""
    if group_presence_rate > GROUP_PRESENCE_AUTO:
        return "auto"
    if group_presence_rate >= GROUP_PRESENCE_OPTIONAL:
        return "optional"
    return ""


def build_grouping_candidates(
    cleaned_rows: list[dict[str, str]],
    core_service_rows: list[dict[str, str]],
    optional_rows: list[dict[str, Any]],
    rate_lookup: dict[str, RateRow],
) -> list[dict[str, Any]]:
    default_codes = {normalize_code(row.get("item_code")) for row in core_service_rows}
    optional_codes = {row["item_code"] for row in optional_rows}
    grouped: dict[str, dict[str, Any]] = {}
    for row in dedupe_service_rows(cleaned_rows):
        code = normalize_code(row.get("item_code"))
        if not code or code in LOGIC_CODES:
            continue
        if is_cath_lab_slot_service(code=code, grouping=row.get("grouping"), service_name=row.get("item_name")):
            continue
        if code not in default_codes and code not in optional_codes:
            continue
        grouping = normalize_text(row.get("grouping")) or "Ungrouped"
        fc_bucket = normalize_summary_bucket(row.get("fc_estimate_bucket", ""))
        presence = as_float(row.get("case_presence_rate"))
        rate = preferred_general_rate(rate_lookup.get(code))
        qty_p25 = as_float(row.get("quantity_p25"))
        qty_p50 = as_float(row.get("quantity_p50"))
        qty_p75 = as_float(row.get("quantity_p75"))
        state = grouped.setdefault(
            grouping,
            {
                "grouping": grouping,
                "fc_bucket": fc_bucket,
                "group_presence_rate": 0.0,
                "group_amount_p25_exact": 0.0,
                "group_amount_p50_exact": 0.0,
                "group_amount_p75_exact": 0.0,
                "group_amount_captured_by_default_rows": 0.0,
            },
        )
        state["group_presence_rate"] = max(state["group_presence_rate"], presence)
        state["group_amount_p25_exact"] += qty_p25 * rate
        state["group_amount_p50_exact"] += qty_p50 * rate
        state["group_amount_p75_exact"] += qty_p75 * rate
        if code in default_codes:
            state["group_amount_captured_by_default_rows"] += qty_p50 * rate
    rows: list[dict[str, Any]] = []
    for grouping, state in sorted(grouped.items()):
        residual_p50 = max(0.0, state["group_amount_p50_exact"] - state["group_amount_captured_by_default_rows"])
        band = classify_group_residual_band(state["group_presence_rate"], residual_p50)
        if not band:
            continue
        state["group_residual_band"] = band
        rows.append(state)
    return rows


def clamp_ratio(value: float) -> float:
    return max(0.0, min(1.0, value))


def build_pharmacy_variance_shortlist(
    pharmacy_rows: list[dict[str, str]],
    *,
    fc_bucket: str,
    present_key: str,
    qty_key: str,
    amount_key: str,
    baseline_p25: float,
    baseline_p50: float,
    baseline_p75: float,
) -> list[dict[str, Any]]:
    eligible: list[dict[str, Any]] = []
    for row in pharmacy_rows:
        if normalize_text(row.get("fc_bucket")) != fc_bucket:
            continue
        if normalize_text(row.get(present_key)) != "Yes":
            continue
        presence_rate = as_float(row.get("case_presence_rate"))
        typical_qty = as_float(row.get(qty_key))
        typical_amount = as_float(row.get(amount_key))
        if presence_rate <= 0 or typical_amount <= 0:
            continue
        if presence_rate >= 70:
            continue
        typical_rate = typical_amount / typical_qty if typical_qty > 0 else as_float(row.get("observed_rate_p50"))
        expected_contribution = (presence_rate / 100.0) * typical_amount
        eligible.append(
            {
                "item_code": normalize_code(row.get("item_code")),
                "item_name": normalize_text(row.get("item_name")),
                "presence_rate": presence_rate,
                "typical_qty": typical_qty,
                "typical_rate": typical_rate,
                "typical_amount": typical_amount,
                "expected_contribution": expected_contribution,
            }
        )
    if not eligible:
        return []

    eligible.sort(key=lambda row: (-row["expected_contribution"], row["item_name"], row["item_code"]))
    total_expected = sum(row["expected_contribution"] for row in eligible)
    shortlist: list[dict[str, Any]] = []
    cumulative_expected = 0.0
    for row in eligible:
        shortlist.append(dict(row))
        cumulative_expected += row["expected_contribution"]
        if len(shortlist) >= PHARMACY_SHORTLIST_MAX_ROWS:
            break
        if total_expected > 0 and (cumulative_expected / total_expected) >= PHARMACY_SHORTLIST_TARGET_CUM_SHARE:
            break

    running = 0.0
    for row in shortlist:
        share = row["expected_contribution"] / total_expected if total_expected > 0 else 0.0
        running += share
        row["share"] = share
        row["cumulative_share"] = running

    if baseline_p75 > baseline_p25:
        target_ratio = clamp_ratio((baseline_p50 - baseline_p25) / (baseline_p75 - baseline_p25))
    else:
        target_ratio = 0.0
    running_selected = 0.0
    crossed = False
    for index, row in enumerate(shortlist):
        include = False
        if target_ratio > 0 and not crossed:
            include = True
            running_selected += row["share"]
            if running_selected >= target_ratio or index == len(shortlist) - 1:
                crossed = True
        row["default_selected"] = include
    return shortlist


def load_ip_pharmacy_per_day_percentiles(path: Path) -> dict[str, tuple[float, float, float]]:
    mapping: dict[str, tuple[float, float, float]] = {}
    for row in load_csv_rows(path):
        metric = normalize_text(row.get("metric"))
        mapping[metric] = (as_float(row.get("p25")), as_float(row.get("p50")), as_float(row.get("p75")))
    return mapping


def whole_day_metric(value: float, *, minimum: int = 0) -> int:
    return max(minimum, int(round(as_float(value))))


def write_reference_sheet(
    ws,
    rate_lookup: dict[str, RateRow],
    ip_pharmacy_percentiles: dict[str, tuple[float, float, float]],
    room_metrics: dict[str, Any],
    service_line_metrics: dict[str, Any],
    cath_lab_metrics: tuple[float, float, float],
    grouping_rows: list[dict[str, Any]],
    payer_basis_resolution_rows: list[dict[str, str]],
    pf_payor_summary_rows: list[dict[str, str]],
) -> dict[str, str]:
    supported_basis_options = [AUTO_BASIS, *supported_basis_options_from_resolution_rows(payer_basis_resolution_rows)]
    ws.sheet_state = "hidden"
    ws["A1"] = "item_code"
    ws["B1"] = "general"
    ws["C1"] = "twin"
    ws["D1"] = "single"
    ws["E1"] = "icu"
    ws["F1"] = "item_name"
    row = 2
    for code in sorted(rate_lookup.keys()):
        rate_row = rate_lookup[code]
        ws[f"A{row}"] = code
        ws[f"B{row}"] = rate_row.general
        ws[f"C{row}"] = rate_row.twin
        ws[f"D{row}"] = rate_row.single
        ws[f"E{row}"] = rate_row.icu
        ws[f"F{row}"] = rate_row.item_name
        row += 1

    ws["H1"] = "mode"
    ws["H2"] = MODE_LOW
    ws["H3"] = MODE_TYPICAL
    ws["H4"] = MODE_HIGH
    ws["I1"] = "room"
    ws["I2"] = ROOM_GENERAL
    ws["I3"] = ROOM_TWIN
    ws["I4"] = ROOM_SINGLE
    ws["I5"] = ROOM_ICU
    ws["J1"] = "basis"
    ws["J2"] = DRIVER_AUTO
    ws["J3"] = DRIVER_P25
    ws["J4"] = DRIVER_P50
    ws["J5"] = DRIVER_P75
    ws["J6"] = DRIVER_MANUAL
    ws["K1"] = "selection"
    ws["K2"] = INCLUDE
    ws["K3"] = EXCLUDE

    metric_map = room_metrics.get("metrics", {})
    los_metric = metric_map.get("effective_los_days") or metric_map.get("los_days") or {}
    ws["M1"] = "driver"
    ws["N1"] = "p25"
    ws["O1"] = "p50"
    ws["P1"] = "p75"
    driver_rows = [
        ("los_days", los_metric, 2, 1),
        ("icu_days", metric_map.get("icu_days") or {}, 3, 0),
        ("ward_days", metric_map.get("ward_days") or {}, 4, 0),
    ]
    for metric_name, metric, target_row, minimum in driver_rows:
        ws[f"M{target_row}"] = metric_name
        ws[f"N{target_row}"] = whole_day_metric(metric.get("p25"), minimum=minimum)
        ws[f"O{target_row}"] = whole_day_metric(metric.get("p50"), minimum=minimum)
        ws[f"P{target_row}"] = whole_day_metric(metric.get("p75"), minimum=minimum)

    service_count = service_line_metrics.get("cleaned_distinct_service_line_count") or {}
    ws["R1"] = "service_line_metric"
    ws["S1"] = "value"
    ws["R2"] = "p25"
    ws["R3"] = "p50"
    ws["R4"] = "p75"
    ws["S2"] = as_float(service_count.get("p25"))
    ws["S3"] = as_float(service_count.get("p50"))
    ws["S4"] = as_float(service_count.get("p75"))

    ws["U1"] = "ip_pharmacy_metric"
    ws["V1"] = "p25"
    ws["W1"] = "p50"
    ws["X1"] = "p75"
    metric_rows = {
        "ip_drugs_per_day": 2,
        "ip_treatment_supplies_per_day": 3,
    }
    for metric_name, target_row in metric_rows.items():
        q1, q2, q3 = ip_pharmacy_percentiles.get(metric_name, (0.0, 0.0, 0.0))
        ws[f"U{target_row}"] = metric_name
        ws[f"V{target_row}"] = q1
        ws[f"W{target_row}"] = q2
        ws[f"X{target_row}"] = q3

    ws["U5"] = "cath_lab_amount_net"
    ws["V5"] = cath_lab_metrics[0]
    ws["W5"] = cath_lab_metrics[1]
    ws["X5"] = cath_lab_metrics[2]

    ws["Z1"] = "grouping"
    ws["AA1"] = "fc_bucket"
    ws["AB1"] = "presence_rate"
    ws["AC1"] = "group_amount_p25_exact"
    ws["AD1"] = "group_amount_p50_exact"
    ws["AE1"] = "group_amount_p75_exact"
    ws["AF1"] = "captured_by_default"
    ws["AG1"] = "group_residual_band"
    for row_idx, row in enumerate(grouping_rows, start=2):
        ws[f"Z{row_idx}"] = row["grouping"]
        ws[f"AA{row_idx}"] = row["fc_bucket"]
        ws[f"AB{row_idx}"] = row["group_presence_rate"] / 100.0
        ws[f"AC{row_idx}"] = row["group_amount_p25_exact"]
        ws[f"AD{row_idx}"] = row["group_amount_p50_exact"]
        ws[f"AE{row_idx}"] = row["group_amount_p75_exact"]
        ws[f"AF{row_idx}"] = row["group_amount_captured_by_default_rows"]
        ws[f"AG{row_idx}"] = row["group_residual_band"]

    ws["AI1"] = "payer_basis_option"
    for row_idx, label in enumerate(supported_basis_options, start=2):
        ws[f"AI{row_idx}"] = label

    ws["AK1"] = "component"
    ws["AL1"] = "target_payor_bucket"
    ws["AM1"] = "selected_basis"
    ws["AN1"] = "selected_case_count"
    ws["AO1"] = "selection_reason"
    for row_idx, row in enumerate(payer_basis_resolution_rows, start=2):
        ws[f"AK{row_idx}"] = normalize_text(row.get("component"))
        ws[f"AL{row_idx}"] = normalize_text(row.get("target_payor_bucket"))
        ws[f"AM{row_idx}"] = normalize_text(row.get("selected_basis"))
        ws[f"AN{row_idx}"] = as_float(row.get("selected_case_count"))
        ws[f"AO{row_idx}"] = normalize_text(row.get("selection_reason"))

    ws["AQ1"] = "payor_bucket"
    ws["AR1"] = "admission_count"
    ws["AS1"] = "pf_collectible_historical_total_p25"
    ws["AT1"] = "pf_collectible_historical_total_p50"
    ws["AU1"] = "pf_collectible_historical_total_p75"
    ws["AV1"] = "pf_named_total_p50"
    ws["AW1"] = "pf_general_needed_total_p50"
    ws["AX1"] = "surgeon_named_total_p50"
    ws["AY1"] = "assistant_surgeon_named_total_p50"
    ws["AZ1"] = "anesthetist_named_total_p50"
    ws["BA1"] = "assistant_anesthetist_named_total_p50"
    ws["BB1"] = "consultant_or_physician_named_total_p50"
    ws["BC1"] = "dominant_pf_shape"
    for row_idx, row in enumerate(pf_payor_summary_rows, start=2):
        ws[f"AQ{row_idx}"] = normalize_text(row.get("payor_bucket"))
        ws[f"AR{row_idx}"] = as_float(row.get("admission_count"))
        ws[f"AS{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p25"))
        ws[f"AT{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p50"))
        ws[f"AU{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p75"))
        ws[f"AV{row_idx}"] = as_float(row.get("pf_named_total_p50"))
        ws[f"AW{row_idx}"] = as_float(row.get("pf_general_needed_total_p50"))
        ws[f"AX{row_idx}"] = as_float(row.get("surgeon_named_total_p50"))
        ws[f"AY{row_idx}"] = as_float(row.get("assistant_surgeon_named_total_p50"))
        ws[f"AZ{row_idx}"] = as_float(row.get("anesthetist_named_total_p50"))
        ws[f"BA{row_idx}"] = as_float(row.get("assistant_anesthetist_named_total_p50"))
        ws[f"BB{row_idx}"] = as_float(row.get("consultant_or_physician_named_total_p50"))
        ws[f"BC{row_idx}"] = normalize_text(row.get("dominant_pf_shape"))

    return {
        "los_p25": "Reference!N2",
        "los_p50": "Reference!O2",
        "los_p75": "Reference!P2",
        "icu_p25": "Reference!N3",
        "icu_p50": "Reference!O3",
        "icu_p75": "Reference!P3",
        "ward_p25": "Reference!N4",
        "ward_p50": "Reference!O4",
        "ward_p75": "Reference!P4",
        "service_line_p25": "Reference!S2",
        "service_line_p50": "Reference!S3",
        "service_line_p75": "Reference!S4",
        "ip_drugs_per_day_p25": "Reference!V2",
        "ip_drugs_per_day_p50": "Reference!W2",
        "ip_drugs_per_day_p75": "Reference!X2",
        "ip_supplies_per_day_p25": "Reference!V3",
        "ip_supplies_per_day_p50": "Reference!W3",
        "ip_supplies_per_day_p75": "Reference!X3",
        "cath_lab_p25": "Reference!V5",
        "cath_lab_p50": "Reference!W5",
        "cath_lab_p75": "Reference!X5",
        "payer_basis_options": f"Reference!AI2:AI{len(supported_basis_options) + 1}",
    }


def group_metric_formula(metric_col: str, grouping_ref: str) -> str:
    return f'=IFERROR(INDEX({SHEET_REFERENCE}!${metric_col}:${metric_col},MATCH({grouping_ref},{SHEET_REFERENCE}!$Z:$Z,0)),0)'


def write_grouping_review_sheet(ws, grouping_rows: list[dict[str, Any]]) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Grouping Review"
    ws["A2"] = "This review sheet highlights only high-presence service groups where the current default GMM estimate does not fully capture the common-case grouped amount."
    style_row(ws, 1, 1, 8, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_row(ws, 2, 1, 8, fill=REFERENCE_FILL, align="left", wrap=True)
    headers = ["Grouping", "FC Bucket", "Presence Rate", "Group Amount P25 Exact", "Group Amount P50 Exact", "Group Amount P75 Exact", "Captured by Default", "Residual Band"]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=4, column=idx, value=value)
    style_row(ws, 4, 1, 8, fill=SUBHEADER_FILL, bold=True)
    for row_idx, row in enumerate(grouping_rows, start=5):
        ws[f"A{row_idx}"] = row["grouping"]
        ws[f"B{row_idx}"] = row["fc_bucket"]
        ws[f"C{row_idx}"] = row["group_presence_rate"] / 100.0
        ws[f"D{row_idx}"] = row["group_amount_p25_exact"]
        ws[f"E{row_idx}"] = row["group_amount_p50_exact"]
        ws[f"F{row_idx}"] = row["group_amount_p75_exact"]
        ws[f"G{row_idx}"] = row["group_amount_captured_by_default_rows"]
        ws[f"H{row_idx}"] = row["group_residual_band"]
        style_row(ws, row_idx, 1, 8, fill=FORMULA_FILL, align="left", wrap=True)
        ws[f"C{row_idx}"].number_format = '0.00%'
        for col in "DEFG":
            ws[f"{col}{row_idx}"].number_format = '#,##0.00'
    set_widths(ws, {"A": 28, "B": 22, "C": 14, "D": 16, "E": 16, "F": 16, "G": 16, "H": 14})


def write_grouped_adjustments_sheet(ws, grouping_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Grouped Adjustments"
    ws["A2"] = "Grouped adjustments complete common service groups without double counting. Selected optional service add-ons from the same grouping reduce the residual automatically."
    style_row(ws, 1, 1, 14, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_row(ws, 2, 1, 14, fill=REFERENCE_FILL, align="left", wrap=True)
    ws["A4"], ws["B4"], ws["C4"], ws["D4"], ws["E4"] = "Grouped Adjustments", "Low", "Typical", "High", "Included Rows"
    style_row(ws, 4, 1, 5, fill=SUBHEADER_FILL, bold=True, align="left")
    headers = [
        "Grouping", "FC Bucket", "Group Presence Rate", "Group Amount P25 Exact", "Group Amount P50 Exact", "Group Amount P75 Exact",
        "Captured By Default", "Selected Add-On Amount", "Net Residual Low", "Net Residual Typical", "Net Residual High", "Selected", "Selected Amount", "Why",
    ]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=6, column=idx, value=value)
    style_row(ws, 6, 1, 14, fill=SUBHEADER_FILL, bold=True, wrap=True)
    dropdown = DataValidation(type="list", formula1=f'"{INCLUDE},{EXCLUDE}"', allow_blank=False)
    ws.add_data_validation(dropdown)
    mode_ref = f"'{SHEET_BUILDER}'!$B$6"
    rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(grouping_rows, start=7):
        ws[f"A{row_idx}"] = row["grouping"]
        ws[f"B{row_idx}"] = row["fc_bucket"]
        ws[f"C{row_idx}"] = row["group_presence_rate"] / 100.0
        ws[f"D{row_idx}"] = row["group_amount_p25_exact"]
        ws[f"E{row_idx}"] = row["group_amount_p50_exact"]
        ws[f"F{row_idx}"] = row["group_amount_p75_exact"]
        ws[f"G{row_idx}"] = row["group_amount_captured_by_default_rows"]
        ws[f"O{row_idx}"] = f'=SUMPRODUCT((\'{SHEET_SERVICE_ADDONS}\'!$E$5:$E$999=$A{row_idx})*(\'{SHEET_SERVICE_ADDONS}\'!$A$5:$A$999="{INCLUDE}")*(\'{SHEET_SERVICE_ADDONS}\'!$G$5:$G$999)*(\'{SHEET_SERVICE_ADDONS}\'!$J$5:$J$999))'
        ws[f"P{row_idx}"] = f'=SUMPRODUCT((\'{SHEET_SERVICE_ADDONS}\'!$E$5:$E$999=$A{row_idx})*(\'{SHEET_SERVICE_ADDONS}\'!$A$5:$A$999="{INCLUDE}")*(\'{SHEET_SERVICE_ADDONS}\'!$H$5:$H$999)*(\'{SHEET_SERVICE_ADDONS}\'!$J$5:$J$999))'
        ws[f"Q{row_idx}"] = f'=SUMPRODUCT((\'{SHEET_SERVICE_ADDONS}\'!$E$5:$E$999=$A{row_idx})*(\'{SHEET_SERVICE_ADDONS}\'!$A$5:$A$999="{INCLUDE}")*(\'{SHEET_SERVICE_ADDONS}\'!$I$5:$I$999)*(\'{SHEET_SERVICE_ADDONS}\'!$J$5:$J$999))'
        ws[f"H{row_idx}"] = choose_amount_formula(f'"{ROOM_GENERAL}"', mode_ref, row_idx).replace("P", "O").replace("Q", "P").replace("R", "Q").replace("S", "O").replace("T", "P").replace("U", "Q").replace("V", "O").replace("W", "P").replace("X", "Q").replace("Y", "O").replace("Z", "P").replace("AA", "Q")
        ws[f"I{row_idx}"] = f"=MAX(0,D{row_idx}-G{row_idx}-O{row_idx})"
        ws[f"J{row_idx}"] = f"=MAX(0,E{row_idx}-G{row_idx}-P{row_idx})"
        ws[f"K{row_idx}"] = f"=MAX(0,F{row_idx}-G{row_idx}-Q{row_idx})"
        ws[f"L{row_idx}"] = INCLUDE if row["group_residual_band"] == "auto" else EXCLUDE
        dropdown.add(ws[f"L{row_idx}"])
        ws[f"M{row_idx}"] = f'=IF(L{row_idx}="{INCLUDE}",IF({mode_ref}="{MODE_LOW}",I{row_idx},IF({mode_ref}="{MODE_HIGH}",K{row_idx},J{row_idx})),0)'
        ws[f"N{row_idx}"] = "Auto common-case residual" if row["group_residual_band"] == "auto" else "Optional common-case residual"
        style_row(ws, row_idx, 1, 14, fill=FORMULA_FILL, align="left", wrap=True)
        ws[f"L{row_idx}"].fill = INPUT_FILL
        ws[f"C{row_idx}"].number_format = '0.00%'
        for col in "DEFGHIJKM":
            ws[f"{col}{row_idx}"].number_format = '#,##0.00'
        rows.append({"sheet_row": row_idx, **row})
    end = max(7, 6 + len(grouping_rows))
    ws["B5"] = f'=SUMIF(L7:L{end},"{INCLUDE}",I7:I{end})' if grouping_rows else 0
    ws["C5"] = f'=SUMIF(L7:L{end},"{INCLUDE}",J7:J{end})' if grouping_rows else 0
    ws["D5"] = f'=SUMIF(L7:L{end},"{INCLUDE}",K7:K{end})' if grouping_rows else 0
    ws["E5"] = f'=COUNTIF(L7:L{end},"{INCLUDE}")' if grouping_rows else 0
    style_row(ws, 5, 1, 5, fill=RESULT_FILL, bold=True, align="left")
    for col in ["O", "P", "Q"]:
        ws.column_dimensions[col].hidden = True
    set_widths(ws, {"A": 28, "B": 22, "C": 14, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 16, "K": 16, "L": 12, "M": 16, "N": 30})
    return rows


def write_builder_sheet(ws, refs: dict[str, str]) -> dict[str, str]:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"
    ws.merge_cells("A1:G1")
    ws["A1"] = "FC Estimate Builder - General Medical Management (Non-Daycare)"
    style_row(ws, 1, 1, 7, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")

    ws["A3"] = "Template"
    ws["B3"] = "General Medical Management (Non-Daycare)"
    ws["A4"] = "Payor"
    ws["B4"] = "Cash (TR1)"
    ws["A5"] = "Selected Room Type"
    ws["B5"] = ROOM_TWIN
    ws["A6"] = "Selected Estimate Mode"
    ws["B6"] = MODE_TYPICAL
    for row in range(3, 7):
        style_cell(ws[f"A{row}"], fill=SUBHEADER_FILL, bold=True, align="left")
        style_cell(ws[f"B{row}"], fill=INPUT_FILL if row in {5, 6} else FORMULA_FILL, align="left")

    ws["D3"] = "Historical Payer Basis"
    ws["E3"] = AUTO_BASIS
    service_resolution_formula = selection_lookup_formula('"service_basis"', '"Cash"', "AM", component_col="AK", target_payor_col="AL").lstrip("=")
    pharmacy_resolution_formula = selection_lookup_formula('"pharmacy_basis"', '"Cash"', "AM", component_col="AK", target_payor_col="AL").lstrip("=")
    pf_resolution_formula = selection_lookup_formula('"pf_basis"', '"Cash"', "AM", component_col="AK", target_payor_col="AL").lstrip("=")
    resolver_note_formula = selection_lookup_formula('"service_basis"', '"Cash"', "AO", component_col="AK", target_payor_col="AL").lstrip("=")
    ws["D4"] = "Resolved Service Basis"
    ws["E4"] = f'=IF(E3<>"{AUTO_BASIS}",E3,{service_resolution_formula})'
    ws["D5"] = "Resolved Pharmacy Basis"
    ws["E5"] = f'=IF(E3<>"{AUTO_BASIS}",E3,{pharmacy_resolution_formula})'
    ws["D6"] = "Resolved PF Basis"
    ws["E6"] = f'=IF(E3<>"{AUTO_BASIS}",E3,{pf_resolution_formula})'
    ws["D7"] = "Resolver Note"
    ws["E7"] = f'=IF(E3<>"{AUTO_BASIS}","Manual historical payer basis override applied",{resolver_note_formula})'
    for row in range(3, 8):
        style_cell(ws[f"D{row}"], fill=SUBHEADER_FILL, bold=True, align="left")
        style_cell(ws[f"E{row}"], fill=INPUT_FILL if row == 3 else (REFERENCE_FILL if row == 7 else FORMULA_FILL), align="left", wrap=(row == 7))

    room_validation = DataValidation(type="list", formula1=f"='{SHEET_REFERENCE}'!$I$2:$I$5", allow_blank=False)
    mode_validation = DataValidation(type="list", formula1=f"='{SHEET_REFERENCE}'!$H$2:$H$4", allow_blank=False)
    payer_basis_validation = DataValidation(type="list", formula1=f"='{SHEET_REFERENCE}'!$AI$2:INDEX('{SHEET_REFERENCE}'!$AI:$AI,COUNTA('{SHEET_REFERENCE}'!$AI:$AI))", allow_blank=False)
    ws.add_data_validation(room_validation)
    ws.add_data_validation(mode_validation)
    ws.add_data_validation(payer_basis_validation)
    room_validation.add(ws["B5"])
    mode_validation.add(ws["B6"])
    payer_basis_validation.add(ws["E3"])

    ws["A7"] = "Clinical Drivers"
    style_row(ws, 7, 1, 7, fill=SUBHEADER_FILL, bold=True, align="left")

    headers = ["Driver", "P25", "P50", "P75", "Selection", "Manual Override", "Selected Value"]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=8, column=idx, value=value)
    style_row(ws, 8, 1, 7, fill=SUBHEADER_FILL, bold=True)

    basis_validation = DataValidation(type="list", formula1=f"='{SHEET_REFERENCE}'!$J$3:$J$6", allow_blank=False)
    ws.add_data_validation(basis_validation)

    driver_specs = [
        ("LOS (Days)", refs["los_p25"], refs["los_p50"], refs["los_p75"], DRIVER_P50),
        ("ICU Days", refs["icu_p25"], refs["icu_p50"], refs["icu_p75"], DRIVER_P50),
        ("Ward Days", refs["ward_p25"], refs["ward_p50"], refs["ward_p75"], DRIVER_P50),
    ]
    cell_refs: dict[str, str] = {}
    for row_num, (label, p25_ref, p50_ref, p75_ref, basis_default) in enumerate(driver_specs, start=9):
        ws[f"A{row_num}"] = label
        ws[f"B{row_num}"] = f"={p25_ref}"
        ws[f"C{row_num}"] = f"={p50_ref}"
        ws[f"D{row_num}"] = f"={p75_ref}"
        ws[f"E{row_num}"] = basis_default
        ws[f"F{row_num}"] = f"=C{row_num}"
        ws[f"G{row_num}"] = (
            f'=IF(E{row_num}="{DRIVER_P25}",B{row_num},IF(E{row_num}="{DRIVER_P50}",C{row_num},IF(E{row_num}="{DRIVER_P75}",D{row_num},F{row_num})))'
        )
        ws[f"H{row_num}"] = (
            f'=IF(E{row_num}="{DRIVER_MANUAL}",F{row_num},IF(E{row_num}="{DRIVER_P25}",B{row_num},IF(E{row_num}="{DRIVER_P50}",C{row_num},IF(E{row_num}="{DRIVER_P75}",D{row_num},B{row_num}))))'
        )
        ws[f"I{row_num}"] = (
            f'=IF(E{row_num}="{DRIVER_MANUAL}",F{row_num},IF(E{row_num}="{DRIVER_P25}",B{row_num},IF(E{row_num}="{DRIVER_P50}",C{row_num},IF(E{row_num}="{DRIVER_P75}",D{row_num},C{row_num}))))'
        )
        ws[f"J{row_num}"] = (
            f'=IF(E{row_num}="{DRIVER_MANUAL}",F{row_num},IF(E{row_num}="{DRIVER_P25}",B{row_num},IF(E{row_num}="{DRIVER_P50}",C{row_num},IF(E{row_num}="{DRIVER_P75}",D{row_num},D{row_num}))))'
        )
        basis_validation.add(ws[f"E{row_num}"])
        style_row(ws, row_num, 1, 7, fill=FORMULA_FILL, align="left")
        style_cell(ws[f"E{row_num}"], fill=INPUT_FILL, align="left")
        style_cell(ws[f"F{row_num}"], fill=INPUT_FILL, align="left")
        style_cell(ws[f"G{row_num}"], fill=FORMULA_FILL, bold=True, align="left")
        for col in "BCDFGHIJ":
            ws[f"{col}{row_num}"].number_format = '#,##0.00'
        key = label.lower().replace(" ", "_").replace("(", "").replace(")", "")
        cell_refs[f"{key}_low"] = f"'{SHEET_BUILDER}'!H{row_num}"
        cell_refs[f"{key}_typical"] = f"'{SHEET_BUILDER}'!I{row_num}"
        cell_refs[f"{key}_high"] = f"'{SHEET_BUILDER}'!J{row_num}"
        cell_refs[f"{key}_selected"] = f"'{SHEET_BUILDER}'!G{row_num}"

    ws["A14"] = "Selected Headline Estimate"
    ws["B14"] = f"='{SHEET_SUMMARY}'!B5"
    style_cell(ws["A14"], fill=SUBHEADER_FILL, bold=True, align="left")
    style_cell(ws["B14"], fill=RESULT_FILL, bold=True, align="left")
    ws["B14"].number_format = '#,##0.00'

    notes_row = 16
    ws[f"A{notes_row}"] = "How To Use"
    ws[f"A{notes_row+1}"] = "1. Choose room type and estimate mode."
    ws[f"A{notes_row+2}"] = "2. Review LOS / ICU / Ward history and override only if needed."
    ws[f"A{notes_row+3}"] = "3. Use Advanced Controls only for pharmacy variance and optional service adjustments."
    ws[f"A{notes_row+4}"] = "4. Read the final estimate on Estimate Summary."
    style_row(ws, notes_row, 1, 7, fill=SUBHEADER_FILL, bold=True, align="left")
    style_row(ws, notes_row + 1, 1, 7, fill=REFERENCE_FILL, align="left", wrap=True)
    style_row(ws, notes_row + 2, 1, 7, fill=REFERENCE_FILL, align="left", wrap=True)
    style_row(ws, notes_row + 3, 1, 7, fill=REFERENCE_FILL, align="left", wrap=True)
    style_row(ws, notes_row + 4, 1, 7, fill=REFERENCE_FILL, align="left", wrap=True)

    ws.column_dimensions["H"].hidden = True
    ws.column_dimensions["I"].hidden = True
    ws.column_dimensions["J"].hidden = True
    set_widths(ws, {"A": 24, "B": 14, "C": 14, "D": 18, "E": 28, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14, "L": 18, "M": 18, "N": 14, "O": 42})
    cell_refs["selected_room"] = f"'{SHEET_BUILDER}'!B5"
    cell_refs["selected_mode"] = f"'{SHEET_BUILDER}'!B6"
    return cell_refs


def write_advanced_controls_sheet(
    ws,
    pharmacy_rows: list[dict[str, str]],
    builder_refs: dict[str, str],
    baseline_metrics: dict[str, float],
    base_service_count: int,
) -> dict[str, str]:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws["A1"] = "Advanced Controls"
    ws["A2"] = "Use these controls only to refine the default estimate. Pharmacy selections move the estimate between historical P25 and P75."
    style_row(ws, 1, 1, 13, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_row(ws, 2, 1, 13, fill=REFERENCE_FILL, align="left", wrap=True)

    selection_validation = DataValidation(type="list", formula1=f"='{SHEET_REFERENCE}'!$K$2:$K$3", allow_blank=False)
    ws.add_data_validation(selection_validation)

    def write_pharmacy_block(
        *,
        title: str,
        start_row: int,
        shortlist: list[dict[str, Any]],
        low_formula: str,
        base_formula: str,
        high_formula: str,
    ) -> dict[str, str]:
        ws[f"A{start_row}"] = title
        ws[f"B{start_row}"] = "Low"
        ws[f"C{start_row}"] = "Typical"
        ws[f"D{start_row}"] = "High"
        style_row(ws, start_row, 1, 4, fill=SUBHEADER_FILL, bold=True, align="left")
        ws[f"B{start_row+1}"] = f"={low_formula}"
        ws[f"D{start_row+1}"] = f"={high_formula}"
        if shortlist:
            shortlist_start = start_row + 3
            shortlist_end = shortlist_start + len(shortlist) - 1
            ws[f"C{start_row+1}"] = (
                f'=IFERROR((SUMIF(H{shortlist_start}:H{shortlist_end},"{INCLUDE}",F{shortlist_start}:F{shortlist_end})/'
                f'SUM(F{shortlist_start}:F{shortlist_end}))*(D{start_row+1}-B{start_row+1})+B{start_row+1},{base_formula})'
            )
        else:
            ws[f"C{start_row+1}"] = f"={base_formula}"
        style_row(ws, start_row + 1, 1, 4, fill=RESULT_FILL, bold=True, align="left")
        for col in "BCD":
            ws[f"{col}{start_row+1}"].number_format = '#,##0.00'

        headers = ["Item", "Typical Qty", "Typical Rate", "Typical Amount", "Presence Rate", "Expected Contribution", "Cumulative Share", "Selected"]
        header_row = start_row + 3
        for idx, value in enumerate(headers, start=1):
            ws.cell(row=header_row, column=idx, value=value)
        style_row(ws, header_row, 1, 8, fill=SUBHEADER_FILL, bold=True, wrap=True)

        if shortlist:
            for idx, row in enumerate(shortlist, start=header_row + 1):
                ws[f"A{idx}"] = row["item_name"]
                ws[f"B{idx}"] = row["typical_qty"]
                ws[f"C{idx}"] = row["typical_rate"]
                ws[f"D{idx}"] = row["typical_amount"]
                ws[f"E{idx}"] = row["presence_rate"] / 100.0
                ws[f"F{idx}"] = row["expected_contribution"]
                ws[f"G{idx}"] = row["cumulative_share"]
                ws[f"H{idx}"] = INCLUDE if row["default_selected"] else EXCLUDE
                selection_validation.add(ws[f"H{idx}"])
                style_row(ws, idx, 1, 8, fill=FORMULA_FILL, align="left", wrap=True)
                ws[f"A{idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws[f"H{idx}"].fill = INPUT_FILL
                for col in "BCDF":
                    ws[f"{col}{idx}"].number_format = '#,##0.00'
                ws[f"E{idx}"].number_format = '0.00%'
                ws[f"G{idx}"].number_format = '0.00%'
            end_row = header_row + len(shortlist)
        else:
            ws[f"A{header_row + 1}"] = "No shortlisted high-variance items found for this bucket in the inpatient cohort."
            style_row(ws, header_row + 1, 1, 8, fill=REFERENCE_FILL, align="left", wrap=True)
            end_row = header_row + 1

        return {
            "low": f"'{SHEET_ADVANCED}'!B{start_row+1}",
            "typical": f"'{SHEET_ADVANCED}'!C{start_row+1}",
            "high": f"'{SHEET_ADVANCED}'!D{start_row+1}",
            "end_row": str(end_row),
        }

    drugs_shortlist = build_pharmacy_variance_shortlist(
        pharmacy_rows,
        fc_bucket=DRUG_CLASS,
        present_key="present_in_ip",
        qty_key="ip_net_quantity_p50",
        amount_key="ip_net_amount_p50",
        baseline_p25=baseline_metrics["los_p50"] * baseline_metrics["ip_drugs_per_day_p25"],
        baseline_p50=baseline_metrics["los_p50"] * baseline_metrics["ip_drugs_per_day_p50"],
        baseline_p75=baseline_metrics["los_p50"] * baseline_metrics["ip_drugs_per_day_p75"],
    )
    drugs_refs = write_pharmacy_block(
        title="IP Drugs",
        start_row=4,
        shortlist=drugs_shortlist,
        low_formula=f"{builder_refs['los_days_selected']}*{builder_refs['ip_drugs_per_day_p25']}",
        base_formula=f"{builder_refs['los_days_selected']}*{builder_refs['ip_drugs_per_day_p50']}",
        high_formula=f"{builder_refs['los_days_selected']}*{builder_refs['ip_drugs_per_day_p75']}",
    )

    supplies_start = int(drugs_refs["end_row"]) + 3
    supplies_shortlist = build_pharmacy_variance_shortlist(
        pharmacy_rows,
        fc_bucket=SUPPLY_CLASS,
        present_key="present_in_ip",
        qty_key="ip_net_quantity_p50",
        amount_key="ip_net_amount_p50",
        baseline_p25=baseline_metrics["los_p50"] * baseline_metrics["ip_supplies_per_day_p25"],
        baseline_p50=baseline_metrics["los_p50"] * baseline_metrics["ip_supplies_per_day_p50"],
        baseline_p75=baseline_metrics["los_p50"] * baseline_metrics["ip_supplies_per_day_p75"],
    )
    supplies_refs = write_pharmacy_block(
        title="IP Treatment Supplies",
        start_row=supplies_start,
        shortlist=supplies_shortlist,
        low_formula=f"{builder_refs['los_days_selected']}*{builder_refs['ip_supplies_per_day_p25']}",
        base_formula=f"{builder_refs['los_days_selected']}*{builder_refs['ip_supplies_per_day_p50']}",
        high_formula=f"{builder_refs['los_days_selected']}*{builder_refs['ip_supplies_per_day_p75']}",
    )

    ws["J4"] = "Service Line Count Alert"
    ws["J5"] = "Historical P25"
    ws["J6"] = "Historical P50"
    ws["J7"] = "Historical P75"
    ws["J8"] = "Base Included Non-Pharmacy Count"
    ws["J9"] = "Selected Optional Count"
    ws["J10"] = "Current Included Non-Pharmacy Count"
    ws["J11"] = "Alert"
    ws["K5"] = f"={builder_refs['service_line_p25']}"
    ws["K6"] = f"={builder_refs['service_line_p50']}"
    ws["K7"] = f"={builder_refs['service_line_p75']}"
    style_row(ws, 4, 10, 11, fill=SUBHEADER_FILL, bold=True, align="left")
    for row in range(5, 12):
        style_row(ws, row, 10, 11, fill=FORMULA_FILL if row < 11 else RESULT_FILL, bold=row >= 8, align="left")

    ws["K8"] = base_service_count
    ws["K9"] = f'=COUNTIF(\'{SHEET_SERVICE_ADDONS}\'!A5:A999,"{INCLUDE}")'
    ws["K10"] = "=K8+K9"
    ws["K11"] = '=IF(K10<K5,"Below historical P25",IF(K10>K7,"Above historical P75","Within historical range"))'

    set_widths(
        ws,
        {
            "A": 14,
            "B": 14,
            "C": 38,
            "D": 30,
            "E": 28,
            "F": 12,
            "G": 11,
            "H": 11,
            "I": 11,
            "J": 12,
            "K": 12,
            "L": 12,
            "M": 12,
        },
    )
    return {
        "ip_drugs_low": drugs_refs["low"],
        "ip_drugs_typical": drugs_refs["typical"],
        "ip_drugs_high": drugs_refs["high"],
        "ip_supplies_low": supplies_refs["low"],
        "ip_supplies_typical": supplies_refs["typical"],
        "ip_supplies_high": supplies_refs["high"],
        "service_line_count_current": f"'{SHEET_ADVANCED}'!K10",
        "service_line_count_alert": f"'{SHEET_ADVANCED}'!K11",
    }


def write_service_addons_sheet(ws, optional_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Service Add-Ons"
    ws["A2"] = "Use this sheet for exact optional service rows. If you include an item here, the grouped residual for the same grouping shrinks automatically on Grouped Adjustments."
    style_row(ws, 1, 1, 13, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_row(ws, 2, 1, 13, fill=REFERENCE_FILL, align="left", wrap=True)
    headers = [
        "Select",
        "Item Code",
        "Item Name",
        "Original FC Bucket",
        "Grouping",
        "Presence %",
        "Qty P25",
        "Qty P50",
        "Qty P75",
        "Rate General",
        "Rate Twin",
        "Rate Single",
        "Rate ICU",
    ]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=4, column=idx, value=value)
    style_row(ws, 4, 1, 13, fill=SUBHEADER_FILL, bold=True)

    selection_validation = DataValidation(type="list", formula1=f"='{SHEET_REFERENCE}'!$K$2:$K$3", allow_blank=False)
    ws.add_data_validation(selection_validation)
    output_rows: list[dict[str, Any]] = []
    current_row = 5
    for row in optional_rows:
        ws[f"A{current_row}"] = EXCLUDE
        selection_validation.add(ws[f"A{current_row}"])
        ws[f"B{current_row}"] = row["item_code"]
        ws[f"C{current_row}"] = row["item_name"]
        ws[f"D{current_row}"] = row["fc_bucket"]
        ws[f"E{current_row}"] = row["grouping"]
        ws[f"F{current_row}"] = row["presence_rate"] / 100.0
        ws[f"G{current_row}"] = row["qty_p25"]
        ws[f"H{current_row}"] = row["qty_p50"]
        ws[f"I{current_row}"] = row["qty_p75"]
        ws[f"J{current_row}"] = row["rate_general"]
        ws[f"K{current_row}"] = row["rate_twin"]
        ws[f"L{current_row}"] = row["rate_single"]
        ws[f"M{current_row}"] = row["rate_icu"]
        style_row(ws, current_row, 1, 13, fill=FORMULA_FILL, align="left", wrap=True)
        style_cell(ws[f"A{current_row}"], fill=INPUT_FILL, align="left")
        ws[f"F{current_row}"].number_format = '0.00%'
        for col in "GHIJKLM":
            ws[f"{col}{current_row}"].number_format = '#,##0.00'
        output_rows.append({"sheet_row": current_row, **row})
        current_row += 1

    set_widths(
        ws,
        {
            "A": 14,
            "B": 14,
            "C": 38,
            "D": 30,
            "E": 28,
            "F": 12,
            "G": 11,
            "H": 11,
            "I": 11,
            "J": 12,
            "K": 12,
            "L": 12,
            "M": 12,
        },
    )
    return output_rows


def write_detail_sheet(
    ws,
    builder_refs: dict[str, str],
    core_service_rows: list[dict[str, str]],
    optional_rows: list[dict[str, Any]],
    grouped_rows: list[dict[str, Any]],
) -> tuple[int, int]:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    headers = [
        "Line Item",
        "Parent Bucket",
        "Sub-Bucket",
        "Source Type",
        "How Calculated",
        "Item Code",
        "Count In Service Line Metric",
        "Qty Low",
        "Qty Typical",
        "Qty High",
        "Selected Qty",
        "General Rate",
        "Twin Rate",
        "Single Rate",
        "ICU Rate",
        "General Low",
        "General Typical",
        "General High",
        "Twin Low",
        "Twin Typical",
        "Twin High",
        "Single Low",
        "Single Typical",
        "Single High",
        "ICU Low",
        "ICU Typical",
        "ICU High",
        "Selected Amount",
    ]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=value)
    style_row(ws, 1, 1, len(headers), fill=HEADER_FILL, bold=True, font_color="FFFFFF")

    selected_room_ref = builder_refs["selected_room"]
    selected_mode_ref = builder_refs["selected_mode"]
    current_row = 2

    def add_row(payload: dict[str, Any]) -> int:
        nonlocal current_row
        row_num = current_row
        ws[f"A{row_num}"] = payload["name"]
        ws[f"B{row_num}"] = payload["parent"]
        ws[f"C{row_num}"] = payload["sub"]
        ws[f"D{row_num}"] = payload["source"]
        ws[f"E{row_num}"] = payload["how"]
        ws[f"F{row_num}"] = payload.get("code", "")
        ws[f"G{row_num}"] = payload.get("countable", "Yes")
        for col, key in zip("HIJK", ["qty_low", "qty_typical", "qty_high", "qty_selected"]):
            ws[f"{col}{row_num}"] = payload.get(key, 0)
        for col, key in zip("LMNO", ["rate_general", "rate_twin", "rate_single", "rate_icu"]):
            ws[f"{col}{row_num}"] = payload.get(key, "")
        ws[f"P{row_num}"] = payload.get("general_low_formula", f'=IF(L{row_num}="",0,H{row_num}*L{row_num})')
        ws[f"Q{row_num}"] = payload.get("general_typical_formula", f'=IF(L{row_num}="",0,I{row_num}*L{row_num})')
        ws[f"R{row_num}"] = payload.get("general_high_formula", f'=IF(L{row_num}="",0,J{row_num}*L{row_num})')
        ws[f"S{row_num}"] = payload.get("twin_low_formula", f'=IF(M{row_num}="",0,H{row_num}*M{row_num})')
        ws[f"T{row_num}"] = payload.get("twin_typical_formula", f'=IF(M{row_num}="",0,I{row_num}*M{row_num})')
        ws[f"U{row_num}"] = payload.get("twin_high_formula", f'=IF(M{row_num}="",0,J{row_num}*M{row_num})')
        ws[f"V{row_num}"] = payload.get("single_low_formula", f'=IF(N{row_num}="",0,H{row_num}*N{row_num})')
        ws[f"W{row_num}"] = payload.get("single_typical_formula", f'=IF(N{row_num}="",0,I{row_num}*N{row_num})')
        ws[f"X{row_num}"] = payload.get("single_high_formula", f'=IF(N{row_num}="",0,J{row_num}*N{row_num})')
        ws[f"Y{row_num}"] = payload.get("icu_low_formula", f'=IF(O{row_num}="",0,H{row_num}*O{row_num})')
        ws[f"Z{row_num}"] = payload.get("icu_typical_formula", f'=IF(O{row_num}="",0,I{row_num}*O{row_num})')
        ws[f"AA{row_num}"] = payload.get("icu_high_formula", f'=IF(O{row_num}="",0,J{row_num}*O{row_num})')
        ws[f"AB{row_num}"] = payload.get("selected_amount_formula", choose_amount_formula(selected_room_ref, selected_mode_ref, row_num))
        style_row(ws, row_num, 1, 28, fill=FORMULA_FILL, align="left", wrap=True)
        for col in "HIJKL MNOPQRSTUVWXYZAB".replace(" ", ""):
            ws[f"{col}{row_num}"].number_format = '#,##0.00'
        current_row += 1
        return row_num

    los_low = builder_refs["los_days_low"]
    los_typ = builder_refs["los_days_typical"]
    los_high = builder_refs["los_days_high"]
    los_sel = builder_refs["los_days_selected"]
    icu_low = builder_refs["icu_days_low"]
    icu_typ = builder_refs["icu_days_typical"]
    icu_high = builder_refs["icu_days_high"]
    icu_sel = builder_refs["icu_days_selected"]
    ward_low = builder_refs["ward_days_low"]
    ward_typ = builder_refs["ward_days_typical"]
    ward_high = builder_refs["ward_days_high"]
    ward_sel = builder_refs["ward_days_selected"]

    # Room / stay logic rows
    add_row(
        {
            "name": "Ward Bed Charges",
            "parent": "Room Charges",
            "sub": "Ward Bed Charges",
            "source": "Logic",
            "how": "Selected ward days multiplied by TR1 bed charge for the compared room category.",
            "code": "ROOM_BED",
            "qty_low": f"={ward_low}",
            "qty_typical": f"={ward_typ}",
            "qty_high": f"={ward_high}",
            "qty_selected": f"={ward_sel}",
            "rate_general": room_rate_cell("ROM0001", ROOM_GENERAL),
            "rate_twin": room_rate_cell("ROM0024", ROOM_TWIN),
            "rate_single": room_rate_cell("ROM0036", ROOM_SINGLE),
            "rate_icu": room_rate_cell("ROM5009", ROOM_ICU),
        }
    )
    add_row(
        {
            "name": "ICU Bed Charges",
            "parent": "Room Charges",
            "sub": "ICU Bed Charges",
            "source": "Logic",
            "how": "Selected ICU days multiplied by TR1 ICU bed charge.",
            "code": "ROM5009",
            "qty_low": f"={icu_low}",
            "qty_typical": f"={icu_typ}",
            "qty_high": f"={icu_high}",
            "qty_selected": f"={icu_sel}",
            "rate_general": room_rate_cell("ROM5009", ROOM_ICU),
            "rate_twin": room_rate_cell("ROM5009", ROOM_ICU),
            "rate_single": room_rate_cell("ROM5009", ROOM_ICU),
            "rate_icu": room_rate_cell("ROM5009", ROOM_ICU),
        }
    )
    for label, code in [("Nursing Charges - Ward", "ROM5189"), ("DMO Charges - Ward", "ROM0093")]:
        add_row(
            {
                "name": label,
                "parent": "Room Charges",
                "sub": label,
                "source": "Logic",
                "how": "Selected ward days multiplied by room-specific TR1 service rate.",
                "code": code,
                "qty_low": f"={ward_low}",
                "qty_typical": f"={ward_typ}",
                "qty_high": f"={ward_high}",
                "qty_selected": f"={ward_sel}",
                "rate_general": room_rate_cell(code, ROOM_GENERAL),
                "rate_twin": room_rate_cell(code, ROOM_TWIN),
                "rate_single": room_rate_cell(code, ROOM_SINGLE),
                "rate_icu": room_rate_cell(code, ROOM_ICU),
            }
        )
    for label, code in [("Nursing Charges - ICU", "ROM5189"), ("DMO Charges - ICU", "ROM0093"), ("Intensivist Per Day", "ICC0002"), ("Assistant Intensivist Per Day", "ICC0001"), ("Monitor Per Day", "EME0019")]:
        add_row(
            {
                "name": label,
                "parent": "Room Charges",
                "sub": label,
                "source": "Logic",
                "how": "Selected ICU days multiplied by TR1 ICU-support rate.",
                "code": code,
                "qty_low": f"={icu_low}",
                "qty_typical": f"={icu_typ}",
                "qty_high": f"={icu_high}",
                "qty_selected": f"={icu_sel}",
                "rate_general": room_rate_cell(code, ROOM_ICU),
                "rate_twin": room_rate_cell(code, ROOM_ICU),
                "rate_single": room_rate_cell(code, ROOM_ICU),
                "rate_icu": room_rate_cell(code, ROOM_ICU),
            }
        )
    add_row(
        {
            "name": "Ward Consumables",
            "parent": "Room Charges",
            "sub": "Ward Consumables",
            "source": "Logic",
            "how": "Total stay days (ward + ICU) multiplied by TR1 ward consumables rate.",
            "code": "HSP5013",
            "qty_low": f"={ward_low}+{icu_low}",
            "qty_typical": f"={ward_typ}+{icu_typ}",
            "qty_high": f"={ward_high}+{icu_high}",
            "qty_selected": f"={ward_sel}+{icu_sel}",
            "rate_general": room_rate_cell("HSP5013", ROOM_GENERAL),
            "rate_twin": room_rate_cell("HSP5013", ROOM_TWIN),
            "rate_single": room_rate_cell("HSP5013", ROOM_SINGLE),
            "rate_icu": room_rate_cell("HSP5013", ROOM_ICU),
        }
    )

    # Core service rows
    for row in core_service_rows:
        code = normalize_code(row.get("item_code"))
        bucket = normalize_summary_bucket(row.get("fc_estimate_bucket", ""))
        rate_general = room_rate_cell(code, ROOM_GENERAL)
        rate_twin = room_rate_cell(code, ROOM_TWIN)
        rate_single = room_rate_cell(code, ROOM_SINGLE)
        rate_icu = room_rate_cell(code, ROOM_ICU)
        add_row(
            {
                "name": normalize_text(row.get("item_name")),
                "parent": bucket,
                "sub": normalize_text(row.get("grouping")),
                "source": "Template Core",
                "how": "Included from cleaned GMM services template using the conservative default rule plus curated medical-core overrides.",
                "code": code,
                "qty_low": as_float(row.get("quantity_p25")) or 1.0,
                "qty_typical": as_float(row.get("quantity_p50")) or 1.0,
                "qty_high": as_float(row.get("quantity_p75")) or 1.0,
                "qty_selected": f'=IF({selected_mode_ref}="{MODE_LOW}",H{current_row},IF({selected_mode_ref}="{MODE_HIGH}",J{current_row},I{current_row}))',
                "rate_general": rate_general,
                "rate_twin": rate_twin,
                "rate_single": rate_single,
                "rate_icu": rate_icu,
            }
        )

    add_row(
        {
            "name": "Cath Lab Charges",
            "parent": "Procedure / OT Charges",
            "sub": "Cath Lab Hours",
            "source": "Historical Cath Lab Family",
            "how": "Actual billed cath-lab slot-family P25 / P50 / P75 from the filtered GMM cash cohort.",
            "countable": "No",
            "general_low_formula": f"={builder_refs['cath_lab_p25']}",
            "general_typical_formula": f"={builder_refs['cath_lab_p50']}",
            "general_high_formula": f"={builder_refs['cath_lab_p75']}",
            "twin_low_formula": f"=P{current_row}",
            "twin_typical_formula": f"=Q{current_row}",
            "twin_high_formula": f"=R{current_row}",
            "single_low_formula": f"=P{current_row}",
            "single_typical_formula": f"=Q{current_row}",
            "single_high_formula": f"=R{current_row}",
            "icu_low_formula": f"=P{current_row}",
            "icu_typical_formula": f"=Q{current_row}",
            "icu_high_formula": f"=R{current_row}",
        }
    )

    # Pharmacy rows
    add_row(
        {
            "name": "IP Drugs",
            "parent": "Pharmacy",
            "sub": "IP Drugs / Medicines / IVs / Nutrition Products",
            "source": "Historical Cash Pharmacy",
            "how": "Inpatient LOS-normalized historical baseline; Advanced Controls shortlist can move this bucket between P25 and P75.",
            "countable": "No",
            "qty_low": f"={los_low}",
            "qty_typical": f"={los_typ}",
            "qty_high": f"={los_high}",
            "qty_selected": f"={los_sel}",
            "general_low_formula": f"={builder_refs['ip_drugs_low']}",
            "general_typical_formula": f"={builder_refs['ip_drugs_typical']}",
            "general_high_formula": f"={builder_refs['ip_drugs_high']}",
            "twin_low_formula": f"=P{current_row}",
            "twin_typical_formula": f"=Q{current_row}",
            "twin_high_formula": f"=R{current_row}",
            "single_low_formula": f"=P{current_row}",
            "single_typical_formula": f"=Q{current_row}",
            "single_high_formula": f"=R{current_row}",
            "icu_low_formula": f"=P{current_row}",
            "icu_typical_formula": f"=Q{current_row}",
            "icu_high_formula": f"=R{current_row}",
        }
    )
    add_row(
        {
            "name": "IP Treatment Supplies",
            "parent": "Pharmacy",
            "sub": "IP Treatment Supplies",
            "source": "Historical Cash Pharmacy",
            "how": "Inpatient LOS-normalized historical baseline; Advanced Controls shortlist can move this bucket between P25 and P75.",
            "countable": "No",
            "qty_low": f"={los_low}",
            "qty_typical": f"={los_typ}",
            "qty_high": f"={los_high}",
            "qty_selected": f"={los_sel}",
            "general_low_formula": f"={builder_refs['ip_supplies_low']}",
            "general_typical_formula": f"={builder_refs['ip_supplies_typical']}",
            "general_high_formula": f"={builder_refs['ip_supplies_high']}",
            "twin_low_formula": f"=P{current_row}",
            "twin_typical_formula": f"=Q{current_row}",
            "twin_high_formula": f"=R{current_row}",
            "single_low_formula": f"=P{current_row}",
            "single_typical_formula": f"=Q{current_row}",
            "single_high_formula": f"=R{current_row}",
            "icu_low_formula": f"=P{current_row}",
            "icu_typical_formula": f"=Q{current_row}",
            "icu_high_formula": f"=R{current_row}",
        }
    )

    # Optional service add-ons
    for row in optional_rows:
        include_ref = f"'{SHEET_SERVICE_ADDONS}'!A{row['sheet_row']}"
        add_row(
            {
                "name": row["item_name"],
                "parent": "Optional Add-Ons",
                "sub": f"{row['fc_bucket']} / {row['grouping']}",
                "source": "Template Optional",
                "how": f'Included only when Advanced Controls marks this add-on as "{INCLUDE}".',
                "code": row["item_code"],
                "qty_low": row["qty_p25"],
                "qty_typical": row["qty_p50"],
                "qty_high": row["qty_p75"],
                "qty_selected": f'=IF({selected_mode_ref}="{MODE_LOW}",H{current_row},IF({selected_mode_ref}="{MODE_HIGH}",J{current_row},I{current_row}))',
                "rate_general": row["rate_general"] if row["rate_general"] is not None else "",
                "rate_twin": row["rate_twin"] if row["rate_twin"] is not None else "",
                "rate_single": row["rate_single"] if row["rate_single"] is not None else "",
                "rate_icu": row["rate_icu"] if row["rate_icu"] is not None else "",
                "general_low_formula": f'=IF({include_ref}="{INCLUDE}",IF(L{current_row}="",0,H{current_row}*L{current_row}),0)',
                "general_typical_formula": f'=IF({include_ref}="{INCLUDE}",IF(L{current_row}="",0,I{current_row}*L{current_row}),0)',
                "general_high_formula": f'=IF({include_ref}="{INCLUDE}",IF(L{current_row}="",0,J{current_row}*L{current_row}),0)',
                "twin_low_formula": f'=IF({include_ref}="{INCLUDE}",IF(M{current_row}="",0,H{current_row}*M{current_row}),0)',
                "twin_typical_formula": f'=IF({include_ref}="{INCLUDE}",IF(M{current_row}="",0,I{current_row}*M{current_row}),0)',
                "twin_high_formula": f'=IF({include_ref}="{INCLUDE}",IF(M{current_row}="",0,J{current_row}*M{current_row}),0)',
                "single_low_formula": f'=IF({include_ref}="{INCLUDE}",IF(N{current_row}="",0,H{current_row}*N{current_row}),0)',
                "single_typical_formula": f'=IF({include_ref}="{INCLUDE}",IF(N{current_row}="",0,I{current_row}*N{current_row}),0)',
                "single_high_formula": f'=IF({include_ref}="{INCLUDE}",IF(N{current_row}="",0,J{current_row}*N{current_row}),0)',
                "icu_low_formula": f'=IF({include_ref}="{INCLUDE}",IF(O{current_row}="",0,H{current_row}*O{current_row}),0)',
                "icu_typical_formula": f'=IF({include_ref}="{INCLUDE}",IF(O{current_row}="",0,I{current_row}*O{current_row}),0)',
                "icu_high_formula": f'=IF({include_ref}="{INCLUDE}",IF(O{current_row}="",0,J{current_row}*O{current_row}),0)',
            }
        )

    for row in grouped_rows:
        include_ref = f"'{SHEET_GROUPED_ADJUSTMENTS}'!L{row['sheet_row']}"
        add_row(
            {
                "name": f'{row["grouping"]} Residual',
                "parent": row["fc_bucket"],
                "sub": row["grouping"],
                "source": "Grouped Residual",
                "how": "Mode-aware grouped residual net of selected optional child add-ons from the same grouping.",
                "countable": "No",
                "general_low_formula": f'=IF({include_ref}="{INCLUDE}",\'{SHEET_GROUPED_ADJUSTMENTS}\'!I{row["sheet_row"]},0)',
                "general_typical_formula": f'=IF({include_ref}="{INCLUDE}",\'{SHEET_GROUPED_ADJUSTMENTS}\'!J{row["sheet_row"]},0)',
                "general_high_formula": f'=IF({include_ref}="{INCLUDE}",\'{SHEET_GROUPED_ADJUSTMENTS}\'!K{row["sheet_row"]},0)',
                "twin_low_formula": f"=P{current_row}",
                "twin_typical_formula": f"=Q{current_row}",
                "twin_high_formula": f"=R{current_row}",
                "single_low_formula": f"=P{current_row}",
                "single_typical_formula": f"=Q{current_row}",
                "single_high_formula": f"=R{current_row}",
                "icu_low_formula": f"=P{current_row}",
                "icu_typical_formula": f"=Q{current_row}",
                "icu_high_formula": f"=R{current_row}",
                "selected_amount_formula": f'=IF({include_ref}="{INCLUDE}",\'{SHEET_GROUPED_ADJUSTMENTS}\'!M{row["sheet_row"]},0)',
            }
        )

    total_row = current_row
    ws[f"A{total_row}"] = "Grand Total"
    for col_pair in [("P", "Q", "R"), ("S", "T", "U"), ("V", "W", "X"), ("Y", "Z", "AA")]:
        for col in col_pair:
            ws[f"{col}{total_row}"] = f"=SUM({col}2:{col}{current_row-1})"
            ws[f"{col}{total_row}"].number_format = '#,##0.00'
    ws[f"AB{total_row}"] = choose_amount_formula(selected_room_ref, selected_mode_ref, total_row)
    style_row(ws, total_row, 1, 28, fill=RESULT_FILL, bold=True, align="left")

    count_row = total_row + 2
    ws[f"A{count_row}"] = "Current Included Non-Pharmacy Service Line Count"
    ws[f"AB{count_row}"] = f"={builder_refs['service_line_count_current']}"
    style_row(ws, count_row, 1, 28, fill=SUBHEADER_FILL, bold=True, align="left")
    ws[f"AB{count_row}"].number_format = '0'

    set_widths(
        ws,
        {
            "A": 28,
            "B": 22,
            "C": 28,
            "D": 18,
            "E": 42,
            "F": 12,
            "G": 14,
            "H": 10,
            "I": 10,
            "J": 10,
            "K": 11,
            "L": 11,
            "M": 11,
            "N": 11,
            "O": 11,
            "P": 12,
            "Q": 12,
            "R": 12,
            "S": 12,
            "T": 12,
            "U": 12,
            "V": 12,
            "W": 12,
            "X": 12,
            "Y": 12,
            "Z": 12,
            "AA": 12,
            "AB": 14,
        }
    )
    return total_row, count_row


def write_summary_sheet(
    ws,
    detail_total_row: int,
    detail_count_row: int,
    builder_refs: dict[str, str],
    pf_payor_summary_rows: list[dict[str, str]],
) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Estimate Summary"
    style_row(ws, 1, 1, 8, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    ws["A3"] = "Selected Room"
    ws["B3"] = f"='{SHEET_BUILDER}'!B5"
    ws["A4"] = "Estimate Mode"
    ws["B4"] = f"='{SHEET_BUILDER}'!B6"
    ws["A5"] = "Historical Payer Basis"
    ws["B5"] = f"='{SHEET_BUILDER}'!E3"
    ws["A6"] = "Resolved Service Basis"
    ws["B6"] = f"='{SHEET_BUILDER}'!E4"
    ws["A7"] = "Resolved Pharmacy Basis"
    ws["B7"] = f"='{SHEET_BUILDER}'!E5"
    ws["A8"] = "Resolved PF Basis"
    ws["B8"] = f"='{SHEET_BUILDER}'!E6"
    ws["A9"] = "Headline Estimate"
    ws["B9"] = f"='{SHEET_DETAIL}'!AB{detail_total_row}"
    for cell in ["A3", "A4", "A5", "A6", "A7", "A8", "A9"]:
        style_cell(ws[cell], fill=SUBHEADER_FILL, bold=True, align="left")
    for cell in ["B3", "B4", "B5", "B6", "B7", "B8"]:
        style_cell(ws[cell], fill=FORMULA_FILL, align="left")
    style_cell(ws["B9"], fill=RESULT_FILL, bold=True, align="left")
    ws["B9"].number_format = '#,##0.00'

    ws["D3"] = "Room Comparison"
    style_row(ws, 3, 4, 7, fill=SUBHEADER_FILL, bold=True)
    ws["D4"] = "Room"
    ws["E4"] = MODE_LOW
    ws["F4"] = MODE_TYPICAL
    ws["G4"] = MODE_HIGH
    style_row(ws, 4, 4, 7, fill=REFERENCE_FILL, bold=True)
    room_to_cols = {
        ROOM_GENERAL: ("P", "Q", "R"),
        ROOM_TWIN: ("S", "T", "U"),
        ROOM_SINGLE: ("V", "W", "X"),
        ROOM_ICU: ("Y", "Z", "AA"),
    }
    row_idx = 5
    for room, cols in room_to_cols.items():
        ws[f"D{row_idx}"] = room
        ws[f"E{row_idx}"] = f"='{SHEET_DETAIL}'!{cols[0]}{detail_total_row}"
        ws[f"F{row_idx}"] = f"='{SHEET_DETAIL}'!{cols[1]}{detail_total_row}"
        ws[f"G{row_idx}"] = f"='{SHEET_DETAIL}'!{cols[2]}{detail_total_row}"
        style_row(ws, row_idx, 4, 7, fill=FORMULA_FILL, align="left")
        for col in "EFG":
            ws[f"{col}{row_idx}"].number_format = '#,##0.00'
        row_idx += 1

    ws["A11"] = "Selected Drivers"
    style_row(ws, 11, 1, 4, fill=SUBHEADER_FILL, bold=True)
    ws["A12"] = "LOS"
    ws["B12"] = f"='{SHEET_BUILDER}'!G9"
    ws["A13"] = "ICU Days"
    ws["B13"] = f"='{SHEET_BUILDER}'!G10"
    ws["A14"] = "Ward Days"
    ws["B14"] = f"='{SHEET_BUILDER}'!G11"
    for row in (12, 13, 14):
        style_row(ws, row, 1, 2, fill=FORMULA_FILL, align="left")
        ws[f"B{row}"].number_format = '#,##0.00'

    ws["D10"] = "Bucket"
    ws["E10"] = "Selected Amount"
    style_row(ws, 10, 4, 5, fill=SUBHEADER_FILL, bold=True)
    row_idx = 11
    for bucket in SUMMARY_BUCKETS:
        ws[f"D{row_idx}"] = bucket
        ws[f"E{row_idx}"] = f"=SUMIF('{SHEET_DETAIL}'!$B$2:$B${detail_total_row-1},D{row_idx},'{SHEET_DETAIL}'!$AB$2:$AB${detail_total_row-1})"
        style_row(ws, row_idx, 4, 5, fill=FORMULA_FILL, align="left")
        ws[f"E{row_idx}"].number_format = '#,##0.00'
        row_idx += 1
    ws[f"D{row_idx}"] = "Grand Total"
    ws[f"E{row_idx}"] = f"='{SHEET_DETAIL}'!AB{detail_total_row}"
    style_row(ws, row_idx, 4, 5, fill=RESULT_FILL, bold=True, align="left")
    ws[f"E{row_idx}"].number_format = '#,##0.00'

    service_metrics_row = row_idx + 2
    ws[f"A{service_metrics_row}"] = "Historical Non-Pharmacy Service Line Count"
    ws[f"B{service_metrics_row}"] = "P25"
    ws[f"C{service_metrics_row}"] = "P50"
    ws[f"D{service_metrics_row}"] = "P75"
    style_row(ws, service_metrics_row, 1, 4, fill=SUBHEADER_FILL, bold=True)
    ws[f"B{service_metrics_row+1}"] = f"={builder_refs['service_line_p25']}"
    ws[f"C{service_metrics_row+1}"] = f"={builder_refs['service_line_p50']}"
    ws[f"D{service_metrics_row+1}"] = f"={builder_refs['service_line_p75']}"
    ws[f"A{service_metrics_row+2}"] = "Current Included Non-Pharmacy Service Line Count"
    ws[f"B{service_metrics_row+2}"] = f"='{SHEET_DETAIL}'!AB{detail_count_row}"
    ws[f"A{service_metrics_row+3}"] = "Alert"
    ws[f"B{service_metrics_row+3}"] = f'=IF(B{service_metrics_row+2}<B{service_metrics_row+1},"Below historical P25",IF(B{service_metrics_row+2}>D{service_metrics_row+1},"Above historical P75","Within historical band"))'
    style_row(ws, service_metrics_row + 1, 1, 4, fill=FORMULA_FILL, align="left")
    style_row(ws, service_metrics_row + 2, 1, 2, fill=FORMULA_FILL, bold=True, align="left")
    style_row(ws, service_metrics_row + 3, 1, 2, fill=FORMULA_FILL, bold=True, align="left")

    pf_lookup = build_pf_summary_lookup(pf_payor_summary_rows)
    ws["I3"] = "Professional Fees by Payer"
    style_row(ws, 3, 9, 13, fill=SUBHEADER_FILL, bold=True, align="left")
    pf_headers = ["Payer", "Cases", "P25", "P50", "P75"]
    for idx, header in enumerate(pf_headers, start=9):
        ws.cell(row=4, column=idx, value=header)
    style_row(ws, 4, 9, 13, fill=REFERENCE_FILL, bold=True, align="left")
    for row_idx, payor_bucket in enumerate(PF_PAYOR_ORDER, start=5):
        row = get_pf_summary_row(pf_lookup, payor_bucket)
        ws[f"I{row_idx}"] = payor_bucket
        ws[f"J{row_idx}"] = as_float(row.get("admission_count"))
        ws[f"K{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p25"))
        ws[f"L{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p50"))
        ws[f"M{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p75"))
        style_row(ws, row_idx, 9, 13, fill=FORMULA_FILL, align="left")
        ws[f"J{row_idx}"].number_format = '#,##0'
        for col in "KLM":
            ws[f"{col}{row_idx}"].number_format = '#,##0.00'

    ws["I12"] = "Selected Basis PF Mix"
    style_row(ws, 12, 9, 10, fill=SUBHEADER_FILL, bold=True, align="left")
    pf_mix_rows = [
        ("Collectible Historical PF", '=IFERROR(INDEX(Reference!$AT:$AT,MATCH(B8,Reference!$AQ:$AQ,0)),0)'),
        ("Named PF", '=IFERROR(INDEX(Reference!$AV:$AV,MATCH(B8,Reference!$AQ:$AQ,0)),0)'),
        ("General Needed PF", '=IFERROR(INDEX(Reference!$AW:$AW,MATCH(B8,Reference!$AQ:$AQ,0)),0)'),
        ("Surgeon Named", '=IFERROR(INDEX(Reference!$AX:$AX,MATCH(B8,Reference!$AQ:$AQ,0)),0)'),
        ("Assistant Surgeon Named", '=IFERROR(INDEX(Reference!$AY:$AY,MATCH(B8,Reference!$AQ:$AQ,0)),0)'),
        ("Anesthetist Named", '=IFERROR(INDEX(Reference!$AZ:$AZ,MATCH(B8,Reference!$AQ:$AQ,0)),0)'),
        ("Assistant Anesthetist Named", '=IFERROR(INDEX(Reference!$BA:$BA,MATCH(B8,Reference!$AQ:$AQ,0)),0)'),
        ("Consultant / Physician Named", '=IFERROR(INDEX(Reference!$BB:$BB,MATCH(B8,Reference!$AQ:$AQ,0)),0)'),
        ("Selected Basis PF Shape", '=IFERROR(INDEX(Reference!$BC:$BC,MATCH(B8,Reference!$AQ:$AQ,0)),"n/a")'),
    ]
    for row_idx, (label, value) in enumerate(pf_mix_rows, start=13):
        ws[f"I{row_idx}"] = label
        ws[f"J{row_idx}"] = value
        style_row(ws, row_idx, 9, 10, fill=FORMULA_FILL, align="left")
        if row_idx < 21:
            ws[f"J{row_idx}"].number_format = '#,##0.00'

    set_widths(ws, {"A": 36, "B": 16, "C": 16, "D": 24, "E": 16, "F": 16, "G": 16, "I": 26, "J": 18, "K": 14, "L": 14, "M": 14})


def write_breakdown_sheet(ws, detail_last_data_row: int) -> None:
    ws.sheet_view.showGridLines = False
    headers = [
        "Line Item",
        "Bucket",
        "Sub-Bucket",
        "Source",
        "How Calculated",
        "Selected Quantity",
        "Rate General",
        "Rate Twin",
        "Rate Single",
        "Rate ICU",
        "Selected Amount",
    ]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=value)
    style_row(ws, 1, 1, len(headers), fill=HEADER_FILL, bold=True, font_color="FFFFFF")
    out_row = 2
    for detail_row in range(2, detail_last_data_row):
        ws[f"A{out_row}"] = f"='{SHEET_DETAIL}'!A{detail_row}"
        ws[f"B{out_row}"] = f"='{SHEET_DETAIL}'!B{detail_row}"
        ws[f"C{out_row}"] = f"='{SHEET_DETAIL}'!C{detail_row}"
        ws[f"D{out_row}"] = f"='{SHEET_DETAIL}'!D{detail_row}"
        ws[f"E{out_row}"] = f"='{SHEET_DETAIL}'!E{detail_row}"
        ws[f"F{out_row}"] = f"='{SHEET_DETAIL}'!K{detail_row}"
        ws[f"G{out_row}"] = f"='{SHEET_DETAIL}'!L{detail_row}"
        ws[f"H{out_row}"] = f"='{SHEET_DETAIL}'!M{detail_row}"
        ws[f"I{out_row}"] = f"='{SHEET_DETAIL}'!N{detail_row}"
        ws[f"J{out_row}"] = f"='{SHEET_DETAIL}'!O{detail_row}"
        ws[f"K{out_row}"] = f"='{SHEET_DETAIL}'!AB{detail_row}"
        style_row(ws, out_row, 1, 11, fill=FORMULA_FILL, align="left", wrap=True)
        for col in "FGHIJK":
            ws[f"{col}{out_row}"].number_format = '#,##0.00'
        out_row += 1
    set_widths(ws, {"A": 30, "B": 24, "C": 28, "D": 18, "E": 42, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14})


def write_table_sheet(ws, rows: list[dict[str, str]], title: str) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    headers = list(rows[0].keys()) if rows else ["No Data"]
    style_row(ws, 1, 1, len(headers), fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    if not rows:
        return
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=idx, value=header)
    style_row(ws, 2, 1, len(headers), fill=SUBHEADER_FILL, bold=True)
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
        style_row(ws, row_idx, 1, len(headers), fill=FORMULA_FILL, align="left", wrap=True)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(rows)+2}"
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(header) + 2, 12), 36)
        for row in rows[:100]:
            width = min(max(width, len(str(row.get(header, ""))) + 2), 40)
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_ip_fc_actuals_sheet(ws, rows: list[dict[str, str]]) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Per-IP Actual FC Bucket Amounts"
    ws["A2"] = (
        "Actual billed amounts for General Medical Management cash admissions after excluding Food & Beverage service rows "
        "and pharmacy returns. This audit layer matches workbook-facing FC buckets and includes LOS / ICU / ward breakup plus "
        "normalized room and IP-pharmacy per-day metrics."
    )
    headers = [
        ("admission_no", "Admission No"),
        ("patient_name", "Patient Name"),
        ("payor_bucket", "Payor Bucket"),
        ("patient_type", "Patient Type"),
        ("organization_name", "Organization"),
        ("surgical_medical", "Management Type"),
        ("short_stay_non_daycare_ip", "Short-Stay Non-Daycare"),
        ("room_category", "Room Category"),
        ("icu_unit_name", "ICU Unit"),
        ("los_days", "LOS Days"),
        ("effective_los_days", "Effective LOS Days"),
        ("icu_days", "ICU Days"),
        ("ward_days", "Ward Days"),
        ("service_line_count", "Service Line Count"),
        ("room_charges", "Room Charges"),
        ("room_charges_per_day", "Room Charges / Day"),
        ("investigations", "Investigations"),
        ("doctor_or_professional_charges", "Professional Fees"),
        ("bedside_services", "Bedside Services"),
        ("comparable_other_services", "Other Services"),
        ("ip_drugs_amount_net", "IP Drugs"),
        ("ip_drugs_per_day", "IP Drugs / Day"),
        ("ip_treatment_supplies_amount_net", "IP Consumables"),
        ("ip_treatment_supplies_per_day", "IP Consumables / Day"),
        ("ot_drugs_amount_net", "OT Drugs"),
        ("ot_treatment_supplies_amount_net", "OT Consumables"),
        ("implants_amount_net", "Implants"),
        ("pharmacy_total_comparable", "Pharmacy Total"),
        ("services_total_excluding_food_and_beverage", "Services Total ex F&B"),
        ("food_and_beverage_services_total_excluded", "Food & Beverage Excluded"),
        ("pharmacy_returns_total", "Pharmacy Returns Excluded"),
        ("total_bill_excluding_food_and_beverage_and_returns", "Total Amount ex F&B and Returns"),
        ("fc_bucket_sum_total", "FC Bucket Sum Total"),
        ("reconciliation_difference", "Reconciliation Difference"),
    ]
    summary_fields = [
        ("los_days", "LOS Days"),
        ("effective_los_days", "Effective LOS Days"),
        ("icu_days", "ICU Days"),
        ("ward_days", "Ward Days"),
        ("service_line_count", "Service Line Count"),
        ("room_charges", "Room Charges"),
        ("room_charges_per_day", "Room Charges / Day"),
        ("investigations", "Investigations"),
        ("doctor_or_professional_charges", "Professional Fees"),
        ("bedside_services", "Bedside Services"),
        ("comparable_other_services", "Other Services"),
        ("ip_drugs_amount_net", "IP Drugs"),
        ("ip_drugs_per_day", "IP Drugs / Day"),
        ("ip_treatment_supplies_amount_net", "IP Consumables"),
        ("ip_treatment_supplies_per_day", "IP Consumables / Day"),
        ("ot_drugs_amount_net", "OT Drugs"),
        ("ot_treatment_supplies_amount_net", "OT Consumables"),
        ("implants_amount_net", "Implants"),
        ("pharmacy_total_comparable", "Pharmacy Total"),
        ("services_total_excluding_food_and_beverage", "Services Total ex F&B"),
        ("food_and_beverage_services_total_excluded", "Food & Beverage Excluded"),
        ("pharmacy_returns_total", "Pharmacy Returns Excluded"),
        ("total_bill_excluding_food_and_beverage_and_returns", "Total Amount ex F&B and Returns"),
        ("fc_bucket_sum_total", "FC Bucket Sum Total"),
        ("reconciliation_difference", "Reconciliation Difference"),
    ]

    for idx, title in enumerate(["Field", "Min", "Max", "Average", "P25", "P50", "P75"], start=1):
        ws.cell(row=4, column=idx, value=title)
    style_row(ws, 1, 1, 32, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_row(ws, 2, 1, 32, fill=REFERENCE_FILL, align="left", wrap=True)
    style_row(ws, 4, 1, 7, fill=SUBHEADER_FILL, bold=True)

    summary_start = 5
    for offset, (field, label) in enumerate(summary_fields):
        row_num = summary_start + offset
        ws.cell(row=row_num, column=1, value=label)
        stats = summarize_numeric_series([as_float(row.get(field)) for row in rows])
        ws.cell(row=row_num, column=2, value=stats["min"])
        ws.cell(row=row_num, column=3, value=stats["max"])
        ws.cell(row=row_num, column=4, value=stats["average"])
        ws.cell(row=row_num, column=5, value=stats["p25"])
        ws.cell(row=row_num, column=6, value=stats["p50"])
        ws.cell(row=row_num, column=7, value=stats["p75"])
        style_row(ws, row_num, 1, 7, fill=FORMULA_FILL, align="left")
        style_cell(ws.cell(row=row_num, column=1), fill=SUBHEADER_FILL, bold=True, align="left")
        for col in range(2, 8):
            ws.cell(row=row_num, column=col).number_format = '#,##0.00'

    header_row = summary_start + len(summary_fields) + 2
    for idx, (_, label) in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=label)
    style_row(ws, header_row, 1, len(headers), fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")

    data_start = header_row + 1
    for row_idx, row in enumerate(rows, start=data_start):
        for col_idx, (field, _) in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
        style_row(ws, row_idx, 1, len(headers), fill=FORMULA_FILL, align="left", wrap=True)
        for col_idx in range(9, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{data_start + max(len(rows) - 1, 0)}"
    set_widths(
        ws,
        {
            "A": 16,
            "B": 28,
            "C": 18,
            "D": 18,
            "E": 24,
            "F": 18,
            "G": 18,
            "H": 14,
            "I": 10,
            "J": 10,
            "K": 10,
            "L": 16,
            "M": 14,
            "N": 16,
            "O": 16,
            "P": 18,
            "Q": 16,
            "R": 16,
            "S": 14,
            "T": 14,
            "U": 16,
            "V": 18,
            "W": 14,
            "X": 16,
            "Y": 12,
            "Z": 12,
            "AA": 16,
            "AB": 18,
            "AC": 18,
            "AD": 20,
            "AE": 18,
            "AF": 18,
        }
    )


def write_pf_review_sheet(ws, args: argparse.Namespace) -> None:
    write_professional_fees_review_sheet(
        ws,
        template_name="General Medical Management",
        estimate_behavior="PF kept review-only; no derived PF branch in estimate body.",
        payor_summary_rows=load_pf_csv_rows(args.pf_payor_summary_csv),
        shape_review=load_pf_json(args.pf_shape_review_json),
        modeled_vs_actual_rows=[],
        header_fill=HEADER_FILL,
        subheader_fill=SUBHEADER_FILL,
        formula_fill=FORMULA_FILL,
        result_fill=RESULT_FILL,
        reference_fill=REFERENCE_FILL,
    )


def build_workbook(args: argparse.Namespace) -> Path:
    pharmacy_template_rows = load_csv_rows(args.pharmacy_template)
    pharmacy_metrics_rows = [row for row in load_csv_rows(args.pharmacy_per_patient) if normalize_text(row.get("payor_bucket")).lower() == "cash"]
    ip_actual_rows = load_csv_rows(args.ip_actuals_csv)
    service_template_rows = filter_out_cath_lab_slot_rows(load_csv_rows(args.cleaned_services))
    default_service_rows = filter_out_cath_lab_slot_rows(load_csv_rows(args.default_services))
    optional_service_source_rows = filter_out_cath_lab_slot_rows(load_csv_rows(args.optional_services))
    service_line_metrics = load_json(args.service_line_count)
    room_metrics = load_json(args.room_metrics)
    rate_lookup = load_rate_lookup(args.rate_csv)
    ip_pharmacy_percentiles = load_ip_pharmacy_per_day_percentiles(args.ip_pharmacy_per_day)
    cath_lab_metrics = load_cath_lab_metrics(args.cath_lab_metrics)
    pf_payor_summary_rows = load_pf_csv_rows(args.pf_payor_summary_csv)
    room_metric_map = room_metrics.get("metrics") or {}
    los_metric = room_metric_map.get("effective_los_days") or room_metric_map.get("los_days") or {}
    baseline_metrics = {
        "los_p50": whole_day_metric(los_metric.get("p50"), minimum=1),
        "ip_drugs_per_day_p25": ip_pharmacy_percentiles.get("ip_drugs_per_day", (0.0, 0.0, 0.0))[0],
        "ip_drugs_per_day_p50": ip_pharmacy_percentiles.get("ip_drugs_per_day", (0.0, 0.0, 0.0))[1],
        "ip_drugs_per_day_p75": ip_pharmacy_percentiles.get("ip_drugs_per_day", (0.0, 0.0, 0.0))[2],
        "ip_supplies_per_day_p25": ip_pharmacy_percentiles.get("ip_treatment_supplies_per_day", (0.0, 0.0, 0.0))[0],
        "ip_supplies_per_day_p50": ip_pharmacy_percentiles.get("ip_treatment_supplies_per_day", (0.0, 0.0, 0.0))[1],
        "ip_supplies_per_day_p75": ip_pharmacy_percentiles.get("ip_treatment_supplies_per_day", (0.0, 0.0, 0.0))[2],
    }

    core_service_rows = build_core_service_rows(default_service_rows, service_template_rows)
    optional_service_rows = build_optional_service_rows(optional_service_source_rows, rate_lookup)
    grouping_rows = build_grouping_candidates(service_template_rows, core_service_rows, optional_service_rows, rate_lookup)
    base_service_count = 10 + len(core_service_rows)

    workbook = Workbook()
    default_ws = workbook.active
    workbook.remove(default_ws)
    builder_ws = workbook.create_sheet(SHEET_BUILDER)
    summary_ws = workbook.create_sheet(SHEET_SUMMARY)
    advanced_ws = workbook.create_sheet(SHEET_ADVANCED)
    service_addons_ws = workbook.create_sheet(SHEET_SERVICE_ADDONS)
    grouped_adjustments_ws = workbook.create_sheet(SHEET_GROUPED_ADJUSTMENTS)
    grouping_review_ws = workbook.create_sheet(SHEET_GROUPING_REVIEW)
    breakdown_ws = workbook.create_sheet(SHEET_BREAKDOWN)
    detail_ws = workbook.create_sheet(SHEET_DETAIL)
    pharmacy_template_ws = workbook.create_sheet(SHEET_PHARMACY_TEMPLATE)
    service_template_ws = workbook.create_sheet(SHEET_SERVICE_TEMPLATE)
    pharmacy_metrics_ws = workbook.create_sheet(SHEET_PHARMACY_METRICS)
    ip_actuals_ws = workbook.create_sheet(SHEET_IP_ACTUALS)
    pf_review_ws = workbook.create_sheet(SHEET_PF_REVIEW)
    reference_ws = workbook.create_sheet(SHEET_REFERENCE)

    refs = write_reference_sheet(reference_ws, rate_lookup, ip_pharmacy_percentiles, room_metrics, service_line_metrics, cath_lab_metrics, grouping_rows, load_resolution_rows(args.payer_basis_resolution_csv), pf_payor_summary_rows)
    builder_refs = write_builder_sheet(builder_ws, refs)
    builder_refs.update(refs)
    control_refs = write_advanced_controls_sheet(
        advanced_ws,
        pharmacy_template_rows,
        builder_refs,
        baseline_metrics,
        base_service_count,
    )
    advanced_rows = write_service_addons_sheet(service_addons_ws, optional_service_rows)
    grouped_rows = write_grouped_adjustments_sheet(grouped_adjustments_ws, grouping_rows)
    write_grouping_review_sheet(grouping_review_ws, grouping_rows)
    builder_refs.update(control_refs)
    detail_total_row, detail_count_row = write_detail_sheet(detail_ws, builder_refs, core_service_rows, advanced_rows, grouped_rows)
    write_summary_sheet(summary_ws, detail_total_row, detail_count_row, refs, pf_payor_summary_rows)
    write_breakdown_sheet(breakdown_ws, detail_total_row)
    write_table_sheet(pharmacy_template_ws, pharmacy_template_rows, "Cash General Medical Management Pharmacy Template")
    write_table_sheet(service_template_ws, service_template_rows, "Cash General Medical Management Service Template")
    write_table_sheet(pharmacy_metrics_ws, pharmacy_metrics_rows, "Cash General Medical Management Pharmacy Metrics")
    write_ip_fc_actuals_sheet(ip_actuals_ws, ip_actual_rows)
    write_pf_review_sheet(pf_review_ws, args)

    for ws in workbook.worksheets:
        ws.sheet_view.showGridLines = False

    apply_default_calc_settings(workbook)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    recalculate_workbook_with_soffice(args.output)
    return args.output


def main() -> None:
    args = parse_args()
    output = build_workbook(args)
    print(f"output={output}")


if __name__ == "__main__":
    main()
