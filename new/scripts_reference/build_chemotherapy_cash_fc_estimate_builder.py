from __future__ import annotations

import argparse
import csv
import json
import math
from dataclasses import dataclass
from pathlib import Path
from statistics import quantiles
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from fc_slot_family import is_cath_lab_slot_service
from fc_payer_basis_resolution import (
    AUTO_BASIS,
    PAYER_BASIS_OPTIONS,
    load_resolution_rows,
    selection_lookup_formula,
    supported_basis_options_from_resolution_rows,
)
from professional_fee_review_workbook import (
    PF_PAYOR_ORDER,
    build_pf_summary_lookup,
    get_pf_summary_row,
    load_csv_rows as load_pf_csv_rows,
    load_json as load_pf_json,
    write_professional_fees_review_sheet,
)
from workbook_postprocess import apply_default_calc_settings, recalculate_workbook_with_soffice


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = REPO_ROOT / "output/chemotherapy_systemic_therapy_infusion_fc"
DEFAULT_OUTPUT = REPO_ROOT / "output/fc_estimate_builder_chemotherapy_systemic_therapy_infusion_cash_tr1.xlsx"
DEFAULT_PHARMACY_TEMPLATE = DEFAULT_OUTPUT_DIR / "01_clean_chemo_pharmacy_template_cash.csv"
DEFAULT_PHARMACY_PER_PATIENT = DEFAULT_OUTPUT_DIR / "02_per_patient_pharmacy_bucket_totals_cash.csv"
DEFAULT_PHARMACY_BUCKETS = DEFAULT_OUTPUT_DIR / "03_bucket_percentile_summary_cash.csv"
DEFAULT_SERVICES_TEMPLATE = DEFAULT_OUTPUT_DIR / "10_clean_chemo_services_template_cash.csv"
DEFAULT_CLEANED_SERVICES = DEFAULT_OUTPUT_DIR / "11_clean_chemo_services_template_for_fc_cash.csv"
DEFAULT_OPTIONAL_SERVICES = DEFAULT_OUTPUT_DIR / "13_optional_service_add_ons_cash.csv"
DEFAULT_SERVICE_LINE_COUNT = DEFAULT_OUTPUT_DIR / "14_service_line_count_metrics_cash.json"
DEFAULT_ROLLUP = DEFAULT_OUTPUT_DIR / "23_all_cash_patients_fc_bucket_rollup_reconciled.csv"
DEFAULT_RATE_CSV = DEFAULT_OUTPUT_DIR / "tr1_cash_rates_chemotherapy_systemic_therapy_infusion_full_codes.csv"
DEFAULT_CATH_LAB_METRICS = DEFAULT_OUTPUT_DIR / "24_cath_lab_metrics_cash.json"
DEFAULT_PAYER_BASIS_RESOLUTION = DEFAULT_OUTPUT_DIR / "30_payer_basis_resolution_summary.csv"

SHEET_BUILDER = "Builder"
SHEET_SUMMARY = "Estimate Summary"
SHEET_ADVANCED = "Advanced Controls"
SHEET_SERVICE_ADDONS = "Service Add-Ons"
SHEET_GROUPED_ADJUSTMENTS = "Grouped Adjustments"
SHEET_GROUPING_REVIEW = "Grouping Review"
SHEET_BREAKDOWN = "Estimate Breakdown"
SHEET_DETAIL = "Line Item Detail"
SHEET_PHARMACY_TEMPLATE = "Pharmacy Template"
SHEET_SERVICE_TEMPLATE = "Service Template"
SHEET_PHARMACY_METRICS = "Pharmacy Metrics"
SHEET_PF_REVIEW = "Professional Fees Review"
SHEET_REFERENCE = "Reference"

THIN_GREY = Side(style="thin", color="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="EAF4EA")
RESULT_FILL = PatternFill("solid", fgColor="FFF2CC")
REFERENCE_FILL = PatternFill("solid", fgColor="F4F6F8")
SPACER_FILL = PatternFill("solid", fgColor="F7F7F7")

MODE_LOW = "Low"
MODE_TYPICAL = "Typical"
MODE_HIGH = "High"
INCLUDE = "Include"
EXCLUDE = "Exclude"
GROUP_PRESENCE_AUTO = 90.0
GROUP_PRESENCE_OPTIONAL = 75.0

LOGIC_CODES = {
    "ROM0010",
    "ROM5189",
    "ROM0093",
    "HSP5013",
    "MSC10",
    "RNS0120",
}

OPTIONAL_LOGIC_ROWS = [
    ("Nursing Charges", "Room Charges", "Nursing Charges", "ROM5189", EXCLUDE, "Optional logic row; include when a non-daycare nursing pattern is expected."),
    ("DMO Charges", "Room Charges", "DMO Charges", "ROM0093", EXCLUDE, "Optional logic row; include when medical officer / DMO billing is expected."),
]


@dataclass
class RateRow:
    item_name: str
    general: float | None
    twin: float | None
    single: float | None
    icu: float | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the chemo cash/TR1 FC estimate workbook.")
    parser.add_argument("--pharmacy-template", type=Path, default=DEFAULT_PHARMACY_TEMPLATE)
    parser.add_argument("--pharmacy-per-patient", type=Path, default=DEFAULT_PHARMACY_PER_PATIENT)
    parser.add_argument("--pharmacy-buckets", type=Path, default=DEFAULT_PHARMACY_BUCKETS)
    parser.add_argument("--services-template", type=Path, default=DEFAULT_SERVICES_TEMPLATE)
    parser.add_argument("--cleaned-services", type=Path, default=DEFAULT_CLEANED_SERVICES)
    parser.add_argument("--optional-services", type=Path, default=DEFAULT_OPTIONAL_SERVICES)
    parser.add_argument("--service-line-count", type=Path, default=DEFAULT_SERVICE_LINE_COUNT)
    parser.add_argument("--rollup", type=Path, default=DEFAULT_ROLLUP)
    parser.add_argument("--rate-csv", type=Path, default=DEFAULT_RATE_CSV)
    parser.add_argument("--cath-lab-metrics", type=Path, default=DEFAULT_CATH_LAB_METRICS)
    parser.add_argument("--payer-basis-resolution-csv", type=Path, default=DEFAULT_PAYER_BASIS_RESOLUTION)
    parser.add_argument("--pf-payor-summary-csv", type=Path)
    parser.add_argument("--pf-shape-review-json", type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_code(value: Any) -> str:
    return normalize_text(value).replace(" ", "").upper()


def maybe_float(value: Any) -> float | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def as_float(value: Any) -> float:
    number = maybe_float(value)
    return 0.0 if number is None else float(number)


def load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def load_cath_lab_metrics(path: Path) -> tuple[float, float, float]:
    payload = load_json(path)
    metrics = payload.get("metrics") or {}
    return (
        as_float(metrics.get("p25")),
        as_float(metrics.get("p50")),
        as_float(metrics.get("p75")),
    )


def filter_out_cath_lab_slot_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [
        row
        for row in rows
        if not is_cath_lab_slot_service(
            code=row.get("item_code"),
            grouping=row.get("grouping"),
            service_name=row.get("item_name"),
        )
    ]


def load_rate_lookup(path: Path) -> dict[str, RateRow]:
    lookup: dict[str, RateRow] = {}
    for row in load_csv_rows(path):
        code = normalize_code(row.get("item_code"))
        if not code:
            continue
        lookup[code] = RateRow(
            item_name=normalize_text(row.get("item_name")),
            general=maybe_float(row.get("general")),
            twin=maybe_float(row.get("twin")),
            single=maybe_float(row.get("single")),
            icu=maybe_float(row.get("icu")),
        )
    return lookup


def preferred_rate(rate_row: RateRow | None) -> float:
    if not rate_row:
        return 0.0
    for value in (rate_row.general, rate_row.twin, rate_row.single, rate_row.icu):
        if value is not None:
            return float(value)
    return 0.0


def style_cell(cell, *, fill=None, bold=False, font_color="000000", align="center", wrap=False):
    if fill is not None:
        cell.fill = fill
    cell.font = Font(bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)


def style_row(ws, row_number: int, start_col: int, end_col: int, *, fill=None, bold=False, font_color="000000", align="center", wrap=False):
    for col in range(start_col, end_col + 1):
        style_cell(ws.cell(row=row_number, column=col), fill=fill, bold=bold, font_color=font_color, align=align, wrap=wrap)


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def compute_quartiles(values: list[float]) -> tuple[float, float, float]:
    cleaned = sorted(float(v) for v in values)
    if not cleaned:
        return 0.0, 0.0, 0.0
    if len(cleaned) == 1:
        return cleaned[0], cleaned[0], cleaned[0]
    q1, q2, q3 = quantiles(cleaned, n=4, method="inclusive")
    return float(q1), float(q2), float(q3)


def mode_pick_formula(mode_ref: str, low_ref: str, typical_ref: str, high_ref: str) -> str:
    return f'=IF({mode_ref}="{MODE_LOW}",{low_ref},IF({mode_ref}="{MODE_HIGH}",{high_ref},{typical_ref}))'


def yes_no_pick_formula(choice_ref: str, include_ref: str, exclude_ref: str) -> str:
    return f'=IF({choice_ref}="{INCLUDE}",{include_ref},{exclude_ref})'


def load_bucket_percentiles(path: Path) -> dict[str, tuple[float, float, float]]:
    mapping: dict[str, tuple[float, float, float]] = {}
    for row in load_csv_rows(path):
        mapping[normalize_text(row.get("metric"))] = (
            as_float(row.get("p25")),
            as_float(row.get("p50")),
            as_float(row.get("p75")),
        )
    return mapping


def load_rollup_rows(path: Path) -> list[dict[str, str]]:
    return load_csv_rows(path)


def compute_rollup_bucket_quartiles(rows: list[dict[str, str]], field: str) -> tuple[float, float, float]:
    return compute_quartiles([as_float(row.get(field)) for row in rows])


def dedupe_service_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    deduped: dict[str, dict[str, str]] = {}
    for row in rows:
        code = normalize_code(row.get("item_code"))
        key = code or normalize_text(row.get("canonical_item_key")) or normalize_text(row.get("item_name"))
        if not key:
            continue
        current = deduped.get(key)
        if current is None:
            deduped[key] = row
            continue
        current_case_count = as_float(current.get("case_count"))
        new_case_count = as_float(row.get("case_count"))
        current_rate_count = as_float(current.get("rate_cash_count"))
        new_rate_count = as_float(row.get("rate_cash_count"))
        if (new_case_count, new_rate_count, normalize_text(row.get("item_name"))) > (
            current_case_count,
            current_rate_count,
            normalize_text(current.get("item_name")),
        ):
            deduped[key] = row
    return sorted(
        deduped.values(),
        key=lambda row: (
            -as_float(row.get("case_count")),
            -as_float(row.get("amount_cash_typical")),
            normalize_text(row.get("item_name")),
        ),
    )


def build_optional_service_rows(
    rows: list[dict[str, str]],
    rate_lookup: dict[str, RateRow],
    max_rows: int = 40,
) -> list[dict[str, Any]]:
    filtered: list[dict[str, Any]] = []
    for row in dedupe_service_rows(rows):
        code = normalize_code(row.get("item_code"))
        if code in LOGIC_CODES:
            continue
        if is_cath_lab_slot_service(code=code, grouping=row.get("grouping"), service_name=row.get("item_name")):
            continue
        rate_row = rate_lookup.get(code)
        tariff_rate = preferred_rate(rate_row)
        if tariff_rate <= 0:
            tariff_rate = as_float(row.get("rate_cash_p50"))
        if tariff_rate <= 0:
            continue
        qty_p25 = as_float(row.get("quantity_p25")) or 1.0
        qty_p50 = as_float(row.get("quantity_p50")) or qty_p25
        qty_p75 = as_float(row.get("quantity_p75")) or qty_p50
        filtered.append(
            {
                "item_code": code,
                "item_name": normalize_text(row.get("item_name")),
                "fc_bucket": normalize_text(row.get("fc_estimate_bucket")),
                "grouping": normalize_text(row.get("grouping")),
                "presence_rate": as_float(row.get("case_presence_rate")),
                "qty_p25": qty_p25,
                "qty_p50": qty_p50,
                "qty_p75": qty_p75,
                "rate": tariff_rate,
            }
        )
    return filtered[:max_rows]


def classify_group_residual_band(group_presence_rate: float, residual_p50: float) -> str:
    if residual_p50 <= 0.01:
        return ""
    if group_presence_rate > GROUP_PRESENCE_AUTO:
        return "auto"
    if group_presence_rate >= GROUP_PRESENCE_OPTIONAL:
        return "optional"
    return ""


def build_grouping_rows(
    cleaned_rows: list[dict[str, str]],
    optional_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    optional_by_code = {row["item_code"]: row for row in optional_rows}
    grouped: dict[str, dict[str, Any]] = {}
    for row in dedupe_service_rows(cleaned_rows):
        code = normalize_code(row.get("item_code"))
        if not code or code in LOGIC_CODES or code not in optional_by_code:
            continue
        if is_cath_lab_slot_service(code=code, grouping=row.get("grouping"), service_name=row.get("item_name")):
            continue
        optional = optional_by_code[code]
        grouping = optional["grouping"] or "Ungrouped"
        state = grouped.setdefault(
            grouping,
            {
                "grouping": grouping,
                "fc_bucket": normalize_text(optional["fc_bucket"]),
                "group_presence_rate": 0.0,
                "group_amount_p25_exact": 0.0,
                "group_amount_p50_exact": 0.0,
                "group_amount_p75_exact": 0.0,
                "group_amount_captured_by_default_rows": 0.0,
            },
        )
        state["group_presence_rate"] = max(state["group_presence_rate"], as_float(row.get("case_presence_rate")))
        state["group_amount_p25_exact"] += optional["qty_p25"] * optional["rate"]
        state["group_amount_p50_exact"] += optional["qty_p50"] * optional["rate"]
        state["group_amount_p75_exact"] += optional["qty_p75"] * optional["rate"]
    rows: list[dict[str, Any]] = []
    for grouping, state in sorted(grouped.items()):
        residual_p50 = max(0.0, state["group_amount_p50_exact"] - state["group_amount_captured_by_default_rows"])
        band = classify_group_residual_band(state["group_presence_rate"], residual_p50)
        if not band:
            continue
        state["group_residual_band"] = band
        rows.append(state)
    return rows


def write_reference_sheet(
    ws,
    pharmacy_template_rows: list[dict[str, str]],
    service_template_rows: list[dict[str, str]],
    service_line_metrics: dict[str, Any],
    rollup_rows: list[dict[str, str]],
    bucket_percentiles: dict[str, tuple[float, float, float]],
    grouping_rows: list[dict[str, Any]],
    payer_basis_resolution_rows: list[dict[str, str]],
    pf_payor_summary_rows: list[dict[str, str]],
) -> dict[str, str]:
    supported_basis_options = [AUTO_BASIS, *supported_basis_options_from_resolution_rows(payer_basis_resolution_rows)]
    ws.sheet_state = "hidden"
    ws["A1"] = "Mode"
    ws["A2"] = MODE_LOW
    ws["A3"] = MODE_TYPICAL
    ws["A4"] = MODE_HIGH
    ws["B1"] = "Selection"
    ws["B2"] = INCLUDE
    ws["B3"] = EXCLUDE

    bucket_row_start = 2
    ws["D1"] = "metric"
    ws["E1"] = "p25"
    ws["F1"] = "p50"
    ws["G1"] = "p75"
    service_metric_refs: dict[str, str] = {}
    for idx, metric in enumerate(
        [
            "room_charges",
            "doctor_or_professional_charges",
            "other_services",
            "bedside_services",
            "procedure_charges",
            "investigations",
            "physiotherapy",
            "emergency",
            "cath_lab",
            "ot",
        ],
        start=bucket_row_start,
    ):
        q1, q2, q3 = compute_rollup_bucket_quartiles(rollup_rows, metric)
        ws[f"D{idx}"] = metric
        ws[f"E{idx}"] = q1
        ws[f"F{idx}"] = q2
        ws[f"G{idx}"] = q3
        service_metric_refs[metric] = f"Reference!F{idx}"

    pharmacy_ref_start = bucket_row_start + 12
    ws[f"D{pharmacy_ref_start-1}"] = "pharmacy_metric"
    ws[f"E{pharmacy_ref_start-1}"] = "p25"
    ws[f"F{pharmacy_ref_start-1}"] = "p50"
    ws[f"G{pharmacy_ref_start-1}"] = "p75"
    pharmacy_metric_refs: dict[str, tuple[str, str, str]] = {}
    for idx, metric in enumerate(
        [
            "ip_drugs_amount_net",
            "ip_treatment_supplies_amount_net",
            "ot_drugs_amount_net",
            "ot_treatment_supplies_amount_net",
            "implants_amount_net",
        ],
        start=pharmacy_ref_start,
    ):
        q1, q2, q3 = bucket_percentiles.get(metric, (0.0, 0.0, 0.0))
        ws[f"D{idx}"] = metric
        ws[f"E{idx}"] = q1
        ws[f"F{idx}"] = q2
        ws[f"G{idx}"] = q3
        pharmacy_metric_refs[metric] = (f"Reference!E{idx}", f"Reference!F{idx}", f"Reference!G{idx}")

    service_count = service_line_metrics.get("cleaned_distinct_service_line_count") or {}
    ws["J1"] = "service_line_metric"
    ws["J2"] = "p25"
    ws["J3"] = "p50"
    ws["J4"] = "p75"
    ws["K2"] = as_float(service_count.get("p25"))
    ws["K3"] = as_float(service_count.get("p50"))
    ws["K4"] = as_float(service_count.get("p75"))

    ws["M1"] = "pharmacy_template_rows"
    ws["M2"] = len(pharmacy_template_rows)
    ws["N1"] = "service_template_rows"
    ws["N2"] = len(service_template_rows)
    ws["P1"] = "grouping"
    ws["Q1"] = "fc_bucket"
    ws["R1"] = "presence_rate"
    ws["S1"] = "group_amount_p25_exact"
    ws["T1"] = "group_amount_p50_exact"
    ws["U1"] = "group_amount_p75_exact"
    ws["V1"] = "captured_by_default"
    ws["W1"] = "group_residual_band"
    for row_idx, row in enumerate(grouping_rows, start=2):
        ws[f"P{row_idx}"] = row["grouping"]
        ws[f"Q{row_idx}"] = row["fc_bucket"]
        ws[f"R{row_idx}"] = row["group_presence_rate"] / 100.0
        ws[f"S{row_idx}"] = row["group_amount_p25_exact"]
        ws[f"T{row_idx}"] = row["group_amount_p50_exact"]
        ws[f"U{row_idx}"] = row["group_amount_p75_exact"]
        ws[f"V{row_idx}"] = row["group_amount_captured_by_default_rows"]
        ws[f"W{row_idx}"] = row["group_residual_band"]
    ws["Y1"] = "payer_basis_option"
    for row_idx, label in enumerate(supported_basis_options, start=2):
        ws[f"Y{row_idx}"] = label
    ws["AA1"] = "component"
    ws["AB1"] = "target_payor_bucket"
    ws["AC1"] = "selected_basis"
    ws["AD1"] = "selected_case_count"
    ws["AE1"] = "selection_reason"
    for row_idx, row in enumerate(payer_basis_resolution_rows, start=2):
        ws[f"AA{row_idx}"] = normalize_text(row.get("component"))
        ws[f"AB{row_idx}"] = normalize_text(row.get("target_payor_bucket"))
        ws[f"AC{row_idx}"] = normalize_text(row.get("selected_basis"))
        ws[f"AD{row_idx}"] = as_float(row.get("selected_case_count"))
        ws[f"AE{row_idx}"] = normalize_text(row.get("selection_reason"))
    ws["AG1"] = "payor_bucket"
    ws["AH1"] = "admission_count"
    ws["AI1"] = "pf_collectible_historical_total_p25"
    ws["AJ1"] = "pf_collectible_historical_total_p50"
    ws["AK1"] = "pf_collectible_historical_total_p75"
    ws["AL1"] = "pf_named_total_p50"
    ws["AM1"] = "pf_general_needed_total_p50"
    ws["AN1"] = "surgeon_named_total_p50"
    ws["AO1"] = "assistant_surgeon_named_total_p50"
    ws["AP1"] = "anesthetist_named_total_p50"
    ws["AQ1"] = "assistant_anesthetist_named_total_p50"
    ws["AR1"] = "consultant_or_physician_named_total_p50"
    ws["AS1"] = "dominant_pf_shape"
    for row_idx, row in enumerate(pf_payor_summary_rows, start=2):
        ws[f"AG{row_idx}"] = normalize_text(row.get("payor_bucket"))
        ws[f"AH{row_idx}"] = as_float(row.get("admission_count"))
        ws[f"AI{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p25"))
        ws[f"AJ{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p50"))
        ws[f"AK{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p75"))
        ws[f"AL{row_idx}"] = as_float(row.get("pf_named_total_p50"))
        ws[f"AM{row_idx}"] = as_float(row.get("pf_general_needed_total_p50"))
        ws[f"AN{row_idx}"] = as_float(row.get("surgeon_named_total_p50"))
        ws[f"AO{row_idx}"] = as_float(row.get("assistant_surgeon_named_total_p50"))
        ws[f"AP{row_idx}"] = as_float(row.get("anesthetist_named_total_p50"))
        ws[f"AQ{row_idx}"] = as_float(row.get("assistant_anesthetist_named_total_p50"))
        ws[f"AR{row_idx}"] = as_float(row.get("consultant_or_physician_named_total_p50"))
        ws[f"AS{row_idx}"] = normalize_text(row.get("dominant_pf_shape"))
    return {
        "service_line_p25": "Reference!K2",
        "service_line_p50": "Reference!K3",
        "service_line_p75": "Reference!K4",
        "room_p25": "Reference!E2",
        "room_typical": service_metric_refs["room_charges"],
        "room_p75": "Reference!G2",
        "doctor_p25": "Reference!E3",
        "doctor_typical": service_metric_refs["doctor_or_professional_charges"],
        "doctor_p75": "Reference!G3",
        "other_p25": "Reference!E4",
        "other_typical": service_metric_refs["other_services"],
        "other_p75": "Reference!G4",
        "bedside_p25": "Reference!E5",
        "bedside_typical": service_metric_refs["bedside_services"],
        "bedside_p75": "Reference!G5",
        "procedure_p25": "Reference!E6",
        "procedure_typical": service_metric_refs["procedure_charges"],
        "procedure_p75": "Reference!G6",
        "investigations_p25": "Reference!E7",
        "investigations_typical": service_metric_refs["investigations"],
        "investigations_p75": "Reference!G7",
        "physio_p25": "Reference!E8",
        "physio_typical": service_metric_refs["physiotherapy"],
        "physio_p75": "Reference!G8",
        "emergency_p25": "Reference!E9",
        "emergency_typical": service_metric_refs["emergency"],
        "emergency_p75": "Reference!G9",
        "cath_typical": service_metric_refs["cath_lab"],
        "ot_typical": service_metric_refs["ot"],
        "cath_p25": "Reference!E10",
        "cath_p50": "Reference!F10",
        "cath_p75": "Reference!G10",
        "ot_p25": "Reference!E11",
        "ot_p75": "Reference!G11",
        "ip_drugs_p25": pharmacy_metric_refs["ip_drugs_amount_net"][0],
        "ip_drugs_p50": pharmacy_metric_refs["ip_drugs_amount_net"][1],
        "ip_drugs_p75": pharmacy_metric_refs["ip_drugs_amount_net"][2],
        "ip_supplies_p25": pharmacy_metric_refs["ip_treatment_supplies_amount_net"][0],
        "ip_supplies_p50": pharmacy_metric_refs["ip_treatment_supplies_amount_net"][1],
        "ip_supplies_p75": pharmacy_metric_refs["ip_treatment_supplies_amount_net"][2],
        "ot_drugs_p25": pharmacy_metric_refs["ot_drugs_amount_net"][0],
        "ot_drugs_p50": pharmacy_metric_refs["ot_drugs_amount_net"][1],
        "ot_drugs_p75": pharmacy_metric_refs["ot_drugs_amount_net"][2],
        "ot_supplies_p25": pharmacy_metric_refs["ot_treatment_supplies_amount_net"][0],
        "ot_supplies_p50": pharmacy_metric_refs["ot_treatment_supplies_amount_net"][1],
        "ot_supplies_p75": pharmacy_metric_refs["ot_treatment_supplies_amount_net"][2],
        "implants_p25": pharmacy_metric_refs["implants_amount_net"][0],
        "implants_p50": pharmacy_metric_refs["implants_amount_net"][1],
        "implants_p75": pharmacy_metric_refs["implants_amount_net"][2],
        "payer_basis_options": f"Reference!Y2:Y{len(supported_basis_options) + 1}",
    }


def write_builder_sheet(ws, refs: dict[str, str]) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.merge_cells("A1:D1")
    ws["A1"] = "FC Estimate Builder - Chemotherapy / Systemic Therapy Infusion (Non-Daycare)"
    style_row(ws, 1, 1, 4, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")

    rows = [
        ("Template", "Chemotherapy / Systemic Therapy Infusion (Non-Daycare)"),
        ("Payor", "Cash (TR1)"),
        ("Historical Payer Basis", AUTO_BASIS),
        ("Estimate Mode", MODE_TYPICAL),
        ("Stay Pattern", "Non-Daycare"),
        ("Optional Services", "Use Advanced Controls"),
        ("Notes", "Default estimate is built from the non-daycare medical cohort using historical non-pharmacy bucket anchors with optional service and grouped-adjustment controls."),
    ]
    start_row = 3
    for idx, (label, value) in enumerate(rows, start=start_row):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = value
        style_cell(ws[f"A{idx}"], fill=SUBHEADER_FILL, bold=True, align="left")
        style_cell(ws[f"B{idx}"], fill=INPUT_FILL if label in {"Estimate Mode", "Historical Payer Basis"} else FORMULA_FILL, align="left", wrap=True)

    mode_validation = DataValidation(type="list", formula1=f"='{SHEET_REFERENCE}'!$A$2:$A$4", allow_blank=False)
    payer_basis_validation = DataValidation(type="list", formula1=f"='{SHEET_REFERENCE}'!$Y$2:INDEX('{SHEET_REFERENCE}'!$Y:$Y,COUNTA('{SHEET_REFERENCE}'!$Y:$Y))", allow_blank=False)
    ws.add_data_validation(mode_validation)
    ws.add_data_validation(payer_basis_validation)
    payer_basis_validation.add(ws["B5"])
    mode_validation.add(ws["B6"])

    service_resolution_formula = selection_lookup_formula('"service_basis"', '"Cash"', "AC", component_col="AA", target_payor_col="AB").lstrip("=")
    pharmacy_resolution_formula = selection_lookup_formula('"pharmacy_basis"', '"Cash"', "AC", component_col="AA", target_payor_col="AB").lstrip("=")
    pf_resolution_formula = selection_lookup_formula('"pf_basis"', '"Cash"', "AC", component_col="AA", target_payor_col="AB").lstrip("=")
    resolver_note_formula = selection_lookup_formula('"service_basis"', '"Cash"', "AE", component_col="AA", target_payor_col="AB").lstrip("=")
    ws["D3"] = "Resolved Service Basis"
    ws["E3"] = f'=IF(B5<>"{AUTO_BASIS}",B5,{service_resolution_formula})'
    ws["D4"] = "Resolved Pharmacy Basis"
    ws["E4"] = f'=IF(B5<>"{AUTO_BASIS}",B5,{pharmacy_resolution_formula})'
    ws["D5"] = "Resolved PF Basis"
    ws["E5"] = f'=IF(B5<>"{AUTO_BASIS}",B5,{pf_resolution_formula})'
    ws["D6"] = "Resolver Note"
    ws["E6"] = f'=IF(B5<>"{AUTO_BASIS}","Manual historical payer basis override applied",{resolver_note_formula})'
    for row in range(3, 7):
        style_cell(ws[f"D{row}"], fill=SUBHEADER_FILL, bold=True, align="left")
        style_cell(ws[f"E{row}"], fill=REFERENCE_FILL if row == 6 else FORMULA_FILL, align="left", wrap=(row == 6))

    ws["A11"] = "Historical Service Line Count"
    ws["B11"] = "P25"
    ws["C11"] = "P50"
    ws["D11"] = "P75"
    style_row(ws, 11, 1, 4, fill=SUBHEADER_FILL, bold=True)
    ws["B12"] = f"={refs['service_line_p25']}"
    ws["C12"] = f"={refs['service_line_p50']}"
    ws["D12"] = f"={refs['service_line_p75']}"
    style_row(ws, 12, 2, 4, fill=FORMULA_FILL)

    ws["A14"] = "Selected Headline Estimate"
    ws["B14"] = f"='{SHEET_SUMMARY}'!B8"
    style_cell(ws["A14"], fill=SUBHEADER_FILL, bold=True, align="left")
    style_cell(ws["B14"], fill=RESULT_FILL, bold=True, align="left")
    ws["B14"].number_format = '#,##0.00'

    set_widths(ws, {"A": 26, "B": 38, "C": 14, "D": 14})


def group_metric_formula(metric_col: str, grouping_ref: str) -> str:
    return f'=IFERROR(INDEX({SHEET_REFERENCE}!${metric_col}:${metric_col},MATCH({grouping_ref},{SHEET_REFERENCE}!$P:$P,0)),0)'


def write_advanced_controls_sheet(ws, rate_lookup: dict[str, RateRow]) -> dict[str, Any]:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws["A1"] = "Advanced Controls"
    ws["A2"] = "Use Service Add-Ons for exact optional service rows and Grouped Adjustments for common-case grouped completion. Grouped residuals shrink automatically when add-ons from the same grouping are included."
    style_row(ws, 1, 1, 10, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_row(ws, 2, 1, 10, fill=REFERENCE_FILL, align="left", wrap=True)

    ws["A4"] = "Optional Logic Rows"
    style_row(ws, 4, 1, 6, fill=SUBHEADER_FILL, bold=True, align="left")
    headers = ["Line Item", "Item Code", "Default", "Selection", "Tariff Rate", "Notes"]
    for col_idx, value in enumerate(headers, start=1):
        ws.cell(row=5, column=col_idx, value=value)
    style_row(ws, 5, 1, 6, fill=REFERENCE_FILL, bold=True)

    dropdown = DataValidation(type="list", formula1=f"='{SHEET_REFERENCE}'!$B$2:$B$3", allow_blank=False)
    ws.add_data_validation(dropdown)

    logic_selection_cells: dict[str, str] = {}
    for offset, (name, _parent, _sub, code, default, notes) in enumerate(OPTIONAL_LOGIC_ROWS, start=6):
        ws[f"A{offset}"] = name
        ws[f"B{offset}"] = code
        ws[f"C{offset}"] = default
        ws[f"D{offset}"] = default
        dropdown.add(ws[f"D{offset}"])
        ws[f"E{offset}"] = preferred_rate(rate_lookup.get(code))
        ws[f"F{offset}"] = notes
        style_row(ws, offset, 1, 6, fill=INPUT_FILL if offset in {5, 6} else FORMULA_FILL, align="left", wrap=True)
        ws[f"E{offset}"].number_format = '#,##0.00'
        logic_selection_cells[code] = f"'{SHEET_ADVANCED}'!D{offset}"

    set_widths(
        ws,
        {
            "A": 12,
            "B": 12,
            "C": 34,
            "D": 24,
            "E": 28,
            "F": 11,
            "G": 10,
            "H": 10,
            "I": 10,
            "J": 12,
            "K": 14,
        },
    )
    return {
        "logic_selection_cells": logic_selection_cells,
    }


def write_service_addons_sheet(ws, optional_service_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws["A1"] = "Service Add-Ons"
    ws["A2"] = "Use this sheet for exact optional chemo service rows. If you include an item here, the grouped residual for the same grouping shrinks automatically."
    style_row(ws, 1, 1, 11, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_row(ws, 2, 1, 11, fill=REFERENCE_FILL, align="left", wrap=True)
    add_headers = [
        "Select",
        "Item Code",
        "Item Name",
        "FC Bucket",
        "Grouping",
        "Presence %",
        "Qty P25",
        "Qty P50",
        "Qty P75",
        "Tariff Rate",
        "Typical Amount",
    ]
    for col_idx, value in enumerate(add_headers, start=1):
        ws.cell(row=3, column=col_idx, value=value)
    style_row(ws, 3, 1, 11, fill=REFERENCE_FILL, bold=True)
    dropdown = DataValidation(type="list", formula1=f"='{SHEET_REFERENCE}'!$B$2:$B$3", allow_blank=False)
    ws.add_data_validation(dropdown)
    service_selection_rows: list[dict[str, Any]] = []
    current_row = 4
    for row in optional_service_rows:
        ws[f"A{current_row}"] = EXCLUDE
        dropdown.add(ws[f"A{current_row}"])
        ws[f"B{current_row}"] = row["item_code"]
        ws[f"C{current_row}"] = row["item_name"]
        ws[f"D{current_row}"] = row["fc_bucket"]
        ws[f"E{current_row}"] = row["grouping"]
        ws[f"F{current_row}"] = row["presence_rate"] / 100.0
        ws[f"G{current_row}"] = row["qty_p25"]
        ws[f"H{current_row}"] = row["qty_p50"]
        ws[f"I{current_row}"] = row["qty_p75"]
        ws[f"J{current_row}"] = row["rate"]
        ws[f"K{current_row}"] = row["qty_p50"] * row["rate"]
        style_row(ws, current_row, 1, 11, fill=FORMULA_FILL, align="left", wrap=True)
        ws[f"F{current_row}"].number_format = '0.00%'
        for col in "GHIJK":
            ws[f"{col}{current_row}"].number_format = '#,##0.00'
        service_selection_rows.append({"source_row": current_row, **row})
        current_row += 1
    set_widths(ws, {"A": 12, "B": 12, "C": 34, "D": 24, "E": 28, "F": 11, "G": 10, "H": 10, "I": 10, "J": 12, "K": 14})
    return service_selection_rows


def write_grouping_review_sheet(ws, grouping_rows: list[dict[str, Any]]) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Grouping Review"
    ws["A2"] = "This audit sheet highlights only high-presence chemo service groups where the default estimate does not fully cover the common-case grouped amount."
    style_row(ws, 1, 1, 8, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_row(ws, 2, 1, 8, fill=REFERENCE_FILL, align="left", wrap=True)
    headers = ["Grouping", "FC Bucket", "Presence Rate", "Group Amount P25 Exact", "Group Amount P50 Exact", "Group Amount P75 Exact", "Captured by Default", "Residual Band"]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=4, column=idx, value=value)
    style_row(ws, 4, 1, 8, fill=SUBHEADER_FILL, bold=True)
    for row_idx, row in enumerate(grouping_rows, start=5):
        ws[f"A{row_idx}"] = row["grouping"]
        ws[f"B{row_idx}"] = row["fc_bucket"]
        ws[f"C{row_idx}"] = row["group_presence_rate"] / 100.0
        ws[f"D{row_idx}"] = row["group_amount_p25_exact"]
        ws[f"E{row_idx}"] = row["group_amount_p50_exact"]
        ws[f"F{row_idx}"] = row["group_amount_p75_exact"]
        ws[f"G{row_idx}"] = row["group_amount_captured_by_default_rows"]
        ws[f"H{row_idx}"] = row["group_residual_band"]
        style_row(ws, row_idx, 1, 8, fill=FORMULA_FILL, align="left", wrap=True)
        ws[f"C{row_idx}"].number_format = '0.00%'
        for col in "DEFG":
            ws[f"{col}{row_idx}"].number_format = '#,##0.00'
    set_widths(ws, {"A": 28, "B": 24, "C": 14, "D": 16, "E": 16, "F": 16, "G": 16, "H": 14})


def write_grouped_adjustments_sheet(ws, grouping_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Grouped Adjustments"
    ws["A2"] = "Grouped adjustments auto-complete high-presence chemo service groups without double counting. Selected service add-ons from the same grouping reduce the residual automatically."
    style_row(ws, 1, 1, 14, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_row(ws, 2, 1, 14, fill=REFERENCE_FILL, align="left", wrap=True)
    ws["A4"], ws["B4"], ws["C4"], ws["D4"], ws["E4"] = "Grouped Adjustments", "Low", "Typical", "High", "Included Rows"
    style_row(ws, 4, 1, 5, fill=SUBHEADER_FILL, bold=True, align="left")
    headers = [
        "Grouping", "FC Bucket", "Group Presence Rate", "Group Amount P25 Exact", "Group Amount P50 Exact", "Group Amount P75 Exact",
        "Captured By Default", "Selected Add-On Amount", "Net Residual Low", "Net Residual Typical", "Net Residual High", "Selected", "Selected Amount", "Why",
    ]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=6, column=idx, value=value)
    style_row(ws, 6, 1, 14, fill=SUBHEADER_FILL, bold=True, wrap=True)
    dropdown = DataValidation(type="list", formula1=f'"{INCLUDE},{EXCLUDE}"', allow_blank=False)
    ws.add_data_validation(dropdown)
    mode_ref = f"'{SHEET_BUILDER}'!$B$5"
    rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(grouping_rows, start=7):
        ws[f"A{row_idx}"] = row["grouping"]
        ws[f"B{row_idx}"] = row["fc_bucket"]
        ws[f"C{row_idx}"] = row["group_presence_rate"] / 100.0
        ws[f"D{row_idx}"] = row["group_amount_p25_exact"]
        ws[f"E{row_idx}"] = row["group_amount_p50_exact"]
        ws[f"F{row_idx}"] = row["group_amount_p75_exact"]
        ws[f"G{row_idx}"] = row["group_amount_captured_by_default_rows"]
        ws[f"O{row_idx}"] = f'=SUMPRODUCT((\'{SHEET_SERVICE_ADDONS}\'!$E$4:$E$999=$A{row_idx})*(\'{SHEET_SERVICE_ADDONS}\'!$A$4:$A$999="{INCLUDE}")*(\'{SHEET_SERVICE_ADDONS}\'!$G$4:$G$999)*(\'{SHEET_SERVICE_ADDONS}\'!$J$4:$J$999))'
        ws[f"P{row_idx}"] = f'=SUMPRODUCT((\'{SHEET_SERVICE_ADDONS}\'!$E$4:$E$999=$A{row_idx})*(\'{SHEET_SERVICE_ADDONS}\'!$A$4:$A$999="{INCLUDE}")*(\'{SHEET_SERVICE_ADDONS}\'!$H$4:$H$999)*(\'{SHEET_SERVICE_ADDONS}\'!$J$4:$J$999))'
        ws[f"Q{row_idx}"] = f'=SUMPRODUCT((\'{SHEET_SERVICE_ADDONS}\'!$E$4:$E$999=$A{row_idx})*(\'{SHEET_SERVICE_ADDONS}\'!$A$4:$A$999="{INCLUDE}")*(\'{SHEET_SERVICE_ADDONS}\'!$I$4:$I$999)*(\'{SHEET_SERVICE_ADDONS}\'!$J$4:$J$999))'
        ws[f"H{row_idx}"] = mode_pick_formula(mode_ref, f"O{row_idx}", f"P{row_idx}", f"Q{row_idx}")
        ws[f"I{row_idx}"] = f"=MAX(0,D{row_idx}-G{row_idx}-O{row_idx})"
        ws[f"J{row_idx}"] = f"=MAX(0,E{row_idx}-G{row_idx}-P{row_idx})"
        ws[f"K{row_idx}"] = f"=MAX(0,F{row_idx}-G{row_idx}-Q{row_idx})"
        ws[f"L{row_idx}"] = INCLUDE if row["group_residual_band"] == "auto" else EXCLUDE
        dropdown.add(ws[f"L{row_idx}"])
        ws[f"M{row_idx}"] = f'=IF(L{row_idx}="{INCLUDE}",IF({mode_ref}="{MODE_LOW}",I{row_idx},IF({mode_ref}="{MODE_HIGH}",K{row_idx},J{row_idx})),0)'
        ws[f"N{row_idx}"] = "Auto common-case residual" if row["group_residual_band"] == "auto" else "Optional common-case residual"
        style_row(ws, row_idx, 1, 14, fill=FORMULA_FILL, align="left", wrap=True)
        ws[f"L{row_idx}"].fill = INPUT_FILL
        ws[f"C{row_idx}"].number_format = '0.00%'
        for col in "DEFGHIJKM":
            ws[f"{col}{row_idx}"].number_format = '#,##0.00'
        rows.append({"sheet_row": row_idx, **row})
    end = max(7, 6 + len(grouping_rows))
    ws["B5"] = f'=SUMIF(L7:L{end},"{INCLUDE}",I7:I{end})' if grouping_rows else 0
    ws["C5"] = f'=SUMIF(L7:L{end},"{INCLUDE}",J7:J{end})' if grouping_rows else 0
    ws["D5"] = f'=SUMIF(L7:L{end},"{INCLUDE}",K7:K{end})' if grouping_rows else 0
    ws["E5"] = f'=COUNTIF(L7:L{end},"{INCLUDE}")' if grouping_rows else 0
    style_row(ws, 5, 1, 5, fill=RESULT_FILL, bold=True, align="left")
    for col in ["O", "P", "Q"]:
        ws.column_dimensions[col].hidden = True
    set_widths(ws, {"A": 28, "B": 24, "C": 14, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 16, "K": 16, "L": 12, "M": 16, "N": 30})
    return rows


def write_detail_sheet(
    ws,
    refs: dict[str, str],
    advanced_refs: dict[str, Any],
    service_selection_rows: list[dict[str, Any]],
    grouped_rows: list[dict[str, Any]],
) -> int:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    headers = [
        "Line Item",
        "Parent Bucket",
        "Sub-Bucket",
        "Source Type",
        "How Calculated",
        "Item Code",
        "Count In Service Line Metric",
        "Qty Low",
        "Qty Typical",
        "Qty High",
        "Selected Qty",
        "Rate Low",
        "Rate Typical",
        "Rate High",
        "Amount Low",
        "Amount Typical",
        "Amount High",
        "Selected Amount",
    ]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=value)
    style_row(ws, 1, 1, len(headers), fill=HEADER_FILL, bold=True, font_color="FFFFFF")

    mode_ref = f"'{SHEET_BUILDER}'!B6"
    current_row = 2

    def write_row(values: dict[str, Any]) -> int:
        nonlocal current_row
        row_num = current_row
        ws[f"A{row_num}"] = values["name"]
        ws[f"B{row_num}"] = values["parent"]
        ws[f"C{row_num}"] = values["sub"]
        ws[f"D{row_num}"] = values["source"]
        ws[f"E{row_num}"] = values["how"]
        ws[f"F{row_num}"] = values.get("code", "")
        ws[f"G{row_num}"] = values.get("countable", "Yes")
        ws[f"H{row_num}"] = values.get("qty_low", 1)
        ws[f"I{row_num}"] = values.get("qty_typical", values.get("qty_low", 1))
        ws[f"J{row_num}"] = values.get("qty_high", values.get("qty_typical", values.get("qty_low", 1)))
        ws[f"K{row_num}"] = values.get("selected_qty_formula") or mode_pick_formula(mode_ref, f"H{row_num}", f"I{row_num}", f"J{row_num}")
        ws[f"L{row_num}"] = values.get("rate_low", 0)
        ws[f"M{row_num}"] = values.get("rate_typical", values.get("rate_low", 0))
        ws[f"N{row_num}"] = values.get("rate_high", values.get("rate_typical", values.get("rate_low", 0)))
        ws[f"O{row_num}"] = values.get("amount_low_formula") or f"=H{row_num}*L{row_num}"
        ws[f"P{row_num}"] = values.get("amount_typical_formula") or f"=I{row_num}*M{row_num}"
        ws[f"Q{row_num}"] = values.get("amount_high_formula") or f"=J{row_num}*N{row_num}"
        ws[f"R{row_num}"] = values.get("selected_amount_formula") or mode_pick_formula(mode_ref, f"O{row_num}", f"P{row_num}", f"Q{row_num}")
        style_row(ws, row_num, 1, 18, fill=FORMULA_FILL, align="left", wrap=True)
        for col in "HIJKLMNOPQR":
            ws[f"{col}{row_num}"].number_format = '#,##0.00'
        current_row += 1
        return row_num

    base_bucket_rows = [
        (
            "Room Charges",
            "Room Charges",
            "Historical Non-Daycare Room Bucket",
            refs["room_p25"],
            refs["room_typical"],
            refs["room_p75"],
            "Historical cash bucket",
            "Historical non-daycare room-charge P25 / P50 / P75 from reconciled actuals.",
        ),
        (
            "Investigations",
            "Investigations",
            "Historical Investigation Bucket",
            refs["investigations_p25"],
            refs["investigations_typical"],
            refs["investigations_p75"],
            "Historical cash bucket",
            "Historical non-daycare investigation P25 / P50 / P75 from reconciled actuals.",
        ),
        (
            "Bedside Services",
            "Bedside Services",
            "Historical Bedside Services Bucket",
            refs["bedside_p25"],
            refs["bedside_typical"],
            refs["bedside_p75"],
            "Historical cash bucket",
            "Historical non-daycare bedside-services P25 / P50 / P75 from reconciled actuals.",
        ),
        (
            "Other Services",
            "Other Services",
            "Historical Other Services Bucket",
            refs["other_p25"],
            refs["other_typical"],
            refs["other_p75"],
            "Historical cash bucket",
            "Historical non-daycare other-services P25 / P50 / P75 from reconciled actuals.",
        ),
    ]
    for name, parent, sub, p25_ref, p50_ref, p75_ref, source, how in base_bucket_rows:
        write_row(
            {
                "name": name,
                "parent": parent,
                "sub": sub,
                "source": source,
                "how": how,
                "countable": "No",
                "qty_low": 1,
                "qty_typical": 1,
                "qty_high": 1,
                "rate_low": 0,
                "rate_typical": 0,
                "rate_high": 0,
                "amount_low_formula": f"={p25_ref}",
                "amount_typical_formula": f"={p50_ref}",
                "amount_high_formula": f"={p75_ref}",
            }
        )
    write_row(
        {
            "name": "Doctors / Professionals",
            "parent": "Doctors / Professionals",
            "sub": "Historical Physician / Professional Bucket",
            "source": "Historical Bucket",
            "how": "Cash cohort direct billed professional bucket quartiles from reconciled actuals.",
            "code": "",
            "qty_low": 1,
            "qty_typical": 1,
            "qty_high": 1,
            "rate_low": 0,
            "rate_typical": 0,
            "rate_high": 0,
            "amount_low_formula": f"={refs['doctor_p25']}",
            "amount_typical_formula": f"={refs['doctor_typical']}",
            "amount_high_formula": f"={refs['doctor_p75']}",
        }
    )

    nursing_rate = 550.0
    dmo_rate = 1250.0
    logic_row_map = {
        "ROM5189": ("Nursing Charges", "Room Charges", "Nursing Charges", nursing_rate),
        "ROM0093": ("DMO Charges", "Room Charges", "DMO Charges", dmo_rate),
    }
    for code, (name, parent, sub, rate) in logic_row_map.items():
        select_ref = advanced_refs["logic_selection_cells"][code]
        write_row(
            {
                "name": name,
                "parent": parent,
                "sub": sub,
                "source": "Logic",
                "how": f'Optional logic row; amount included only when Advanced Controls marks it as "{INCLUDE}".',
                "code": code,
                "qty_low": 1,
                "qty_typical": 1,
                "qty_high": 1,
                "rate_low": rate,
                "rate_typical": rate,
                "rate_high": rate,
                "amount_low_formula": f'=IF({select_ref}="{INCLUDE}",H{current_row}*L{current_row},0)',
                "amount_typical_formula": f'=IF({select_ref}="{INCLUDE}",I{current_row}*M{current_row},0)',
                "amount_high_formula": f'=IF({select_ref}="{INCLUDE}",J{current_row}*N{current_row},0)',
                "selected_amount_formula": mode_pick_formula(mode_ref, f"O{current_row}", f"P{current_row}", f"Q{current_row}"),
            }
        )

    write_row(
        {
            "name": "Cath Lab Charges",
            "parent": "Procedure / OT Charges",
            "sub": "Cath Lab Hours",
            "source": "Historical Cath Lab Family",
            "how": "Actual billed cath-lab slot-family P25 / P50 / P75 from the filtered cash chemo cohort.",
            "countable": "No",
            "qty_low": 1,
            "qty_typical": 1,
            "qty_high": 1,
            "rate_low": 0,
            "rate_typical": 0,
            "rate_high": 0,
            "amount_low_formula": f"={refs['cath_p25']}",
            "amount_typical_formula": f"={refs['cath_p50']}",
            "amount_high_formula": f"={refs['cath_p75']}",
        }
    )

    pharmacy_rows = [
        ("IP Drugs", "Pharmacy", "IP Drugs / Medicines / IVs / Nutrition Products", refs["ip_drugs_p25"], refs["ip_drugs_p50"], refs["ip_drugs_p75"]),
        ("IP Consumables", "Pharmacy", "IP Treatment Supplies", refs["ip_supplies_p25"], refs["ip_supplies_p50"], refs["ip_supplies_p75"]),
        ("OT Drugs", "Pharmacy", "OT Drugs / Medicines / IVs / Nutrition Products", refs["ot_drugs_p25"], refs["ot_drugs_p50"], refs["ot_drugs_p75"]),
        ("OT Consumables", "Pharmacy", "OT Treatment Supplies", refs["ot_supplies_p25"], refs["ot_supplies_p50"], refs["ot_supplies_p75"]),
        ("Implants", "Pharmacy", "Implants / Stents", refs["implants_p25"], refs["implants_p50"], refs["implants_p75"]),
    ]
    for name, parent, sub, p25_ref, p50_ref, p75_ref in pharmacy_rows:
        write_row(
            {
                "name": name,
                "parent": parent,
                "sub": sub,
                "source": "Historical Cash Pharmacy Bucket",
                "how": "Direct cash patient bucket P25 / P50 / P75 totals from cleaned net pharmacy actuals.",
                "countable": "No",
                "qty_low": 1,
                "qty_typical": 1,
                "qty_high": 1,
                "rate_low": 0,
                "rate_typical": 0,
                "rate_high": 0,
                "amount_low_formula": f"={p25_ref}",
                "amount_typical_formula": f"={p50_ref}",
                "amount_high_formula": f"={p75_ref}",
            }
        )

    for row in service_selection_rows:
        source_row = row["source_row"]
        include_ref = f"'{SHEET_SERVICE_ADDONS}'!A{source_row}"
        write_row(
            {
                "name": row["item_name"],
                "parent": row["fc_bucket"],
                "sub": row["grouping"],
                "source": "Template",
                "how": f'Cash service-template add-on; included only when Advanced Controls marks it as "{INCLUDE}".',
                "code": row["item_code"],
                "qty_low": row["qty_p25"],
                "qty_typical": row["qty_p50"],
                "qty_high": row["qty_p75"],
                "rate_low": row["rate"],
                "rate_typical": row["rate"],
                "rate_high": row["rate"],
                "amount_low_formula": f'=IF({include_ref}="{INCLUDE}",H{current_row}*L{current_row},0)',
                "amount_typical_formula": f'=IF({include_ref}="{INCLUDE}",I{current_row}*M{current_row},0)',
                "amount_high_formula": f'=IF({include_ref}="{INCLUDE}",J{current_row}*N{current_row},0)',
                "selected_amount_formula": mode_pick_formula(mode_ref, f"O{current_row}", f"P{current_row}", f"Q{current_row}"),
            }
        )

    for row in grouped_rows:
        source_row = row["sheet_row"]
        include_ref = f"'{SHEET_GROUPED_ADJUSTMENTS}'!L{source_row}"
        write_row(
            {
                "name": f'{row["grouping"]} Residual',
                "parent": row["fc_bucket"],
                "sub": row["grouping"],
                "source": "Grouped Residual",
                "how": "Mode-aware grouped residual net of selected optional child add-ons from the same grouping.",
                "countable": "No",
                "amount_low_formula": f'=IF({include_ref}="{INCLUDE}",\'{SHEET_GROUPED_ADJUSTMENTS}\'!I{source_row},0)',
                "amount_typical_formula": f'=IF({include_ref}="{INCLUDE}",\'{SHEET_GROUPED_ADJUSTMENTS}\'!J{source_row},0)',
                "amount_high_formula": f'=IF({include_ref}="{INCLUDE}",\'{SHEET_GROUPED_ADJUSTMENTS}\'!K{source_row},0)',
                "selected_amount_formula": f'=IF({include_ref}="{INCLUDE}",\'{SHEET_GROUPED_ADJUSTMENTS}\'!M{source_row},0)',
            }
        )

    total_row = current_row + 1
    ws[f"A{total_row}"] = "Grand Total"
    ws[f"O{total_row}"] = f"=SUM(O2:O{current_row-1})"
    ws[f"P{total_row}"] = f"=SUM(P2:P{current_row-1})"
    ws[f"Q{total_row}"] = f"=SUM(Q2:Q{current_row-1})"
    ws[f"R{total_row}"] = mode_pick_formula(mode_ref, f"O{total_row}", f"P{total_row}", f"Q{total_row}")
    style_row(ws, total_row, 1, 18, fill=RESULT_FILL, bold=True, align="left")
    for col in "OPQR":
        ws[f"{col}{total_row}"].number_format = '#,##0.00'

    count_row = total_row + 2
    ws[f"A{count_row}"] = "Current Included Non-Pharmacy Service Line Count"
    ws[f"R{count_row}"] = f'=SUMPRODUCT((B2:B{current_row-1}<>"Pharmacy")*(G2:G{current_row-1}="Yes")*(R2:R{current_row-1}>0))'
    style_row(ws, count_row, 1, 18, fill=SUBHEADER_FILL, bold=True, align="left")
    ws[f"R{count_row}"].number_format = '0'

    set_widths(
        ws,
        {
            "A": 30,
            "B": 22,
            "C": 28,
            "D": 18,
            "E": 42,
            "F": 12,
            "G": 14,
            "H": 10,
            "I": 10,
            "J": 10,
            "K": 12,
            "L": 12,
            "M": 12,
            "N": 12,
            "O": 13,
            "P": 13,
            "Q": 13,
            "R": 13,
        },
    )
    return total_row


def write_summary_sheet(
    ws,
    detail_last_row: int,
    refs: dict[str, str],
    pf_payor_summary_rows: list[dict[str, str]],
) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Estimate Summary"
    style_row(ws, 1, 1, 6, fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    rows = [
        ("Historical Payer Basis", f"='{SHEET_BUILDER}'!B5"),
        ("Resolved Service Basis", f"='{SHEET_BUILDER}'!E3"),
        ("Resolved Pharmacy Basis", f"='{SHEET_BUILDER}'!E4"),
        ("Resolved PF Basis", f"='{SHEET_BUILDER}'!E5"),
        ("Estimate Mode", f"='{SHEET_BUILDER}'!B6"),
        ("Headline Estimate", f"='{SHEET_DETAIL}'!R{detail_last_row}"),
    ]
    for idx, (label, formula) in enumerate(rows, start=3):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = formula
        style_cell(ws[f"A{idx}"], fill=SUBHEADER_FILL, bold=True, align="left")
        style_cell(ws[f"B{idx}"], fill=RESULT_FILL if idx == 8 else FORMULA_FILL, bold=(idx == 8), align="left")
    ws["B8"].number_format = '#,##0.00'

    ws["A8"] = "Estimate Band"
    style_row(ws, 8, 1, 4, fill=SUBHEADER_FILL, bold=True)
    ws["A9"] = MODE_LOW
    ws["B9"] = f"='{SHEET_DETAIL}'!O{detail_last_row}"
    ws["A10"] = MODE_TYPICAL
    ws["B10"] = f"='{SHEET_DETAIL}'!P{detail_last_row}"
    ws["A11"] = MODE_HIGH
    ws["B11"] = f"='{SHEET_DETAIL}'!Q{detail_last_row}"
    for row in (9, 10, 11):
        style_row(ws, row, 1, 2, fill=FORMULA_FILL)
        ws[f"B{row}"].number_format = '#,##0.00'

    ws["D8"] = "Bucket"
    ws["E8"] = "Selected Amount"
    style_row(ws, 8, 4, 5, fill=SUBHEADER_FILL, bold=True)
    summary_buckets = [
        "Room Charges",
        "Investigations",
        "Doctors / Professionals",
        "Bedside Services",
        "Procedure Charges",
        "Other Services",
        "Physiotherapy",
        "Emergency",
        "Cath Lab",
        "OT",
        "Pharmacy",
    ]
    row_num = 9
    for bucket in summary_buckets:
        ws[f"D{row_num}"] = bucket
        ws[f"E{row_num}"] = (
            f"=SUMIF('{SHEET_DETAIL}'!$B$2:$B${detail_last_row-2},D{row_num},'{SHEET_DETAIL}'!$R$2:$R${detail_last_row-2})"
        )
        style_row(ws, row_num, 4, 5, fill=FORMULA_FILL, align="left")
        ws[f"E{row_num}"].number_format = '#,##0.00'
        row_num += 1

    ws["A22"] = "Historical Service Line Count"
    ws["B22"] = "P25"
    ws["C22"] = "P50"
    ws["D22"] = "P75"
    style_row(ws, 22, 1, 4, fill=SUBHEADER_FILL, bold=True)
    ws["B23"] = f"={refs['service_line_p25']}"
    ws["C23"] = f"={refs['service_line_p50']}"
    ws["D23"] = f"={refs['service_line_p75']}"
    ws["A24"] = "Current Included Non-Pharmacy Service Line Count"
    ws["B24"] = f"='{SHEET_DETAIL}'!R{detail_last_row+2}"
    ws["A25"] = "Alert"
    ws["B25"] = '=IF(B24<B23,"Below historical P25",IF(B24>D23,"Above historical P75","Within historical band"))'
    style_row(ws, 23, 1, 4, fill=FORMULA_FILL, align="left")
    style_row(ws, 24, 1, 2, fill=FORMULA_FILL, bold=True, align="left")
    style_row(ws, 25, 1, 2, fill=FORMULA_FILL, bold=True, align="left")

    pf_lookup = build_pf_summary_lookup(pf_payor_summary_rows)
    ws["G3"] = "Professional Fees by Payer"
    style_row(ws, 3, 7, 11, fill=SUBHEADER_FILL, bold=True, align="left")
    for idx, header in enumerate(["Payer", "Cases", "P25", "P50", "P75"], start=7):
        ws.cell(row=4, column=idx, value=header)
    style_row(ws, 4, 7, 11, fill=REFERENCE_FILL, bold=True, align="left")
    for row_idx, payor_bucket in enumerate(PF_PAYOR_ORDER, start=5):
        row = get_pf_summary_row(pf_lookup, payor_bucket)
        ws[f"G{row_idx}"] = payor_bucket
        ws[f"H{row_idx}"] = as_float(row.get("admission_count"))
        ws[f"I{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p25"))
        ws[f"J{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p50"))
        ws[f"K{row_idx}"] = as_float(row.get("pf_collectible_historical_total_p75"))
        style_row(ws, row_idx, 7, 11, fill=FORMULA_FILL, align="left")
        ws[f"H{row_idx}"].number_format = '#,##0'
        for col in "IJK":
            ws[f"{col}{row_idx}"].number_format = '#,##0.00'

    ws["G12"] = "Selected Basis PF Mix"
    style_row(ws, 12, 7, 8, fill=SUBHEADER_FILL, bold=True, align="left")
    pf_mix_rows = [
        ("Collectible Historical PF", '=IFERROR(INDEX(Reference!$AJ:$AJ,MATCH(B6,Reference!$AG:$AG,0)),0)'),
        ("Named PF", '=IFERROR(INDEX(Reference!$AL:$AL,MATCH(B6,Reference!$AG:$AG,0)),0)'),
        ("General Needed PF", '=IFERROR(INDEX(Reference!$AM:$AM,MATCH(B6,Reference!$AG:$AG,0)),0)'),
        ("Surgeon Named", '=IFERROR(INDEX(Reference!$AN:$AN,MATCH(B6,Reference!$AG:$AG,0)),0)'),
        ("Assistant Surgeon Named", '=IFERROR(INDEX(Reference!$AO:$AO,MATCH(B6,Reference!$AG:$AG,0)),0)'),
        ("Anesthetist Named", '=IFERROR(INDEX(Reference!$AP:$AP,MATCH(B6,Reference!$AG:$AG,0)),0)'),
        ("Assistant Anesthetist Named", '=IFERROR(INDEX(Reference!$AQ:$AQ,MATCH(B6,Reference!$AG:$AG,0)),0)'),
        ("Consultant / Physician Named", '=IFERROR(INDEX(Reference!$AR:$AR,MATCH(B6,Reference!$AG:$AG,0)),0)'),
        ("Selected Basis PF Shape", '=IFERROR(INDEX(Reference!$AS:$AS,MATCH(B6,Reference!$AG:$AG,0)),"n/a")'),
    ]
    for row_idx, (label, value) in enumerate(pf_mix_rows, start=13):
        ws[f"G{row_idx}"] = label
        ws[f"H{row_idx}"] = value
        style_row(ws, row_idx, 7, 8, fill=FORMULA_FILL, align="left")
        if row_idx < 21:
            ws[f"H{row_idx}"].number_format = '#,##0.00'

    set_widths(ws, {"A": 34, "B": 16, "C": 16, "D": 26, "E": 16, "G": 28, "H": 18, "I": 14, "J": 14, "K": 14})


def write_breakdown_sheet(ws, detail_ws, last_detail_row: int) -> None:
    ws.sheet_view.showGridLines = False
    headers = ["Line Item", "Bucket", "Sub-Bucket", "Source", "How Calculated", "Selected Amount"]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=value)
    style_row(ws, 1, 1, len(headers), fill=HEADER_FILL, bold=True, font_color="FFFFFF")

    out_row = 2
    for detail_row in range(2, last_detail_row):
        ws[f"A{out_row}"] = f"='{SHEET_DETAIL}'!A{detail_row}"
        ws[f"B{out_row}"] = f"='{SHEET_DETAIL}'!B{detail_row}"
        ws[f"C{out_row}"] = f"='{SHEET_DETAIL}'!C{detail_row}"
        ws[f"D{out_row}"] = f"='{SHEET_DETAIL}'!D{detail_row}"
        ws[f"E{out_row}"] = f"='{SHEET_DETAIL}'!E{detail_row}"
        ws[f"F{out_row}"] = f"='{SHEET_DETAIL}'!R{detail_row}"
        style_row(ws, out_row, 1, 6, fill=FORMULA_FILL, align="left", wrap=True)
        for col in "F":
            ws[f"{col}{out_row}"].number_format = '#,##0.00'
        out_row += 1
    set_widths(ws, {"A": 30, "B": 24, "C": 28, "D": 18, "E": 42, "F": 14})


def write_table_sheet(ws, rows: list[dict[str, str]], title: str) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    style_row(ws, 1, 1, max(1, len(rows[0]) if rows else 1), fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    if not rows:
        return
    headers = list(rows[0].keys())
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=idx, value=header)
    style_row(ws, 2, 1, len(headers), fill=SUBHEADER_FILL, bold=True)
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
        style_row(ws, row_idx, 1, len(headers), fill=FORMULA_FILL, align="left", wrap=True)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(rows)+2}"
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(header) + 2, 12), 36)
        for row in rows[:100]:
            width = min(max(width, len(str(row.get(header, ""))) + 2), 40)
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_pf_review_sheet(ws, args: argparse.Namespace) -> None:
    write_professional_fees_review_sheet(
        ws,
        template_name="Chemotherapy / Systemic Therapy Infusion",
        estimate_behavior="PF kept review-only; service-driven/historical insight only.",
        payor_summary_rows=load_pf_csv_rows(args.pf_payor_summary_csv),
        shape_review=load_pf_json(args.pf_shape_review_json),
        modeled_vs_actual_rows=[],
        header_fill=HEADER_FILL,
        subheader_fill=SUBHEADER_FILL,
        formula_fill=FORMULA_FILL,
        result_fill=RESULT_FILL,
        reference_fill=REFERENCE_FILL,
    )


def build_workbook(args: argparse.Namespace) -> Path:
    pharmacy_template_rows = load_csv_rows(args.pharmacy_template)
    pharmacy_metrics_rows = [row for row in load_csv_rows(args.pharmacy_per_patient) if normalize_text(row.get("payor_bucket")).lower() == "cash"]
    service_template_rows = filter_out_cath_lab_slot_rows(load_csv_rows(args.cleaned_services))
    optional_service_rows = filter_out_cath_lab_slot_rows(load_csv_rows(args.optional_services))
    bucket_percentiles = load_bucket_percentiles(args.pharmacy_buckets)
    service_line_metrics = load_json(args.service_line_count)
    rollup_rows = load_rollup_rows(args.rollup)
    rate_lookup = load_rate_lookup(args.rate_csv)
    pf_payor_summary_rows = load_pf_csv_rows(args.pf_payor_summary_csv)
    _cath_lab_metrics = load_cath_lab_metrics(args.cath_lab_metrics)
    built_optional_service_rows = build_optional_service_rows(optional_service_rows, rate_lookup)
    grouping_rows = build_grouping_rows(service_template_rows, built_optional_service_rows)

    workbook = Workbook()
    default_ws = workbook.active
    workbook.remove(default_ws)
    builder_ws = workbook.create_sheet(SHEET_BUILDER)
    summary_ws = workbook.create_sheet(SHEET_SUMMARY)
    advanced_ws = workbook.create_sheet(SHEET_ADVANCED)
    service_addons_ws = workbook.create_sheet(SHEET_SERVICE_ADDONS)
    grouped_adjustments_ws = workbook.create_sheet(SHEET_GROUPED_ADJUSTMENTS)
    grouping_review_ws = workbook.create_sheet(SHEET_GROUPING_REVIEW)
    breakdown_ws = workbook.create_sheet(SHEET_BREAKDOWN)
    detail_ws = workbook.create_sheet(SHEET_DETAIL)
    pharmacy_template_ws = workbook.create_sheet(SHEET_PHARMACY_TEMPLATE)
    service_template_ws = workbook.create_sheet(SHEET_SERVICE_TEMPLATE)
    pharmacy_metrics_ws = workbook.create_sheet(SHEET_PHARMACY_METRICS)
    pf_review_ws = workbook.create_sheet(SHEET_PF_REVIEW)
    reference_ws = workbook.create_sheet(SHEET_REFERENCE)

    refs = write_reference_sheet(
        reference_ws,
        pharmacy_template_rows,
        service_template_rows,
        service_line_metrics,
        rollup_rows,
        bucket_percentiles,
        grouping_rows,
        load_resolution_rows(args.payer_basis_resolution_csv),
        pf_payor_summary_rows,
    )
    write_builder_sheet(builder_ws, refs)
    advanced_refs = write_advanced_controls_sheet(
        advanced_ws,
        rate_lookup,
    )
    service_selection_rows = write_service_addons_sheet(service_addons_ws, built_optional_service_rows)
    grouped_rows = write_grouped_adjustments_sheet(grouped_adjustments_ws, grouping_rows)
    write_grouping_review_sheet(grouping_review_ws, grouping_rows)
    detail_last_row = write_detail_sheet(detail_ws, refs, advanced_refs, service_selection_rows, grouped_rows)
    write_summary_sheet(summary_ws, detail_last_row, refs, pf_payor_summary_rows)
    write_breakdown_sheet(breakdown_ws, detail_ws, detail_last_row)
    write_table_sheet(pharmacy_template_ws, pharmacy_template_rows, "Cash Chemo Pharmacy Template")
    write_table_sheet(service_template_ws, service_template_rows, "Cash Chemo Service Template")
    write_table_sheet(pharmacy_metrics_ws, pharmacy_metrics_rows, "Cash Chemo Pharmacy Metrics")
    write_pf_review_sheet(pf_review_ws, args)

    for ws in workbook.worksheets:
        ws.sheet_view.showGridLines = False

    apply_default_calc_settings(workbook)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    recalculate_workbook_with_soffice(args.output)
    return args.output


def main() -> None:
    args = parse_args()
    output = build_workbook(args)
    print(f"output={output}")


if __name__ == "__main__":
    main()
