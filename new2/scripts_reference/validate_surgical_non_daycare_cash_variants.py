from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from scripts.etl.fc_estimate.builder_input_pack import (
    apply_builder_filters,
    fetch_template_main_table_rows,
    normalize_text,
)
from scripts.etl.fc_estimate.surgical_workbook_standard import (
    inspect_generated_workbook,
    write_surgical_review_artifacts,
)
from scripts.etl.fc_estimate.variant_manifest import BuilderVariant, get_variant
from scripts.export_robotic_tkr_fc_estimate_builder import validate_generated_surgical_workbook


REPO_ROOT = Path("/Users/reyvanttambi/Documents/New project 3")
OUTPUT_ROOT = REPO_ROOT / "output"
DEFAULT_SUMMARY_OUTPUT = OUTPUT_ROOT / "fc_surgical_non_daycare_cash_validation_summary.json"
NON_GATING_ANALYTICAL_METRICS = {
    "Professional Fees (Calculated)",
    "Grand Total (Calculated PF)",
}


@dataclass(frozen=True)
class SurgicalValidationTarget:
    variant: BuilderVariant
    wrapper_script: Path
    workbook_output: Path
    workbook_slug: str

    @property
    def output_dir(self) -> Path:
        return self.variant.output_dir

    @property
    def cohort_audit_output(self) -> Path:
        return self.output_dir / "00_fc_cohort_audit_non_daycare_surgical_cash.json"

    @property
    def validation_output(self) -> Path:
        return self.output_dir / "38_surgical_workbook_validation_summary.json"

    @property
    def review_output(self) -> Path:
        return self.output_dir / "39_surgical_cash_review_pack.json"

    @property
    def ux_contract_output(self) -> Path:
        return self.output_dir / "33_surgical_workbook_ux_contract.json"

    @property
    def sheet_map_output(self) -> Path:
        return self.output_dir / "34_surgical_workbook_sheet_map.csv"

    @property
    def field_source_output(self) -> Path:
        return self.output_dir / "35_surgical_workbook_field_source_map.csv"

    @property
    def legacy_removals_output(self) -> Path:
        return self.output_dir / "36_surgical_workbook_removed_legacy_controls.json"

    @property
    def acceptance_checklist_output(self) -> Path:
        return self.output_dir / "37_surgical_workbook_acceptance_checklist.md"


TARGETS: list[SurgicalValidationTarget] = [
    SurgicalValidationTarget(
        variant=get_variant("robotic_tkr_unilateral_right_cash"),
        wrapper_script=REPO_ROOT / "scripts" / "export_robotic_tkr_unilateral_fc_estimate_builder.py",
        workbook_output=OUTPUT_ROOT / "fc_estimate_builder_robotic_tkr_unilateral_right_cash_tr1.xlsx",
        workbook_slug="robotic_tkr_unilateral_right_cash_tr1",
    ),
    SurgicalValidationTarget(
        variant=get_variant("robotic_tkr_unilateral_left_cash"),
        wrapper_script=REPO_ROOT / "scripts" / "export_robotic_tkr_unilateral_left_fc_estimate_builder.py",
        workbook_output=OUTPUT_ROOT / "fc_estimate_builder_robotic_tkr_unilateral_left_cash_tr1.xlsx",
        workbook_slug="robotic_tkr_unilateral_left_cash_tr1",
    ),
    SurgicalValidationTarget(
        variant=get_variant("robotic_tkr_bilateral_cash"),
        wrapper_script=REPO_ROOT / "scripts" / "export_robotic_tkr_bilateral_fc_estimate_builder.py",
        workbook_output=OUTPUT_ROOT / "fc_estimate_builder_robotic_tkr_bilateral_cash_tr1.xlsx",
        workbook_slug="robotic_tkr_bilateral_cash_tr1",
    ),
    SurgicalValidationTarget(
        variant=get_variant("robotic_tkr_single_cash"),
        wrapper_script=REPO_ROOT / "scripts" / "export_tkr_robotic_single_fc_estimate_builder.py",
        workbook_output=OUTPUT_ROOT / "fc_estimate_builder_tkr_robotic_single_cash_tr1.xlsx",
        workbook_slug="tkr_robotic_single_cash_tr1",
    ),
    SurgicalValidationTarget(
        variant=get_variant("thr_hemiarthroplasty_cash"),
        wrapper_script=REPO_ROOT / "scripts" / "export_total_hip_replacement_thr_hemiarthroplasty_fc_estimate_builder.py",
        workbook_output=OUTPUT_ROOT / "fc_estimate_builder_total_hip_replacement_thr_hemiarthroplasty_cash_tr1.xlsx",
        workbook_slug="total_hip_replacement_thr_hemiarthroplasty_cash_tr1",
    ),
]


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_json_if_exists(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def load_builder_input_pack_rows(output_dir: Path) -> list[dict[str, Any]]:
    matches = sorted(output_dir.glob("31_builder_input_pack*.json"))
    if not matches:
        return []
    payload = load_json_if_exists(matches[0])
    return payload.get("ip_actual_benchmark_rows") or []


def workbook_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value)
    if not text:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def run_wrapper(target: SurgicalValidationTarget) -> None:
    env = dict(os.environ)
    env["PYTHONPATH"] = f"{REPO_ROOT / 'scripts'}:{REPO_ROOT / 'scripts' / 'etl'}:{REPO_ROOT}"
    env["PYTHONPYCACHEPREFIX"] = str(REPO_ROOT / "tmp" / "pycache")
    subprocess.run(
        [sys.executable, str(target.wrapper_script)],
        cwd=REPO_ROOT,
        env=env,
        check=True,
    )


def parse_labeled_value_block(ws, start_row: int, label_col: str, value_col: str, end_row: int) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for row_idx in range(start_row, end_row + 1):
        label = normalize_text(ws[f"{label_col}{row_idx}"].value)
        if not label:
            continue
        output[label] = ws[f"{value_col}{row_idx}"].value
    return output


def read_builder_snapshot(workbook_path: Path) -> dict[str, Any]:
    calc_wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        ws = calc_wb["Builder"]
        return {
            "procedure": normalize_text(ws["B2"].value),
            "pricing_mode": normalize_text(ws["E2"].value),
            "historical_payer_basis": normalize_text(ws["E3"].value),
            "selected_room_type": normalize_text(ws["B4"].value),
            "estimate_mode": normalize_text(ws["B5"].value),
            "emergency_ot": normalize_text(ws["B6"].value),
            "mlc": normalize_text(ws["E6"].value),
            "resolved_payor_bucket": normalize_text(ws["E4"].value),
            "resolved_tariff_name": normalize_text(ws["G4"].value),
            "resolved_tariff_code": normalize_text(ws["E5"].value),
            "resolved_pharmacy_basis": normalize_text(ws["G5"].value),
            "resolved_service_basis": normalize_text(ws["G6"].value),
            "resolved_pf_basis": normalize_text(ws["G7"].value),
            "basis_resolver_note": normalize_text(ws["G8"].value),
            "drivers": {
                "los_p25": workbook_number(ws["B10"].value),
                "los_p50": workbook_number(ws["C10"].value),
                "los_p75": workbook_number(ws["D10"].value),
                "selected_los": workbook_number(ws["G10"].value),
                "icu_p25": workbook_number(ws["B11"].value),
                "icu_p50": workbook_number(ws["C11"].value),
                "icu_p75": workbook_number(ws["D11"].value),
                "selected_icu_days": workbook_number(ws["G11"].value),
                "ward_p25": workbook_number(ws["B12"].value),
                "ward_p50": workbook_number(ws["C12"].value),
                "ward_p75": workbook_number(ws["D12"].value),
                "selected_ward_days": workbook_number(ws["G12"].value),
                "ot_p25": workbook_number(ws["B13"].value),
                "ot_p50": workbook_number(ws["C13"].value),
                "ot_p75": workbook_number(ws["D13"].value),
                "selected_ot_hours": workbook_number(ws["G13"].value),
            },
            "resolved_ot": {
                "hours": workbook_number(ws["B14"].value),
                "code": normalize_text(ws["B15"].value),
                "label": normalize_text(ws["B16"].value),
                "type": normalize_text(ws["B17"].value),
            },
        }
    finally:
        calc_wb.close()


def read_summary_snapshot(workbook_path: Path) -> dict[str, Any]:
    calc_wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        ws = calc_wb["Estimate Summary"]
        selected_controls = parse_labeled_value_block(ws, 2, "A", "B", 11)
        bucket_rows = {
            normalize_text(ws[f"A{row_idx}"].value): workbook_number(ws[f"B{row_idx}"].value)
            for row_idx in range(13, 24)
            if normalize_text(ws[f"A{row_idx}"].value)
        }
        return {
            "selected_controls": selected_controls,
            "bucket_rows": bucket_rows,
            "final_estimate": workbook_number(ws["E2"].value),
        }
    finally:
        calc_wb.close()


def read_comparison_snapshot(workbook_path: Path) -> dict[str, Any]:
    calc_wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        ws = calc_wb["Estimate vs IP FC Actuals"]
        metric_rows: list[dict[str, Any]] = []
        row_idx = 5
        while True:
            label = normalize_text(ws[f"A{row_idx}"].value)
            if not label:
                break
            metric_rows.append(
                {
                    "metric": label,
                    "basis": normalize_text(ws[f"B{row_idx}"].value),
                    "selected_estimate": workbook_number(ws[f"C{row_idx}"].value),
                    "actual_p25": workbook_number(ws[f"D{row_idx}"].value),
                    "actual_p50": workbook_number(ws[f"E{row_idx}"].value),
                    "actual_p75": workbook_number(ws[f"F{row_idx}"].value),
                    "delta_vs_p25": workbook_number(ws[f"G{row_idx}"].value),
                    "delta_vs_p50": workbook_number(ws[f"H{row_idx}"].value),
                    "delta_vs_p75": workbook_number(ws[f"I{row_idx}"].value),
                    "status": normalize_text(ws[f"J{row_idx}"].value),
                }
            )
            row_idx += 1

        driver_rows: list[dict[str, Any]] = []
        row_idx += 4
        while True:
            label = normalize_text(ws[f"A{row_idx}"].value)
            if not label:
                break
            driver_rows.append(
                {
                    "driver": label,
                    "basis": normalize_text(ws[f"B{row_idx}"].value),
                    "selected": workbook_number(ws[f"C{row_idx}"].value),
                    "actual_p25": workbook_number(ws[f"D{row_idx}"].value),
                    "actual_p50": workbook_number(ws[f"E{row_idx}"].value),
                    "actual_p75": workbook_number(ws[f"F{row_idx}"].value),
                    "status": normalize_text(ws[f"G{row_idx}"].value),
                }
            )
            row_idx += 1

        row_idx += 4
        cohort_rows: list[dict[str, Any]] = []
        while True:
            label = normalize_text(ws[f"A{row_idx}"].value)
            if not label:
                break
            cohort_rows.append(
                {
                    "component": label,
                    "resolved_basis": normalize_text(ws[f"B{row_idx}"].value),
                    "cases_used": workbook_number(ws[f"C{row_idx}"].value),
                    "cash_cases": workbook_number(ws[f"D{row_idx}"].value),
                    "gipsa_cases": workbook_number(ws[f"E{row_idx}"].value),
                    "non_gipsa_cases": workbook_number(ws[f"F{row_idx}"].value),
                    "corporate_cases": workbook_number(ws[f"G{row_idx}"].value),
                    "scope": normalize_text(ws[f"H{row_idx}"].value),
                }
            )
            row_idx += 1

        material_rows = [row for row in metric_rows if row["status"] == "Material Gap"]
        out_of_range_rows = [
            row
            for row in metric_rows
            if row["status"] in {"Material Gap", "Above Range", "Below Range"}
        ]
        gating_material_rows = [
            row for row in material_rows if row["metric"] not in NON_GATING_ANALYTICAL_METRICS
        ]
        top_deltas = sorted(
            metric_rows,
            key=lambda row: abs(row["delta_vs_p50"] or 0.0),
            reverse=True,
        )[:8]
        return {
            "metric_rows": metric_rows,
            "driver_rows": driver_rows,
            "cohort_rows": cohort_rows,
            "material_gap_rows": material_rows,
            "gating_material_gap_rows": gating_material_rows,
            "out_of_range_rows": out_of_range_rows,
            "top_delta_vs_p50_rows": top_deltas,
        }
    finally:
        calc_wb.close()


def read_ip_actual_admissions(workbook_path: Path) -> list[str]:
    calc_wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        ws = calc_wb["IP FC Actuals"]
        admissions: list[str] = []
        row_idx = 5
        while True:
            admission = normalize_text(ws[f"A{row_idx}"].value)
            if not admission:
                break
            admissions.append(admission)
            row_idx += 1
        return admissions
    finally:
        calc_wb.close()


def build_cohort_review(target: SurgicalValidationTarget, workbook_admissions: list[str]) -> dict[str, Any]:
    source_rows = fetch_template_main_table_rows(target.variant.template_registry_id)
    benchmark_rows = load_builder_input_pack_rows(target.output_dir)
    if not benchmark_rows:
        benchmark_rows, _benchmark_excluded = apply_builder_filters(
            source_rows,
            payor_scope=target.variant.target_payor_bucket,
            management_type=target.variant.management_type,
            stay_type=target.variant.stay_type,
        )
    included_rows, excluded_rows = apply_builder_filters(
        source_rows,
        payor_scope=target.variant.target_payor_bucket,
        management_type=target.variant.management_type,
        stay_type=target.variant.stay_type,
    )
    included_admissions = {normalize_text(row.get("admission_no")) for row in included_rows}
    benchmark_admissions = {normalize_text(row.get("admission_no")) for row in benchmark_rows}
    workbook_admission_set = set(workbook_admissions)
    missing_from_workbook = sorted(included_admissions - workbook_admission_set)
    extra_in_workbook = sorted(workbook_admission_set - included_admissions)
    missing_from_benchmark = sorted(benchmark_admissions - workbook_admission_set)
    extra_from_benchmark = sorted(workbook_admission_set - benchmark_admissions)
    quality_counts = Counter(normalize_text(row.get("fc_actual_quality_level")) or "ok" for row in included_rows)
    payor_counts = Counter(normalize_text(row.get("payor_bucket")) or "Unknown" for row in included_rows)
    return {
        "source_row_count": len(source_rows),
        "included_row_count": len(included_rows),
        "benchmark_row_count": len(benchmark_rows),
        "excluded_row_count": len(excluded_rows),
        "included_payor_mix": dict(payor_counts),
        "included_quality_mix": dict(quality_counts),
        "workbook_ip_actual_row_count": len(workbook_admissions),
        "workbook_matches_included_rows": not missing_from_workbook and not extra_in_workbook,
        "workbook_matches_benchmark_rows": not missing_from_benchmark and not extra_from_benchmark,
        "missing_from_workbook_sample": missing_from_workbook[:10],
        "extra_in_workbook_sample": extra_in_workbook[:10],
        "missing_from_benchmark_sample": missing_from_benchmark[:10],
        "extra_from_benchmark_sample": extra_from_benchmark[:10],
    }


def build_consistency_checks(
    builder_snapshot: dict[str, Any],
    summary_snapshot: dict[str, Any],
    comparison_snapshot: dict[str, Any],
) -> dict[str, Any]:
    drivers = builder_snapshot["drivers"]
    los_matches = (
        (drivers["selected_los"] or 0.0)
        == round((drivers["selected_icu_days"] or 0.0) + (drivers["selected_ward_days"] or 0.0), 6)
    )
    ot_selected = drivers["selected_ot_hours"] or 0.0
    ot_resolved = builder_snapshot["resolved_ot"]["hours"] or 0.0
    return {
        "los_matches_icu_plus_ward": los_matches,
        "selected_los": drivers["selected_los"],
        "selected_icu_days": drivers["selected_icu_days"],
        "selected_ward_days": drivers["selected_ward_days"],
        "selected_ot_hours": ot_selected,
        "resolved_ot_hours": ot_resolved,
        "resolved_ot_code_present": bool(builder_snapshot["resolved_ot"]["code"]),
        "resolved_ot_label_present": bool(builder_snapshot["resolved_ot"]["label"]),
        "final_estimate": summary_snapshot["final_estimate"],
        "comparison_metric_count": len(comparison_snapshot["metric_rows"]),
        "material_gap_count": len(comparison_snapshot["material_gap_rows"]),
        "gating_material_gap_count": len(comparison_snapshot["gating_material_gap_rows"]),
        "out_of_range_count": len(comparison_snapshot["out_of_range_rows"]),
    }


def build_review_pack(target: SurgicalValidationTarget) -> dict[str, Any]:
    write_surgical_review_artifacts(
        variant_key=target.variant.variant_key,
        template_registry_id=target.variant.template_registry_id,
        template_name=target.variant.template_name,
        ux_contract_output=target.ux_contract_output,
        sheet_map_output=target.sheet_map_output,
        field_source_map_output=target.field_source_output,
        legacy_removals_output=target.legacy_removals_output,
        acceptance_checklist_output=target.acceptance_checklist_output,
    )
    validation_summary = validate_generated_surgical_workbook(target.workbook_output)
    validation_summary["workbook_shape"] = inspect_generated_workbook(target.workbook_output)
    write_json(target.validation_output, validation_summary)

    builder_snapshot = read_builder_snapshot(target.workbook_output)
    summary_snapshot = read_summary_snapshot(target.workbook_output)
    comparison_snapshot = read_comparison_snapshot(target.workbook_output)
    workbook_admissions = read_ip_actual_admissions(target.workbook_output)
    cohort_review = build_cohort_review(target, workbook_admissions)
    cohort_audit = load_json_if_exists(target.cohort_audit_output)
    consistency_checks = build_consistency_checks(builder_snapshot, summary_snapshot, comparison_snapshot)
    structural_shape = validation_summary.get("workbook_shape") or {}
    freeze_violations = [
        sheet_name
        for sheet_name, details in (structural_shape.get("sheet_dimensions") or {}).items()
        if details.get("freeze_panes") is not None
    ]

    review = {
        "variant_key": target.variant.variant_key,
        "template_name": target.variant.template_name,
        "template_registry_id": target.variant.template_registry_id,
        "workbook_path": str(target.workbook_output),
        "wrapper_script": str(target.wrapper_script),
        "validation_summary": validation_summary,
        "structural_checks": {
            "sheet_order_matches": validation_summary.get("sheet_order") == inspect_generated_workbook(target.workbook_output).get("sheet_order"),
            "freeze_panes_removed": not freeze_violations,
            "freeze_pane_violations": freeze_violations,
            "merged_ranges_removed": bool(validation_summary.get("merged_ranges_removed")),
        },
        "builder_snapshot": builder_snapshot,
        "estimate_summary_snapshot": summary_snapshot,
        "comparison_snapshot": comparison_snapshot,
        "cohort_review": cohort_review,
        "cohort_audit": cohort_audit,
        "consistency_checks": consistency_checks,
        "material_mismatches": comparison_snapshot["material_gap_rows"],
        "gating_material_mismatches": comparison_snapshot["gating_material_gap_rows"],
        "top_estimate_bucket_vs_actual_p50_deltas": comparison_snapshot["top_delta_vs_p50_rows"],
    }
    write_json(target.review_output, review)
    return review


def build_cross_summary(reviews: list[dict[str, Any]]) -> dict[str, Any]:
    passing_variants: list[str] = []
    failing_variants: list[dict[str, Any]] = []
    common_issues: Counter[str] = Counter()

    for review in reviews:
        checks = review["consistency_checks"]
        structural = review["structural_checks"]
        cohort = review["cohort_review"]
        issues: list[str] = []
        if not structural["freeze_panes_removed"]:
            issues.append("freeze_panes_present")
        if not structural["merged_ranges_removed"]:
            issues.append("merged_ranges_present")
        if not checks["los_matches_icu_plus_ward"]:
            issues.append("los_driver_mismatch")
        if not checks["resolved_ot_code_present"] or not checks["resolved_ot_label_present"]:
            issues.append("ot_slot_resolution_missing")
        if cohort["extra_from_benchmark_sample"]:
            issues.append("cohort_row_mismatch")
        if checks["gating_material_gap_count"] > 0:
            issues.append("material_bucket_gaps_present")

        if issues:
            for issue in issues:
                common_issues[issue] += 1
            failing_variants.append(
                {
                    "variant_key": review["variant_key"],
                    "template_name": review["template_name"],
                    "issues": issues,
                    "review_output": str(review["workbook_path"]).replace(".xlsx", ""),
                }
            )
        else:
            passing_variants.append(review["variant_key"])

    return {
        "validated_variant_count": len(reviews),
        "passing_variant_keys": passing_variants,
        "failing_variants": failing_variants,
        "common_issue_counts": dict(common_issues),
        "shell_ready_for_cash_tr1": not failing_variants,
        "variant_summaries": [
            {
                "variant_key": review["variant_key"],
                "template_name": review["template_name"],
                "final_estimate": review["consistency_checks"]["final_estimate"],
                "material_gap_count": review["consistency_checks"]["material_gap_count"],
                "gating_material_gap_count": review["consistency_checks"]["gating_material_gap_count"],
                "out_of_range_count": review["consistency_checks"]["out_of_range_count"],
                "included_row_count": review["cohort_review"]["included_row_count"],
                "benchmark_row_count": review["cohort_review"]["benchmark_row_count"],
                "payor_mix": review["cohort_review"]["included_payor_mix"],
                "resolved_bases": {
                    "service": review["builder_snapshot"]["resolved_service_basis"],
                    "pharmacy": review["builder_snapshot"]["resolved_pharmacy_basis"],
                    "pf": review["builder_snapshot"]["resolved_pf_basis"],
                },
                "top_delta_rows": review["top_estimate_bucket_vs_actual_p50_deltas"][:5],
                "review_output": str(Path(review["workbook_path"]).with_suffix("")),
            }
            for review in reviews
        ],
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build and validate surgical non-daycare cash FC workbooks.")
    parser.add_argument("--skip-build", action="store_true", help="Reuse existing workbook outputs instead of rebuilding them first.")
    parser.add_argument("--summary-output", type=Path, default=DEFAULT_SUMMARY_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    reviews: list[dict[str, Any]] = []
    for target in TARGETS:
        if not args.skip_build:
            run_wrapper(target)
        review = build_review_pack(target)
        reviews.append(review)
        print(f"validated_variant={target.variant.variant_key}")
        print(f"review_output={target.review_output}")

    cross_summary = build_cross_summary(reviews)
    write_json(args.summary_output, cross_summary)
    print(f"summary_output={args.summary_output}")


if __name__ == "__main__":
    main()
