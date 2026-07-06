from __future__ import annotations

import argparse
import csv
import json
import math
import os
import statistics
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import psycopg
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from fc_slot_family import is_cath_lab_slot_service
from fc_payer_basis_resolution import (
    AUTO_BASIS,
    PAYER_BASIS_OPTIONS as RESOLVER_PAYER_BASIS_OPTIONS,
    normalize_text as resolver_normalize_text,
    supported_basis_options_from_resolution_rows,
)
from professional_fee_review_workbook import (
    load_csv_rows as load_pf_csv_rows,
    load_json as load_pf_json,
    write_professional_fees_review_sheet,
)
from workbook_postprocess import apply_default_calc_settings, recalculate_workbook_with_soffice


DEFAULT_TEMPLATE_WORKBOOK = Path(
    "/Users/reyvanttambi/Documents/New project 3/RawData/FC Estimate Builder/FC estimate builder - robotic unilateral TKR right.xlsx"
)
DEFAULT_QUARTILES_JSON = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/17_los_icu_ward_ot_quartiles.json"
)
DEFAULT_PER_IP_BUCKETS = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/09_per_ip_bucket_totals_from_classification.csv"
)
DEFAULT_PHARMACY_TEMPLATE = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/10_classified_pharmacy_template.csv"
)
DEFAULT_IMPLANT_HIERARCHY = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/16_implant_family_brand_item_presence.csv"
)
DEFAULT_IMPLANT_DETAIL = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/12_implant_combinations_per_ip.csv"
)
DEFAULT_SERVICES_TEMPLATE = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_services_template.csv"
)
DEFAULT_CLEANED_SERVICES_OUTPUT = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_services_template_cleaned_for_fc.csv"
)
DEFAULT_SERVICE_LINE_COUNT_METRICS = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_service_line_count_metrics.json"
)
DEFAULT_IP_PHARMACY_PER_DAY_METRICS = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_ip_pharmacy_per_day_metrics.json"
)
DEFAULT_RATE_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/tr1_cash_rates_robotic_tkr_unilateral_right_full_codes.csv"
)
DEFAULT_OT_SLOT_RATE_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/18_ot_slot_rates_tr1.csv"
)
DEFAULT_CATH_LAB_METRICS_JSON = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/22_cath_lab_metrics_cash.json"
)
DEFAULT_CATH_LAB_SLOT_RATE_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/23_cath_lab_slot_rates_tr1.csv"
)
DEFAULT_PAYER_BASIS_SUMMARY_JSON = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/19_payer_basis_summary.json"
)
DEFAULT_PAYER_BASIS_SERVICE_METRICS_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/20_payer_basis_service_metrics.csv"
)
DEFAULT_PAYER_BASIS_PHARMACY_METRICS_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/21_payer_basis_pharmacy_metrics.csv"
)
DEFAULT_ORG_TARIFF_REFERENCE_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/24_insurance_org_tariff_reference.csv"
)
DEFAULT_TARIFF_RATE_MATRIX_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/25_tariff_service_rate_matrix.csv"
)
DEFAULT_TARIFF_OT_SLOT_RATE_MATRIX_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/26_tariff_ot_slot_rate_matrix.csv"
)
DEFAULT_INSURANCE_POLICY_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/27_insurance_fc_policy.csv"
)
DEFAULT_GROUPING_GAP_SUMMARY_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/fc_grouping_gap_audit_cash_only/robotic_tkr_unilateral_right_cash/grouping_gap_summary.csv"
)
DEFAULT_GROUPING_GAP_CHILD_DETAIL_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/fc_grouping_gap_audit_cash_only/robotic_tkr_unilateral_right_cash/grouping_gap_child_detail.csv"
)
DEFAULT_PAYER_BASIS_RESOLUTION_CSV = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/robotic_tkr_unilateral_right_three_csvs/30_payer_basis_resolution_summary.csv"
)
DEFAULT_OUTPUT = Path(
    "/Users/reyvanttambi/Documents/New project 3/output/fc_estimate_builder_robotic_tkr_unilateral_right_cash_tr1.xlsx"
)
DEFAULT_DB_URL = "postgresql://postgres:postgres@127.0.0.1:54322/postgres"

DEFAULT_SHEET1_TEMPLATE_NAME = "Robotic TKR Unilateral - Right"
SHEET1_PAYOR_TEXT = "Cash (Tarriff Code Tr1)"
DEFAULT_SHEET1_LOS_GUIDANCE = ""
DEFAULT_SHEET1_MANAGEMENT_TYPE = "Surgical"

DEFAULT_PROFESSIONAL_FEES = {
    33: 65000.0,
    34: 9750.0,
    35: 16250.0,
    36: 4062.5,
}

OT_CONSUMABLES_P25_SHARE_THRESHOLD = 0.30
OT_CONSUMABLES_P50_SHARE_THRESHOLD = 0.50

BASE_ESTIMATE_BUILDER_TEMPLATE_ROWS = {
    2: "XRY5090",
    8: "RNS5005",
    10: "PHY5082",
    11: "PAT0045",
    12: "PAT0042",
    13: "OTI0098",
    21: "EME0087",
    23: "EME0017",
    24: "DIE0001",
    25: "CAS0007",
    26: "CAR5341",
    27: "BIO0162",
    28: "BIO0004",
    29: "BIO0003",
    30: "BIO0002",
    31: "BIO0001",
}

ESTIMATE_BUILDER_ALWAYS_ONE_ROWS = {
    9: "RNS0120",
    14: "OTI0018",
    15: "OTI0015",
    16: "OTC5005",
    17: "OTC0010",
}

ESTIMATE_BUILDER_LOGIC_ROWS = {
    3: ("ward_days", "ROM5189", False),
    4: ("icu_days", "ROM5189", True),
    5: ("ward_days", "ROM0093", False),
    6: ("icu_days", "ROM5009", True),
    18: ("icu_days", "ICC0002", True),
    19: ("icu_days", "ICC0001", True),
    20: ("los_days", "HSP5013", False),
    22: ("icu_days", "EME0019", False),
}

DEFAULT_SERVICES_SELECTION_CODES = [
    "BIO0106",
    "BIO0181",
    "BIO0121",
    "BLD0024",
    "EME5045",
    "EME0020",
    "EME0016",
    "EME0052",
    "BIO0098",
    "EME0088",
    "BIO0097",
    "BIO0064",
    "XRY044",
    "PHY5076",
]

SERVICES_SIGNIFICANCE_AMOUNT_THRESHOLD = 1000.0

FIXED_ESTIMATE_TEMPLATE_SERVICE_CODES = {
    "XRY5090",
    "PHY5082",
    "PAT0045",
    "PAT0042",
    "OTI0098",
    "EME0087",
    "EME0017",
    "DIE0001",
    "CAS0007",
    "CAR5341",
    "BIO0162",
    "BIO0004",
    "BIO0003",
    "BIO0002",
    "BIO0001",
}

LOGIC_DRIVEN_SERVICE_CODES = {
    "ROM5189",
    "ROM0093",
    "ROM5009",
    "ROM0001",
    "ROM0024",
    "ROM0036",
    "ICC0002",
    "ICC0001",
    "HSP5013",
    "EME0019",
    "OTC0010",
    "RNS0120",
    "RNS5005",
    "OTI0018",
    "OTI0015",
    "OTC5005",
    "HSP0047",
}

TEMPLATE_EXCLUDED_SERVICE_CODES = FIXED_ESTIMATE_TEMPLATE_SERVICE_CODES | LOGIC_DRIVEN_SERVICE_CODES
UNILATERAL_TR1_BASE_CODES = sorted(TEMPLATE_EXCLUDED_SERVICE_CODES)

DEFAULT_OT_CONSUMABLE_SHORTLIST = [
    ("PURESIGHT ARRAY SET KNEE(4515-70-011-DEPUY)", True),
    ("STERILE DRAPE KIT(4515-70-018-DEPUY)", False),
    ("STERILE DRAPE KIT(4515-70-008-DEPUY)", True),
    ("OSCILLATING SAW BLADE 85MMX19X2MM(4515-70-000-DEPUY)", False),
    ("MICS SAGITTAL BLADE STANDARD(116170-STRYKER)", False),
]

PHARMACY_VARIANCE_OT_RESULT_CELL = "F3"
PHARMACY_VARIANCE_IMPLANT_RESULT_CELL = "N3"
SELECTION_INCLUDE = "Include"
SELECTION_EXCLUDE = "Exclude"
SHEET_BUILDER = "Builder"
SHEET_SUMMARY = "Estimate Summary"
SHEET_ESTIMATE_VS_ACTUAL = "Estimate vs IP FC Actuals"
SHEET_ADVANCED = "Advanced Controls"
SHEET_SERVICE_ADDONS = "Service Add-Ons"
SHEET_GROUPED_ADJUSTMENTS = "Grouped Adjustments"
SHEET_IMPLANTS_SELECT = "Implant Selection"
SHEET_BREAKDOWN = "Estimate Breakdown"
SHEET_DETAIL = "Line Item Detail"
SHEET_PHARMACY_TEMPLATE = "Pharmacy Template"
SHEET_SERVICE_TEMPLATE = "Service Template"
SHEET_GROUPING_REVIEW = "Grouping Review"
SHEET_PHARMACY_METRICS = "Pharmacy Metrics"
SHEET_IP_ACTUALS = "IP FC Actuals"
SHEET_PF_REVIEW = "Professional Fees Review"
SHEET_REFERENCE = "Reference"
PAYER_BASIS_OPTIONS = [
    *RESOLVER_PAYER_BASIS_OPTIONS,
]
PRICING_MODE_OPTIONS = [
    "Cash / TR1",
    "Insurance / Org Tariff",
]
OT_SLOT_REFERENCE_START_ROW = 300
OT_SLOT_REFERENCE_END_ROW = 331
OT_SLOT_REFERENCE_COLS = {
    "tariff_code": "J",
    "hours": "K",
    "mode": "L",
    "code": "M",
    "name": "N",
    "general": "O",
    "twin": "P",
    "single": "Q",
    "icu": "R",
}
CATH_LAB_REFERENCE_COLS = {"label": "S", "p25": "T", "p50": "U", "p75": "V"}
PAYER_BASIS_SUMMARY_START_ROW = 2
PAYER_BASIS_SUMMARY_COLS = {
    "basis_label": "AZ",
    "cohort_size": "BA",
    "cash_count": "BB",
    "gipsa_count": "BC",
    "non_gipsa_count": "BD",
    "corporate_count": "BE",
    "los_p25": "BF",
    "los_p50": "BG",
    "los_p75": "BH",
    "icu_p25": "BI",
    "icu_p50": "BJ",
    "icu_p75": "BK",
    "ward_p25": "BL",
    "ward_p50": "BM",
    "ward_p75": "BN",
    "ot_p25": "BO",
    "ot_p50": "BP",
    "ot_p75": "BQ",
    "service_line_p25": "BR",
    "service_line_p50": "BS",
    "service_line_p75": "BT",
    "ip_drugs_p25": "BU",
    "ip_drugs_p50": "BV",
    "ip_drugs_p75": "BW",
    "ip_consumables_p25": "BX",
    "ip_consumables_p50": "BY",
    "ip_consumables_p75": "BZ",
    "ot_drugs_p25": "CA",
    "ot_drugs_p50": "CB",
    "ot_drugs_p75": "CC",
    "ot_consumables_p25": "CD",
    "ot_consumables_p50": "CE",
    "ot_consumables_p75": "CF",
    "implants_p25": "CG",
    "implants_p50": "CH",
    "implants_p75": "CI",
    "ip_drugs_day_p25": "CJ",
    "ip_drugs_day_p50": "CK",
    "ip_drugs_day_p75": "CL",
    "ip_consumables_day_p25": "CM",
    "ip_consumables_day_p50": "CN",
    "ip_consumables_day_p75": "CO",
    "cath_lab_p25": "CP",
    "cath_lab_p50": "CQ",
    "cath_lab_p75": "CR",
}
PAYER_BASIS_SERVICE_START_ROW = 2
PAYER_BASIS_SERVICE_COLS = {
    "key": "CQ",
    "basis_label": "CR",
    "item_code": "CS",
    "item_name": "CT",
    "fc_estimate_bucket": "CU",
    "grouping": "CV",
    "case_presence_rate": "CW",
    "quantity_p25": "CX",
    "quantity_p50": "CY",
    "quantity_p75": "CZ",
    "amount_cash_typical": "DA",
    "tariff_general": "DB",
    "tariff_twin": "DC",
    "tariff_single": "DD",
    "tariff_icu": "DE",
}
PAYER_BASIS_PHARMACY_START_ROW = 2
PAYER_BASIS_PHARMACY_COLS = {
    "key": "DG",
    "basis_label": "DH",
    "item_code": "DI",
    "item_name": "DJ",
    "classification": "DK",
    "case_presence_rate": "DL",
    "present_in_ip_pharmacy": "DM",
    "present_in_ot_pharmacy": "DN",
    "ot_quantity_typical_cleaned": "DO",
    "ot_amount_typical_cleaned": "DP",
    "overall_amount_typical_cleaned": "DQ",
    "basis_name_key": "DR",
}
ORG_TARIFF_REFERENCE_START_ROW = 2
ORG_TARIFF_REFERENCE_COLS = {
    "payor_bucket": "DT",
    "organization_cd": "DU",
    "organization_name": "DV",
    "organization_label": "DW",
    "tariff_code": "DX",
    "tariff_name": "DY",
    "case_count": "DZ",
}
TARIFF_RATE_MATRIX_START_ROW = 2
TARIFF_RATE_MATRIX_COLS = {
    "matrix_key": "EB",
    "tariff_code": "EC",
    "tariff_name": "ED",
    "item_code": "EE",
    "item_name": "EF",
    "general": "EG",
    "twin": "EH",
    "single": "EI",
    "icu": "EJ",
}
TARIFF_OT_SLOT_MATRIX_START_ROW = 2
TARIFF_OT_SLOT_MATRIX_COLS = {
    "matrix_key": "EL",
    "tariff_code": "EM",
    "tariff_name": "EN",
    "ot_slot_hours": "EO",
    "ot_mode": "EP",
    "item_code": "EQ",
    "item_name": "ER",
    "general": "ES",
    "twin": "ET",
    "single": "EU",
    "icu": "EV",
}
INSURANCE_POLICY_START_ROW = 2
INSURANCE_POLICY_COLS = {
    "item_code": "EX",
    "policy_scope": "EY",
    "exclude_from_insurance_estimate": "EZ",
    "note": "FA",
}
PF_PAYOR_SUMMARY_START_ROW = 2
PF_PAYOR_SUMMARY_COLS = {
    "payor_bucket": "FC",
    "admission_count": "FD",
    "pf_collectible_historical_total_p25": "FE",
    "pf_collectible_historical_total_p50": "FF",
    "pf_collectible_historical_total_p75": "FG",
    "pf_named_total_p25": "FH",
    "pf_named_total_p50": "FI",
    "pf_named_total_p75": "FJ",
    "pf_general_needed_total_p25": "FK",
    "pf_general_needed_total_p50": "FL",
    "pf_general_needed_total_p75": "FM",
    "surgeon_named_total_p25": "FN",
    "surgeon_named_total_p50": "FO",
    "surgeon_named_total_p75": "FP",
    "assistant_surgeon_named_total_p25": "FQ",
    "assistant_surgeon_named_total_p50": "FR",
    "assistant_surgeon_named_total_p75": "FS",
    "anesthetist_named_total_p25": "FT",
    "anesthetist_named_total_p50": "FU",
    "anesthetist_named_total_p75": "FV",
    "assistant_anesthetist_named_total_p25": "FW",
    "assistant_anesthetist_named_total_p50": "FX",
    "assistant_anesthetist_named_total_p75": "FY",
    "consultant_or_physician_named_total_p25": "FZ",
    "consultant_or_physician_named_total_p50": "GA",
    "consultant_or_physician_named_total_p75": "GB",
    "dominant_pf_shape": "GC",
}
PAYER_BASIS_RESOLUTION_START_ROW = 2
PAYER_BASIS_RESOLUTION_COLS = {
    "lookup_key": "GD",
    "component": "GE",
    "target_payor_bucket": "GF",
    "basis": "GG",
    "case_count": "GH",
    "anchor_p25": "GI",
    "anchor_p50": "GJ",
    "anchor_p75": "GK",
    "variability_score": "GL",
    "spread_vs_insurance_all": "GM",
    "spread_vs_all_payers": "GN",
    "recommended_status": "GO",
    "selected_basis": "GP",
    "selected_case_count": "GQ",
    "confidence": "GR",
    "selection_reason": "GS",
}
ACTUAL_BASIS_METRIC_START_ROW = 2
ACTUAL_BASIS_METRIC_COLS = {
    "lookup_key": "HA",
    "basis_label": "HB",
    "field_key": "HC",
    "field_label": "HD",
    "min": "HE",
    "max": "HF",
    "average": "HG",
    "p25": "HH",
    "p50": "HI",
    "p75": "HJ",
}

IMPLANT_FAMILY_ORDER = [
    "Femoral Component",
    "Tibial Insert / Bearing",
    "Bone Cement",
    "Tibial Baseplate",
    "Stem / Extension",
    "Screw",
    "Pin",
]

IMPLANTS_SHEET_COLUMNS = [
    "implant_family",
    "implant_family_distinct_ip_count",
    "implant_family_presence_rate",
    "implant_family_quantity_p25",
    "implant_family_quantity_p50",
    "implant_family_quantity_p75",
    "implant_family_rate_p25",
    "implant_family_rate_p50",
    "implant_family_rate_p75",
    "brand_family",
    "brand_distinct_ip_count",
    "brand_presence_rate_within_implant_family",
    "brand_rate_p25",
    "brand_rate_p50",
    "brand_rate_p75",
    "item_code",
    "item_name",
    "item_quantity_p75",
    "item_rate_p25",
    "item_distinct_ip_count",
    "item_presence_rate",
    "item_quantity_p25",
    "item_quantity_p50",
    "typical_rate_p50",
    "item_rate_p75",
]


@dataclass
class RateRow:
    item_name: str
    general: float | None
    twin: float | None
    single: float | None
    icu: float | None


@dataclass
class OtSlotRateRow:
    tariff_code: str
    ot_slot_hours: float
    ot_mode: str
    item_code: str
    item_name: str
    general: float | None
    twin: float | None
    single: float | None
    icu: float | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Rebuild the Robotic TKR unilateral right FC estimate builder workbook."
    )
    parser.add_argument("--builder-input-pack-json", type=Path, default=None)
    parser.add_argument("--template-workbook", type=Path, default=DEFAULT_TEMPLATE_WORKBOOK)
    parser.add_argument("--quartiles-json", type=Path, default=DEFAULT_QUARTILES_JSON)
    parser.add_argument("--per-ip-buckets", type=Path, default=DEFAULT_PER_IP_BUCKETS)
    parser.add_argument("--pharmacy-template", type=Path, default=DEFAULT_PHARMACY_TEMPLATE)
    parser.add_argument("--implant-hierarchy", type=Path, default=DEFAULT_IMPLANT_HIERARCHY)
    parser.add_argument("--implant-detail", type=Path, default=DEFAULT_IMPLANT_DETAIL)
    parser.add_argument("--services-template", type=Path, default=DEFAULT_SERVICES_TEMPLATE)
    parser.add_argument("--cleaned-services-output", type=Path, default=DEFAULT_CLEANED_SERVICES_OUTPUT)
    parser.add_argument("--service-line-count-metrics", type=Path, default=DEFAULT_SERVICE_LINE_COUNT_METRICS)
    parser.add_argument("--ip-pharmacy-per-day-metrics", type=Path, default=DEFAULT_IP_PHARMACY_PER_DAY_METRICS)
    parser.add_argument("--rate-csv", type=Path, default=DEFAULT_RATE_CSV)
    parser.add_argument("--ot-slot-rate-csv", type=Path, default=DEFAULT_OT_SLOT_RATE_CSV)
    parser.add_argument("--cath-lab-metrics-json", type=Path, default=DEFAULT_CATH_LAB_METRICS_JSON)
    parser.add_argument("--cath-lab-slot-rate-csv", type=Path, default=DEFAULT_CATH_LAB_SLOT_RATE_CSV)
    parser.add_argument("--payer-basis-summary-json", type=Path, default=DEFAULT_PAYER_BASIS_SUMMARY_JSON)
    parser.add_argument("--payer-basis-service-metrics-csv", type=Path, default=DEFAULT_PAYER_BASIS_SERVICE_METRICS_CSV)
    parser.add_argument("--payer-basis-pharmacy-metrics-csv", type=Path, default=DEFAULT_PAYER_BASIS_PHARMACY_METRICS_CSV)
    parser.add_argument("--org-tariff-reference-csv", type=Path, default=DEFAULT_ORG_TARIFF_REFERENCE_CSV)
    parser.add_argument("--tariff-rate-matrix-csv", type=Path, default=DEFAULT_TARIFF_RATE_MATRIX_CSV)
    parser.add_argument("--tariff-ot-slot-rate-matrix-csv", type=Path, default=DEFAULT_TARIFF_OT_SLOT_RATE_MATRIX_CSV)
    parser.add_argument("--insurance-policy-csv", type=Path, default=DEFAULT_INSURANCE_POLICY_CSV)
    parser.add_argument("--grouping-gap-summary-csv", type=Path, default=DEFAULT_GROUPING_GAP_SUMMARY_CSV)
    parser.add_argument("--grouping-gap-child-detail-csv", type=Path, default=DEFAULT_GROUPING_GAP_CHILD_DETAIL_CSV)
    parser.add_argument("--payer-basis-resolution-csv", type=Path, default=DEFAULT_PAYER_BASIS_RESOLUTION_CSV)
    parser.add_argument("--pf-payor-summary-csv", type=Path)
    parser.add_argument("--pf-shape-review-json", type=Path)
    parser.add_argument("--pf-modeled-vs-actual-csv", type=Path)
    parser.add_argument("--payor-label", default=SHEET1_PAYOR_TEXT)
    parser.add_argument("--sheet1-template-name", default=DEFAULT_SHEET1_TEMPLATE_NAME)
    parser.add_argument("--sheet1-los-guidance", default=DEFAULT_SHEET1_LOS_GUIDANCE)
    parser.add_argument("--sheet1-management-type", default=DEFAULT_SHEET1_MANAGEMENT_TYPE)
    parser.add_argument("--procedure-code", default="OTI0098")
    parser.add_argument("--procedure-label", default="ROBO (TKR) - UNILATERAL")
    parser.add_argument("--include-procedure-row", choices=("yes", "no"), default="yes")
    parser.add_argument("--robotic-default-mode", choices=("auto", "yes", "no"), default="auto")
    parser.add_argument("--robotic-presence-threshold", type=float, default=90.0)
    parser.add_argument("--surgeon-fee", type=float, default=DEFAULT_PROFESSIONAL_FEES[33])
    parser.add_argument("--assistant-surgeon-fee", type=float, default=DEFAULT_PROFESSIONAL_FEES[34])
    parser.add_argument("--anesthetist-fee", type=float, default=DEFAULT_PROFESSIONAL_FEES[35])
    parser.add_argument("--assistant-anesthetist-fee", type=float, default=DEFAULT_PROFESSIONAL_FEES[36])
    parser.add_argument(
        "--ot-consumable-shortlist-mode",
        choices=("fixed", "derived"),
        default="derived",
    )
    parser.add_argument("--ot-consumable-shortlist-count", type=int, default=10)
    parser.add_argument(
        "--services-selection-mode",
        choices=("fixed", "derived"),
        default="derived",
    )
    parser.add_argument("--services-selection-count", type=int, default=200)
    parser.add_argument("--service-rate-field", default="rate_cash_p50")
    parser.add_argument("--validation-output-json", type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def load_builder_runtime_inputs(args: argparse.Namespace) -> dict[str, Any]:
    if args.builder_input_pack_json:
        payload = load_json(args.builder_input_pack_json)
        historical_actual_metrics = payload.get("historical_actual_metrics") or {}
        implant_template_rows = payload.get("implant_template_rows") or {}
        payer_basis_metrics = payload.get("payer_basis_metrics") or {}
        bucket_quartiles = {
            key: (
                as_float((metrics or {}).get("p25")),
                as_float((metrics or {}).get("p50")),
                as_float((metrics or {}).get("p75")),
            )
            for key, metrics in (historical_actual_metrics.get("bucket_quartiles") or {}).items()
        }
        return {
            "quartiles_json": payload.get("historical_driver_metrics") or {},
            "bucket_quartiles": bucket_quartiles,
            "ip_pharmacy_per_day_metrics": historical_actual_metrics.get("ip_pharmacy_per_day_metrics") or {},
            "payer_basis_summary": payer_basis_metrics.get("summary") or {},
            "payer_basis_service_rows": payer_basis_metrics.get("service_rows") or [],
            "payer_basis_pharmacy_rows": payer_basis_metrics.get("pharmacy_rows") or [],
            "service_rows": payload.get("service_template_rows") or [],
            "pharmacy_rows": payload.get("pharmacy_template_rows") or [],
            "per_ip_rows": historical_actual_metrics.get("per_ip_compat_rows") or [],
            "ip_actual_rows": payload.get("ip_actual_benchmark_rows") or [],
            "implant_rows": implant_template_rows.get("hierarchy_rows") or [],
            "implant_detail_rows": implant_template_rows.get("detail_rows") or [],
        }

    quartiles_json = load_json(args.quartiles_json)
    bucket_quartiles = compute_bucket_quartiles(args.per_ip_buckets)
    ip_pharmacy_per_day_metrics = load_json(args.ip_pharmacy_per_day_metrics)
    service_rows = load_csv_rows(args.services_template)
    pharmacy_rows = load_csv_rows(args.pharmacy_template)
    per_ip_rows = load_csv_rows(args.per_ip_buckets)
    for row in per_ip_rows:
        los_days = as_float(row.get("los_days"))
        ip_drugs_amount = as_float(row.get("total_amount_ip_drugs_medicines_ivs_nutrition_products"))
        ip_consumables_amount = as_float(row.get("total_amount_ip_treatment_supplies"))
        row["ip_drugs_per_los_day"] = (ip_drugs_amount / los_days) if los_days > 0 else ""
        row["ip_consumables_per_los_day"] = (ip_consumables_amount / los_days) if los_days > 0 else ""
    ip_actual_rows = build_ip_fc_actual_rows(per_ip_rows, service_rows)
    if should_synthesize_local_basis_artifacts(args):
        payer_basis_summary = synthesize_local_cash_basis_summary(
            quartiles_json=quartiles_json,
            bucket_quartiles=bucket_quartiles,
            ip_pharmacy_per_day_metrics=ip_pharmacy_per_day_metrics,
            service_line_count_metrics=load_json(args.service_line_count_metrics),
            cath_lab_metrics=load_json(args.cath_lab_metrics_json),
            ip_actual_rows=ip_actual_rows,
        )
        payer_basis_service_rows = synthesize_local_basis_service_rows(
            filter_out_cath_lab_slot_rows(service_rows)
        )
        payer_basis_pharmacy_rows = synthesize_local_basis_pharmacy_rows(pharmacy_rows)
    else:
        payer_basis_summary = load_json(args.payer_basis_summary_json)
        payer_basis_service_rows = filter_out_cath_lab_slot_rows(load_csv_rows(args.payer_basis_service_metrics_csv))
        payer_basis_pharmacy_rows = load_csv_rows(args.payer_basis_pharmacy_metrics_csv)
    return {
        "quartiles_json": quartiles_json,
        "bucket_quartiles": bucket_quartiles,
        "ip_pharmacy_per_day_metrics": ip_pharmacy_per_day_metrics,
        "payer_basis_summary": payer_basis_summary,
        "payer_basis_service_rows": payer_basis_service_rows,
        "payer_basis_pharmacy_rows": payer_basis_pharmacy_rows,
        "service_rows": service_rows,
        "pharmacy_rows": pharmacy_rows,
        "per_ip_rows": per_ip_rows,
        "ip_actual_rows": ip_actual_rows,
        "implant_rows": load_implant_rows(args.implant_hierarchy),
        "implant_detail_rows": load_csv_rows(args.implant_detail),
    }


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_code(value: Any) -> str:
    return normalize_text(value).replace(" ", "").upper()


def filter_out_cath_lab_slot_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [
        row
        for row in rows
        if not is_cath_lab_slot_service(
            code=row.get("item_code"),
            grouping=row.get("grouping"),
            service_name=row.get("item_name"),
        )
    ]


def should_synthesize_local_basis_artifacts(args: argparse.Namespace) -> bool:
    if getattr(args, "builder_input_pack_json", None):
        return False
    return (
        args.payer_basis_summary_json == DEFAULT_PAYER_BASIS_SUMMARY_JSON
        and args.services_template != DEFAULT_SERVICES_TEMPLATE
    )


def synthesize_local_cash_basis_summary(
    *,
    quartiles_json: dict[str, Any],
    bucket_quartiles: dict[str, tuple[float, float, float]],
    ip_pharmacy_per_day_metrics: dict[str, Any],
    service_line_count_metrics: dict[str, Any],
    cath_lab_metrics: dict[str, Any],
    ip_actual_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    payor_counts = Counter(normalize_text(row.get("payor_bucket")) or "Cash" for row in ip_actual_rows)
    shared_metrics = {
        "cohort_size": len(ip_actual_rows),
        "payor_counts": dict(payor_counts),
        "clinical_drivers": quartiles_json.get("metrics") or {},
        "bucket_quartiles": {
            bucket: {"p25": p25, "p50": p50, "p75": p75}
            for bucket, (p25, p50, p75) in bucket_quartiles.items()
        },
        "ip_pharmacy_per_day": ip_pharmacy_per_day_metrics or {},
        "service_line_count": (service_line_count_metrics or {}).get("cleaned_distinct_service_line_count") or {},
        "cath_lab_amount": (cath_lab_metrics or {}).get("amount_metrics") or (cath_lab_metrics or {}).get("metrics") or {},
    }
    return {
        "basis_order": ["Cash", "All Payers"],
        "basis_metrics": {
            "Cash": shared_metrics,
            "All Payers": shared_metrics,
        },
    }


def synthesize_local_basis_service_rows(service_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for basis_label in ["Cash", "All Payers"]:
        for row in service_rows:
            item_code = normalize_code(row.get("item_code"))
            row_copy = dict(row)
            row_copy["basis_label"] = basis_label
            row_copy["basis_item_key"] = f"{basis_label}|{item_code}"
            output.append(row_copy)
    return output


def synthesize_local_basis_pharmacy_rows(pharmacy_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for basis_label in ["Cash", "All Payers"]:
        for row in pharmacy_rows:
            item_code = normalize_code(row.get("item_code"))
            item_name = normalize_text(row.get("item_name"))
            row_copy = dict(row)
            row_copy["basis_label"] = basis_label
            row_copy["basis_item_key"] = f"{basis_label}|{item_code}|{item_name}"
            output.append(row_copy)
    return output


def as_float(value: Any) -> float:
    text = normalize_text(value)
    if not text:
        return 0.0
    return float(text)


def maybe_float(value: Any) -> float | None:
    text = normalize_text(value)
    if not text:
        return None
    return float(text)


def inclusive_quartiles(values: list[float]) -> tuple[float, float, float]:
    ordered = sorted(values)
    if not ordered:
        return 0.0, 0.0, 0.0
    if len(ordered) == 1:
        return ordered[0], ordered[0], ordered[0]
    q1, q2, q3 = statistics.quantiles(ordered, n=4, method="inclusive")
    return float(q1), float(q2), float(q3)


def round_display_quantity(value: float) -> float:
    whole = math.floor(value)
    fraction = value - whole
    if fraction > 0.3:
        return float(whole + 1)
    return float(whole)


ROOM_PRECEDENCE = [
    "MICU",
    "SICU",
    "ICCU",
    "ICU",
    "HDU",
    "SINGLE",
    "DELUXE",
    "TWIN SHARING",
    "GENERAL WARD",
    "DAYCARE",
]

COMMERCIAL_ROOM_PRECEDENCE = [
    "DELUXE",
    "SINGLE",
    "TWIN SHARING",
    "GENERAL WARD",
    "DAYCARE",
]


THIN_GREY = Side(style="thin", color="D9D9D9")
MEDIUM_BLUE = Side(style="medium", color="4F81BD")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
OT_FILL = PatternFill("solid", fgColor="FCE4D6")
IMPLANT_FILL = PatternFill("solid", fgColor="E2F0D9")
RESULT_FILL = PatternFill("solid", fgColor="FFF2CC")
SELECTION_FILL = PatternFill("solid", fgColor="F4F6F8")
FORMULA_GREEN_FILL = PatternFill("solid", fgColor="EAF4EA")
FORMULA_BLUE_FILL = PatternFill("solid", fgColor="EAF1FB")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
SPACER_FILL = PatternFill("solid", fgColor="F7F7F7")
SECTION_BORDER = Border(left=MEDIUM_BLUE, right=MEDIUM_BLUE, top=MEDIUM_BLUE, bottom=MEDIUM_BLUE)


def apply_cell_style(cell, *, fill=None, bold=False, font_color=None, align="center", wrap=False, border=True):
    if fill is not None:
        cell.fill = fill
    cell.font = Font(bold=bold, color=font_color or "000000")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)


def style_range(ws, cell_range: str, *, fill=None, bold=False, font_color=None, align="center", wrap=False, border=True):
    for row in ws[cell_range]:
        for cell in row:
            apply_cell_style(cell, fill=fill, bold=bold, font_color=font_color, align=align, wrap=wrap, border=border)


def set_number_format(ws, cell_range: str, fmt: str) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = fmt


def style_row(ws, row_number: int, start_col: int, end_col: int, *, fill=None, bold=False, font_color=None, align="center", wrap=False, border=True):
    for col_idx in range(start_col, end_col + 1):
        apply_cell_style(
            ws.cell(row=row_number, column=col_idx),
            fill=fill,
            bold=bold,
            font_color=font_color,
            align=align,
            wrap=wrap,
            border=border,
        )


def load_json(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def load_csv_rows_if_exists(path: Path | None) -> list[dict[str, str]]:
    if not path or not path.exists():
        return []
    return load_csv_rows(path)


def build_reference_basis_lookup_formula(target_col: str, basis_ref: str = "Builder!G6") -> str:
    return (
        f'=IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${target_col}${PAYER_BASIS_SUMMARY_START_ROW}:'
        f'${target_col}$500, MATCH({basis_ref}, \'{SHEET_REFERENCE}\'!${PAYER_BASIS_SUMMARY_COLS["basis_label"]}${PAYER_BASIS_SUMMARY_START_ROW}:'
        f'${PAYER_BASIS_SUMMARY_COLS["basis_label"]}$500, 0)), 0)'
    )


def build_resolution_lookup_formula(target_col: str, component_ref: str, target_payor_ref: str, fallback: str = '""') -> str:
    key_col = PAYER_BASIS_RESOLUTION_COLS["lookup_key"]
    return (
        f'=IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${target_col}${PAYER_BASIS_RESOLUTION_START_ROW}:${target_col}$800, '
        f'MATCH({component_ref}&"|"&{target_payor_ref}, '
        f'\'{SHEET_REFERENCE}\'!${key_col}${PAYER_BASIS_RESOLUTION_START_ROW}:${key_col}$800, 0)), {fallback})'
    )


def build_pf_payor_lookup_formula(target_col: str, payor_ref: str, fallback: str = "0") -> str:
    return (
        f'=IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${target_col}${PF_PAYOR_SUMMARY_START_ROW}:${target_col}$500, '
        f'MATCH({payor_ref}, '
        f'\'{SHEET_REFERENCE}\'!${PF_PAYOR_SUMMARY_COLS["payor_bucket"]}${PF_PAYOR_SUMMARY_START_ROW}:'
        f'${PF_PAYOR_SUMMARY_COLS["payor_bucket"]}$500, 0)), {fallback})'
    )


def resolve_ot_consumables_band_value(p25: float, p50: float, p75: float, selected_expected_share: float) -> float:
    if selected_expected_share <= OT_CONSUMABLES_P25_SHARE_THRESHOLD:
        return p25
    if selected_expected_share <= OT_CONSUMABLES_P50_SHARE_THRESHOLD:
        return p50
    return p75


def build_ot_consumables_piecewise_formula(
    *,
    p25_ref: str,
    p50_ref: str,
    p75_ref: str,
    selected_flag_range: str,
    expected_contribution_range: str,
) -> str:
    selected_share = (
        f'IFERROR(SUMIF({selected_flag_range},"{SELECTION_INCLUDE}",{expected_contribution_range})/'
        f"SUM({expected_contribution_range}),0)"
    )
    return (
        f"=IF({selected_share}<={OT_CONSUMABLES_P25_SHARE_THRESHOLD},{p25_ref},"
        f'IF({selected_share}<={OT_CONSUMABLES_P50_SHARE_THRESHOLD},{p50_ref},{p75_ref}))'
    )


def build_component_mix_total_actual_formula(point_key: str, *, pf_mode: str = "calculated") -> str:
    metric_col = ACTUAL_BASIS_METRIC_COLS[point_key]
    service_parts = [
        build_actual_basis_metric_lookup_formula(metric_col, "room_charges", "Builder!G6").lstrip("="),
        build_actual_basis_metric_lookup_formula(metric_col, "investigations", "Builder!G6").lstrip("="),
        build_actual_basis_metric_lookup_formula(metric_col, "procedure_ot_charges", "Builder!G6").lstrip("="),
        build_actual_basis_metric_lookup_formula(metric_col, "bedside_services", "Builder!G6").lstrip("="),
    ]
    pharmacy_parts = [
        build_actual_basis_metric_lookup_formula(metric_col, "pharmacy_total", "Builder!G5").lstrip("="),
        build_actual_basis_metric_lookup_formula(metric_col, "drug_administration_charges", "Builder!G5").lstrip("="),
    ]
    if pf_mode == "historic":
        pf_target_col = {
            "p25": PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p25"],
            "p50": PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p50"],
            "p75": PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p75"],
        }[point_key]
        pf_part = build_pf_payor_lookup_formula(pf_target_col, "Builder!G7").lstrip("=")
    else:
        pf_part = build_actual_basis_metric_lookup_formula(metric_col, "professional_fees", "Builder!G7").lstrip("=")
    return f'=({" + ".join(service_parts + pharmacy_parts + [pf_part])})'


def build_actual_basis_metric_lookup_formula(stat_col: str, field_key: str, basis_ref: str) -> str:
    key_col = ACTUAL_BASIS_METRIC_COLS["lookup_key"]
    return (
        f'=IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${stat_col}${ACTUAL_BASIS_METRIC_START_ROW}:${stat_col}$1000, '
        f'MATCH({basis_ref}&"|"&"{field_key}", '
        f'\'{SHEET_REFERENCE}\'!${key_col}${ACTUAL_BASIS_METRIC_START_ROW}:${key_col}$1000, 0)), 0)'
    )


def build_reference_service_lookup_formula(item_code_ref: str, target_col: str, fallback: str = "0", basis_ref: str = "Builder!G6") -> str:
    key_col = PAYER_BASIS_SERVICE_COLS["key"]
    return (
        f'=IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${target_col}${PAYER_BASIS_SERVICE_START_ROW}:'
        f'${target_col}$4000, MATCH({basis_ref}&"|"&{item_code_ref}, '
        f'\'{SHEET_REFERENCE}\'!${key_col}${PAYER_BASIS_SERVICE_START_ROW}:${key_col}$4000, 0)), {fallback})'
    )


def build_reference_pharmacy_lookup_formula(
    item_name_ref: str,
    item_code_ref: str,
    target_col: str,
    fallback: str = "0",
    basis_ref: str = "Builder!G5",
) -> str:
    key_col = PAYER_BASIS_PHARMACY_COLS["key"]
    return (
        f'=IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${target_col}${PAYER_BASIS_PHARMACY_START_ROW}:'
        f'${target_col}$6000, MATCH({basis_ref}&"|"&{item_code_ref}&"|"&{item_name_ref}, '
        f'\'{SHEET_REFERENCE}\'!${key_col}${PAYER_BASIS_PHARMACY_START_ROW}:${key_col}$6000, 0)), {fallback})'
    )


def build_reference_pharmacy_name_lookup_formula(
    item_name_ref: str,
    target_col: str,
    fallback: str = "0",
    basis_ref: str = "Builder!G5",
) -> str:
    key_col = PAYER_BASIS_PHARMACY_COLS["basis_name_key"]
    return (
        f'=IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${target_col}${PAYER_BASIS_PHARMACY_START_ROW}:'
        f'${target_col}$6000, MATCH({basis_ref}&"|"&{item_name_ref}, '
        f'\'{SHEET_REFERENCE}\'!${key_col}${PAYER_BASIS_PHARMACY_START_ROW}:${key_col}$6000, 0)), {fallback})'
    )


def build_org_reference_lookup_formula(org_code_ref: str, target_col: str, fallback: str = '""') -> str:
    key_col = ORG_TARIFF_REFERENCE_COLS["organization_cd"]
    return (
        f'=IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${target_col}${ORG_TARIFF_REFERENCE_START_ROW}:'
        f'${target_col}$1000, MATCH({org_code_ref}, '
        f'\'{SHEET_REFERENCE}\'!${key_col}${ORG_TARIFF_REFERENCE_START_ROW}:${key_col}$1000, 0)), {fallback})'
    )


def build_tariff_rate_lookup_formula(code_ref: str, target_col: str, fallback: str = "0", tariff_ref: str = "Builder!E5") -> str:
    key_col = TARIFF_RATE_MATRIX_COLS["matrix_key"]
    return (
        f'=0+IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${target_col}${TARIFF_RATE_MATRIX_START_ROW}:'
        f'${target_col}$8000, MATCH({tariff_ref}&"|"&{code_ref}, '
        f'\'{SHEET_REFERENCE}\'!${key_col}${TARIFF_RATE_MATRIX_START_ROW}:${key_col}$8000, 0)), {fallback})'
    )


def build_is_insurance_mode_formula(pricing_mode_ref: str = "Builder!E2") -> str:
    return f'{pricing_mode_ref}="Insurance / Org Tariff"'


def build_insurance_policy_lookup_formula(code_ref: str, target_col: str, fallback: str = '""') -> str:
    key_col = INSURANCE_POLICY_COLS["item_code"]
    return (
        f'=IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${target_col}${INSURANCE_POLICY_START_ROW}:'
        f'${target_col}$1000, MATCH({code_ref}, '
        f'\'{SHEET_REFERENCE}\'!${key_col}${INSURANCE_POLICY_START_ROW}:${key_col}$1000, 0)), {fallback})'
    )


def build_is_insurance_excluded_formula(code_ref: str) -> str:
    exclusion_lookup = build_insurance_policy_lookup_formula(
        code_ref,
        INSURANCE_POLICY_COLS["exclude_from_insurance_estimate"],
        '"No"',
    ).lstrip("=")
    return f'AND({build_is_insurance_mode_formula()},{exclusion_lookup}="Yes")'


def wrap_insurance_exclusion(code_ref: str, expression: str) -> str:
    inner = expression[1:] if expression.startswith("=") else expression
    return f'=IF({build_is_insurance_excluded_formula(code_ref)},0,{inner})'


def load_rate_lookup(path: Path) -> dict[str, RateRow]:
    lookup: dict[str, RateRow] = {}
    for row in load_csv_rows(path):
        code = normalize_code(row.get("item_code"))
        lookup[code] = RateRow(
            item_name=normalize_text(row.get("item_name")),
            general=maybe_float(row.get("general")),
            twin=maybe_float(row.get("twin")),
            single=maybe_float(row.get("single")),
            icu=maybe_float(row.get("icu")),
        )
    return lookup


def load_ot_slot_rate_rows(path: Path) -> list[OtSlotRateRow]:
    rows: list[OtSlotRateRow] = []
    for row in load_csv_rows(path):
        slot_hours = maybe_float(row.get("ot_slot_hours"))
        if slot_hours is None:
            continue
        rows.append(
            OtSlotRateRow(
                tariff_code=normalize_text(row.get("tariff_code")) or "TR1",
                ot_slot_hours=slot_hours,
                ot_mode=normalize_text(row.get("ot_mode")).lower() or "normal",
                item_code=normalize_code(row.get("item_code")),
                item_name=normalize_text(row.get("item_name")),
                general=maybe_float(row.get("general")),
                twin=maybe_float(row.get("twin")),
                single=maybe_float(row.get("single")),
                icu=maybe_float(row.get("icu")),
            )
        )
    rows.sort(key=lambda row: (row.ot_slot_hours, 0 if row.ot_mode == "normal" else 1, row.item_code))
    return rows


def load_service_lookup(path: Path) -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    for row in load_csv_rows(path):
        code = normalize_code(row.get("item_code"))
        if code and code not in lookup:
            lookup[code] = row
    return lookup


def build_service_bucket_lookups(service_rows: list[dict[str, str]]) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    by_code: dict[str, dict[str, str]] = {}
    by_name: dict[str, dict[str, str]] = {}
    for row in service_rows:
        code = normalize_code(row.get("item_code"))
        name = normalize_text(row.get("item_name"))
        if code and code not in by_code:
            by_code[code] = row
        if name and name not in by_name:
            by_name[name] = row
    return by_code, by_name


def connect_db() -> psycopg.Connection:
    return psycopg.connect(os.getenv("SUPABASE_DB_URL", DEFAULT_DB_URL))


def service_rule_text(service_row: dict[str, str] | None) -> str:
    if not service_row:
        return "Template"
    presence_rate = as_float(service_row.get("case_presence_rate"))
    typical_amount = as_float(service_row.get("amount_cash_typical"))
    if presence_rate > 90.0:
        return "Historic Presence Rate > 90%"
    if presence_rate >= 75.0 and typical_amount <= SERVICES_SIGNIFICANCE_AMOUNT_THRESHOLD:
        return "Historic Presence Rate >= 75% & Typical Amount <= Rs 1,000"
    return "Template Override"


def is_food_or_beverage_row(item: dict[str, Any]) -> bool:
    signals = [
        normalize_text(item.get("department_name")),
        normalize_text(item.get("service_group_name")),
        normalize_text(item.get("service_name")),
    ]
    upper = [signal.upper() for signal in signals if signal]
    keywords = ["FOOD", "BEVERAGE", "FOOD AND BEVERAGES", "TEA", "COFFEE", "JUICE", "SOUP"]
    return any(keyword in signal for signal in upper for keyword in keywords)


def map_actual_service_bucket(
    service_code: str,
    service_name: str,
    raw_bucket: str,
    *,
    service_type: str = "",
) -> str:
    bucket = normalize_text(raw_bucket)
    code = normalize_code(service_code)
    name = normalize_text(service_name)
    service_type_normalized = normalize_text(service_type)
    if code == "MSC10" or name == "MEDICAL RECORDS-1 DAY":
        return "Other Services"
    if service_type_normalized in {"Professional", "Consultations"}:
        return "Professional Fees"
    if bucket == "Room Charges - Remove":
        return "Room Charges"
    if bucket == "Administrative Charges - Remove":
        return "Procedure / OT Charges"
    if bucket in {"Doctors & Professionals - Name Wise - Remove", "Doctors & Professionals - General - Remove"}:
        return "Professional Fees"
    if bucket in {"Anesthetist - Name Wise - Remove", "Anesthetist - General - Remove"}:
        return "Professional Fees"
    if bucket == "OT - Remove":
        return "Procedure / OT Charges"
    if bucket == "N/A - Remove":
        return "Procedure / OT Charges"
    if code in {"ROM5189", "ROM0093", "ROM5009", "ROM0001", "ROM0024", "ROM0036", "ICC0001", "ICC0002", "HSP5013", "EME0019"}:
        return "Room Charges"
    if name in {"NURSING CHARGES", "DMO CHARGES", "WARD CONSUMABLES", "MONITOR PER DAY"}:
        return "Room Charges"
    return bucket_label(bucket or "Procedure / OT Charges")


def collapse_actual_display_bucket(bucket: str) -> str:
    normalized = normalize_text(bucket)
    if normalized in {
        "Procedure / OT Charges",
        "Procedure Charges",
        "OT",
        "Physiotherapy",
        "Other Services",
        "Administrative Charges",
        "Cath Lab",
    }:
        return "Procedure / OT Charges"
    if normalized in {
        "Professional Fees",
        "Professional Fee",
        "Doctors & Professionals - General - Needed",
        "Doctors & Professionals - Name Wise - Needed",
        "Doctors & Professionals - General - Remove",
        "Doctors & Professionals - Name Wise - Remove",
        "Anesthetist - General - Needed",
        "Anesthetist - Name Wise - Needed",
        "Anesthetist - General - Remove",
        "Anesthetist - Name Wise - Remove",
    }:
        return "Professional Fees"
    if normalized == "Bedside services":
        return "Bedside Services"
    return normalized or "Procedure / OT Charges"


def count_distinct_non_food_service_lines(services_json: list[dict[str, Any]]) -> int:
    seen: set[str] = set()
    for item in services_json or []:
        if is_food_or_beverage_row(item):
            continue
        code = normalize_code(item.get("service_code"))
        name = normalize_text(item.get("service_name"))
        key = code or name
        if key:
            seen.add(key)
    return len(seen)


def infer_room_category(service_name: str | None, ward_name: str | None) -> str | None:
    value = f"{service_name or ''} {ward_name or ''}".upper().strip()
    if any(token in value for token in ["MICU", "SICU", "PICU", "NICU", "ICCU", "ICU"]):
        return "icu"
    if "HDU" in value:
        return "hdu"
    if "SINGLE" in value:
        return "single"
    if "TWIN" in value:
        return "twin"
    if "GENERAL" in value:
        return "general"
    if "DAY CARE" in value or "DAYCARE" in value:
        return "daycare"
    if "DELUXE" in value:
        return "deluxe"
    return None


def derive_room_labels(services_json: list[dict[str, Any]]) -> list[str]:
    raw_labels: set[str] = set()
    for item in services_json or []:
        ward_name = normalize_text(item.get("ward_name")).upper()
        if ward_name:
            raw_labels.add(ward_name)
        inferred = infer_room_category(item.get("service_name"), item.get("ward_name"))
        if inferred == "icu":
            raw_labels.add("ICU")
        elif inferred == "hdu":
            raw_labels.add("HDU")
        elif inferred == "single":
            raw_labels.add("SINGLE")
        elif inferred == "twin":
            raw_labels.add("TWIN SHARING")
        elif inferred == "general":
            raw_labels.add("GENERAL WARD")
        elif inferred == "daycare":
            raw_labels.add("DAYCARE")
        elif inferred == "deluxe":
            raw_labels.add("DELUXE")
    return sorted(raw_labels, key=lambda label: (ROOM_PRECEDENCE.index(label) if label in ROOM_PRECEDENCE else len(ROOM_PRECEDENCE), label))


def derive_primary_room_category(services_json: list[dict[str, Any]]) -> str:
    labels = derive_room_labels(services_json)
    if not labels:
        return ""
    for preferred in ROOM_PRECEDENCE:
        if preferred in labels:
            return preferred
    return labels[0]


def derive_primary_commercial_room_category(services_json: list[dict[str, Any]]) -> str:
    labels = derive_room_labels(services_json)
    if not labels:
        return ""
    for preferred in COMMERCIAL_ROOM_PRECEDENCE:
        if preferred in labels:
            return preferred
    return ""


def derive_icu_unit_name(services_json: list[dict[str, Any]]) -> str:
    labels = derive_room_labels(services_json)
    for preferred in ["SICU", "MICU", "ICCU", "ICU", "HDU"]:
        if preferred in labels:
            return preferred
    return ""


def derive_icu_and_ward_days(services_json: list[dict[str, Any]], los_days: float) -> tuple[float, float]:
    if los_days <= 0:
        return 0.0, 0.0
    total_days = max(1, int(math.ceil(los_days)))
    critical_days_observed = 0
    ward_days_observed = 0
    daycare_days_observed = 0
    for item in services_json or []:
        if normalize_text(item.get("service_type")).upper() != "WARD CHARGES":
            continue
        room_category = infer_room_category(item.get("service_name"), item.get("ward_name"))
        try:
            quantity = max(float(item.get("quantity") or 0), 0)
        except (TypeError, ValueError):
            quantity = 0.0
        rounded_quantity = int(round(quantity))
        if room_category in {"icu", "hdu"}:
            critical_days_observed += rounded_quantity
        elif room_category in {"general", "single", "twin", "deluxe"}:
            ward_days_observed += rounded_quantity
        elif room_category == "daycare":
            daycare_days_observed += rounded_quantity
    icu_days = min(critical_days_observed, total_days)
    residual_days = max(total_days - icu_days - min(daycare_days_observed, max(total_days - icu_days, 0)), 0)
    ward_days = min(ward_days_observed, residual_days) if ward_days_observed > 0 else residual_days
    return float(icu_days), float(ward_days)


def summarize_numeric_series(values: list[float]) -> dict[str, float]:
    cleaned = [float(value) for value in values]
    if not cleaned:
        return {"min": 0.0, "max": 0.0, "average": 0.0, "p25": 0.0, "p50": 0.0, "p75": 0.0}
    p25, p50, p75 = inclusive_quartiles(cleaned)
    return {
        "min": min(cleaned),
        "max": max(cleaned),
        "average": float(statistics.mean(cleaned)),
        "p25": p25,
        "p50": p50,
        "p75": p75,
    }


def build_actual_basis_metric_rows(actual_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    basis_filters = [
        ("Cash", lambda row: normalize_text(row.get("payor_bucket")) == "Cash"),
        ("GIPSA Insurance", lambda row: normalize_text(row.get("payor_bucket")) == "GIPSA Insurance"),
        ("Non-GIPSA Insurance", lambda row: normalize_text(row.get("payor_bucket")) == "Non-GIPSA Insurance"),
        ("Corporate", lambda row: normalize_text(row.get("payor_bucket")) == "Corporate"),
        (
            "Insurance All",
            lambda row: normalize_text(row.get("payor_bucket")) in {"GIPSA Insurance", "Non-GIPSA Insurance"},
        ),
        ("All Payers", lambda row: True),
    ]
    rows: list[dict[str, Any]] = []
    for basis_label, predicate in basis_filters:
        basis_rows = [row for row in actual_rows if predicate(row)]
        for field_key, field_label in ACTUAL_IP_SUMMARY_FIELDS:
            stats = summarize_numeric_series([as_float(row.get(field_key)) for row in basis_rows])
            rows.append(
                {
                    "lookup_key": f"{basis_label}|{field_key}",
                    "basis_label": basis_label,
                    "field_key": field_key,
                    "field_label": field_label,
                    "min": stats["min"],
                    "max": stats["max"],
                    "average": stats["average"],
                    "p25": stats["p25"],
                    "p50": stats["p50"],
                    "p75": stats["p75"],
                }
            )
    return rows


def fetch_main_rows_for_admissions(admission_nos: list[str]) -> list[dict[str, Any]]:
    query = """
    select
        admission_no,
        patient_name,
        los_days,
        payor_bucket,
        organization_name,
        patient_type,
        organization_cd,
        surgical_medical,
        services_json,
        pharmacy_json
    from mart.main_table
    where admission_no = any(%s)
    order by admission_no
    """
    with connect_db() as conn:
        with conn.cursor() as cur:
            cur.execute(query, (admission_nos,))
            rows = cur.fetchall()
    return [
        {
            "admission_no": normalize_text(admission_no),
            "patient_name": normalize_text(patient_name),
            "los_days": float(los_days or 0),
            "payor_bucket": normalize_text(payor_bucket),
            "organization_name": normalize_text(organization_name),
            "patient_type": normalize_text(patient_type),
            "organization_cd": normalize_text(organization_cd),
            "surgical_medical": normalize_text(surgical_medical),
            "services_json": services_json or [],
            "pharmacy_json": pharmacy_json or {},
        }
        for admission_no, patient_name, los_days, payor_bucket, organization_name, patient_type, organization_cd, surgical_medical, services_json, pharmacy_json in rows
    ]


def build_ip_fc_actual_rows(
    per_ip_rows: list[dict[str, str]],
    service_rows: list[dict[str, str]],
) -> list[dict[str, Any]]:
    admissions = [normalize_text(row.get("admission_no")) for row in per_ip_rows if normalize_text(row.get("admission_no"))]
    main_rows = fetch_main_rows_for_admissions(admissions)
    main_by_admission = {row["admission_no"]: row for row in main_rows}
    service_by_code, service_by_name = build_service_bucket_lookups(service_rows)
    output_rows: list[dict[str, Any]] = []
    for per_ip in per_ip_rows:
        admission_no = normalize_text(per_ip.get("admission_no"))
        if not admission_no or admission_no not in main_by_admission:
            continue
        main_row = main_by_admission[admission_no]
        service_bucket_amounts = {
            "Room Charges": 0.0,
            "Investigations": 0.0,
            "Procedure / OT Charges": 0.0,
            "Bedside Services": 0.0,
            "Professional Fees": 0.0,
            "Pharmacy": 0.0,
        }
        services_total_ex_fnb = 0.0
        food_beverage_excluded = 0.0
        for item in main_row["services_json"] or []:
            amount = as_float(item.get("amount"))
            if is_food_or_beverage_row(item):
                food_beverage_excluded += amount
                continue
            services_total_ex_fnb += amount
            service_code = normalize_code(item.get("service_code"))
            service_name = normalize_text(item.get("service_name"))
            mapping = service_by_code.get(service_code) or service_by_name.get(service_name) or {}
            bucket = collapse_actual_display_bucket(
                map_actual_service_bucket(
                    service_code,
                    service_name,
                    mapping.get("fc_estimate_bucket", ""),
                    service_type=normalize_text(item.get("service_type")),
                )
            )
            service_bucket_amounts.setdefault(bucket, 0.0)
            service_bucket_amounts[bucket] += amount

        ip_drugs = as_float(per_ip.get("total_amount_ip_drugs_medicines_ivs_nutrition_products"))
        ip_supplies = as_float(per_ip.get("total_amount_ip_treatment_supplies"))
        ot_drugs = as_float(per_ip.get("total_amount_ot_drugs_medicines_ivs_nutrition_products"))
        ot_supplies = as_float(per_ip.get("total_amount_ot_treatment_supplies"))
        implants = as_float(per_ip.get("total_amount_implants"))
        pharmacy_total = ip_drugs + ip_supplies + ot_drugs + ot_supplies + implants
        drug_administration_charges = 0.125 * pharmacy_total
        los_days = as_float(per_ip.get("los_days")) or main_row["los_days"]
        pharmacy_returns_excluded = sum(
            as_float(item.get("return_amount"))
            for item in ((main_row["pharmacy_json"] if isinstance(main_row["pharmacy_json"], dict) else {}).get("returns", []) or [])
        )
        service_line_count = count_distinct_non_food_service_lines(main_row["services_json"])
        room_category = derive_primary_commercial_room_category(main_row["services_json"])
        icu_unit_name = derive_icu_unit_name(main_row["services_json"])
        icu_days, ward_days = derive_icu_and_ward_days(main_row["services_json"], los_days)
        room_charges_per_day = service_bucket_amounts.get("Room Charges", 0.0) / los_days if los_days > 0 else 0.0
        ip_drugs_per_day = ip_drugs / los_days if los_days > 0 else 0.0
        ip_consumables_per_day = ip_supplies / los_days if los_days > 0 else 0.0
        total_amount = services_total_ex_fnb + pharmacy_total + drug_administration_charges
        output_rows.append(
            {
                "admission_no": admission_no,
                "patient_name": normalize_text(per_ip.get("patient_name")) or main_row["patient_name"],
                "payor_bucket": main_row["payor_bucket"],
                "patient_type": main_row["patient_type"],
                "organization_name": main_row["organization_name"],
                "surgical_medical": main_row["surgical_medical"],
                "room_category": room_category,
                "icu_unit_name": icu_unit_name,
                "los_days": los_days,
                "icu_days": icu_days,
                "ward_days": ward_days,
                "ot_hours": as_float(per_ip.get("ot_hours")),
                "service_line_count": service_line_count,
                "room_charges": service_bucket_amounts.get("Room Charges", 0.0),
                "room_charges_per_day": room_charges_per_day,
                "investigations": service_bucket_amounts.get("Investigations", 0.0),
                "procedure_ot_charges": service_bucket_amounts.get("Procedure / OT Charges", 0.0),
                "bedside_services": service_bucket_amounts.get("Bedside Services", 0.0),
                "professional_fees": service_bucket_amounts.get("Professional Fees", 0.0),
                "ip_drugs": ip_drugs,
                "ip_drugs_per_day": ip_drugs_per_day,
                "ip_consumables": ip_supplies,
                "ip_consumables_per_day": ip_consumables_per_day,
                "ot_drugs": ot_drugs,
                "ot_consumables": ot_supplies,
                "implants": implants,
                "pharmacy_total": pharmacy_total,
                "drug_administration_charges": drug_administration_charges,
                "services_total_excluding_food_and_beverage": services_total_ex_fnb,
                "food_and_beverage_excluded": food_beverage_excluded,
                "pharmacy_returns_excluded": pharmacy_returns_excluded,
                "total_amount_excluding_food_and_beverage_and_returns_plus_drug_admin": total_amount,
            }
        )
    return output_rows


def is_template_default_included(service_row: dict[str, str]) -> bool:
    presence_rate = as_float(service_row.get("case_presence_rate"))
    typical_amount = as_float(service_row.get("amount_cash_typical"))
    return presence_rate > 90.0 or (
        presence_rate >= 75.0 and typical_amount <= SERVICES_SIGNIFICANCE_AMOUNT_THRESHOLD
    )


def clean_services_for_fc(service_rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    cleaned_rows: list[dict[str, str]] = []
    auto_included_rows: list[dict[str, str]] = []
    optional_rows: list[dict[str, str]] = []
    for row in service_rows:
        bucket = normalize_text(row.get("fc_estimate_bucket")).lower()
        code = normalize_code(row.get("item_code"))
        if "remove" in bucket:
            continue
        if code in TEMPLATE_EXCLUDED_SERVICE_CODES:
            continue
        cleaned_rows.append(row)
        if is_template_default_included(row):
            auto_included_rows.append(row)
        else:
            optional_rows.append(row)
    return cleaned_rows, auto_included_rows, optional_rows


def write_cleaned_services_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "item_code",
        "item_name",
        "fc_estimate_bucket",
        "grouping",
        "case_count",
        "case_presence_rate",
        "quantity_p25",
        "quantity_p50",
        "quantity_p75",
        "tariff_code",
        "tariff_general",
        "tariff_twin",
        "tariff_single",
        "tariff_icu",
        "rate_cash_p25",
        "rate_cash_p50",
        "rate_cash_p75",
        "amount_cash_typical",
        "room_category_dependent",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def tariff_rate_for_add_on(rate_row: RateRow | None) -> float | None:
    if not rate_row:
        return None
    for value in [rate_row.single, rate_row.twin, rate_row.general, rate_row.icu]:
        if value is not None:
            return value
    return None


def expected_add_on_contribution(service_row: dict[str, str], rate_lookup: dict[str, RateRow]) -> float:
    quantity_p50 = as_float(service_row.get("quantity_p50"))
    rate_row = rate_lookup.get(normalize_code(service_row.get("item_code")))
    rate_value = tariff_rate_for_add_on(rate_row)
    if rate_value is None and quantity_p50 > 0:
        rate_value = as_float(service_row.get("amount_cash_typical")) / quantity_p50
    presence_rate = as_float(service_row.get("case_presence_rate"))
    return quantity_p50 * rate_value * presence_rate / 100.0 if rate_value else 0.0


def prioritize_optional_service_rows(
    optional_service_rows: list[dict[str, str]],
    rate_lookup: dict[str, RateRow],
) -> list[dict[str, str]]:
    return sorted(
        optional_service_rows,
        key=lambda row: (
            -expected_add_on_contribution(row, rate_lookup),
            -as_float(row.get("case_presence_rate")),
            -(tariff_rate_for_add_on(rate_lookup.get(normalize_code(row.get("item_code")))) or 0.0),
            normalize_text(row.get("grouping")),
            normalize_text(row.get("item_name")),
            normalize_code(row.get("item_code")),
        ),
    )


def is_robotic_service_row(service_row: dict[str, Any]) -> bool:
    code = normalize_code(service_row.get("item_code"))
    item_name = normalize_text(service_row.get("item_name")).upper()
    grouping = normalize_text(service_row.get("grouping")).upper()
    bucket = normalize_text(service_row.get("fc_estimate_bucket")).upper()
    return any(
        marker in text
        for text in [code, item_name, grouping, bucket]
        for marker in ["ROBO", "ROBOTIC"]
    )


def split_robotic_optional_service_rows(
    optional_service_rows: list[dict[str, str]],
    *,
    procedure_code: str,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    normalized_procedure_code = normalize_code(procedure_code)
    standard_rows: list[dict[str, str]] = []
    robotic_rows: list[dict[str, str]] = []
    for row in optional_service_rows:
        code = normalize_code(row.get("item_code"))
        if is_robotic_service_row(row):
            if code and code == normalized_procedure_code:
                continue
            robotic_rows.append(row)
            continue
        standard_rows.append(row)
    return standard_rows, robotic_rows


def collect_robotic_service_rows(
    service_rows: list[dict[str, str]],
    *,
    procedure_code: str,
    include_procedure_row: str,
) -> list[dict[str, str]]:
    normalized_procedure_code = normalize_code(procedure_code)
    seen_keys: set[tuple[str, str]] = set()
    robotic_rows: list[dict[str, str]] = []
    for row in service_rows:
        code = normalize_code(row.get("item_code"))
        key = (code, normalize_text(row.get("item_name")))
        if key in seen_keys:
            continue
        if not is_robotic_service_row(row):
            continue
        if normalize_text(include_procedure_row).lower() != "no" and code and code == normalized_procedure_code:
            seen_keys.add(key)
            robotic_rows.append(row)
            continue
        seen_keys.add(key)
        robotic_rows.append(row)
    return robotic_rows


def compute_robotic_charge_presence_rate(robotic_service_rows: list[dict[str, str]]) -> float:
    if not robotic_service_rows:
        return 0.0
    return max(as_float(row.get("case_presence_rate")) for row in robotic_service_rows)


def collect_robotic_presence_signal_rows(
    service_rows: list[dict[str, str]],
    *,
    procedure_code: str,
) -> list[dict[str, str]]:
    signal_rows: list[dict[str, str]] = []
    normalized_procedure_code = normalize_code(procedure_code)
    seen_keys: set[tuple[str, str]] = set()
    for row in service_rows:
        bucket = normalize_text(row.get("fc_estimate_bucket")).lower()
        code = normalize_code(row.get("item_code"))
        key = (code, normalize_text(row.get("item_name")))
        if "remove" in bucket or key in seen_keys:
            continue
        if not is_robotic_service_row(row):
            continue
        if code and code == normalized_procedure_code:
            seen_keys.add(key)
            signal_rows.append(row)
            continue
        if code in TEMPLATE_EXCLUDED_SERVICE_CODES:
            continue
        seen_keys.add(key)
        signal_rows.append(row)
    return signal_rows


def resolve_robotic_default_selection(
    *,
    default_mode: str,
    presence_rate: float,
    presence_threshold: float,
) -> str:
    normalized_mode = normalize_text(default_mode).lower()
    if normalized_mode == "yes":
        return "Yes"
    if normalized_mode == "no":
        return "No"
    return "Yes" if presence_rate > presence_threshold else ""


def load_pharmacy_lookup(path: Path) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    by_code: dict[str, dict[str, str]] = {}
    by_name: dict[str, dict[str, str]] = {}
    for row in load_csv_rows(path):
        code = normalize_code(row.get("item_code"))
        name = normalize_text(row.get("item_name"))
        if code:
            by_code[code] = row
        if name:
            by_name[name] = row
    return by_code, by_name


def load_implant_rows(path: Path) -> list[dict[str, str]]:
    return load_csv_rows(path)


def load_existing_ot_selection_state(ws) -> dict[str, bool]:
    selections: dict[str, bool] = {}
    for row_number in range(1, ws.max_row + 1):
        item_name = normalize_text(ws[f"A{row_number}"].value)
        selected_value = normalize_text(ws[f"H{row_number}"].value)
        if not item_name:
            continue
        selections[item_name] = selected_value == SELECTION_INCLUDE
    return selections


def compute_ot_expected_contribution(row: dict[str, str]) -> float:
    quantity = as_float(row.get("ot_quantity_typical_cleaned"))
    amount = as_float(row.get("ot_amount_typical_cleaned"))
    presence_rate = as_float(row.get("case_presence_rate"))
    rate = amount / quantity if quantity else 0.0
    return (presence_rate * quantity * rate) / 100.0


def build_ot_consumable_shortlist(
    pharmacy_rows: list[dict[str, str]],
    *,
    max_count: int,
    cumulative_target: float = 0.80,
) -> list[dict[str, str]]:
    eligible_rows = [
        row
        for row in pharmacy_rows
        if normalize_text(row.get("classification")) == "Treatment Supplies"
        and normalize_text(row.get("present_in_ot_pharmacy")).lower() == "true"
        and as_float(row.get("case_presence_rate")) < 70.0
    ]
    ranked = sorted(
        eligible_rows,
        key=lambda row: (
            -compute_ot_expected_contribution(row),
            -as_float(row.get("case_presence_rate")),
            -as_float(row.get("ot_amount_typical_cleaned")),
            normalize_text(row.get("item_name")),
        ),
    )
    if not ranked:
        return []
    total_expected = sum(compute_ot_expected_contribution(row) for row in ranked)
    shortlist: list[dict[str, str]] = []
    running = 0.0
    for row in ranked:
        shortlist.append(row)
        running += compute_ot_expected_contribution(row)
        if len(shortlist) >= max_count:
            break
        if total_expected > 0 and (running / total_expected) >= cumulative_target:
            break
    return shortlist


def default_ot_shortlist_selection_indices(shortlist_rows: list[dict[str, str]]) -> set[int]:
    expected_values = [compute_ot_expected_contribution(row) for row in shortlist_rows]
    total_expected = sum(expected_values)
    if total_expected <= 0:
        return set()

    running = 0.0
    best_prefix_len = 0
    best_distance = float("inf")
    preferred_target = 0.40
    lower_target = 0.30
    upper_target = 0.50

    for idx, value in enumerate(expected_values, start=1):
        running += value
        share = running / total_expected
        distance = abs(share - preferred_target)
        if distance < best_distance:
            best_distance = distance
            best_prefix_len = idx
        if lower_target < share <= upper_target:
            return set(range(idx))

    if best_prefix_len <= 0:
        return set()
    return set(range(best_prefix_len))


def percentile_from_values(values: list[float], point: float) -> float:
    cleaned = sorted(float(v) for v in values if v is not None)
    if not cleaned:
        return 0.0
    if len(cleaned) == 1:
        return cleaned[0]
    if point <= 0:
        return cleaned[0]
    if point >= 1:
        return cleaned[-1]
    return float(statistics.quantiles(cleaned, n=100, method="inclusive")[int(point * 100) - 1]) if point not in {0.25, 0.5, 0.75} else float(
        statistics.quantiles(cleaned, n=4, method="inclusive")[0 if point == 0.25 else 1 if point == 0.5 else 2]
    )


def build_implant_selection_records(
    implant_rows: list[dict[str, str]],
    implant_detail_rows: list[dict[str, str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    item_lookup: dict[str, dict[str, str]] = {}
    family_records: dict[str, dict[str, Any]] = {}
    brand_records: dict[tuple[str, str], dict[str, Any]] = {}

    for row in implant_rows:
        code = normalize_code(row.get("item_code"))
        family = normalize_text(row.get("implant_family"))
        brand = normalize_text(row.get("brand_family"))
        if code and code not in item_lookup:
            item_lookup[code] = row
        if family and family not in family_records:
            family_qty = as_float(row.get("implant_family_quantity_p50"))
            family_rate = as_float(row.get("implant_family_rate_p50"))
            family_records[family] = {
                "family": family,
                "presence_rate": as_float(row.get("implant_family_presence_rate")),
                "quantity_p50": family_qty,
                "rate_p50": family_rate,
                "amount_p50": family_qty * family_rate,
            }
        if family and brand and (family, brand) not in brand_records:
            brand_records[(family, brand)] = {
                "family": family,
                "brand": brand,
                "presence_rate": as_float(row.get("family_brand_presence_rate") or row.get("brand_presence_rate_within_implant_family")),
                "quantity_p50": 0.0,
                "rate_p50": 0.0,
                "amount_p50": 0.0,
            }

    family_amounts_by_ip: dict[str, dict[str, float]] = {}
    brand_amounts_by_ip: dict[tuple[str, str], dict[str, float]] = {}
    brand_rates_by_key: dict[tuple[str, str], list[float]] = {}
    brand_quantities_by_key: dict[tuple[str, str], dict[str, float]] = {}
    item_amounts_by_ip: dict[str, dict[str, float]] = {}

    for detail in implant_detail_rows:
        code = normalize_code(detail.get("item_code"))
        ip = normalize_text(detail.get("admission_no"))
        mapping = item_lookup.get(code)
        if not code or not ip or not mapping:
            continue
        family = normalize_text(mapping.get("implant_family"))
        brand = normalize_text(mapping.get("brand_family"))
        qty = as_float(detail.get("quantity"))
        rate = as_float(detail.get("sale_rate"))
        amount = as_float(detail.get("gross_amount"))

        family_amounts_by_ip.setdefault(family, {})
        family_amounts_by_ip[family][ip] = family_amounts_by_ip[family].get(ip, 0.0) + amount

        brand_key = (family, brand)
        brand_amounts_by_ip.setdefault(brand_key, {})
        brand_amounts_by_ip[brand_key][ip] = brand_amounts_by_ip[brand_key].get(ip, 0.0) + amount
        brand_quantities_by_key.setdefault(brand_key, {})
        brand_quantities_by_key[brand_key][ip] = brand_quantities_by_key[brand_key].get(ip, 0.0) + qty
        brand_rates_by_key.setdefault(brand_key, []).append(rate)

        item_amounts_by_ip.setdefault(code, {})
        item_amounts_by_ip[code][ip] = item_amounts_by_ip[code].get(ip, 0.0) + amount

    for family, record in family_records.items():
        amount_values = list(family_amounts_by_ip.get(family, {}).values())
        if amount_values:
            record["amount_p50"] = inclusive_quartiles(amount_values)[1]

    for brand_key, record in brand_records.items():
        amount_values = list(brand_amounts_by_ip.get(brand_key, {}).values())
        qty_values = list(brand_quantities_by_key.get(brand_key, {}).values())
        rate_values = brand_rates_by_key.get(brand_key, [])
        record["quantity_p50"] = inclusive_quartiles(qty_values)[1] if qty_values else 0.0
        record["rate_p50"] = inclusive_quartiles(rate_values)[1] if rate_values else 0.0
        record["amount_p50"] = inclusive_quartiles(amount_values)[1] if amount_values else record["quantity_p50"] * record["rate_p50"]

    item_records: list[dict[str, Any]] = []
    seen_items: set[str] = set()
    for row in implant_rows:
        code = normalize_code(row.get("item_code"))
        if not code or code in seen_items:
            continue
        seen_items.add(code)
        qty = as_float(row.get("item_quantity_p50"))
        rate = as_float(row.get("typical_rate_p50"))
        amount_values = list(item_amounts_by_ip.get(code, {}).values())
        amount_p50 = inclusive_quartiles(amount_values)[1] if amount_values else qty * rate
        item_records.append(
            {
                "family": normalize_text(row.get("implant_family")),
                "brand": normalize_text(row.get("brand_family")),
                "item_code": code,
                "item_name": normalize_text(row.get("item_name")),
                "presence_rate": as_float(row.get("item_presence_rate")),
                "quantity_p50": qty,
                "rate_p50": rate,
                "amount_p50": amount_p50,
            }
        )

    ordered_families = sorted(
        family_records.values(),
        key=lambda rec: (IMPLANT_FAMILY_ORDER.index(rec["family"]) if rec["family"] in IMPLANT_FAMILY_ORDER else 999, rec["family"]),
    )
    ordered_brands = sorted(
        brand_records.values(),
        key=lambda rec: (
            IMPLANT_FAMILY_ORDER.index(rec["family"]) if rec["family"] in IMPLANT_FAMILY_ORDER else 999,
            -rec["presence_rate"],
            rec["brand"],
        ),
    )
    ordered_items = sorted(
        item_records,
        key=lambda rec: (
            IMPLANT_FAMILY_ORDER.index(rec["family"]) if rec["family"] in IMPLANT_FAMILY_ORDER else 999,
            rec["brand"],
            -rec["presence_rate"],
            rec["item_name"],
        ),
    )
    return ordered_families, ordered_brands, ordered_items


def build_implant_template_records(
    implant_rows: list[dict[str, str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    family_records: dict[str, dict[str, Any]] = {}
    brand_records: dict[tuple[str, str], dict[str, Any]] = {}
    item_records: dict[str, dict[str, Any]] = {}

    for row in implant_rows:
        family = normalize_text(row.get("implant_family"))
        brand = normalize_text(row.get("brand_family"))
        code = normalize_code(row.get("item_code"))
        if family and family not in family_records:
            family_records[family] = {
                "family": family,
                "presence_rate": as_float(row.get("implant_family_presence_rate")),
                "quantity_p25": as_float(row.get("implant_family_quantity_p25")),
                "quantity_p50": as_float(row.get("implant_family_quantity_p50")),
                "quantity_p75": as_float(row.get("implant_family_quantity_p75")),
                "rate_p25": as_float(row.get("implant_family_rate_p25")),
                "rate_p50": as_float(row.get("implant_family_rate_p50")),
                "rate_p75": as_float(row.get("implant_family_rate_p75")),
            }
            family_records[family]["amount_p50"] = family_records[family]["quantity_p50"] * family_records[family]["rate_p50"]
        if family and brand and (family, brand) not in brand_records:
            brand_records[(family, brand)] = {
                "family": family,
                "brand": brand,
                "presence_rate": as_float(row.get("family_brand_presence_rate") or row.get("brand_presence_rate_within_implant_family") or row.get("brand_presence_rate")),
                "quantity_p25": as_float(row.get("brand_quantity_p25")),
                "quantity_p50": as_float(row.get("brand_quantity_p50")),
                "quantity_p75": as_float(row.get("brand_quantity_p75")),
                "rate_p25": as_float(row.get("brand_rate_p25")),
                "rate_p50": as_float(row.get("brand_rate_p50")),
                "rate_p75": as_float(row.get("brand_rate_p75")),
            }
            brand_records[(family, brand)]["amount_p50"] = brand_records[(family, brand)]["quantity_p50"] * brand_records[(family, brand)]["rate_p50"]
        if code and code not in item_records:
            item_records[code] = {
                "family": family,
                "brand": brand,
                "item_code": code,
                "item_name": normalize_text(row.get("item_name")),
                "presence_rate": as_float(row.get("item_presence_rate")),
                "quantity_p25": as_float(row.get("item_quantity_p25")),
                "quantity_p50": as_float(row.get("item_quantity_p50")),
                "quantity_p75": as_float(row.get("item_quantity_p75")),
                "rate_p25": as_float(row.get("item_rate_p25")),
                "rate_p50": as_float(row.get("typical_rate_p50")),
                "rate_p75": as_float(row.get("item_rate_p75")),
            }
            item_records[code]["amount_p50"] = item_records[code]["quantity_p50"] * item_records[code]["rate_p50"]

    def family_sort_key(record: dict[str, Any]) -> tuple[int, str]:
        family = record["family"]
        return (IMPLANT_FAMILY_ORDER.index(family) if family in IMPLANT_FAMILY_ORDER else 999, family)

    ordered_families = sorted(family_records.values(), key=family_sort_key)
    ordered_brands = sorted(
        brand_records.values(),
        key=lambda rec: (
            IMPLANT_FAMILY_ORDER.index(rec["family"]) if rec["family"] in IMPLANT_FAMILY_ORDER else 999,
            rec["family"],
            rec["brand"],
        ),
    )
    ordered_items = sorted(
        item_records.values(),
        key=lambda rec: (
            IMPLANT_FAMILY_ORDER.index(rec["family"]) if rec["family"] in IMPLANT_FAMILY_ORDER else 999,
            rec["family"],
            rec["brand"],
            rec["item_name"],
        ),
    )
    return ordered_families, ordered_brands, ordered_items


def compute_bucket_quartiles(path: Path) -> dict[str, tuple[float, float, float]]:
    rows = load_csv_rows(path)
    key_map = {
        "ip_drugs": "total_amount_ip_drugs_medicines_ivs_nutrition_products",
        "ip_consumables": "total_amount_ip_treatment_supplies",
        "ot_drugs": "total_amount_ot_drugs_medicines_ivs_nutrition_products",
        "ot_consumables": "total_amount_ot_treatment_supplies",
        "implants": "total_amount_implants",
    }
    quartiles: dict[str, tuple[float, float, float]] = {}
    for label, field in key_map.items():
        values = [as_float(row.get(field)) for row in rows]
        quartiles[label] = inclusive_quartiles(values)
    return quartiles


def get_metric(quartiles_json: dict[str, Any], metric_name: str, point: str) -> float:
    return float(quartiles_json["metrics"][metric_name][point])


def get_service_line_count_metric(payload: dict[str, Any], point: str) -> float:
    return float((payload.get("cleaned_distinct_service_line_count") or {}).get(point) or 0.0)


def get_ip_pharmacy_per_day_metric(payload: dict[str, Any], metric_name: str, point: str) -> float:
    return float((((payload.get("metrics") or {}).get(metric_name) or {}).get(point)) or 0.0)


def derive_payer_type(payor_label: str) -> str:
    label = normalize_text(payor_label).lower()
    return "cash" if "cash" in label else "insurance"


def get_professional_fee_multipliers(payor_label: str) -> dict[str, float]:
    payer_type = derive_payer_type(payor_label)
    if payer_type == "cash":
        return {
            "surgeon": 0.25,
            "assistant_surgeon": 0.15,
            "anesthetist": 0.25,
            "assistant_anesthetist": 0.25,
        }
    return {
        "surgeon": 0.35,
        "assistant_surgeon": 0.35,
        "anesthetist": 0.45,
        "assistant_anesthetist": 0.0,
    }


def build_pf_base_formula(ws, total_col: str) -> str:
    cross_consult_refs: list[str] = []
    for row_number in range(2, 43):
        item_name = normalize_text(ws[f"A{row_number}"].value).lower()
        item_code = normalize_code(ws[f"F{row_number}"].value)
        if "cross consult" in item_name or "cross consultation" in item_name or item_code.startswith("CC"):
            cross_consult_refs.append(f"{total_col}{row_number}")
    if not cross_consult_refs:
        return f"{total_col}42"
    return f"{total_col}42-SUM({','.join(cross_consult_refs)})"


def build_estimate_builder_template_rows(args: argparse.Namespace) -> dict[int, str]:
    rows = dict(BASE_ESTIMATE_BUILDER_TEMPLATE_ROWS)
    rows[13] = normalize_code(args.procedure_code)
    return rows


def derive_los_guidance(quartiles_json: dict[str, Any], override: str) -> str:
    explicit = normalize_text(override)
    if explicit:
        return explicit

    los_p25 = round_display_quantity(get_metric(quartiles_json, "los_days", "p25"))
    los_p75 = round_display_quantity(get_metric(quartiles_json, "los_days", "p75"))
    icu_p25 = round_display_quantity(get_metric(quartiles_json, "icu_days", "p25"))
    icu_p75 = round_display_quantity(get_metric(quartiles_json, "icu_days", "p75"))
    ward_p25 = round_display_quantity(get_metric(quartiles_json, "ward_days", "p25"))
    ward_p75 = round_display_quantity(get_metric(quartiles_json, "ward_days", "p75"))

    def display_range(low: float, high: float) -> str:
        if low == high:
            return str(int(low)) if float(low).is_integer() else str(low)
        low_text = str(int(low)) if float(low).is_integer() else str(low)
        high_text = str(int(high)) if float(high).is_integer() else str(high)
        return f"{low_text}-{high_text}"

    return f"{display_range(los_p25, los_p75)} Days ({display_range(icu_p25, icu_p75)} ICU and {display_range(ward_p25, ward_p75)} Room)"


def write_sheet1(workbook, quartiles_json: dict[str, Any], args: argparse.Namespace) -> None:
    ws = workbook["Sheet1"]
    ws["B1"] = args.sheet1_template_name
    ws["B2"] = args.payor_label
    ws["B3"] = derive_los_guidance(quartiles_json, args.sheet1_los_guidance)
    ws["B4"] = args.sheet1_management_type

    ws["B8"] = get_metric(quartiles_json, "los_days", "p50")
    ws["C8"] = get_metric(quartiles_json, "los_days", "p25")
    ws["D8"] = get_metric(quartiles_json, "los_days", "p75")

    ws["B9"] = get_metric(quartiles_json, "icu_days", "p50")
    ws["C9"] = get_metric(quartiles_json, "icu_days", "p25")
    ws["D9"] = get_metric(quartiles_json, "icu_days", "p75")

    ws["B10"] = get_metric(quartiles_json, "ward_days", "p50")
    ws["C10"] = get_metric(quartiles_json, "ward_days", "p25")
    ws["D10"] = get_metric(quartiles_json, "ward_days", "p75")

    ws["B12"] = get_metric(quartiles_json, "ot_hours", "p50")
    ws["C12"] = get_metric(quartiles_json, "ot_hours", "p25")
    ws["D12"] = get_metric(quartiles_json, "ot_hours", "p75")


def fill_rate_cells(ws, row_number: int, rate_row: RateRow | None) -> None:
    ws[f"U{row_number}"] = rate_row.icu if rate_row else None
    ws[f"V{row_number}"] = rate_row.general if rate_row else None
    ws[f"W{row_number}"] = rate_row.twin if rate_row else None
    ws[f"X{row_number}"] = rate_row.single if rate_row else None


def compute_row_totals(
    rate_row: RateRow | None,
    quantity: float | None,
    icu_only: bool = False,
) -> tuple[float | None, float | None, float | None]:
    if quantity in (None, 0):
        return None, None, None
    if not rate_row:
        return None, None, None
    if icu_only:
        icu_rate = rate_row.icu
        if icu_rate is None:
            return None, None, None
        total = icu_rate * quantity
        return total, total, total
    general = rate_row.general * quantity if rate_row.general is not None else None
    twin = rate_row.twin * quantity if rate_row.twin is not None else None
    single = rate_row.single * quantity if rate_row.single is not None else None
    return general, twin, single


def write_estimate_builder(
    workbook,
    quartiles_json: dict[str, Any],
    bucket_quartiles: dict[str, tuple[float, float, float]],
    service_lookup: dict[str, dict[str, str]],
    rate_lookup: dict[str, RateRow],
    args: argparse.Namespace,
) -> None:
    ws = workbook["Estimate Builder"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    estimate_builder_template_rows = build_estimate_builder_template_rows(args)
    pf_multipliers = get_professional_fee_multipliers(args.payor_label)
    payer_type = derive_payer_type(args.payor_label)

    los_p50_display = round_display_quantity(get_metric(quartiles_json, "los_days", "p50"))
    ward_p50_display = round_display_quantity(get_metric(quartiles_json, "ward_days", "p50"))
    icu_p50_display = round_display_quantity(get_metric(quartiles_json, "icu_days", "p50"))

    ws["A13"] = args.procedure_label
    ws["F13"] = normalize_code(args.procedure_code)

    for row_number, code in estimate_builder_template_rows.items():
        service_row = service_lookup.get(normalize_code(code), {})
        rate_row = rate_lookup.get(normalize_code(code))
        default_quantity = 1.0 if row_number == 13 else None
        ws[f"E{row_number}"] = service_rule_text(service_row)
        ws[f"H{row_number}"] = maybe_float(service_row.get("quantity_p50")) if service_row else default_quantity
        ws[f"I{row_number}"] = maybe_float(service_row.get("quantity_p25")) if service_row else None
        ws[f"J{row_number}"] = maybe_float(service_row.get("quantity_p75")) if service_row else None
        fill_rate_cells(ws, row_number, rate_row)
        qty = maybe_float(service_row.get("quantity_p50")) if service_row else default_quantity
        total_general, total_twin, total_single = compute_row_totals(rate_row, qty)
        ws[f"Y{row_number}"] = total_general
        ws[f"Z{row_number}"] = total_twin
        ws[f"AA{row_number}"] = total_single

    for row_number, code in ESTIMATE_BUILDER_ALWAYS_ONE_ROWS.items():
        rate_row = rate_lookup.get(normalize_code(code))
        ws[f"H{row_number}"] = 1.0
        ws[f"I{row_number}"] = None
        ws[f"J{row_number}"] = None
        fill_rate_cells(ws, row_number, rate_row)
        total_general, total_twin, total_single = compute_row_totals(rate_row, 1.0)
        ws[f"Y{row_number}"] = total_general
        ws[f"Z{row_number}"] = total_twin
        ws[f"AA{row_number}"] = total_single

    for row_number, (metric_name, code, icu_only) in ESTIMATE_BUILDER_LOGIC_ROWS.items():
        rate_row = rate_lookup.get(normalize_code(code))
        if metric_name == "los_days":
            quantity = los_p50_display
        elif metric_name == "ward_days":
            quantity = ward_p50_display
        else:
            quantity = icu_p50_display
        ws[f"H{row_number}"] = quantity
        ws[f"I{row_number}"] = None
        ws[f"J{row_number}"] = None
        fill_rate_cells(ws, row_number, rate_row)
        total_general, total_twin, total_single = compute_row_totals(rate_row, quantity, icu_only=icu_only)
        ws[f"Y{row_number}"] = total_general
        ws[f"Z{row_number}"] = total_twin
        ws[f"AA{row_number}"] = total_single

    ws["H7"] = ward_p50_display
    ws["I7"] = None
    ws["J7"] = None
    general_bed = rate_lookup.get("ROM0001")
    twin_bed = rate_lookup.get("ROM0024")
    single_bed = rate_lookup.get("ROM0036")
    ws["U7"] = None
    ws["V7"] = general_bed.general if general_bed else None
    ws["W7"] = twin_bed.general if twin_bed else None
    ws["X7"] = single_bed.general if single_bed else None
    ws["Y7"] = (general_bed.general if general_bed and general_bed.general is not None else 0.0) * ward_p50_display
    ws["Z7"] = (twin_bed.general if twin_bed and twin_bed.general is not None else 0.0) * ward_p50_display
    ws["AA7"] = (single_bed.general if single_bed and single_bed.general is not None else 0.0) * ward_p50_display
    ws["E20"] = "LOS Days x Rate"

    ip_drugs_p25, ip_drugs_p50, ip_drugs_p75 = bucket_quartiles["ip_drugs"]
    ip_consumables_p25, ip_consumables_p50, ip_consumables_p75 = bucket_quartiles["ip_consumables"]
    ot_drugs_p25, ot_drugs_p50, ot_drugs_p75 = bucket_quartiles["ot_drugs"]
    ot_consumables_p25, ot_consumables_p50, ot_consumables_p75 = bucket_quartiles["ot_consumables"]
    implants_p25, implants_p50, implants_p75 = bucket_quartiles["implants"]

    ws["H37"] = los_p50_display
    ws["I37"] = None
    ws["J37"] = None
    ws["N37"] = ip_drugs_p50 / los_p50_display if los_p50_display else None
    ws["O37"] = ip_drugs_p25 / los_p50_display if los_p50_display else None
    ws["P37"] = ip_drugs_p75 / los_p50_display if los_p50_display else None
    ws["Q37"] = ip_drugs_p50
    ws["R37"] = ip_drugs_p25
    ws["S37"] = ip_drugs_p75
    ws["Y37"] = ip_drugs_p50
    ws["Z37"] = ip_drugs_p50
    ws["AA37"] = ip_drugs_p50

    ws["H38"] = los_p50_display
    ws["I38"] = None
    ws["J38"] = None
    ws["N38"] = ip_consumables_p50 / los_p50_display if los_p50_display else None
    ws["O38"] = ip_consumables_p25 / los_p50_display if los_p50_display else None
    ws["P38"] = ip_consumables_p75 / los_p50_display if los_p50_display else None
    ws["Q38"] = ip_consumables_p50
    ws["R38"] = ip_consumables_p25
    ws["S38"] = ip_consumables_p75
    ws["Y38"] = ip_consumables_p50
    ws["Z38"] = ip_consumables_p50
    ws["AA38"] = ip_consumables_p50

    ws["H39"] = None
    ws["I39"] = None
    ws["J39"] = None
    ws["Q39"] = ot_drugs_p50
    ws["R39"] = ot_drugs_p25
    ws["S39"] = ot_drugs_p75
    ws["Y39"] = ot_drugs_p50
    ws["Z39"] = ot_drugs_p50
    ws["AA39"] = ot_drugs_p50

    ws["H40"] = None
    ws["I40"] = None
    ws["J40"] = None
    ws["Q40"] = ot_consumables_p50
    ws["R40"] = ot_consumables_p25
    ws["S40"] = ot_consumables_p75
    ws["Y40"] = f"='Pharmacy Variance'!{PHARMACY_VARIANCE_OT_RESULT_CELL}"
    ws["Z40"] = f"='Pharmacy Variance'!{PHARMACY_VARIANCE_OT_RESULT_CELL}"
    ws["AA40"] = f"='Pharmacy Variance'!{PHARMACY_VARIANCE_OT_RESULT_CELL}"

    ws["H41"] = None
    ws["I41"] = None
    ws["J41"] = None
    ws["Q41"] = implants_p50
    ws["R41"] = implants_p25
    ws["S41"] = implants_p75
    ws["Y41"] = f"='Pharmacy Variance'!{PHARMACY_VARIANCE_IMPLANT_RESULT_CELL}"
    ws["Z41"] = f"='Pharmacy Variance'!{PHARMACY_VARIANCE_IMPLANT_RESULT_CELL}"
    ws["AA41"] = f"='Pharmacy Variance'!{PHARMACY_VARIANCE_IMPLANT_RESULT_CELL}"

    drug_admin_p50 = (ip_drugs_p50 + ip_consumables_p50 + ot_drugs_p50 + ot_consumables_p50 + implants_p50) * 0.125
    drug_admin_p25 = (ip_drugs_p25 + ip_consumables_p25 + ot_drugs_p25 + ot_consumables_p25 + implants_p25) * 0.125
    drug_admin_p75 = (ip_drugs_p75 + ip_consumables_p75 + ot_drugs_p75 + ot_consumables_p75 + implants_p75) * 0.125
    ws["Q32"] = drug_admin_p50
    ws["R32"] = drug_admin_p25
    ws["S32"] = drug_admin_p75
    ws["Y32"] = "=0.125*SUM(Y37:Y41)"
    ws["Z32"] = "=0.125*SUM(Z37:Z41)"
    ws["AA32"] = "=0.125*SUM(AA37:AA41)"

    for row_number in [32, 37, 38, 39, 40, 41]:
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "Q", "R", "S", "Y", "Z", "AA"]:
            ws[f"{col}{row_number}"].fill = FORMULA_BLUE_FILL if row_number in {32, 40, 41} and col in {"Y", "Z", "AA"} else SELECTION_FILL
            ws[f"{col}{row_number}"].border = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)
    for row_number in [37, 38, 39, 40, 41]:
        ws[f"A{row_number}"].font = Font(bold=True)
        ws[f"Y{row_number}"].fill = FORMULA_GREEN_FILL
        ws[f"Z{row_number}"].fill = FORMULA_GREEN_FILL
        ws[f"AA{row_number}"].fill = FORMULA_GREEN_FILL
    ws["A40"].comment = Comment("Driven by OT consumables selection on the Pharmacy Variance sheet.", "Codex")
    ws["A41"].comment = Comment("Defaults to implant P50 unless family/brand/item selections override it on the Pharmacy Variance sheet.", "Codex")
    ws["A32"].comment = Comment("Calculated as 12.5% of the current pharmacy total (rows 37 to 41).", "Codex")

    ws["Y42"] = "=SUM(Y2:Y41)"
    ws["Z42"] = "=SUM(Z2:Z41)"
    ws["AA42"] = "=SUM(AA2:AA41)"

    ws["E33"] = f"{payer_type.title()} PF % of pre-PF subtotal"
    ws["E34"] = f"{payer_type.title()} PF % of surgeon fee"
    ws["E35"] = f"{payer_type.title()} PF % of surgeon fee"
    ws["E36"] = (
        f"{payer_type.title()} PF % of anesthetist fee"
        if pf_multipliers["assistant_anesthetist"] > 0
        else "Not applicable for insurance"
    )

    pf_base_general = build_pf_base_formula(ws, "Y")
    pf_base_twin = build_pf_base_formula(ws, "Z")
    pf_base_single = build_pf_base_formula(ws, "AA")

    ws["AB33"] = f"={pf_multipliers['surgeon']}*({pf_base_general})"
    ws["AC33"] = f"={pf_multipliers['surgeon']}*({pf_base_twin})"
    ws["AD33"] = f"={pf_multipliers['surgeon']}*({pf_base_single})"
    ws["AB34"] = f"={pf_multipliers['assistant_surgeon']}*AB33"
    ws["AC34"] = f"={pf_multipliers['assistant_surgeon']}*AC33"
    ws["AD34"] = f"={pf_multipliers['assistant_surgeon']}*AD33"
    ws["AB35"] = f"={pf_multipliers['anesthetist']}*AB33"
    ws["AC35"] = f"={pf_multipliers['anesthetist']}*AC33"
    ws["AD35"] = f"={pf_multipliers['anesthetist']}*AD33"
    if pf_multipliers["assistant_anesthetist"] > 0:
        ws["AB36"] = f"={pf_multipliers['assistant_anesthetist']}*AB35"
        ws["AC36"] = f"={pf_multipliers['assistant_anesthetist']}*AC35"
        ws["AD36"] = f"={pf_multipliers['assistant_anesthetist']}*AD35"
    else:
        ws["AB36"] = 0
        ws["AC36"] = 0
        ws["AD36"] = 0

    ws["AB43"] = "=Y42+AB33+AB34+AB35+AB36"
    ws["AC43"] = "=Z42+AC33+AC34+AC35+AC36"
    ws["AD43"] = "=AA42+AD33+AD34+AD35+AD36"

    pharmacy_section_rows = [32, 37, 38, 39, 40, 41, 42]
    for row_number in pharmacy_section_rows:
        style_row(ws, row_number, 1, 30, fill=SELECTION_FILL, wrap=True)
        for col in ["Y", "Z", "AA"]:
            ws[f"{col}{row_number}"].fill = FORMULA_GREEN_FILL
        for col in ["Q", "R", "S"]:
            ws[f"{col}{row_number}"].fill = RESULT_FILL
        ws[f"A{row_number}"].font = Font(bold=True)
        ws.row_dimensions[row_number].height = 22

    for row_number in [37, 38, 39, 40, 41]:
        for col in ["H", "N", "O", "P", "Q", "R", "S", "Y", "Z", "AA"]:
            ws[f"{col}{row_number}"].border = SECTION_BORDER
        ws[f"H{row_number}"].fill = INPUT_FILL if row_number in {37, 38} else SELECTION_FILL

    for col in ["Y", "Z", "AA"]:
        ws[f"{col}32"].fill = FORMULA_BLUE_FILL
    style_row(ws, 33, 1, 30, fill=SELECTION_FILL, wrap=True)
    style_row(ws, 34, 1, 30, fill=SELECTION_FILL, wrap=True)
    style_row(ws, 35, 1, 30, fill=SELECTION_FILL, wrap=True)
    style_row(ws, 36, 1, 30, fill=SELECTION_FILL, wrap=True)
    for row_number in [33, 34, 35, 36]:
        ws[f"A{row_number}"].font = Font(bold=True)
    for col in ["AB", "AC", "AD"]:
        for row_number in [33, 34, 35, 36]:
            ws[f"{col}{row_number}"].fill = FORMULA_GREEN_FILL
            ws[f"{col}{row_number}"].border = SECTION_BORDER
            ws[f"{col}{row_number}"].font = Font(bold=True)
        ws[f"{col}43"].fill = RESULT_FILL
        ws[f"{col}43"].font = Font(bold=True)

    set_number_format(ws, "H37:H38", '#,##0.00')
    set_number_format(ws, "N37:P38", '#,##0.00')
    set_number_format(ws, "Q32:S41", '#,##0.00')
    set_number_format(ws, "Y32:AD43", '#,##0.00')

    ws["A37"].comment = Comment("Baseline IP drugs estimate from cohort pharmacy bucket quartiles.", "Codex")
    ws["A38"].comment = Comment("Baseline IP treatment supplies estimate from cohort pharmacy bucket quartiles.", "Codex")
    ws["A39"].comment = Comment("Baseline OT drugs estimate from cohort pharmacy bucket quartiles.", "Codex")
    ws["A42"].comment = Comment("Subtotal of hospital estimate rows before professional fees.", "Codex")
    ws["A33"].comment = Comment("Calculated as a payer-based percentage of the pre-professional-fee subtotal, excluding cross consultation only if such a row exists.", "Codex")
    ws["A34"].comment = Comment("Calculated from surgeon fee using payer-specific assistant-surgeon percentage.", "Codex")
    ws["A35"].comment = Comment("Calculated from surgeon fee using payer-specific anesthetist percentage.", "Codex")
    ws["A36"].comment = Comment("Calculated from anesthetist fee for cash; not applicable for insurance.", "Codex")


def write_pharmacy_variance(
    workbook,
    bucket_quartiles: dict[str, tuple[float, float, float]],
    pharmacy_rows: list[dict[str, str]],
    implant_rows: list[dict[str, str]],
    implant_detail_rows: list[dict[str, str]],
    args: argparse.Namespace,
) -> None:
    ws = workbook["Pharmacy Variance"]
    existing_ot_selections = load_existing_ot_selection_state(ws)

    for row_number in range(1, max(ws.max_row, 160) + 1):
        for col_idx in range(1, 21):
            ws.cell(row=row_number, column=col_idx).value = None
            ws.cell(row=row_number, column=col_idx).fill = PatternFill(fill_type=None)
            ws.cell(row=row_number, column=col_idx).border = Border()
            ws.cell(row=row_number, column=col_idx).font = Font(color="000000", bold=False)
            ws.cell(row=row_number, column=col_idx).alignment = Alignment(horizontal="general", vertical="bottom")
            ws.cell(row=row_number, column=col_idx).number_format = "General"

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.column_dimensions["I"].width = 3
    ws.column_dimensions["I"].hidden = True

    ot_consumables_p25, ot_consumables_p50, ot_consumables_p75 = bucket_quartiles["ot_consumables"]
    ws["A1"] = "Pharmacy"
    ws["B2"] = "P50"
    ws["C2"] = "P25"
    ws["D2"] = "P75"
    ws["F2"] = "Derived Estimate"
    ws["A3"] = "OT Consumables"
    ws["B3"] = ot_consumables_p50
    ws["C3"] = ot_consumables_p25
    ws["D3"] = ot_consumables_p75
    ws["A4"] = "Select high-variance OT treatment supplies to move the estimate from P25 toward P75."
    ws["J4"] = "Implants default to P50. Select at family, brand, or exact item level to override."

    if args.ot_consumable_shortlist_mode == "fixed":
        pharmacy_by_name = {
            normalize_text(row.get("item_name")): row for row in pharmacy_rows if normalize_text(row.get("item_name"))
        }
        shortlist_rows = [pharmacy_by_name[normalize_text(item_name)] for item_name, _ in DEFAULT_OT_CONSUMABLE_SHORTLIST if normalize_text(item_name) in pharmacy_by_name]
    else:
        shortlist_rows = build_ot_consumable_shortlist(
            pharmacy_rows,
            max_count=args.ot_consumable_shortlist_count,
        )

    ws["A5"] = "Item"
    ws["B5"] = "Typical Quantity"
    ws["C5"] = "Typical Rate"
    ws["D5"] = "Typical Amount"
    ws["E5"] = "Presence Rate"
    ws["F5"] = "Expected Contribution"
    ws["G5"] = "Cumulative Share"
    ws["H5"] = "Selected"

    ot_row_start = 6
    total_expected_formula_parts: list[str] = []
    for offset, row in enumerate(shortlist_rows, start=ot_row_start):
        item_name = normalize_text(row.get("item_name"))
        quantity = as_float(row.get("ot_quantity_typical_cleaned"))
        amount = as_float(row.get("ot_amount_typical_cleaned"))
        rate = amount / quantity if quantity else 0.0
        presence_rate = as_float(row.get("case_presence_rate"))
        selected = existing_ot_selections.get(item_name, False)
        ws[f"A{offset}"] = item_name
        ws[f"B{offset}"] = quantity
        ws[f"C{offset}"] = rate
        ws[f"D{offset}"] = amount
        ws[f"E{offset}"] = presence_rate
        ws[f"F{offset}"] = f"=E{offset}*B{offset}*C{offset}/100"
        total_expected_formula_parts.append(f"F{offset}")
        if offset == ot_row_start:
            ws[f"G{offset}"] = f"=F{offset}/SUM($F${ot_row_start}:$F${ot_row_start + len(shortlist_rows) - 1})"
        else:
            ws[f"G{offset}"] = f"=G{offset-1}+F{offset}/SUM($F${ot_row_start}:$F${ot_row_start + len(shortlist_rows) - 1})"
        ws[f"H{offset}"] = SELECTION_INCLUDE if selected else SELECTION_EXCLUDE

    if shortlist_rows:
        ot_row_end = ot_row_start + len(shortlist_rows) - 1
        ws[PHARMACY_VARIANCE_OT_RESULT_CELL] = (
            f'=IFERROR((SUMIF(H{ot_row_start}:H{ot_row_end},"{SELECTION_INCLUDE}",F{ot_row_start}:F{ot_row_end})/'
            f'SUM(F{ot_row_start}:F{ot_row_end}))*(D3-C3)+C3,C3)'
        )
    else:
        ws[PHARMACY_VARIANCE_OT_RESULT_CELL] = ot_consumables_p50

    implants_p25, implants_p50, implants_p75 = bucket_quartiles["implants"]
    ws["J1"] = "Implants"
    ws["J2"] = "P50"
    ws["K2"] = "P25"
    ws["L2"] = "P75"
    ws["N2"] = "Resolved Estimate"
    ws["J3"] = implants_p50
    ws["K3"] = implants_p25
    ws["L3"] = implants_p75

    family_records, brand_records, item_records = build_implant_selection_records(implant_rows, implant_detail_rows)
    ws["J5"] = "Level"
    ws["K5"] = "Family"
    ws["L5"] = "Brand"
    ws["M5"] = "Item Code"
    ws["N5"] = "Item Name"
    ws["O5"] = "Presence Rate"
    ws["P5"] = "Qty P50"
    ws["Q5"] = "Rate P50"
    ws["R5"] = "Amount P50"
    ws["S5"] = "Selected"
    ws["T5"] = "Effective Amount"

    current_row = 6
    family_row_map: dict[str, int] = {}
    brand_row_map: dict[tuple[str, str], int] = {}
    item_row_map: dict[str, int] = {}
    family_brand_rows: dict[tuple[str, str], list[int]] = {}
    family_item_rows: dict[str, list[int]] = {}
    for family in family_records:
        family_row_map[family["family"]] = current_row
        ws[f"J{current_row}"] = "Family"
        ws[f"K{current_row}"] = family["family"]
        ws[f"O{current_row}"] = family["presence_rate"]
        ws[f"P{current_row}"] = family["quantity_p50"]
        ws[f"Q{current_row}"] = family["rate_p50"]
        ws[f"R{current_row}"] = family["amount_p50"]
        ws[f"S{current_row}"] = SELECTION_EXCLUDE
        current_row += 1
        matching_brands = [brand for brand in brand_records if brand["family"] == family["family"]]
        for brand in matching_brands:
            brand_key = (brand["family"], brand["brand"])
            brand_row_map[brand_key] = current_row
            family_brand_rows.setdefault(brand_key, [])
            ws[f"J{current_row}"] = "Brand"
            ws[f"K{current_row}"] = brand["family"]
            ws[f"L{current_row}"] = brand["brand"]
            ws[f"O{current_row}"] = brand["presence_rate"]
            ws[f"P{current_row}"] = brand["quantity_p50"]
            ws[f"Q{current_row}"] = brand["rate_p50"]
            ws[f"R{current_row}"] = brand["amount_p50"]
            ws[f"S{current_row}"] = SELECTION_EXCLUDE
            current_row += 1
            matching_items = [item for item in item_records if item["family"] == brand["family"] and item["brand"] == brand["brand"]]
            for item in matching_items:
                item_row_map[item["item_code"]] = current_row
                family_brand_rows.setdefault(brand_key, []).append(current_row)
                family_item_rows.setdefault(item["family"], []).append(current_row)
                ws[f"J{current_row}"] = "Item"
                ws[f"K{current_row}"] = item["family"]
                ws[f"L{current_row}"] = item["brand"]
                ws[f"M{current_row}"] = item["item_code"]
                ws[f"N{current_row}"] = item["item_name"]
                ws[f"O{current_row}"] = item["presence_rate"]
                ws[f"P{current_row}"] = item["quantity_p50"]
                ws[f"Q{current_row}"] = item["rate_p50"]
                ws[f"R{current_row}"] = item["amount_p50"]
                ws[f"S{current_row}"] = SELECTION_EXCLUDE
                current_row += 1

    family_brand_row_lookup: dict[str, list[int]] = {}
    family_direct_brand_rows: dict[str, list[int]] = {}
    for (family_name, brand_name), row_number in brand_row_map.items():
        family_direct_brand_rows.setdefault(family_name, []).append(row_number)
        family_brand_row_lookup[f"{family_name}|{brand_name}"] = family_brand_rows.get((family_name, brand_name), [])

    for item in item_records:
        row_number = item_row_map[item["item_code"]]
        ws[f"T{row_number}"] = f'=IF(S{row_number}="{SELECTION_INCLUDE}",R{row_number},0)'

    for brand in brand_records:
        row_number = brand_row_map[(brand["family"], brand["brand"])]
        item_rows = family_brand_row_lookup.get(f"{brand['family']}|{brand['brand']}", [])
        if item_rows:
            item_selected_formula = "+".join(f'--(S{item_row}="{SELECTION_INCLUDE}")' for item_row in item_rows)
            ws[f"T{row_number}"] = f'=IF(AND(S{row_number}="{SELECTION_INCLUDE}",({item_selected_formula})=0),R{row_number},0)'
        else:
            ws[f"T{row_number}"] = f'=IF(S{row_number}="{SELECTION_INCLUDE}",R{row_number},0)'

    for family in family_records:
        row_number = family_row_map[family["family"]]
        brand_rows = family_direct_brand_rows.get(family["family"], [])
        item_rows = family_item_rows.get(family["family"], [])
        parts: list[str] = []
        if brand_rows:
            parts.append("+".join(f'--(S{brand_row}="{SELECTION_INCLUDE}")' for brand_row in brand_rows))
        if item_rows:
            parts.append("+".join(f'--(S{item_row}="{SELECTION_INCLUDE}")' for item_row in item_rows))
        if parts:
            ws[f"T{row_number}"] = f'=IF(AND(S{row_number}="{SELECTION_INCLUDE}",({" + ".join(parts)})=0),R{row_number},0)'
        else:
            ws[f"T{row_number}"] = f'=IF(S{row_number}="{SELECTION_INCLUDE}",R{row_number},0)'

    implant_row_end = current_row - 1
    if implant_row_end >= 6:
        ws[PHARMACY_VARIANCE_IMPLANT_RESULT_CELL] = (
            f'=IF(COUNTIF(S6:S{implant_row_end},"{SELECTION_INCLUDE}")=0,J3,SUM(T6:T{implant_row_end}))'
        )
    else:
        ws[PHARMACY_VARIANCE_IMPLANT_RESULT_CELL] = implants_p50

    ot_selection_validation = DataValidation(type="list", formula1=f'"{SELECTION_INCLUDE},{SELECTION_EXCLUDE}"', allow_blank=False)
    ot_selection_validation.prompt = "Choose whether to include this OT consumable in the variance-driven estimate."
    ot_selection_validation.promptTitle = "OT Consumable Selection"
    ws.add_data_validation(ot_selection_validation)
    if shortlist_rows:
        ot_selection_validation.add(f"H{ot_row_start}:H{ot_row_end}")

    implant_selection_validation = DataValidation(type="list", formula1=f'"{SELECTION_INCLUDE},{SELECTION_EXCLUDE}"', allow_blank=False)
    implant_selection_validation.prompt = "Select Family, Brand, or exact Item. Precedence is Item > Brand > Family."
    implant_selection_validation.promptTitle = "Implant Selection"
    ws.add_data_validation(implant_selection_validation)
    if implant_row_end >= 6:
        implant_selection_validation.add(f"S6:S{implant_row_end}")

    style_range(ws, "A1:H1", fill=HEADER_FILL, bold=True, font_color="FFFFFF")
    style_range(ws, "J1:T1", fill=HEADER_FILL, bold=True, font_color="FFFFFF")
    style_range(ws, "A2:H2", fill=SUBHEADER_FILL, bold=True)
    style_range(ws, "J2:T2", fill=SUBHEADER_FILL, bold=True)
    style_range(ws, "A5:H5", fill=OT_FILL, bold=True, wrap=True)
    style_range(ws, "J5:T5", fill=IMPLANT_FILL, bold=True, wrap=True)
    style_range(ws, "A3:F3", fill=RESULT_FILL, bold=True)
    style_range(ws, "J3:N3", fill=RESULT_FILL, bold=True)
    style_range(ws, "A4:H4", fill=SELECTION_FILL, wrap=True, align="left")
    style_range(ws, "J4:T4", fill=SELECTION_FILL, wrap=True, align="left")
    style_range(ws, "I1:I160", fill=SPACER_FILL, border=False)

    for cell_ref in ["F3", "N3"]:
        ws[cell_ref].border = SECTION_BORDER
    ws["F3"].comment = Comment(
        "Derived OT consumables estimate. It starts at P25 and moves toward P75 based on which shortlist items are marked Include.",
        "Codex",
    )
    ws["N3"].comment = Comment(
        "Resolved implants estimate. If no selections are made it stays at implant P50. Exact Item selections override Brand, which overrides Family.",
        "Codex",
    )
    ws["H5"].comment = Comment("Use the dropdown to Include or Exclude each OT consumable from the variance calculation.", "Codex")
    ws["S5"].comment = Comment("Use the dropdown to Include a Family, Brand, or exact Item. Precedence is Item > Brand > Family.", "Codex")

    if shortlist_rows:
        style_range(ws, f"A{ot_row_start}:H{ot_row_end}", wrap=True, align="center")
        set_number_format(ws, f"B{ot_row_start}:F{ot_row_end}", '#,##0.00')
        set_number_format(ws, f"G{ot_row_start}:G{ot_row_end}", '0.0%')
        for row in range(ot_row_start, ot_row_end + 1):
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"H{row}"].fill = FORMULA_BLUE_FILL
            if row % 2 == 0:
                for col in range(1, 9):
                    if col != 8:
                        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FBFBFB")

    if implant_row_end >= 6:
        style_range(ws, f"J6:T{implant_row_end}", wrap=True, align="center")
        set_number_format(ws, f"O6:R{implant_row_end}", '#,##0.00')
        for row in range(6, implant_row_end + 1):
            ws[f"N{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"S{row}"].fill = FORMULA_BLUE_FILL
            ws[f"T{row}"].fill = FORMULA_GREEN_FILL
            level = normalize_text(ws[f"J{row}"].value)
            if level == "Family":
                style_row(ws, row, 10, 20, fill=PatternFill("solid", fgColor="EAF4EA"), bold=True, wrap=True)
                ws[f"S{row}"].fill = FORMULA_BLUE_FILL
                ws[f"T{row}"].fill = FORMULA_GREEN_FILL
            elif level == "Brand":
                style_row(ws, row, 10, 20, fill=PatternFill("solid", fgColor="F4FAEE"), wrap=True)
                ws[f"L{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
                ws[f"S{row}"].fill = FORMULA_BLUE_FILL
                ws[f"T{row}"].fill = FORMULA_GREEN_FILL
            elif level == "Item":
                ws[f"N{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
                ws[f"L{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
                ws[f"M{row}"].font = Font(name="Calibri", size=10)
                if row % 2 == 0:
                    for col in range(10, 21):
                        if col not in {19, 20}:
                            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FCFCFC")

    for cell in [PHARMACY_VARIANCE_OT_RESULT_CELL, PHARMACY_VARIANCE_IMPLANT_RESULT_CELL]:
        ws[cell].number_format = '#,##0.00'
        ws[cell].fill = RESULT_FILL
        ws[cell].font = Font(bold=True)

    green_fill = PatternFill("solid", fgColor="E2F0D9")
    ws.conditional_formatting.add(
        f"H{ot_row_start}:H{ot_row_end if shortlist_rows else ot_row_start}",
        FormulaRule(formula=[f'H{ot_row_start}="{SELECTION_INCLUDE}"'], fill=green_fill),
    )
    if implant_row_end >= 6:
        ws.conditional_formatting.add(
            f"S6:S{implant_row_end}",
            FormulaRule(formula=[f'S6="{SELECTION_INCLUDE}"'], fill=green_fill),
        )

    column_widths = {
        "A": 42, "B": 14, "C": 14, "D": 16, "E": 14, "F": 18, "G": 14, "H": 14,
        "J": 12, "K": 24, "L": 20, "M": 14, "N": 46, "O": 14, "P": 12, "Q": 12, "R": 14, "S": 14, "T": 16,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    for row in [1, 2, 3, 4, 5]:
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[4].height = 34
    for row in range(6, implant_row_end + 1 if implant_row_end >= 6 else 6):
        ws.row_dimensions[row].height = 20


def write_implants_sheet(workbook, implant_rows: list[dict[str, str]]) -> None:
    ws = workbook["Implants"]
    existing_code_order = [
        normalize_code(ws[f"P{row_number}"].value)
        for row_number in range(2, ws.max_row + 1)
        if normalize_code(ws[f"P{row_number}"].value)
    ]
    by_code = {normalize_code(row.get("item_code")): row for row in implant_rows}
    ordered_rows: list[dict[str, str]] = []
    seen_codes: set[str] = set()

    for code in existing_code_order:
        row = by_code.get(code)
        if row:
            ordered_rows.append(row)
            seen_codes.add(code)

    for row in implant_rows:
        code = normalize_code(row.get("item_code"))
        if code and code not in seen_codes:
            ordered_rows.append(row)
            seen_codes.add(code)

    for row_number in range(2, ws.max_row + 1):
        for col_idx in range(1, len(IMPLANTS_SHEET_COLUMNS) + 1):
            ws.cell(row=row_number, column=col_idx, value=None)

    for row_number, row in enumerate(ordered_rows, start=2):
        for col_idx, fieldname in enumerate(IMPLANTS_SHEET_COLUMNS, start=1):
            raw = row.get(fieldname, "")
            value = maybe_float(raw) if normalize_text(raw).replace(".", "", 1).replace("-", "", 1).isdigit() else raw
            ws.cell(row=row_number, column=col_idx, value=value)


def write_services_selection(
    workbook,
    service_lookup: dict[str, dict[str, str]],
    optional_service_rows: list[dict[str, str]],
    rate_lookup: dict[str, RateRow],
    service_rate_field: str,
    args: argparse.Namespace,
) -> None:
    ws = workbook["Services Selection"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    selected_rows: list[dict[str, str]] = []
    if args.services_selection_mode == "fixed":
        for code in DEFAULT_SERVICES_SELECTION_CODES:
            row = service_lookup.get(normalize_code(code))
            if row:
                selected_rows.append(row)
    else:
        candidates = []
        for row in optional_service_rows:
            quantity_p50 = maybe_float(row.get("quantity_p50")) or 0.0
            rate_row = rate_lookup.get(normalize_code(row.get("item_code")))
            tariff_rate = tariff_rate_for_add_on(rate_row)
            rate_p50 = tariff_rate if tariff_rate is not None else (maybe_float(row.get(service_rate_field)) or 0.0)
            presence_rate = maybe_float(row.get("case_presence_rate")) or 0.0
            contribution = quantity_p50 * rate_p50 * presence_rate / 100.0 if rate_p50 else 0.0
            candidates.append((contribution, presence_rate, rate_p50, row))
        candidates.sort(
            key=lambda item: (
                -item[0],
                -item[1],
                -item[2],
                normalize_text(item[3].get("item_name")),
            )
        )
        if args.services_selection_count > 0:
            selected_rows = [item[3] for item in candidates[: args.services_selection_count]]
        else:
            selected_rows = [item[3] for item in candidates]

    for row_number in range(4, max(ws.max_row, 4 + len(selected_rows) + 5)):
        for col in "ABCDEFGHIJK":
            ws[f"{col}{row_number}"] = None

    ws["A3"] = "item_code"
    ws["B3"] = "item_name"
    ws["C3"] = "fc_estimate_bucket"
    ws["D3"] = "grouping"
    ws["E3"] = "case_presence_rate"
    ws["F3"] = "quantity_p50 (Typical Quantity)"
    ws["G3"] = "quantity_p25"
    ws["H3"] = "quantity_p75"
    ws["I3"] = "selected_cash_rate_from_tariff"
    ws["J3"] = "Typical gross amount"
    ws["K3"] = "How to organize/priortize"
    ws["A1"] = "Services - Add Ons"
    ws["B1"] = "Use tariff-backed optional add-on services to fine-tune the FC estimate where the default builder does not auto-include them."

    for row_number, row in enumerate(selected_rows, start=4):
        quantity_p50 = maybe_float(row.get("quantity_p50"))
        quantity_p25 = maybe_float(row.get("quantity_p25"))
        quantity_p75 = maybe_float(row.get("quantity_p75"))
        rate_row = rate_lookup.get(normalize_code(row.get("item_code")))
        rate_p50 = tariff_rate_for_add_on(rate_row)
        presence_rate = maybe_float(row.get("case_presence_rate"))

        ws[f"A{row_number}"] = row.get("item_code")
        ws[f"B{row_number}"] = row.get("item_name")
        ws[f"C{row_number}"] = row.get("fc_estimate_bucket")
        ws[f"D{row_number}"] = row.get("grouping")
        ws[f"E{row_number}"] = presence_rate
        ws[f"F{row_number}"] = quantity_p50
        ws[f"G{row_number}"] = quantity_p25
        ws[f"H{row_number}"] = quantity_p75
        ws[f"I{row_number}"] = rate_p50
        ws[f"J{row_number}"] = ((quantity_p50 or 0.0) * rate_p50) if rate_p50 is not None else None
        ws[f"K{row_number}"] = f"=IFERROR(J{row_number}*E{row_number}/100,0)"

    style_range(ws, "A1:K1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left", wrap=True)
    style_range(ws, "A3:K3", fill=SUBHEADER_FILL, bold=True, wrap=True)
    if selected_rows:
        selection_end_row = 3 + len(selected_rows)
        style_range(ws, f"A4:K{selection_end_row}", wrap=True)
        set_number_format(ws, f"E4:E{selection_end_row}", '0.00')
        set_number_format(ws, f"F4:K{selection_end_row}", '#,##0.00')
        for row_number in range(4, selection_end_row + 1):
            ws[f"B{row_number}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"C{row_number}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"D{row_number}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"I{row_number}"].fill = FORMULA_BLUE_FILL
            ws[f"J{row_number}"].fill = FORMULA_GREEN_FILL
            ws[f"K{row_number}"].fill = RESULT_FILL
            if row_number % 2 == 0:
                for col in range(1, 12):
                    if col not in {9, 10, 11}:
                        ws.cell(row=row_number, column=col).fill = PatternFill("solid", fgColor="FBFBFB")

    ws["I3"].comment = Comment("Tariff-driven cash rate for the optional add-on service code.", "Codex")
    ws["J3"].comment = Comment("Typical gross amount = quantity_p50 x tariff-derived cash rate.", "Codex")
    ws["K3"].comment = Comment("Expected contribution proxy = typical gross amount x presence rate.", "Codex")

    column_widths = {
        "A": 14,
        "B": 44,
        "C": 24,
        "D": 28,
        "E": 14,
        "F": 18,
        "G": 12,
        "H": 12,
        "I": 18,
        "J": 18,
        "K": 18,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[3].height = 24


def build_builder_value_formula(choice_ref: str, p25_ref: str, p50_ref: str, p75_ref: str, manual_ref: str) -> str:
    return f'=IF({choice_ref}="P25",{p25_ref},IF({choice_ref}="P50",{p50_ref},IF({choice_ref}="P75",{p75_ref},{manual_ref})))'


def build_day_rounding_formula(expression: str) -> str:
    base_expr = expression[1:] if expression.startswith("=") else expression
    return f"=INT(({base_expr}))+IF(MOD(({base_expr}),1)>0.3,1,0)"


def build_mode_pick_formula(mode_ref: str, low_ref: str, typical_ref: str, high_ref: str) -> str:
    low_expr = low_ref[1:] if isinstance(low_ref, str) and low_ref.startswith("=") else low_ref
    typical_expr = typical_ref[1:] if isinstance(typical_ref, str) and typical_ref.startswith("=") else typical_ref
    high_expr = high_ref[1:] if isinstance(high_ref, str) and high_ref.startswith("=") else high_ref
    return f'=IF({mode_ref}="Low",{low_expr},IF({mode_ref}="Typical",{typical_expr},{high_expr}))'


def build_room_pick_formula(room_ref: str, general_ref: str, twin_ref: str, single_ref: str) -> str:
    general_expr = general_ref[1:] if isinstance(general_ref, str) and general_ref.startswith("=") else general_ref
    twin_expr = twin_ref[1:] if isinstance(twin_ref, str) and twin_ref.startswith("=") else twin_ref
    single_expr = single_ref[1:] if isinstance(single_ref, str) and single_ref.startswith("=") else single_ref
    return f'=IF({room_ref}="General",{general_expr},IF({room_ref}="Twin",{twin_expr},{single_expr}))'


def ot_slot_ref_range(column_key: str) -> str:
    column = OT_SLOT_REFERENCE_COLS[column_key]
    return f"'{SHEET_REFERENCE}'!${column}${OT_SLOT_REFERENCE_START_ROW}:${column}${OT_SLOT_REFERENCE_END_ROW}"


def snap_to_supported_ot_slot_hours(value: float, supported_hours: list[float]) -> float:
    if value <= 0 or not supported_hours:
        return 0.0
    ordered = sorted({float(hours) for hours in supported_hours if float(hours) > 0})
    if not ordered:
        return 0.0
    best = ordered[0]
    best_distance = abs(best - value)
    for candidate in ordered[1:]:
        distance = abs(candidate - value)
        if distance < best_distance or (distance == best_distance and candidate > best):
            best = candidate
            best_distance = distance
    return float(best)


def build_nearest_supported_ot_slot_formula(expression: str) -> str:
    base_expr = expression[1:] if isinstance(expression, str) and expression.startswith("=") else expression
    hours_range = ot_slot_ref_range("hours")
    lower_match = f"MATCH(({base_expr}),{hours_range},1)"
    lower_hours = f"INDEX({hours_range},{lower_match})"
    upper_hours = f"INDEX({hours_range},{lower_match}+1)"
    return (
        f'=IF(({base_expr})<=0,0,'
        f'IF(({base_expr})<=MIN({hours_range}),MIN({hours_range}),'
        f'IF(({base_expr})>=MAX({hours_range}),MAX({hours_range}),'
        f'IF((({base_expr})-{lower_hours})<({upper_hours}-({base_expr})),{lower_hours},{upper_hours}))))'
    )


def build_resolved_ot_slot_formula(duration_ref: str) -> str:
    normalized_formula = build_nearest_supported_ot_slot_formula(duration_ref).lstrip("=")
    return f'=IF(({duration_ref})<=0,"",{normalized_formula})'


def build_resolved_ot_mode_formula(emergency_ref: str) -> str:
    return f'=IF({emergency_ref}="Yes","emergency","normal")'


def build_ot_slot_lookup_formula(duration_ref: str, emergency_ref: str, column_key: str, tariff_ref: str = "Builder!E5") -> str:
    key_col = TARIFF_OT_SLOT_MATRIX_COLS["matrix_key"]
    matrix_field = {
        "code": "item_code",
        "name": "item_name",
        "hours": "ot_slot_hours",
        "general": "general",
        "twin": "twin",
        "single": "single",
        "icu": "icu",
    }[column_key]
    target_col = TARIFF_OT_SLOT_MATRIX_COLS[matrix_field]
    resolved_formula = build_resolved_ot_slot_formula(duration_ref).lstrip("=")
    resolved_mode = build_resolved_ot_mode_formula(emergency_ref).lstrip("=")
    resolved_key = f'{tariff_ref}&"|"&{resolved_mode}&"|"&({resolved_formula})'
    if column_key in {"hours", "general", "twin", "single", "icu"}:
        return (
            f'=IFERROR(0+INDEX(\'{SHEET_REFERENCE}\'!${target_col}${TARIFF_OT_SLOT_MATRIX_START_ROW}:${target_col}$1000,'
            f'MATCH({resolved_key},'
            f'\'{SHEET_REFERENCE}\'!${key_col}${TARIFF_OT_SLOT_MATRIX_START_ROW}:${key_col}$1000,0)),0)'
        )
    return (
        f'=IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${target_col}${TARIFF_OT_SLOT_MATRIX_START_ROW}:${target_col}$1000,'
        f'MATCH({resolved_key},'
        f'\'{SHEET_REFERENCE}\'!${key_col}${TARIFF_OT_SLOT_MATRIX_START_ROW}:${key_col}$1000,0)),"")'
    )


def build_ot_slot_lookup_from_resolved_formula(
    resolved_hours_ref: str,
    emergency_ref: str,
    column_key: str,
    tariff_ref: str = "Builder!E5",
) -> str:
    key_col = TARIFF_OT_SLOT_MATRIX_COLS["matrix_key"]
    matrix_field = {
        "code": "item_code",
        "name": "item_name",
        "hours": "ot_slot_hours",
        "general": "general",
        "twin": "twin",
        "single": "single",
        "icu": "icu",
    }[column_key]
    target_col = TARIFF_OT_SLOT_MATRIX_COLS[matrix_field]
    resolved_mode = build_resolved_ot_mode_formula(emergency_ref).lstrip("=")
    resolved_key = f'{tariff_ref}&"|"&{resolved_mode}&"|"&({resolved_hours_ref})'
    if column_key in {"hours", "general", "twin", "single", "icu"}:
        return (
            f'=IF({resolved_hours_ref}="","",IFERROR(0+INDEX(\'{SHEET_REFERENCE}\'!${target_col}${TARIFF_OT_SLOT_MATRIX_START_ROW}:'
            f'${target_col}$1000,MATCH({resolved_key},'
            f'\'{SHEET_REFERENCE}\'!${key_col}${TARIFF_OT_SLOT_MATRIX_START_ROW}:${key_col}$1000,0)),0))'
        )
    return (
        f'=IF({resolved_hours_ref}="","",IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${target_col}${TARIFF_OT_SLOT_MATRIX_START_ROW}:'
        f'${target_col}$1000,MATCH({resolved_key},'
        f'\'{SHEET_REFERENCE}\'!${key_col}${TARIFF_OT_SLOT_MATRIX_START_ROW}:${key_col}$1000,0)),""))'
    )


def bucket_label(parent_bucket: str) -> str:
    mapping = {
        "Investigation": "Investigations",
        "Investigations": "Investigations",
        "Room Charges": "Room Charges",
        "Procedure Charges": "Procedure / OT Charges",
        "Procedure / OT Charges": "Procedure / OT Charges",
        "OT": "Procedure / OT Charges",
        "Physiotherapy": "Procedure / OT Charges",
        "Cath Lab": "Procedure / OT Charges",
        "Administrative Charges": "Procedure / OT Charges",
        "Emergency": "Procedure / OT Charges",
        "Bedside services": "Bedside Services",
        "Bedside Services": "Bedside Services",
        "Professional Fee": "Professional Fees",
        "Professional Fees": "Professional Fees",
        "Pharmacy": "Pharmacy",
        "Other Services": "Procedure / OT Charges",
        "Optional Add-Ons": "Optional Add-Ons",
        "Drug Administration Charges": "Drug Administration Charges",
    }
    return mapping.get(parent_bucket, parent_bucket)


def grouped_residual_parent_bucket(row: dict[str, Any]) -> str:
    grouping = normalize_text(row.get("grouping"))
    sample_bucket = normalize_text(row.get("sample_fc_estimate_bucket"))
    if grouping == "Consultation Charges" and sample_bucket in {"Physiotherapy", "Professional Fees", "Professional Fee"}:
        return "Professional Fees"
    return bucket_label(sample_bucket)


def core_line_definitions(args: argparse.Namespace) -> list[dict[str, str]]:
    rows = [
        {"name": "X-RAY KNEE JOINT AP & LATERAL VIEW (BEDSIDE)", "parent": "Investigations", "sub": "Investigations", "source": "Template", "how": "Auto-Included", "code": "XRY5090", "kind": "template"},
        {"name": "Nursing - Room", "parent": "Room Charges", "sub": "Ward Care", "source": "Logic", "how": "Ward days x rate", "code": "ROM5189", "kind": "driver", "driver": "ward"},
        {"name": "Nursing - ICU", "parent": "Room Charges", "sub": "Critical Care", "source": "Logic", "how": "ICU days x rate", "code": "ROM5189", "kind": "driver", "driver": "icu", "icu_only": "true"},
        {"name": "DMO", "parent": "Room Charges", "sub": "Ward Care", "source": "Logic", "how": "Ward days x rate", "code": "ROM0093", "kind": "driver", "driver": "ward"},
        {"name": "ICU - Surgical", "parent": "Room Charges", "sub": "Critical Care", "source": "Logic", "how": "ICU days x rate", "code": "ROM5009", "kind": "driver", "driver": "icu", "icu_only": "true"},
        {"name": "Bed Charges - Ward", "parent": "Room Charges", "sub": "Ward Care", "source": "Logic", "how": "Ward days x room rate", "code": "ROOM_BED", "kind": "ward_bed"},
        {"name": "CSSD CHARGES FOR GA", "parent": "Procedure / OT Charges", "sub": "OT Charges", "source": "Template", "how": "Auto-Included", "code": "RNS5005", "kind": "template"},
        {"name": "Medical Records", "parent": "Bedside Services", "sub": "Administrative", "source": "Logic", "how": "Fixed 1", "code": "RNS0120", "kind": "fixed_one"},
        {"name": "PHYSIOTHERAPY PACKAGE 5 VISITS", "parent": "Procedure / OT Charges", "sub": "Physiotherapy", "source": "Template", "how": "Auto-Included", "code": "PHY5082", "kind": "template"},
        {"name": "HAEMOGLOBIN", "parent": "Investigations", "sub": "Investigations", "source": "Template", "how": "Auto-Included", "code": "PAT0045", "kind": "template"},
        {"name": "CBP (COMPLETE BLOOD PICTURE)", "parent": "Investigations", "sub": "Investigations", "source": "Template", "how": "Auto-Included", "code": "PAT0042", "kind": "template"},
        {"name": "Instrument Charges (Major)", "parent": "Procedure / OT Charges", "sub": "OT Charges", "source": "Logic", "how": "Fixed 1", "code": "OTI0018", "kind": "fixed_one"},
        {"name": "OT Disinfection Charges", "parent": "Procedure / OT Charges", "sub": "OT Charges", "source": "Logic", "how": "Fixed 1", "code": "OTI0015", "kind": "fixed_one"},
        {"name": "Post Surgery Recovery Charges", "parent": "Procedure / OT Charges", "sub": "OT Charges", "source": "Logic", "how": "Fixed 1", "code": "OTC5005", "kind": "fixed_one"},
        {"name": "OT Charges", "parent": "Procedure / OT Charges", "sub": "OT Charges", "source": "Logic", "how": "Selected OT duration snapped to the nearest supported tariff OT slot", "code": "", "kind": "ot_hours"},
        {"name": "Cath Lab Charges", "parent": "Procedure / OT Charges", "sub": "Cath Lab Hours", "source": "Historical Cath Lab Family", "how": "Actual billed cath-lab slot-family P25 / P50 / P75 from the selected historical payer basis.", "code": "", "kind": "cath_lab_history"},
        {"name": "Intensivist Per Day", "parent": "Room Charges", "sub": "Critical Care", "source": "Logic", "how": "ICU days x rate", "code": "ICC0002", "kind": "driver", "driver": "icu", "icu_only": "true"},
        {"name": "Assistant Intensivist Per Day", "parent": "Room Charges", "sub": "Critical Care", "source": "Logic", "how": "ICU days x rate", "code": "ICC0001", "kind": "driver", "driver": "icu", "icu_only": "true"},
        {"name": "Ward Consumables", "parent": "Room Charges", "sub": "Ward Care", "source": "Logic", "how": "LOS days x rate", "code": "HSP5013", "kind": "driver", "driver": "los"},
        {"name": "Warmer", "parent": "Bedside Services", "sub": "Bedside", "source": "Template", "how": "Auto-Included", "code": "EME0087", "kind": "template"},
        {"name": "Monitor Per Day", "parent": "Room Charges", "sub": "Critical Care", "source": "Logic", "how": "ICU days x rate", "code": "EME0019", "kind": "driver", "driver": "icu", "icu_only": "true"},
        {"name": "Oxygen Per Hour", "parent": "Bedside Services", "sub": "Bedside", "source": "Template", "how": "Auto-Included", "code": "EME0017", "kind": "template"},
        {"name": "Diet Consultation", "parent": "Bedside Services", "sub": "Consultation", "source": "Template", "how": "Auto-Included", "code": "DIE0001", "kind": "template"},
        {"name": "Dressing - Minor", "parent": "Bedside Services", "sub": "Bedside", "source": "Template", "how": "Auto-Included", "code": "CAS0007", "kind": "template"},
        {"name": "Bedside ECG", "parent": "Bedside Services", "sub": "Bedside", "source": "Template", "how": "Auto-Included", "code": "CAR5341", "kind": "template"},
        {"name": "Albumin", "parent": "Investigations", "sub": "Investigations", "source": "Template", "how": "Auto-Included", "code": "BIO0162", "kind": "template"},
        {"name": "Sodium", "parent": "Investigations", "sub": "Investigations", "source": "Template", "how": "Auto-Included", "code": "BIO0004", "kind": "template"},
        {"name": "Electrolytes", "parent": "Investigations", "sub": "Investigations", "source": "Template", "how": "Auto-Included", "code": "BIO0003", "kind": "template"},
        {"name": "Creatinine", "parent": "Investigations", "sub": "Investigations", "source": "Template", "how": "Auto-Included", "code": "BIO0002", "kind": "template"},
        {"name": "Urea", "parent": "Investigations", "sub": "Investigations", "source": "Template", "how": "Auto-Included", "code": "BIO0001", "kind": "template"},
        {"name": "MLC Charges", "parent": "Bedside Services", "sub": "Administrative", "source": "Logic", "how": "Applied only when MLC input is Yes", "code": "HSP0047", "kind": "mlc_charge"},
        {"name": "Drug Administration Charges", "parent": "Drug Administration Charges", "sub": "Pharmacy Related", "source": "Logic", "how": "12.5% of pharmacy total", "code": "", "kind": "drug_admin"},
        {"name": "Surgeon", "parent": "Professional Fees", "sub": "Professional Fees", "source": "Logic", "how": "Cash PF % of pre-PF subtotal", "code": "", "kind": "pf_surgeon"},
        {"name": "Assistant Surgeon", "parent": "Professional Fees", "sub": "Professional Fees", "source": "Logic", "how": "Cash PF % of surgeon fee", "code": "", "kind": "pf_asst_surgeon"},
        {"name": "Anesthetist", "parent": "Professional Fees", "sub": "Professional Fees", "source": "Logic", "how": "Cash PF % of surgeon fee", "code": "", "kind": "pf_anesthetist"},
        {"name": "Assistant Anesthetist", "parent": "Professional Fees", "sub": "Professional Fees", "source": "Logic", "how": "Cash PF % of anesthetist fee", "code": "", "kind": "pf_asst_anesthetist"},
        {"name": "IP Drugs & Medications", "parent": "Pharmacy", "sub": "IP Pharmacy", "source": "History", "how": "Bucket quartiles", "code": "", "kind": "pharmacy_ip_drugs"},
        {"name": "IP Consumables", "parent": "Pharmacy", "sub": "IP Pharmacy", "source": "History", "how": "Bucket quartiles", "code": "", "kind": "pharmacy_ip_consumables"},
        {"name": "OT Drugs & Medications", "parent": "Pharmacy", "sub": "OT Pharmacy", "source": "History", "how": "Bucket quartiles", "code": "", "kind": "pharmacy_ot_drugs"},
        {"name": "OT Consumables", "parent": "Pharmacy", "sub": "OT Pharmacy", "source": "Advanced", "how": "OT consumables variance controls", "code": "", "kind": "pharmacy_ot_consumables"},
        {"name": "Implants", "parent": "Pharmacy", "sub": "Implants", "source": "Advanced", "how": "Implant variance controls", "code": "", "kind": "pharmacy_implants"},
    ]
    if normalize_text(getattr(args, "include_procedure_row", "yes")).lower() != "no":
        is_robotic_procedure = is_robotic_service_row(
            {
                "item_code": args.procedure_code,
                "item_name": args.procedure_label,
                "grouping": "Robotic Charges" if "ROBO" in normalize_text(args.procedure_label).upper() else "",
            }
        )
        rows.insert(
            11,
            {
                "name": args.procedure_label,
                "parent": "Procedure / OT Charges",
                "sub": "OT Charges",
                "source": "Template",
                "how": "Auto-Included",
                "code": normalize_code(args.procedure_code),
                "kind": "template",
                "robotic_controlled": "true" if is_robotic_procedure else "false",
            },
        )
    return rows


def write_reference_sheet(
    ws,
    quartiles_json: dict[str, Any],
    bucket_quartiles: dict[str, tuple[float, float, float]],
    ip_pharmacy_per_day_metrics: dict[str, Any],
    cath_lab_metrics: dict[str, Any],
    service_line_count_metrics: dict[str, Any],
    cleaned_service_rows: list[dict[str, str]],
    optional_service_rows: list[dict[str, str]],
    rate_lookup: dict[str, RateRow],
    implant_rows: list[dict[str, str]],
    ot_slot_rows: list[OtSlotRateRow],
    org_tariff_reference_rows: list[dict[str, str]],
    tariff_rate_matrix_rows: list[dict[str, str]],
    tariff_ot_slot_matrix_rows: list[dict[str, str]],
    insurance_policy_rows: list[dict[str, str]],
    payer_basis_summary: dict[str, Any],
    payer_basis_service_rows: list[dict[str, str]],
    payer_basis_pharmacy_rows: list[dict[str, str]],
    payer_basis_resolution_rows: list[dict[str, str]],
    pf_payor_summary_rows: list[dict[str, str]],
    actual_rows: list[dict[str, Any]],
) -> None:
    payer_basis_service_numeric_fields = {
        "case_presence_rate",
        "quantity_p25",
        "quantity_p50",
        "quantity_p75",
        "amount_cash_typical",
        "tariff_general",
        "tariff_twin",
        "tariff_single",
        "tariff_icu",
    }
    payer_basis_pharmacy_numeric_fields = {
        "case_presence_rate",
        "ot_quantity_typical_cleaned",
        "ot_amount_typical_cleaned",
        "overall_amount_typical_cleaned",
    }
    org_tariff_numeric_fields = {"case_count"}
    tariff_rate_numeric_fields = {"general", "twin", "single", "icu"}
    tariff_ot_numeric_fields = {"ot_slot_hours", "general", "twin", "single", "icu"}
    ws["A1"] = "Reference Data"
    apply_cell_style(ws["A1"], fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")

    row = 3
    ws[f"A{row}"] = "LOS / ICU / Ward / OT Quartiles"
    apply_cell_style(ws[f"A{row}"], fill=SUBHEADER_FILL, bold=True, align="left")
    row += 1
    ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"], ws[f"D{row}"] = "metric", "p25", "p50", "p75"
    apply_cell_style(ws[f"A{row}"], fill=SELECTION_FILL, bold=True)
    apply_cell_style(ws[f"B{row}"], fill=SELECTION_FILL, bold=True)
    apply_cell_style(ws[f"C{row}"], fill=SELECTION_FILL, bold=True)
    apply_cell_style(ws[f"D{row}"], fill=SELECTION_FILL, bold=True)
    row += 1
    for metric in ["los_days", "icu_days", "ward_days", "ot_hours"]:
        ws[f"A{row}"] = metric
        ws[f"B{row}"] = get_metric(quartiles_json, metric, "p25")
        ws[f"C{row}"] = get_metric(quartiles_json, metric, "p50")
        ws[f"D{row}"] = get_metric(quartiles_json, metric, "p75")
        row += 1

    row += 1
    ws[f"A{row}"] = "Pharmacy Bucket Quartiles"
    apply_cell_style(ws[f"A{row}"], fill=SUBHEADER_FILL, bold=True, align="left")
    row += 1
    for idx, header in enumerate(["bucket", "p25", "p50", "p75"], start=1):
        cell = ws.cell(row=row, column=idx, value=header)
        apply_cell_style(cell, fill=SELECTION_FILL, bold=True)
    row += 1
    for bucket_name in ["ip_drugs", "ip_consumables", "ot_drugs", "ot_consumables", "implants"]:
        p25, p50, p75 = bucket_quartiles[bucket_name]
        ws[f"A{row}"] = bucket_name
        ws[f"B{row}"] = p25
        ws[f"C{row}"] = p50
        ws[f"D{row}"] = p75
        row += 1

    row += 1
    ws[f"A{row}"] = "IP Pharmacy Per LOS Day Quartiles"
    apply_cell_style(ws[f"A{row}"], fill=SUBHEADER_FILL, bold=True, align="left")
    row += 1
    ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"], ws[f"D{row}"] = "metric", "p25", "p50", "p75"
    apply_cell_style(ws[f"A{row}"], fill=SELECTION_FILL, bold=True)
    apply_cell_style(ws[f"B{row}"], fill=SELECTION_FILL, bold=True)
    apply_cell_style(ws[f"C{row}"], fill=SELECTION_FILL, bold=True)
    apply_cell_style(ws[f"D{row}"], fill=SELECTION_FILL, bold=True)
    row += 1
    for metric_name in ["ip_drugs_per_los_day", "ip_consumables_per_los_day"]:
        ws[f"A{row}"] = metric_name
        ws[f"B{row}"] = get_ip_pharmacy_per_day_metric(ip_pharmacy_per_day_metrics, metric_name, "p25")
        ws[f"C{row}"] = get_ip_pharmacy_per_day_metric(ip_pharmacy_per_day_metrics, metric_name, "p50")
        ws[f"D{row}"] = get_ip_pharmacy_per_day_metric(ip_pharmacy_per_day_metrics, metric_name, "p75")
        row += 1

    row += 1
    ws[f"A{row}"] = "Cleaned Service Line Count Quartiles"
    apply_cell_style(ws[f"A{row}"], fill=SUBHEADER_FILL, bold=True, align="left")
    row += 1
    ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"], ws[f"D{row}"] = "metric", "p25", "p50", "p75"
    apply_cell_style(ws[f"A{row}"], fill=SELECTION_FILL, bold=True)
    apply_cell_style(ws[f"B{row}"], fill=SELECTION_FILL, bold=True)
    apply_cell_style(ws[f"C{row}"], fill=SELECTION_FILL, bold=True)
    apply_cell_style(ws[f"D{row}"], fill=SELECTION_FILL, bold=True)
    row += 1
    ws[f"A{row}"] = "cleaned_distinct_service_line_count"
    ws[f"B{row}"] = get_service_line_count_metric(service_line_count_metrics, "p25")
    ws[f"C{row}"] = get_service_line_count_metric(service_line_count_metrics, "p50")
    ws[f"D{row}"] = get_service_line_count_metric(service_line_count_metrics, "p75")
    row += 1

    row += 1
    ws[f"A{row}"] = "Cleaned Services Template"
    apply_cell_style(ws[f"A{row}"], fill=SUBHEADER_FILL, bold=True, align="left")
    row += 1
    service_headers = ["item_code", "item_name", "fc_estimate_bucket", "grouping", "case_presence_rate", "quantity_p25", "quantity_p50", "quantity_p75", "amount_cash_typical"]
    for idx, header in enumerate(service_headers, start=1):
        cell = ws.cell(row=row, column=idx, value=header)
        apply_cell_style(cell, fill=SELECTION_FILL, bold=True)
    row += 1
    for service_row in cleaned_service_rows:
        for idx, header in enumerate(service_headers, start=1):
            ws.cell(row=row, column=idx, value=service_row.get(header))
        row += 1

    row += 1
    ws[f"A{row}"] = "Optional Service Rows"
    apply_cell_style(ws[f"A{row}"], fill=SUBHEADER_FILL, bold=True, align="left")
    row += 1
    for idx, header in enumerate(service_headers, start=1):
        cell = ws.cell(row=row, column=idx, value=header)
        apply_cell_style(cell, fill=SELECTION_FILL, bold=True)
    row += 1
    for service_row in optional_service_rows:
        for idx, header in enumerate(service_headers, start=1):
            ws.cell(row=row, column=idx, value=service_row.get(header))
        row += 1

    row += 1
    ws[f"A{row}"] = "TR1 Tariff Rates"
    apply_cell_style(ws[f"A{row}"], fill=SUBHEADER_FILL, bold=True, align="left")
    row += 1
    rate_headers = ["item_code", "item_name", "general", "twin", "single", "icu"]
    for idx, header in enumerate(rate_headers, start=1):
        cell = ws.cell(row=row, column=idx, value=header)
        apply_cell_style(cell, fill=SELECTION_FILL, bold=True)
    row += 1
    for code in sorted(rate_lookup):
        rate = rate_lookup[code]
        values = [code, rate.item_name, rate.general, rate.twin, rate.single, rate.icu]
        for idx, value in enumerate(values, start=1):
            ws.cell(row=row, column=idx, value=value)
        row += 1

    row += 1
    ws[f"A{row}"] = "Implant Reference"
    apply_cell_style(ws[f"A{row}"], fill=SUBHEADER_FILL, bold=True, align="left")
    row += 1
    implant_headers = ["implant_family", "brand_family", "item_code", "item_name", "item_presence_rate", "item_quantity_p50", "typical_rate_p50"]
    for idx, header in enumerate(implant_headers, start=1):
        cell = ws.cell(row=row, column=idx, value=header)
        apply_cell_style(cell, fill=SELECTION_FILL, bold=True)
    row += 1
    for implant_row in implant_rows:
        for idx, header in enumerate(implant_headers, start=1):
            ws.cell(row=row, column=idx, value=implant_row.get(header))
        row += 1

    ot_header_row = OT_SLOT_REFERENCE_START_ROW - 2
    ws[f"J{ot_header_row}"] = "OT Tariff Slot Reference"
    apply_cell_style(ws[f"J{ot_header_row}"], fill=SUBHEADER_FILL, bold=True, align="left")
    ws[f"J{ot_header_row+1}"], ws[f"K{ot_header_row+1}"], ws[f"L{ot_header_row+1}"], ws[f"M{ot_header_row+1}"], ws[f"N{ot_header_row+1}"], ws[f"O{ot_header_row+1}"], ws[f"P{ot_header_row+1}"], ws[f"Q{ot_header_row+1}"], ws[f"R{ot_header_row+1}"] = (
        "tariff_code",
        "ot_slot_hours",
        "ot_mode",
        "item_code",
        "item_name",
        "general",
        "twin",
        "single",
        "icu",
    )
    for col in ["J", "K", "L", "M", "N", "O", "P", "Q", "R"]:
        apply_cell_style(ws[f"{col}{ot_header_row+1}"], fill=SELECTION_FILL, bold=True)
    current_ot_row = OT_SLOT_REFERENCE_START_ROW
    for slot in ot_slot_rows[: OT_SLOT_REFERENCE_END_ROW - OT_SLOT_REFERENCE_START_ROW + 1]:
        ws[f"J{current_ot_row}"] = slot.tariff_code
        ws[f"K{current_ot_row}"] = slot.ot_slot_hours
        ws[f"L{current_ot_row}"] = slot.ot_mode
        ws[f"M{current_ot_row}"] = slot.item_code
        ws[f"N{current_ot_row}"] = slot.item_name
        ws[f"O{current_ot_row}"] = slot.general
        ws[f"P{current_ot_row}"] = slot.twin
        ws[f"Q{current_ot_row}"] = slot.single
        ws[f"R{current_ot_row}"] = slot.icu
        current_ot_row += 1

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col != 2 else 36
    for col, width in {"J": 12, "K": 14, "L": 12, "M": 12, "N": 28, "O": 12, "P": 12, "Q": 12, "R": 12}.items():
        ws.column_dimensions[col].width = width

    cath_header_row = 2
    ws[f"{CATH_LAB_REFERENCE_COLS['label']}{cath_header_row}"] = "Cath Lab Family Metrics"
    apply_cell_style(ws[f"{CATH_LAB_REFERENCE_COLS['label']}{cath_header_row}"], fill=SELECTION_FILL, bold=True)
    ws[f"{CATH_LAB_REFERENCE_COLS['label']}{cath_header_row + 1}"] = "metric"
    ws[f"{CATH_LAB_REFERENCE_COLS['p25']}{cath_header_row + 1}"] = "p25"
    ws[f"{CATH_LAB_REFERENCE_COLS['p50']}{cath_header_row + 1}"] = "p50"
    ws[f"{CATH_LAB_REFERENCE_COLS['p75']}{cath_header_row + 1}"] = "p75"
    for col in CATH_LAB_REFERENCE_COLS.values():
        apply_cell_style(ws[f"{col}{cath_header_row + 1}"], fill=SELECTION_FILL, bold=True)
    cath_metrics = cath_lab_metrics.get("metrics") or {}
    ws[f"{CATH_LAB_REFERENCE_COLS['label']}{cath_header_row + 2}"] = "cath_lab_amount_net"
    ws[f"{CATH_LAB_REFERENCE_COLS['p25']}{cath_header_row + 2}"] = as_float(cath_metrics.get("p25"))
    ws[f"{CATH_LAB_REFERENCE_COLS['p50']}{cath_header_row + 2}"] = as_float(cath_metrics.get("p50"))
    ws[f"{CATH_LAB_REFERENCE_COLS['p75']}{cath_header_row + 2}"] = as_float(cath_metrics.get("p75"))
    for col in CATH_LAB_REFERENCE_COLS.values():
        ws.column_dimensions[col].width = 18

    supported_ot_slot_hours = [
        hours
        for hours in (
            maybe_float(row.get("ot_slot_hours"))
            for row in tariff_ot_slot_matrix_rows
        )
        if hours is not None
    ]
    basis_headers = {field: field for field in PAYER_BASIS_SUMMARY_COLS}
    header_row = PAYER_BASIS_SUMMARY_START_ROW - 1
    for field, col in PAYER_BASIS_SUMMARY_COLS.items():
        ws[f"{col}{header_row}"] = basis_headers[field]
        apply_cell_style(ws[f"{col}{header_row}"], fill=SELECTION_FILL, bold=True)
    basis_order = payer_basis_summary.get("basis_order", PAYER_BASIS_OPTIONS)
    for idx, basis_label in enumerate(basis_order, start=PAYER_BASIS_SUMMARY_START_ROW):
        metrics = (payer_basis_summary.get("basis_metrics") or {}).get(basis_label, {})
        payor_counts = metrics.get("payor_counts") or {}
        clinical = metrics.get("clinical_drivers") or {}
        buckets = metrics.get("bucket_quartiles") or {}
        ip_day = metrics.get("ip_pharmacy_per_day") or {}
        svc = metrics.get("service_line_count") or {}
        cath_lab = metrics.get("cath_lab_amount") or {}
        values = {
            "basis_label": basis_label,
            "cohort_size": metrics.get("cohort_size", 0),
            "cash_count": payor_counts.get("Cash", 0),
            "gipsa_count": payor_counts.get("GIPSA Insurance", 0),
            "non_gipsa_count": payor_counts.get("Non-GIPSA Insurance", 0),
            "corporate_count": payor_counts.get("Corporate", 0),
            "los_p25": ((clinical.get("los_days") or {}).get("p25", 0)),
            "los_p50": ((clinical.get("los_days") or {}).get("p50", 0)),
            "los_p75": ((clinical.get("los_days") or {}).get("p75", 0)),
            "icu_p25": ((clinical.get("icu_days") or {}).get("p25", 0)),
            "icu_p50": ((clinical.get("icu_days") or {}).get("p50", 0)),
            "icu_p75": ((clinical.get("icu_days") or {}).get("p75", 0)),
            "ward_p25": ((clinical.get("ward_days") or {}).get("p25", 0)),
            "ward_p50": ((clinical.get("ward_days") or {}).get("p50", 0)),
            "ward_p75": ((clinical.get("ward_days") or {}).get("p75", 0)),
            "ot_p25": snap_to_supported_ot_slot_hours(((clinical.get("ot_hours") or {}).get("p25", 0)), supported_ot_slot_hours),
            "ot_p50": snap_to_supported_ot_slot_hours(((clinical.get("ot_hours") or {}).get("p50", 0)), supported_ot_slot_hours),
            "ot_p75": snap_to_supported_ot_slot_hours(((clinical.get("ot_hours") or {}).get("p75", 0)), supported_ot_slot_hours),
            "service_line_p25": svc.get("p25", 0),
            "service_line_p50": svc.get("p50", 0),
            "service_line_p75": svc.get("p75", 0),
            "ip_drugs_p25": ((buckets.get("ip_drugs") or {}).get("p25", 0)),
            "ip_drugs_p50": ((buckets.get("ip_drugs") or {}).get("p50", 0)),
            "ip_drugs_p75": ((buckets.get("ip_drugs") or {}).get("p75", 0)),
            "ip_consumables_p25": ((buckets.get("ip_consumables") or {}).get("p25", 0)),
            "ip_consumables_p50": ((buckets.get("ip_consumables") or {}).get("p50", 0)),
            "ip_consumables_p75": ((buckets.get("ip_consumables") or {}).get("p75", 0)),
            "ot_drugs_p25": ((buckets.get("ot_drugs") or {}).get("p25", 0)),
            "ot_drugs_p50": ((buckets.get("ot_drugs") or {}).get("p50", 0)),
            "ot_drugs_p75": ((buckets.get("ot_drugs") or {}).get("p75", 0)),
            "ot_consumables_p25": ((buckets.get("ot_consumables") or {}).get("p25", 0)),
            "ot_consumables_p50": ((buckets.get("ot_consumables") or {}).get("p50", 0)),
            "ot_consumables_p75": ((buckets.get("ot_consumables") or {}).get("p75", 0)),
            "implants_p25": ((buckets.get("implants") or {}).get("p25", 0)),
            "implants_p50": ((buckets.get("implants") or {}).get("p50", 0)),
            "implants_p75": ((buckets.get("implants") or {}).get("p75", 0)),
            "ip_drugs_day_p25": ((ip_day.get("ip_drugs_per_los_day") or {}).get("p25", 0)),
            "ip_drugs_day_p50": ((ip_day.get("ip_drugs_per_los_day") or {}).get("p50", 0)),
            "ip_drugs_day_p75": ((ip_day.get("ip_drugs_per_los_day") or {}).get("p75", 0)),
            "ip_consumables_day_p25": ((ip_day.get("ip_consumables_per_los_day") or {}).get("p25", 0)),
            "ip_consumables_day_p50": ((ip_day.get("ip_consumables_per_los_day") or {}).get("p50", 0)),
            "ip_consumables_day_p75": ((ip_day.get("ip_consumables_per_los_day") or {}).get("p75", 0)),
            "cath_lab_p25": cath_lab.get("p25", 0),
            "cath_lab_p50": cath_lab.get("p50", 0),
            "cath_lab_p75": cath_lab.get("p75", 0),
        }
        for field, col in PAYER_BASIS_SUMMARY_COLS.items():
            ws[f"{col}{idx}"] = values.get(field, "")

    svc_header_row = PAYER_BASIS_SERVICE_START_ROW - 1
    for field, col in PAYER_BASIS_SERVICE_COLS.items():
        ws[f"{col}{svc_header_row}"] = field
        apply_cell_style(ws[f"{col}{svc_header_row}"], fill=SELECTION_FILL, bold=True)
    for idx, row_data in enumerate(payer_basis_service_rows, start=PAYER_BASIS_SERVICE_START_ROW):
        for field, col in PAYER_BASIS_SERVICE_COLS.items():
            source_field = "basis_item_key" if field == "key" else field
            value: Any = row_data.get(source_field, "")
            if field in payer_basis_service_numeric_fields:
                number = maybe_float(value)
                value = number if number is not None else ""
            ws[f"{col}{idx}"] = value

    pharm_header_row = PAYER_BASIS_PHARMACY_START_ROW - 1
    for field, col in PAYER_BASIS_PHARMACY_COLS.items():
        ws[f"{col}{pharm_header_row}"] = field
        apply_cell_style(ws[f"{col}{pharm_header_row}"], fill=SELECTION_FILL, bold=True)
    for idx, row_data in enumerate(payer_basis_pharmacy_rows, start=PAYER_BASIS_PHARMACY_START_ROW):
        row_copy = dict(row_data)
        row_copy["basis_name_key"] = f"{row_data.get('basis_label','')}|{row_data.get('item_name','')}"
        for field, col in PAYER_BASIS_PHARMACY_COLS.items():
            source_field = "basis_item_key" if field == "key" else field
            value = row_copy.get(source_field, "")
            if field in payer_basis_pharmacy_numeric_fields:
                number = maybe_float(value)
                value = number if number is not None else ""
            ws[f"{col}{idx}"] = value

    org_header_row = ORG_TARIFF_REFERENCE_START_ROW - 1
    for field, col in ORG_TARIFF_REFERENCE_COLS.items():
        ws[f"{col}{org_header_row}"] = field
        apply_cell_style(ws[f"{col}{org_header_row}"], fill=SELECTION_FILL, bold=True)
    for idx, row_data in enumerate(org_tariff_reference_rows, start=ORG_TARIFF_REFERENCE_START_ROW):
        for field, col in ORG_TARIFF_REFERENCE_COLS.items():
            value = row_data.get(field, "")
            if field in org_tariff_numeric_fields:
                number = maybe_float(value)
                value = number if number is not None else ""
            ws[f"{col}{idx}"] = value

    tariff_header_row = TARIFF_RATE_MATRIX_START_ROW - 1
    for field, col in TARIFF_RATE_MATRIX_COLS.items():
        ws[f"{col}{tariff_header_row}"] = field
        apply_cell_style(ws[f"{col}{tariff_header_row}"], fill=SELECTION_FILL, bold=True)
    for idx, row_data in enumerate(tariff_rate_matrix_rows, start=TARIFF_RATE_MATRIX_START_ROW):
        for field, col in TARIFF_RATE_MATRIX_COLS.items():
            value = row_data.get(field, "")
            if field in tariff_rate_numeric_fields:
                number = maybe_float(value)
                value = number if number is not None else ""
            ws[f"{col}{idx}"] = value

    tariff_ot_header_row = TARIFF_OT_SLOT_MATRIX_START_ROW - 1
    for field, col in TARIFF_OT_SLOT_MATRIX_COLS.items():
        ws[f"{col}{tariff_ot_header_row}"] = field
        apply_cell_style(ws[f"{col}{tariff_ot_header_row}"], fill=SELECTION_FILL, bold=True)
    for idx, row_data in enumerate(tariff_ot_slot_matrix_rows, start=TARIFF_OT_SLOT_MATRIX_START_ROW):
        for field, col in TARIFF_OT_SLOT_MATRIX_COLS.items():
            value = row_data.get(field, "")
            if field in tariff_ot_numeric_fields:
                number = maybe_float(value)
                value = number if number is not None else ""
            ws[f"{col}{idx}"] = value

    insurance_header_row = INSURANCE_POLICY_START_ROW - 1
    for field, col in INSURANCE_POLICY_COLS.items():
        ws[f"{col}{insurance_header_row}"] = field
        apply_cell_style(ws[f"{col}{insurance_header_row}"], fill=SELECTION_FILL, bold=True)
    for idx, row_data in enumerate(insurance_policy_rows, start=INSURANCE_POLICY_START_ROW):
        for field, col in INSURANCE_POLICY_COLS.items():
            ws[f"{col}{idx}"] = row_data.get(field, "")

    pf_header_row = PF_PAYOR_SUMMARY_START_ROW - 1
    for field, col in PF_PAYOR_SUMMARY_COLS.items():
        ws[f"{col}{pf_header_row}"] = field
        apply_cell_style(ws[f"{col}{pf_header_row}"], fill=SELECTION_FILL, bold=True)
    for idx, row_data in enumerate(pf_payor_summary_rows, start=PF_PAYOR_SUMMARY_START_ROW):
        for field, col in PF_PAYOR_SUMMARY_COLS.items():
            value = row_data.get(field, "")
            if field != "payor_bucket" and field != "dominant_pf_shape":
                number = maybe_float(value)
                value = number if number is not None else value
            ws[f"{col}{idx}"] = value

    resolution_header_row = PAYER_BASIS_RESOLUTION_START_ROW - 1
    for field, col in PAYER_BASIS_RESOLUTION_COLS.items():
        ws[f"{col}{resolution_header_row}"] = field
        apply_cell_style(ws[f"{col}{resolution_header_row}"], fill=SELECTION_FILL, bold=True)
    for idx, row_data in enumerate(payer_basis_resolution_rows, start=PAYER_BASIS_RESOLUTION_START_ROW):
        row_copy = dict(row_data)
        row_copy["lookup_key"] = f"{row_copy.get('component', '')}|{row_copy.get('target_payor_bucket', '')}"
        for field, col in PAYER_BASIS_RESOLUTION_COLS.items():
            value = row_copy.get(field, "")
            if field in {
                "case_count",
                "anchor_p25",
                "anchor_p50",
                "anchor_p75",
                "variability_score",
                "spread_vs_insurance_all",
                "spread_vs_all_payers",
                "selected_case_count",
            }:
                number = maybe_float(value)
                value = number if number is not None else value
            ws[f"{col}{idx}"] = value

    actual_basis_rows = build_actual_basis_metric_rows(actual_rows)
    actual_basis_header_row = ACTUAL_BASIS_METRIC_START_ROW - 1
    for field, col in ACTUAL_BASIS_METRIC_COLS.items():
        ws[f"{col}{actual_basis_header_row}"] = field
        apply_cell_style(ws[f"{col}{actual_basis_header_row}"], fill=SELECTION_FILL, bold=True)
    for idx, row_data in enumerate(actual_basis_rows, start=ACTUAL_BASIS_METRIC_START_ROW):
        for field, col in ACTUAL_BASIS_METRIC_COLS.items():
            value = row_data.get(field, "")
            if field not in {"lookup_key", "basis_label", "field_key", "field_label"}:
                number = maybe_float(value)
                value = number if number is not None else value
            ws[f"{col}{idx}"] = value


def finalize_filterable_sheet(
    ws,
    *,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    last_col: int,
    widths: dict[str, float],
    currency_ranges: list[str] | None = None,
    quantity_ranges: list[str] | None = None,
    percent_ranges: list[str] | None = None,
) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{data_start_row}"
    style_range(ws, f"A1:{get_column_letter(last_col)}1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_range(ws, f"A{header_row}:{get_column_letter(last_col)}{header_row}", fill=SUBHEADER_FILL, bold=True, wrap=True)
    if data_end_row >= data_start_row:
        style_range(ws, f"A{data_start_row}:{get_column_letter(last_col)}{data_end_row}", fill=SELECTION_FILL, wrap=True)
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{data_end_row}"
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for cell_range in currency_ranges or []:
        set_number_format(ws, cell_range, '#,##0.00')
    for cell_range in quantity_ranges or []:
        set_number_format(ws, cell_range, '#,##0.00')
    for cell_range in percent_ranges or []:
        set_number_format(ws, cell_range, '0.00')


def write_pharmacy_template_sheet(ws, pharmacy_rows: list[dict[str, str]]) -> None:
    ws["A1"] = "Cleaned Pharmacy Template"
    headers = [
        ("item_code", "Item Code"),
        ("item_name", "Item Name"),
        ("classification", "Classification"),
        ("present_in_ot_pharmacy", "OT Present"),
        ("present_in_ip_pharmacy", "IP Present"),
        ("present_in_returns", "Returns Present"),
        ("case_count", "Case Count"),
        ("case_presence_rate", "Presence Rate"),
        ("ot_quantity_typical_cleaned", "OT Qty Typical"),
        ("ip_quantity_typical_cleaned", "IP Qty Typical"),
        ("overall_quantity_typical_cleaned", "Overall Qty Typical"),
        ("ot_amount_typical_cleaned", "OT Amount Typical"),
        ("ip_amount_typical_cleaned", "IP Amount Typical"),
        ("overall_amount_typical_cleaned", "Overall Amount Typical"),
        ("observed_any_sale_rate_count", "Observed Rate Count"),
        ("observed_any_sale_rate_min", "Observed Rate Min"),
        ("observed_any_sale_rate_max", "Observed Rate Max"),
        ("observed_any_sale_rate_values", "Observed Rate Values"),
    ]
    header_row = 3
    for idx, (_, label) in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=label)
    data_start = header_row + 1
    for row_idx, row in enumerate(pharmacy_rows, start=data_start):
        for col_idx, (field, _) in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(field))
    data_end = data_start + len(pharmacy_rows) - 1
    finalize_filterable_sheet(
        ws,
        header_row=header_row,
        data_start_row=data_start,
        data_end_row=data_end,
        last_col=len(headers),
        widths={
            "A": 14, "B": 38, "C": 26, "D": 10, "E": 10, "F": 12, "G": 10, "H": 12,
            "I": 14, "J": 14, "K": 16, "L": 16, "M": 16, "N": 18, "O": 16, "P": 14, "Q": 14, "R": 28,
        },
        currency_ranges=[f"L{data_start}:Q{data_end}"] if data_end >= data_start else [],
        quantity_ranges=[f"I{data_start}:K{data_end}"] if data_end >= data_start else [],
        percent_ranges=[f"H{data_start}:H{data_end}"] if data_end >= data_start else [],
    )


def write_pf_review_sheet(ws, args: argparse.Namespace) -> None:
    payor_summary_rows = load_pf_csv_rows(args.pf_payor_summary_csv)
    shape_review = load_pf_json(args.pf_shape_review_json)
    modeled_vs_actual_rows = load_pf_csv_rows(args.pf_modeled_vs_actual_csv)
    estimate_behavior = (
        "Cash formula in estimate body; historical PF shown as review context."
        if derive_payer_type(args.payor_label) == "cash"
        else "Insurance mode keeps derived PF disabled; PF shown as review-only context."
    )
    write_professional_fees_review_sheet(
        ws,
        template_name=normalize_text(args.sheet1_template_name) or normalize_text(args.procedure_label),
        estimate_behavior=estimate_behavior,
        payor_summary_rows=payor_summary_rows,
        shape_review=shape_review,
        modeled_vs_actual_rows=modeled_vs_actual_rows,
        header_fill=HEADER_FILL,
        subheader_fill=SUBHEADER_FILL,
        formula_fill=FORMULA_GREEN_FILL,
        result_fill=RESULT_FILL,
        reference_fill=SELECTION_FILL,
    )


def workbook_numeric(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def workbook_percent_value(value: Any) -> float | None:
    numeric = workbook_numeric(value)
    if numeric is None:
        return None
    if numeric > 1:
        return numeric / 100.0
    return numeric


def find_labeled_row(ws, label: str, column: str = "A", max_row: int | None = None) -> int | None:
    upper_label = normalize_text(label).upper()
    upper_bound = max_row or ws.max_row
    for row_idx in range(1, upper_bound + 1):
        cell_text = normalize_text(ws[f"{column}{row_idx}"].value).upper()
        if cell_text == upper_label:
            return row_idx
    return None


def validate_generated_surgical_workbook(path: Path) -> dict[str, Any]:
    expected_sheet_order = [
        SHEET_BUILDER,
        SHEET_SUMMARY,
        SHEET_ESTIMATE_VS_ACTUAL,
        SHEET_ADVANCED,
        SHEET_SERVICE_ADDONS,
        SHEET_GROUPED_ADJUSTMENTS,
        SHEET_GROUPING_REVIEW,
        SHEET_IMPLANTS_SELECT,
        SHEET_BREAKDOWN,
        SHEET_DETAIL,
        SHEET_PHARMACY_TEMPLATE,
        SHEET_SERVICE_TEMPLATE,
        SHEET_PHARMACY_METRICS,
        SHEET_IP_ACTUALS,
        SHEET_PF_REVIEW,
        SHEET_REFERENCE,
    ]
    structure_wb = load_workbook(path, data_only=False, read_only=False)
    calc_wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if structure_wb.sheetnames != expected_sheet_order:
            raise ValueError(
                f"Workbook sheet order mismatch. Expected {expected_sheet_order}, found {structure_wb.sheetnames}."
            )

        freeze_pane_violations = [
            ws.title for ws in structure_wb.worksheets if ws.freeze_panes is not None
        ]
        if freeze_pane_violations:
            raise ValueError(f"Freeze panes still present on sheets: {freeze_pane_violations}.")

        merged_range_violations = {
            ws.title: [str(cell_range) for cell_range in ws.merged_cells.ranges]
            for ws in structure_wb.worksheets
            if ws.merged_cells.ranges
        }
        if merged_range_violations:
            raise ValueError(f"Merged ranges still present in standardized workbook: {merged_range_violations}.")

        builder_ws = calc_wb[SHEET_BUILDER]
        summary_ws = calc_wb[SHEET_SUMMARY]
        breakdown_ws = calc_wb[SHEET_BREAKDOWN]
        detail_ws = calc_wb[SHEET_DETAIL]

        final_estimate = workbook_numeric(summary_ws["E2"].value)
        if final_estimate is None:
            raise ValueError("Estimate Summary final estimate is blank after workbook recalculation.")

        breakdown_total = 0.0
        for row_idx in range(1, breakdown_ws.max_row + 1):
            amount = workbook_numeric(breakdown_ws[f"J{row_idx}"].value)
            if amount is not None:
                breakdown_total += amount
        if abs(breakdown_total - final_estimate) > 0.5:
            raise ValueError(
                f"Estimate Breakdown total ({breakdown_total:.2f}) does not match final estimate ({final_estimate:.2f})."
            )

        selected_ot_hours = workbook_numeric(builder_ws["G13"].value) or 0.0
        resolved_ot_hours = workbook_numeric(builder_ws["B14"].value)
        resolved_ot_code = normalize_text(builder_ws["B15"].value)
        resolved_ot_label = normalize_text(builder_ws["B16"].value)
        if selected_ot_hours > 0:
            if resolved_ot_hours is None or resolved_ot_hours <= 0:
                raise ValueError("Resolved OT slot hours are blank or zero despite a positive selected OT Hours value.")
            if not resolved_ot_code:
                raise ValueError("Resolved OT slot code is blank despite a positive selected OT Hours value.")
            if not resolved_ot_label:
                raise ValueError("Resolved OT slot label is blank despite a positive selected OT Hours value.")

        ot_row = find_labeled_row(detail_ws, "OT Charges")
        if ot_row is None:
            raise ValueError("Line Item Detail is missing the OT Charges row.")
        selected_ot_amount = workbook_numeric(detail_ws[f"Y{ot_row}"].value)
        if selected_ot_hours > 0 and (selected_ot_amount is None or selected_ot_amount <= 0):
            raise ValueError("Selected OT Charges amount is blank or zero despite a positive selected OT Hours value.")

        procedure_summary_row = find_labeled_row(summary_ws, "Procedure / OT Charges")
        if procedure_summary_row is None:
            raise ValueError("Estimate Summary is missing the Procedure / OT Charges bucket row.")
        procedure_bucket_amount = workbook_numeric(summary_ws[f"B{procedure_summary_row}"].value)
        if (
            selected_ot_amount is not None
            and procedure_bucket_amount is not None
            and selected_ot_amount - procedure_bucket_amount > 0.5
        ):
            raise ValueError(
                "Selected OT Charges exceeds the Procedure / OT Charges summary bucket, which indicates a broken rollup."
            )

        return {
            "sheet_order": structure_wb.sheetnames,
            "sheet_count": len(structure_wb.sheetnames),
            "freeze_panes_removed": True,
            "merged_ranges_removed": True,
            "final_estimate": round(final_estimate, 2),
            "estimate_breakdown_total": round(breakdown_total, 2),
            "selected_ot_hours": round(selected_ot_hours, 2),
            "resolved_ot_hours": round(resolved_ot_hours or 0.0, 2),
            "resolved_ot_code": resolved_ot_code,
            "resolved_ot_label": resolved_ot_label,
        }
    finally:
        calc_wb.close()
        structure_wb.close()


def write_service_template_sheet(ws, cleaned_service_rows: list[dict[str, str]]) -> None:
    ws["A1"] = "Cleaned Service Template"
    headers = [
        ("item_code", "Item Code"),
        ("item_name", "Item Name"),
        ("fc_estimate_bucket", "FC Bucket"),
        ("grouping", "Grouping"),
        ("case_count", "Case Count"),
        ("case_presence_rate", "Presence Rate"),
        ("quantity_p25", "Qty P25"),
        ("quantity_p50", "Qty P50"),
        ("quantity_p75", "Qty P75"),
        ("tariff_code", "Tariff Code"),
        ("tariff_general", "Tariff General"),
        ("tariff_twin", "Tariff Twin"),
        ("tariff_single", "Tariff Single"),
        ("tariff_icu", "Tariff ICU"),
        ("amount_cash_typical", "Typical Amount"),
        ("room_category_dependent", "Room Dependent"),
    ]
    header_row = 3
    for idx, (_, label) in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=label)
    data_start = header_row + 1
    for row_idx, row in enumerate(cleaned_service_rows, start=data_start):
        for col_idx, (field, _) in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(field))
    data_end = data_start + len(cleaned_service_rows) - 1
    finalize_filterable_sheet(
        ws,
        header_row=header_row,
        data_start_row=data_start,
        data_end_row=data_end,
        last_col=len(headers),
        widths={
            "A": 14, "B": 40, "C": 24, "D": 24, "E": 10, "F": 12, "G": 10, "H": 10, "I": 10,
            "J": 10, "K": 14, "L": 14, "M": 14, "N": 14, "O": 16, "P": 14,
        },
        currency_ranges=[f"K{data_start}:O{data_end}"] if data_end >= data_start else [],
        quantity_ranges=[f"G{data_start}:I{data_end}"] if data_end >= data_start else [],
        percent_ranges=[f"F{data_start}:F{data_end}"] if data_end >= data_start else [],
    )
    for row_idx in range(data_start, data_end + 1):
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"C{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_grouping_review_sheet(
    ws,
    grouping_summary_rows: list[dict[str, str]],
    grouping_child_rows: list[dict[str, str]],
) -> None:
    ws["A1"] = "Grouping Review"
    ws["A2"] = "This review sheet highlights only high-presence service groups where the current FC default estimate does not fully capture the group’s exact P50 case amount. It is an audit layer only and does not change estimate totals."
    style_range(ws, "A1:P1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_range(ws, "A2:P2", fill=SELECTION_FILL, align="left", wrap=True)

    filtered_summary_rows = [
        row for row in grouping_summary_rows
        if normalize_text(row.get("status")) == "material_gap"
    ]
    filtered_summary_rows.sort(
        key=lambda row: (
            -(as_float(row.get("group_amount_left_out_vs_p50")) or 0.0),
            -(as_float(row.get("group_presence_rate")) or 0.0),
            normalize_text(row.get("grouping")),
        )
    )
    flagged_groupings = {normalize_text(row.get("grouping")) for row in filtered_summary_rows}
    filtered_child_rows = [
        row for row in grouping_child_rows
        if normalize_text(row.get("grouping")) in flagged_groupings
    ]
    filtered_child_rows.sort(
        key=lambda row: (
            normalize_text(row.get("grouping")),
            normalize_text(row.get("made_it_to_fc_default")) != "Yes",
            -(as_float(row.get("case_presence_rate")) or 0.0),
            normalize_text(row.get("item_name")),
        )
    )

    summary_headers = [
        ("grouping", "Grouping"),
        ("sample_fc_estimate_bucket", "FC Bucket"),
        ("group_presence_rate", "Group Presence Rate"),
        ("group_amount_p50_exact", "Group Amount P50 Exact"),
        ("group_amount_captured_by_default_rows", "Captured by Default"),
        ("group_amount_left_out_vs_p50", "Left Out vs P50"),
        ("status", "Status"),
    ]
    summary_header_row = 4
    for idx, (_, label) in enumerate(summary_headers, start=1):
        ws.cell(row=summary_header_row, column=idx, value=label)
    summary_data_start = summary_header_row + 1
    for row_idx, row in enumerate(filtered_summary_rows, start=summary_data_start):
        for col_idx, (field, _) in enumerate(summary_headers, start=1):
            value: Any = row.get(field)
            if field == "group_presence_rate":
                value = workbook_percent_value(value)
            ws.cell(row=row_idx, column=col_idx, value=value)
    summary_data_end = summary_data_start + len(filtered_summary_rows) - 1
    finalize_filterable_sheet(
        ws,
        header_row=summary_header_row,
        data_start_row=summary_data_start,
        data_end_row=summary_data_end,
        last_col=len(summary_headers),
        widths={
            "A": 32, "B": 24, "C": 14, "D": 18, "E": 18, "F": 16, "G": 20,
        },
        currency_ranges=[f"D{summary_data_start}:F{summary_data_end}"] if summary_data_end >= summary_data_start else [],
        percent_ranges=[f"C{summary_data_start}:C{summary_data_end}"] if summary_data_end >= summary_data_start else [],
    )
    for row_idx in range(summary_data_start, summary_data_end + 1):
        for col in ["A", "B", "G"]:
            ws[f"{col}{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    child_start_row = max(summary_data_end + 4, 10)
    ws[f"A{child_start_row}"] = "Grouping Child Detail"
    ws[f"A{child_start_row+1}"] = "These child rows are shown only for flagged groups so you can see what made it into the FC default and what stayed out."
    style_range(ws, f"A{child_start_row}:P{child_start_row}", fill=SUBHEADER_FILL, bold=True, align="left")
    style_range(ws, f"A{child_start_row+1}:P{child_start_row+1}", fill=SELECTION_FILL, align="left", wrap=True)
    child_headers = [
        ("grouping", "Grouping"),
        ("item_code", "Item Code"),
        ("item_name", "Item Name"),
        ("case_presence_rate", "Presence Rate"),
        ("amount_cash_typical", "Typical Amount"),
        ("made_it_to_fc_default", "Made It To FC Default"),
        ("why_not_default", "Why Not Default"),
    ]
    child_header_row = child_start_row + 2
    for idx, (_, label) in enumerate(child_headers, start=1):
        ws.cell(row=child_header_row, column=idx, value=label)
    child_data_start = child_header_row + 1
    for row_idx, row in enumerate(filtered_child_rows, start=child_data_start):
        for col_idx, (field, _) in enumerate(child_headers, start=1):
            value: Any = row.get(field)
            if field == "case_presence_rate":
                value = workbook_percent_value(value)
            ws.cell(row=row_idx, column=col_idx, value=value)
    child_data_end = child_data_start + len(filtered_child_rows) - 1
    finalize_filterable_sheet(
        ws,
        header_row=child_header_row,
        data_start_row=child_data_start,
        data_end_row=child_data_end,
        last_col=len(child_headers),
        widths={
            "A": 32, "B": 14, "C": 44, "D": 14, "E": 16, "F": 16, "G": 36,
        },
        currency_ranges=[f"E{child_data_start}:E{child_data_end}"] if child_data_end >= child_data_start else [],
        percent_ranges=[f"D{child_data_start}:D{child_data_end}"] if child_data_end >= child_data_start else [],
    )
    for row_idx in range(child_data_start, child_data_end + 1):
        for col in ["A", "C", "F", "G"]:
            ws[f"{col}{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_grouped_residual_candidates(grouping_summary_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    def has_positive_residual(row: dict[str, str]) -> bool:
        return as_float(row.get("suggested_group_residual_p50")) > 0 and as_float(row.get("group_amount_left_out_vs_p50")) > 0

    def is_valid_banded_group(row: dict[str, str]) -> bool:
        residual_band = normalize_text(row.get("group_residual_band"))
        presence_rate = as_float(row.get("group_presence_rate"))
        if residual_band == "auto":
            return presence_rate > 90.0 and has_positive_residual(row)
        if residual_band == "optional":
            return 75.0 <= presence_rate <= 90.0 and has_positive_residual(row)
        return False

    def should_promote_investigation_group(row: dict[str, str]) -> bool:
        if normalize_text(row.get("sample_fc_estimate_bucket")) != "Investigations":
            return False
        if normalize_text(row.get("group_residual_band")) in {"auto", "optional"}:
            return False
        if as_float(row.get("group_presence_rate")) < 50.0:
            return False
        if as_float(row.get("suggested_group_residual_p50")) < 1000.0:
            return False
        if as_float(row.get("group_amount_left_out_vs_p50")) <= 0:
            return False
        return int(as_float(row.get("optional_child_count"))) >= 1

    candidates = [
        row for row in grouping_summary_rows
        if is_valid_banded_group(row)
    ]
    for row in grouping_summary_rows:
        if should_promote_investigation_group(row):
            promoted = dict(row)
            promoted["group_residual_band"] = "auto"
            promoted["eligible_group_residual"] = "Yes"
            candidates.append(promoted)
    candidates.sort(
        key=lambda row: (
            -(as_float(row.get("suggested_group_residual_p50")) or 0.0),
            -(as_float(row.get("group_presence_rate")) or 0.0),
            normalize_text(row.get("grouping")),
        )
    )
    return candidates


def build_insurance_excluded_groupings(
    grouping_child_rows: list[dict[str, str]],
    insurance_policy_rows: list[dict[str, str]],
) -> set[str]:
    grouped_codes: dict[str, set[str]] = {}
    for row in grouping_child_rows:
        grouping = normalize_text(row.get("grouping"))
        code = normalize_code(row.get("item_code"))
        if not grouping or not code:
            continue
        grouped_codes.setdefault(grouping, set()).add(code)
    excluded_codes = {normalize_code(row.get("item_code")) for row in insurance_policy_rows}
    return {
        grouping
        for grouping, codes in grouped_codes.items()
        if codes and codes.issubset(excluded_codes)
    }


def write_grouped_adjustments_sheet(
    ws,
    grouped_candidates: list[dict[str, str]],
    insurance_excluded_groupings: set[str],
) -> dict[str, str]:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Grouped Adjustments"
    ws["A2"] = "Grouped adjustments complete common high-presence service groups without double counting. Auto groups (>90% presence) start included; optional groups (75% to <90%) start excluded. Any selected child add-ons from the same grouping automatically reduce the grouped residual."
    style_range(ws, "A1:N1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_range(ws, "A2:N2", fill=SELECTION_FILL, align="left", wrap=True)

    ws["A4"], ws["B4"], ws["C4"], ws["D4"] = "Grouped Adjustments", "Low", "Typical", "High"
    style_range(ws, "A4:D4", fill=SUBHEADER_FILL, bold=True)

    headers = [
        "Grouping",
        "FC Bucket",
        "Group Presence Rate",
        "Group Amount P25 Exact",
        "Group Amount P50 Exact",
        "Group Amount P75 Exact",
        "Captured By Default",
        "Selected Add-On Amount",
        "Net Residual Low",
        "Net Residual Typical",
        "Net Residual High",
        "Selected",
        "Selected Amount",
        "Why",
    ]
    header_row = 6
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=header)
    style_range(ws, f"A{header_row}:N{header_row}", fill=SUBHEADER_FILL, bold=True, wrap=True)

    data_start = header_row + 1
    for idx, row in enumerate(grouped_candidates, start=data_start):
        residual_band = normalize_text(row.get("group_residual_band"))
        mode_pick_net = build_mode_pick_formula(f"'{SHEET_BUILDER}'!B5", f"I{idx}", f"J{idx}", f"K{idx}").lstrip("=")
        ws[f"A{idx}"] = normalize_text(row.get("grouping"))
        ws[f"B{idx}"] = normalize_text(row.get("sample_fc_estimate_bucket"))
        ws[f"C{idx}"] = workbook_percent_value(row.get("group_presence_rate"))
        ws[f"D{idx}"] = as_float(row.get("group_amount_p25_exact"))
        ws[f"E{idx}"] = as_float(row.get("group_amount_p50_exact"))
        ws[f"F{idx}"] = as_float(row.get("group_amount_p75_exact"))
        ws[f"G{idx}"] = as_float(row.get("group_amount_captured_by_default_rows"))
        ws[f"O{idx}"] = (
            f'=SUMIFS(\'{SHEET_SERVICE_ADDONS}\'!$J:$J,'
            f'\'{SHEET_SERVICE_ADDONS}\'!$B:$B,$A{idx},'
            f'\'{SHEET_SERVICE_ADDONS}\'!$I:$I,"{SELECTION_INCLUDE}")'
        )
        ws[f"P{idx}"] = (
            f'=SUMIFS(\'{SHEET_SERVICE_ADDONS}\'!$K:$K,'
            f'\'{SHEET_SERVICE_ADDONS}\'!$B:$B,$A{idx},'
            f'\'{SHEET_SERVICE_ADDONS}\'!$I:$I,"{SELECTION_INCLUDE}")'
        )
        ws[f"Q{idx}"] = (
            f'=SUMIFS(\'{SHEET_SERVICE_ADDONS}\'!$L:$L,'
            f'\'{SHEET_SERVICE_ADDONS}\'!$B:$B,$A{idx},'
            f'\'{SHEET_SERVICE_ADDONS}\'!$I:$I,"{SELECTION_INCLUDE}")'
        )
        ws[f"H{idx}"] = build_mode_pick_formula(f"'{SHEET_BUILDER}'!B5", f"O{idx}", f"P{idx}", f"Q{idx}")
        ws[f"I{idx}"] = f"=MAX(0,D{idx}-G{idx}-O{idx})"
        ws[f"J{idx}"] = f"=MAX(0,E{idx}-G{idx}-P{idx})"
        ws[f"K{idx}"] = f"=MAX(0,F{idx}-G{idx}-Q{idx})"
        ws[f"L{idx}"] = SELECTION_INCLUDE if residual_band == "auto" else SELECTION_EXCLUDE
        if normalize_text(row.get("grouping")) in insurance_excluded_groupings:
            ws[f"M{idx}"] = f'=IF({build_is_insurance_mode_formula()},0,IF(L{idx}="{SELECTION_INCLUDE}",{mode_pick_net},0))'
            ws[f"N{idx}"] = (
                "Fully excluded for insurance; zeroed in insurance mode"
                if residual_band == "auto"
                else "Optional residual; fully excluded for insurance"
            )
        else:
            ws[f"M{idx}"] = f'=IF(L{idx}="{SELECTION_INCLUDE}",{mode_pick_net},0)'
            ws[f"N{idx}"] = (
                "Auto common-case residual"
                if residual_band == "auto"
                else "Optional common-case residual"
            )
    data_end = data_start + len(grouped_candidates) - 1

    ws["B5"] = f'=SUMIF(L{data_start}:L{data_end},"{SELECTION_INCLUDE}",I{data_start}:I{data_end})' if data_end >= data_start else 0
    ws["C5"] = f'=SUMIF(L{data_start}:L{data_end},"{SELECTION_INCLUDE}",J{data_start}:J{data_end})' if data_end >= data_start else 0
    ws["D5"] = f'=SUMIF(L{data_start}:L{data_end},"{SELECTION_INCLUDE}",K{data_start}:K{data_end})' if data_end >= data_start else 0
    ws["E5"] = f'=COUNTIF(L{data_start}:L{data_end},"{SELECTION_INCLUDE}")' if data_end >= data_start else 0
    style_range(ws, "A5:D5", fill=RESULT_FILL, bold=True)
    ws["E4"] = "Included Rows"
    ws["E5"].fill = RESULT_FILL
    ws["E5"].font = Font(bold=True)

    if data_end >= data_start:
        selection_validation = DataValidation(type="list", formula1=f'"{SELECTION_INCLUDE},{SELECTION_EXCLUDE}"', allow_blank=False)
        ws.add_data_validation(selection_validation)
        selection_validation.add(f"L{data_start}:L{data_end}")
        style_range(ws, f"A{data_start}:N{data_end}", fill=SELECTION_FILL, wrap=True)
        set_number_format(ws, f"C{data_start}:M{data_end}", '#,##0.00')
        set_number_format(ws, f"C{data_start}:C{data_end}", '0.0%')
        for row_num in range(data_start, data_end + 1):
            for col in ["A", "B", "N"]:
                ws[f"{col}{row_num}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"L{row_num}"].fill = INPUT_FILL
            ws[f"H{row_num}"].fill = FORMULA_BLUE_FILL
            ws[f"I{row_num}"].fill = FORMULA_BLUE_FILL
            ws[f"J{row_num}"].fill = FORMULA_GREEN_FILL
            ws[f"K{row_num}"].fill = FORMULA_BLUE_FILL
            ws[f"M{row_num}"].fill = FORMULA_GREEN_FILL

    for col, width in {
        "A": 30, "B": 22, "C": 14, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 16, "K": 16, "L": 12, "M": 16, "N": 34,
    }.items():
        ws.column_dimensions[col].width = width
    for hidden_col in ["O", "P", "Q"]:
        ws.column_dimensions[hidden_col].hidden = True

    return {
        "grouped_low": f"'{SHEET_GROUPED_ADJUSTMENTS}'!B5",
        "grouped_typical": f"'{SHEET_GROUPED_ADJUSTMENTS}'!C5",
        "grouped_high": f"'{SHEET_GROUPED_ADJUSTMENTS}'!D5",
        "grouped_selected_count": f"'{SHEET_GROUPED_ADJUSTMENTS}'!E5",
        "grouped_start": str(data_start),
        "grouped_end": str(data_end),
    }


def write_pharmacy_metrics_sheet(ws, per_ip_rows: list[dict[str, str]]) -> None:
    ws["A1"] = "Pharmacy Metrics Across Cohort IPs"
    headers = [
        ("admission_no", "Admission No"),
        ("patient_name", "Patient Name"),
        ("los_days", "LOS Days"),
        ("ot_hours", "OT Hours"),
        ("ip_drugs_per_los_day", "IP Drugs / LOS Day"),
        ("ip_consumables_per_los_day", "IP Consumables / LOS Day"),
        ("total_amount_implants", "Implant Amount"),
        ("line_item_count_implants", "Implant Line Count"),
        ("total_amount_ip_drugs_medicines_ivs_nutrition_products", "IP Drugs Amount"),
        ("line_item_count_ip_drugs_medicines_ivs_nutrition_products", "IP Drugs Line Count"),
        ("total_amount_ip_treatment_supplies", "IP Treatment Supplies Amount"),
        ("line_item_count_ip_treatment_supplies", "IP Treatment Supplies Line Count"),
        ("total_amount_ot_drugs_medicines_ivs_nutrition_products", "OT Drugs Amount"),
        ("line_item_count_ot_drugs_medicines_ivs_nutrition_products", "OT Drugs Line Count"),
        ("total_amount_ot_treatment_supplies", "OT Treatment Supplies Amount"),
        ("line_item_count_ot_treatment_supplies", "OT Treatment Supplies Line Count"),
        ("gross_return_quantity_total", "Gross Return Qty"),
        ("unclassified_item_count", "Unclassified Item Count"),
    ]
    header_row = 3
    for idx, (_, label) in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=label)
    data_start = header_row + 1
    for row_idx, row in enumerate(per_ip_rows, start=data_start):
        for col_idx, (field, _) in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(field))
    data_end = data_start + len(per_ip_rows) - 1
    finalize_filterable_sheet(
        ws,
        header_row=header_row,
        data_start_row=data_start,
        data_end_row=data_end,
        last_col=len(headers),
        widths={
            "A": 16, "B": 28, "C": 10, "D": 10, "E": 14, "F": 18, "G": 14, "H": 12,
            "I": 14, "J": 12, "K": 18, "L": 12, "M": 14, "N": 12, "O": 18, "P": 12, "Q": 12, "R": 14,
        },
        currency_ranges=[f"G{data_start}:O{data_end}"] if data_end >= data_start else [],
        quantity_ranges=[f"C{data_start}:F{data_end}", f"Q{data_start}:R{data_end}"] if data_end >= data_start else [],
    )
    for row_idx in range(data_start, data_end + 1):
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


ACTUAL_IP_HEADERS = [
        ("admission_no", "Admission No"),
        ("patient_name", "Patient Name"),
        ("payor_bucket", "Payor Bucket"),
        ("patient_type", "Payer / Patient Type"),
        ("organization_name", "Organization"),
        ("surgical_medical", "Management Type"),
        ("room_category", "Room Category"),
        ("icu_unit_name", "ICU Unit"),
        ("los_days", "LOS Days"),
        ("icu_days", "ICU Days"),
        ("ward_days", "Ward Days"),
        ("ot_hours", "OT Hours"),
        ("service_line_count", "Service Line Count"),
        ("room_charges", "Room Charges"),
        ("room_charges_per_day", "Room Charges / Day"),
        ("investigations", "Investigations"),
        ("procedure_ot_charges", "Procedure / OT Charges"),
        ("bedside_services", "Bedside Services"),
        ("professional_fees", "Professional Fees"),
        ("ip_drugs", "IP Drugs"),
        ("ip_drugs_per_day", "IP Drugs / Day"),
        ("ip_consumables", "IP Consumables"),
        ("ip_consumables_per_day", "IP Consumables / Day"),
        ("ot_drugs", "OT Drugs"),
        ("ot_consumables", "OT Consumables"),
        ("implants", "Implants"),
        ("pharmacy_total", "Pharmacy Total"),
        ("drug_administration_charges", "Drug Administration Charges"),
        ("services_total_excluding_food_and_beverage", "Services Total ex F&B"),
        ("food_and_beverage_excluded", "Food & Beverage Excluded"),
        ("pharmacy_returns_excluded", "Pharmacy Returns Excluded"),
        ("total_amount_excluding_food_and_beverage_and_returns_plus_drug_admin", "Total Amount ex F&B, Returns, plus Drug Admin"),
]

ACTUAL_IP_SUMMARY_FIELDS = [
    ("los_days", "LOS Days"),
    ("icu_days", "ICU Days"),
    ("ward_days", "Ward Days"),
    ("ot_hours", "OT Hours"),
    ("service_line_count", "Service Line Count"),
    ("room_charges", "Room Charges"),
    ("room_charges_per_day", "Room Charges / Day"),
    ("investigations", "Investigations"),
    ("procedure_ot_charges", "Procedure / OT Charges"),
    ("bedside_services", "Bedside Services"),
    ("professional_fees", "Professional Fees"),
    ("ip_drugs", "IP Drugs"),
    ("ip_drugs_per_day", "IP Drugs / Day"),
    ("ip_consumables", "IP Consumables"),
    ("ip_consumables_per_day", "IP Consumables / Day"),
    ("ot_drugs", "OT Drugs"),
    ("ot_consumables", "OT Consumables"),
    ("implants", "Implants"),
    ("pharmacy_total", "Pharmacy Total"),
    ("drug_administration_charges", "Drug Administration Charges"),
    ("services_total_excluding_food_and_beverage", "Services Total ex F&B"),
    ("food_and_beverage_excluded", "Food & Beverage Excluded"),
    ("pharmacy_returns_excluded", "Pharmacy Returns Excluded"),
    ("total_amount_excluding_food_and_beverage_and_returns_plus_drug_admin", "Total Amount ex F&B, Returns, plus Drug Admin"),
]


def actual_basis_ref_for_field(field_key: str) -> str:
    if field_key in {
        "ip_drugs",
        "ip_drugs_per_day",
        "ip_consumables",
        "ip_consumables_per_day",
        "ot_drugs",
        "ot_consumables",
        "implants",
        "pharmacy_total",
        "drug_administration_charges",
        "pharmacy_returns_excluded",
    }:
        return "Builder!G5"
    if field_key == "professional_fees":
        return "Builder!G7"
    return "Builder!G6"


def write_selected_basis_actuals_snapshot(ws, *, start_row: int, title: str) -> int:
    metric_rows = [
        ("los_days", "LOS Days"),
        ("icu_days", "ICU Days"),
        ("ward_days", "Ward Days"),
        ("ot_hours", "OT Hours"),
        ("service_line_count", "Service Line Count"),
        ("room_charges", "Room Charges"),
        ("room_charges_per_day", "Room Charges / Day"),
        ("investigations", "Investigations"),
        ("procedure_ot_charges", "Procedure / OT Charges"),
        ("bedside_services", "Bedside Services"),
        ("ip_drugs", "IP Drugs"),
        ("ip_drugs_per_day", "IP Drugs / Day"),
        ("ip_consumables", "IP Consumables"),
        ("ip_consumables_per_day", "IP Consumables / Day"),
        ("ot_drugs", "OT Drugs"),
        ("ot_consumables", "OT Consumables"),
        ("implants", "Implants"),
        ("pharmacy_total", "Pharmacy Total"),
        ("drug_administration_charges", "Drug Administration Charges"),
        ("professional_fees", "Professional Fees"),
        ("professional_fees_historic_basis", "Professional Fees (Historic Basis PF)"),
        ("total_amount_excluding_food_and_beverage_and_returns_plus_drug_admin", "Total Amount ex F&B, Returns, plus Drug Admin"),
    ]
    ws[f"A{start_row}"] = title
    style_range(ws, f"A{start_row}:F{start_row}", fill=SUBHEADER_FILL, bold=True, align="left")
    header_row = start_row + 1
    headers = ["Field", "Basis", "P25", "P50", "P75", "Notes"]
    for idx, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=label)
    style_range(ws, f"A{header_row}:F{header_row}", fill=SELECTION_FILL, bold=True)
    current_row = header_row + 1
    for field_key, label in metric_rows:
        basis_ref = actual_basis_ref_for_field(field_key)
        ws[f"A{current_row}"] = label
        if field_key == "total_amount_excluding_food_and_beverage_and_returns_plus_drug_admin":
            ws[f"B{current_row}"] = '=IF(AND(Builder!G5=Builder!G6,Builder!G6=Builder!G7),Builder!G6,"Component Mix")'
            ws[f"C{current_row}"] = (
                '=IF(AND(Builder!G5=Builder!G6,Builder!G6=Builder!G7),'
                + build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p25"], field_key, "Builder!G6").lstrip("=")
                + f',{build_component_mix_total_actual_formula("p25").lstrip("=")})'
            )
            ws[f"D{current_row}"] = (
                '=IF(AND(Builder!G5=Builder!G6,Builder!G6=Builder!G7),'
                + build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p50"], field_key, "Builder!G6").lstrip("=")
                + f',{build_component_mix_total_actual_formula("p50").lstrip("=")})'
            )
            ws[f"E{current_row}"] = (
                '=IF(AND(Builder!G5=Builder!G6,Builder!G6=Builder!G7),'
                + build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p75"], field_key, "Builder!G6").lstrip("=")
                + f',{build_component_mix_total_actual_formula("p75").lstrip("=")})'
            )
            ws[f"F{current_row}"] = "Component-mix comparator when auto basis differs across service, pharmacy, and PF."
        elif field_key == "professional_fees_historic_basis":
            ws[f"B{current_row}"] = "=Builder!G7"
            ws[f"C{current_row}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p25"], "Builder!G7")
            ws[f"D{current_row}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p50"], "Builder!G7")
            ws[f"E{current_row}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p75"], "Builder!G7")
            ws[f"F{current_row}"] = "Collectible historical PF reference from the selected PF payer basis."
        else:
            ws[f"B{current_row}"] = f"={basis_ref}"
            ws[f"C{current_row}"] = build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p25"], field_key, basis_ref)
            ws[f"D{current_row}"] = build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p50"], field_key, basis_ref)
            ws[f"E{current_row}"] = build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p75"], field_key, basis_ref)
            ws[f"F{current_row}"] = ""
        current_row += 1
    data_end = current_row - 1
    style_range(ws, f"A{header_row+1}:F{data_end}", fill=SELECTION_FILL)
    for row_idx in range(header_row + 1, data_end + 1):
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"B{row_idx}"].fill = FORMULA_GREEN_FILL
        ws[f"C{row_idx}"].fill = FORMULA_BLUE_FILL
        ws[f"D{row_idx}"].fill = FORMULA_GREEN_FILL
        ws[f"E{row_idx}"].fill = FORMULA_BLUE_FILL
        ws[f"F{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    set_number_format(ws, f"C{header_row+1}:E{data_end}", "#,##0.00")
    return data_end


def write_ip_fc_actuals_sheet(ws, actual_rows: list[dict[str, Any]]) -> None:
    ws["A1"] = "Per-IP Actual FC Bucket Amounts"
    ws["A2"] = "Actual billed amounts by IP after excluding Food & Beverage service rows and pharmacy returns, with a separate 12.5% drug administration field added on top of pharmacy. This audit layer also shows management type, ICU unit, LOS breakup, and per-day normalized room/IP-pharmacy fields."
    header_row = 4
    for idx, (_, label) in enumerate(ACTUAL_IP_HEADERS, start=1):
        ws.cell(row=header_row, column=idx, value=label)
    data_start = header_row + 1
    for row_idx, row in enumerate(actual_rows, start=data_start):
        for col_idx, (field, _) in enumerate(ACTUAL_IP_HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(field))
    data_end = data_start + len(actual_rows) - 1
    last_col_letter = get_column_letter(len(ACTUAL_IP_HEADERS))
    style_range(ws, f"A1:{last_col_letter}1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_range(ws, f"A2:{last_col_letter}2", fill=SELECTION_FILL, align="left", wrap=True)
    finalize_filterable_sheet(
        ws,
        header_row=header_row,
        data_start_row=data_start,
        data_end_row=data_end,
        last_col=len(ACTUAL_IP_HEADERS),
        widths={
            "A": 16, "B": 28, "C": 18, "D": 24, "E": 16, "F": 14, "G": 10, "H": 10, "I": 10, "J": 10, "K": 14,
            "L": 14, "M": 16, "N": 14, "O": 18, "P": 16, "Q": 16, "R": 12, "S": 14, "T": 14, "U": 16,
            "V": 12, "W": 14, "X": 14, "Y": 18, "Z": 18, "AA": 18, "AB": 20, "AC": 18,
        },
        currency_ranges=[f"L{data_start}:AC{data_end}"] if data_end >= data_start else [],
        quantity_ranges=[f"G{data_start}:K{data_end}"] if data_end >= data_start else [],
    )
    for row_idx in range(data_start, data_end + 1):
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_estimate_vs_actual_sheet(ws, summary_ws) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Estimate vs IP FC Actuals"
    ws["A2"] = (
        "Compares the current estimate against basis-aligned historical actuals. "
        "Service buckets use the resolved service basis, pharmacy and drug administration use the resolved pharmacy basis, "
        "and professional fees use the resolved PF basis."
    )
    style_range(ws, "A1:J1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_range(ws, "A2:J2", fill=SELECTION_FILL, align="left", wrap=True)

    ws["A4"], ws["B4"], ws["C4"], ws["D4"], ws["E4"], ws["F4"], ws["G4"], ws["H4"], ws["I4"], ws["J4"] = (
        "Metric",
        "Comparison Basis",
        "Selected Estimate",
        "Actual P25",
        "Actual P50",
        "Actual P75",
        "Delta vs P25",
        "Delta vs P50",
        "Delta vs P75",
        "Status",
    )
    style_range(ws, "A4:J4", fill=SUBHEADER_FILL, bold=True)
    last_detail_row = summary_ws.parent[SHEET_DETAIL].max_row - 2

    def selected_line_item_formula(line_item_name: str) -> str:
        return build_room_pick_formula(
            "Builder!B4",
            f'SUMIF(\'{SHEET_DETAIL}\'!$A$2:$A${last_detail_row},"{line_item_name}",\'{SHEET_DETAIL}\'!$W$2:$W${last_detail_row})',
            f'SUMIF(\'{SHEET_DETAIL}\'!$A$2:$A${last_detail_row},"{line_item_name}",\'{SHEET_DETAIL}\'!$X$2:$X${last_detail_row})',
            f'SUMIF(\'{SHEET_DETAIL}\'!$A$2:$A${last_detail_row},"{line_item_name}",\'{SHEET_DETAIL}\'!$Y$2:$Y${last_detail_row})',
        )

    comparison_rows = [
        ("Room Charges", "room_charges", "Builder!G6", f"'{SHEET_SUMMARY}'!B13", "actual"),
        ("Investigations", "investigations", "Builder!G6", f"'{SHEET_SUMMARY}'!B14", "actual"),
        ("Procedure / OT Charges", "procedure_ot_charges", "Builder!G6", f"'{SHEET_SUMMARY}'!B15", "actual"),
        ("Bedside Services", "bedside_services", "Builder!G6", f"'{SHEET_SUMMARY}'!B16", "actual"),
        ("Pharmacy Total", "pharmacy_total", "Builder!G5", f"'{SHEET_SUMMARY}'!B17", "actual"),
        ("IP Drugs", "ip_drugs", "Builder!G5", selected_line_item_formula("IP Drugs & Medications"), "actual"),
        ("IP Drugs / Day", "ip_drugs_per_day", "Builder!G5", f'=IF(Builder!G10<=0,0,({selected_line_item_formula("IP Drugs & Medications").lstrip("=")})/Builder!G10)', "actual"),
        ("IP Consumables", "ip_consumables", "Builder!G5", selected_line_item_formula("IP Consumables"), "actual"),
        ("IP Consumables / Day", "ip_consumables_per_day", "Builder!G5", f'=IF(Builder!G10<=0,0,({selected_line_item_formula("IP Consumables").lstrip("=")})/Builder!G10)', "actual"),
        ("OT Drugs", "ot_drugs", "Builder!G5", selected_line_item_formula("OT Drugs & Medications"), "actual"),
        ("OT Consumables", "ot_consumables", "Builder!G5", selected_line_item_formula("OT Consumables"), "actual"),
        ("Implants", "implants", "Builder!G5", selected_line_item_formula("Implants"), "actual"),
        ("Drug Administration Charges", "drug_administration_charges", "Builder!G5", f"'{SHEET_SUMMARY}'!B18", "actual"),
        ("Professional Fees (Calculated)", "professional_fees", "Builder!G7", f"'{SHEET_SUMMARY}'!B19", "actual"),
        ("Professional Fees (Historic Basis P50)", "professional_fees_historic_basis", "Builder!G7", f"'{SHEET_SUMMARY}'!B22", "pf_historic"),
        ("Grand Total (Calculated PF)", "total_amount_excluding_food_and_beverage_and_returns_plus_drug_admin", "", f"'{SHEET_SUMMARY}'!B21", "grand_total"),
        ("Grand Total (Historic PF)", "total_with_historic_pf", "", f"'{SHEET_SUMMARY}'!B23", "grand_total_historic_pf"),
    ]
    row_lookup: dict[str, int] = {}
    for row_idx, (label, field_key, basis_ref, estimate_ref, row_kind) in enumerate(comparison_rows, start=5):
        row_lookup[field_key] = row_idx
        ws[f"A{row_idx}"] = label
        if row_kind == "grand_total":
            ws[f"B{row_idx}"] = '=IF(AND(Builder!G5=Builder!G6,Builder!G6=Builder!G7),Builder!G6,"Component Mix")'
            ws[f"C{row_idx}"] = f"={estimate_ref}"
            ws[f"D{row_idx}"] = (
                '=IF(AND(Builder!G5=Builder!G6,Builder!G6=Builder!G7),'
                + build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p25"], "total_amount_excluding_food_and_beverage_and_returns_plus_drug_admin", "Builder!G6").lstrip("=")
                + f',{build_component_mix_total_actual_formula("p25").lstrip("=")})'
            )
            ws[f"E{row_idx}"] = (
                '=IF(AND(Builder!G5=Builder!G6,Builder!G6=Builder!G7),'
                + build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p50"], "total_amount_excluding_food_and_beverage_and_returns_plus_drug_admin", "Builder!G6").lstrip("=")
                + f',{build_component_mix_total_actual_formula("p50").lstrip("=")})'
            )
            ws[f"F{row_idx}"] = (
                '=IF(AND(Builder!G5=Builder!G6,Builder!G6=Builder!G7),'
                + build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p75"], "total_amount_excluding_food_and_beverage_and_returns_plus_drug_admin", "Builder!G6").lstrip("=")
                + f',{build_component_mix_total_actual_formula("p75").lstrip("=")})'
            )
        elif row_kind == "grand_total_historic_pf":
            ws[f"B{row_idx}"] = '=IF(AND(Builder!G5=Builder!G6,Builder!G6=Builder!G7),Builder!G6,"Component Mix + Historic PF")'
            ws[f"C{row_idx}"] = f"={estimate_ref}"
            ws[f"D{row_idx}"] = build_component_mix_total_actual_formula("p25", pf_mode="historic")
            ws[f"E{row_idx}"] = build_component_mix_total_actual_formula("p50", pf_mode="historic")
            ws[f"F{row_idx}"] = build_component_mix_total_actual_formula("p75", pf_mode="historic")
        elif row_kind == "pf_historic":
            ws[f"B{row_idx}"] = f"={basis_ref}"
            ws[f"C{row_idx}"] = f"={estimate_ref}"
            ws[f"D{row_idx}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p25"], basis_ref)
            ws[f"E{row_idx}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p50"], basis_ref)
            ws[f"F{row_idx}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p75"], basis_ref)
        else:
            ws[f"B{row_idx}"] = f"={basis_ref}"
            ws[f"C{row_idx}"] = f"={estimate_ref}"
            ws[f"D{row_idx}"] = build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p25"], field_key, basis_ref)
            ws[f"E{row_idx}"] = build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p50"], field_key, basis_ref)
            ws[f"F{row_idx}"] = build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p75"], field_key, basis_ref)
        ws[f"G{row_idx}"] = f"=C{row_idx}-D{row_idx}"
        ws[f"H{row_idx}"] = f"=C{row_idx}-E{row_idx}"
        ws[f"I{row_idx}"] = f"=C{row_idx}-F{row_idx}"
        ws[f"J{row_idx}"] = (
            f'=IF(AND(OR(C{row_idx}<D{row_idx},C{row_idx}>F{row_idx}),ABS(H{row_idx})>MAX(5000,0.2*MAX(E{row_idx},1))),'
            f'"Material Gap",IF(C{row_idx}<D{row_idx},"Below Range",IF(C{row_idx}>F{row_idx},"Above Range","Within Range")))'
        )
    metrics_end_row = 4 + len(comparison_rows)
    style_range(ws, f"A5:J{metrics_end_row}", fill=SELECTION_FILL)
    for row_idx in range(5, metrics_end_row + 1):
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"B{row_idx}"].fill = FORMULA_GREEN_FILL
        ws[f"C{row_idx}"].fill = RESULT_FILL
        ws[f"D{row_idx}"].fill = FORMULA_BLUE_FILL
        ws[f"E{row_idx}"].fill = FORMULA_GREEN_FILL
        ws[f"F{row_idx}"].fill = FORMULA_BLUE_FILL
        ws[f"G{row_idx}"].fill = RESULT_FILL
        ws[f"H{row_idx}"].fill = RESULT_FILL
        ws[f"I{row_idx}"].fill = RESULT_FILL
        ws[f"J{row_idx}"].fill = RESULT_FILL
    set_number_format(ws, f"C5:I{metrics_end_row}", "#,##0.00")

    driver_header_row = metrics_end_row + 3
    ws[f"A{driver_header_row}"], ws[f"B{driver_header_row}"], ws[f"C{driver_header_row}"], ws[f"D{driver_header_row}"], ws[f"E{driver_header_row}"], ws[f"F{driver_header_row}"], ws[f"G{driver_header_row}"] = (
        "Driver",
        "Basis",
        "Selected",
        "Actual P25",
        "Actual P50",
        "Actual P75",
        "Status",
    )
    style_range(ws, f"A{driver_header_row}:G{driver_header_row}", fill=SUBHEADER_FILL, bold=True)
    driver_rows = [
        ("LOS", "los_days", "Builder!G6", "Builder!G10"),
        ("ICU Days", "icu_days", "Builder!G6", "Builder!G11"),
        ("Ward Days", "ward_days", "Builder!G6", "Builder!G12"),
        ("OT Hours", "ot_hours", "Builder!G6", "Builder!G13"),
    ]
    driver_start_row = driver_header_row + 1
    driver_end_row = driver_start_row + len(driver_rows) - 1
    for row_idx, (label, field_key, basis_ref, selected_ref) in enumerate(driver_rows, start=driver_start_row):
        ws[f"A{row_idx}"] = label
        ws[f"B{row_idx}"] = f"={basis_ref}"
        ws[f"C{row_idx}"] = f"={selected_ref}"
        ws[f"D{row_idx}"] = build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p25"], field_key, basis_ref)
        ws[f"E{row_idx}"] = build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p50"], field_key, basis_ref)
        ws[f"F{row_idx}"] = build_actual_basis_metric_lookup_formula(ACTUAL_BASIS_METRIC_COLS["p75"], field_key, basis_ref)
        ws[f"G{row_idx}"] = f'=IF(C{row_idx}<D{row_idx},"Below Range",IF(C{row_idx}>F{row_idx},"Above Range","Within Range"))'
    style_range(ws, f"A{driver_start_row}:G{driver_end_row}", fill=SELECTION_FILL)
    set_number_format(ws, f"C{driver_start_row}:F{driver_end_row}", "#,##0.00")

    cohort_header_row = driver_end_row + 3
    ws[f"A{cohort_header_row}"], ws[f"B{cohort_header_row}"], ws[f"C{cohort_header_row}"], ws[f"D{cohort_header_row}"], ws[f"E{cohort_header_row}"], ws[f"F{cohort_header_row}"], ws[f"G{cohort_header_row}"], ws[f"H{cohort_header_row}"] = (
        "Component Cohort",
        "Resolved Basis",
        "Cases Used",
        "Cash",
        "GIPSA",
        "Non-GIPSA",
        "Corporate",
        "Scope",
    )
    style_range(ws, f"A{cohort_header_row}:H{cohort_header_row}", fill=SUBHEADER_FILL, bold=True)
    cohort_rows = [
        ("Service Actual Cohort", "Builder!G6", "Service buckets and LOS/ICU/ward/OT service-context metrics"),
        ("Pharmacy Actual Cohort", "Builder!G5", "Pharmacy buckets, per-day pharmacy metrics, and drug administration"),
        ("PF Actual Cohort", "Builder!G7", "Professional-fee comparison rows, including historic PF reference"),
    ]
    cohort_start_row = cohort_header_row + 1
    cohort_end_row = cohort_start_row + len(cohort_rows) - 1
    for row_idx, (label, basis_ref, scope_note) in enumerate(cohort_rows, start=cohort_start_row):
        ws[f"A{row_idx}"] = label
        ws[f"B{row_idx}"] = f"={basis_ref}"
        ws[f"C{row_idx}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["cohort_size"], basis_ref)
        ws[f"D{row_idx}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["cash_count"], basis_ref)
        ws[f"E{row_idx}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["gipsa_count"], basis_ref)
        ws[f"F{row_idx}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["non_gipsa_count"], basis_ref)
        ws[f"G{row_idx}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["corporate_count"], basis_ref)
        ws[f"H{row_idx}"] = scope_note
    style_range(ws, f"A{cohort_start_row}:H{cohort_end_row}", fill=SELECTION_FILL)
    set_number_format(ws, f"C{cohort_start_row}:G{cohort_end_row}", "#,##0")
    for row_idx in range(cohort_start_row, cohort_end_row + 1):
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"H{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"B{row_idx}"].fill = FORMULA_GREEN_FILL

    notes_header_row = cohort_end_row + 3
    ws[f"A{notes_header_row}"] = "Notes"
    ws[f"A{notes_header_row + 1}"] = "1. Robotic-charge rows are controlled from the Builder, not from Service Add-Ons. Generic non-robotic variants leave that control blank or No unless the user explicitly includes robotic charges."
    ws[f"A{notes_header_row + 2}"] = "2. OT Consumables show historical P25 / P50 / P75 anchors separately. The applied resolved value defaults to historical P50 and only changes when Advanced Controls selections intentionally change it."
    ws[f"A{notes_header_row + 3}"] = "3. Pharmacy comparison is split into IP drugs, IP consumables, OT drugs, OT consumables, implants, and the rolled-up pharmacy total. LOS-linked pharmacy rows also show per-day comparators."
    ws[f"A{notes_header_row + 4}"] = "4. Drug Administration Charges are shown separately so Bedside Services compares cleanly against historical bedside actuals."
    ws[f"A{notes_header_row + 5}"] = "5. Grand Total (Historic PF) swaps the calculated PF estimate for the selected-basis collectible historical PF reference."
    ws[f"A{notes_header_row + 6}"] = "6. Case counts are shown at the service / pharmacy / PF component-basis level because all rows inside each component share the same resolved historical cohort."
    style_range(ws, f"A{notes_header_row}:J{notes_header_row}", fill=SUBHEADER_FILL, bold=True, align="left")
    style_range(ws, f"A{notes_header_row + 1}:J{notes_header_row + 6}", fill=SELECTION_FILL, align="left", wrap=True)

    for col, width in {"A": 28, "B": 22, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 18}.items():
        ws.column_dimensions[col].width = width
        ws[f"C{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"E{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"F{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_estimate_breakdown_sheet(ws, detail_ws, detail_meta: dict[str, int]) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Estimate Breakdown"
    ws["A2"] = "This sheet shows the currently selected estimate only. Use Line Item Detail if you need the full low / typical / high mechanics."
    headers = [
        "Line Item",
        "Summary Bucket",
        "Sub-Bucket",
        "Source Type",
        "How Calculated",
        "Included?",
        "Selected Quantity",
        "Selected Room",
        "Selected Rate",
        "Selected Amount",
    ]
    header_row = 3
    last_detail_row = detail_meta["grand_total_row"] - 2
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=header)

    group_order = [
        "Room Charges",
        "Investigations",
        "Procedure / OT Charges",
        "Bedside Services",
        "Drug Administration Charges",
        "Other Services",
        "Pharmacy",
        "Professional Fees",
        "Optional Add-Ons",
    ]
    group_header_rows: list[int] = []
    grouped_rows: dict[str, list[int]] = {bucket: [] for bucket in group_order}
    for row_idx in range(2, last_detail_row + 1):
        bucket = bucket_label(normalize_text(detail_ws[f"B{row_idx}"].value))
        grouped_rows.setdefault(bucket, []).append(row_idx)

    out_row = header_row + 1
    for bucket in group_order:
        source_rows = grouped_rows.get(bucket, [])
        if not source_rows:
            continue
        ws[f"A{out_row}"] = bucket
        group_header_rows.append(out_row)
        out_row += 1
        for source_row in source_rows:
            refs = {"A": "A", "C": "C", "D": "D", "E": "E"}
            for out_col, detail_col in refs.items():
                ws[f"{out_col}{out_row}"] = f"='{SHEET_DETAIL}'!{detail_col}{source_row}"
            ws[f"B{out_row}"] = bucket
            is_optional = bucket == "Optional Add-Ons"
            is_grouped_residual = f"'{SHEET_DETAIL}'!D{source_row}=\"Grouped Residual\""
            is_insurance_excluded = build_is_insurance_excluded_formula(f"'{SHEET_DETAIL}'!F{source_row}")
            selected_amount_formula = build_room_pick_formula(
                f"'{SHEET_BUILDER}'!B4",
                f"'{SHEET_DETAIL}'!W{source_row}",
                f"'{SHEET_DETAIL}'!X{source_row}",
                f"'{SHEET_DETAIL}'!Y{source_row}",
            )
            selected_amount_expr = selected_amount_formula.lstrip("=")
            selected_rate_formula = build_room_pick_formula(
                f"'{SHEET_BUILDER}'!B4",
                f"'{SHEET_DETAIL}'!K{source_row}",
                f"'{SHEET_DETAIL}'!L{source_row}",
                f"'{SHEET_DETAIL}'!M{source_row}",
            )
            selected_rate_expr = selected_rate_formula.lstrip("=")
            if is_optional:
                ws[f"F{out_row}"] = f'=IF({is_insurance_excluded},"Excluded for Insurance",IF(({selected_amount_expr})=0,"Excluded","Included"))'
                ws[f"G{out_row}"] = f'=IF(({selected_amount_expr})=0,"",\'{SHEET_DETAIL}\'!G{source_row})'
                ws[f"I{out_row}"] = f'=IF(({selected_amount_expr})=0,"",{selected_rate_expr})'
            else:
                ws[f"F{out_row}"] = f'=IF({is_insurance_excluded},"Excluded for Insurance",IF({is_grouped_residual},IF(({selected_amount_expr})=0,"Excluded","Included"),"Included"))'
                ws[f"G{out_row}"] = f'=IF({is_grouped_residual},"",\'{SHEET_DETAIL}\'!G{source_row})'
                ws[f"I{out_row}"] = f'=IF({is_grouped_residual},IF(({selected_amount_expr})=0,"",{selected_rate_expr}),{selected_rate_expr})'
            ws[f"H{out_row}"] = f"='{SHEET_BUILDER}'!B4"
            ws[f"J{out_row}"] = selected_amount_formula
            out_row += 1
        out_row += 1

    data_end = out_row - 2
    style_range(ws, "A1:J1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_range(ws, "A2:J2", fill=SELECTION_FILL, align="left", wrap=True)
    style_range(ws, f"A{header_row}:J{header_row}", fill=SELECTION_FILL, bold=True, wrap=True)
    if data_end >= header_row + 1:
        style_range(ws, f"A{header_row+1}:J{data_end}", fill=SELECTION_FILL, wrap=True)
    for row_idx in group_header_rows:
        style_row(ws, row_idx, 1, len(headers), fill=SUBHEADER_FILL, bold=True, align="left")
    ws.freeze_panes = f"A{header_row+1}"
    ws.auto_filter.ref = f"A{header_row}:J{data_end}" if data_end >= header_row + 1 else f"A{header_row}:J{header_row}"
    set_number_format(ws, f"G{header_row+1}:J{data_end}", '#,##0.00')
    for row_idx in range(header_row + 1, data_end + 1):
        if normalize_text(ws[f"A{row_idx}"].value) in group_order and all(
            not normalize_text(ws[f"{col}{row_idx}"].value) for col in ["B", "C", "D", "E"]
        ):
            continue
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"C{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"E{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col, width in {
        "A": 34, "B": 22, "C": 18, "D": 14, "E": 34, "F": 12,
        "G": 14, "H": 14, "I": 14, "J": 16,
    }.items():
        ws.column_dimensions[col].width = width


def write_implant_selection_sheet(
    ws,
    bucket_quartiles: dict[str, tuple[float, float, float]],
    implant_rows: list[dict[str, str]],
) -> str:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Implant Selection"
    ws["A2"] = "Use this sheet only if you want to override the default implant estimate."
    ws["A3"] = "Default P50 uses the cohort-typical implant estimate. Family, Brand, and Exact Item overrides change only the implant portion of the estimate."
    style_range(ws, "A1:R1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_range(ws, "A2:R3", fill=SELECTION_FILL, align="left", wrap=True)

    families, brands, items = build_implant_template_records(implant_rows)
    ws["A4"] = "Implant Estimate Mode"
    ws["B4"] = "Default P50"
    ws["A5"] = "Selected Family"
    ws["B5"] = "All"
    ws["A6"] = "Selected Brand"
    ws["B6"] = "All"
    ws["A7"] = "Selected Item Code"
    ws["B7"] = "None"
    ws["C7"] = '=IF(B7="None","",IFERROR(INDEX($AN$2:$AN$200,MATCH(B7,$AM$2:$AM$200,0)),""))'
    ws["E4"], ws["F4"], ws["G4"], ws["H4"] = "Implants", "Low", "Typical", "High"
    ws["F5"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["implants_p25"], "Builder!G5")
    ws["G5"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["implants_p50"], "Builder!G5")
    ws["H5"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["implants_p75"], "Builder!G5")
    ws["E6"] = "Resolved Implant Estimate"
    ws["F6"] = (
        '=IF(B4="Default P50",$G$5,'
        'IF(B4="Family Override",IFERROR(INDEX($AB$2:$AB$50,MATCH(B5,$U$2:$U$50,0)),$G$5),'
        'IF(B4="Brand Override",IFERROR(INDEX($AI$2:$AI$100,MATCH(B5&"|"&B6,$AJ$2:$AJ$100,0)),' 
        'IFERROR(INDEX($AB$2:$AB$50,MATCH(B5,$U$2:$U$50,0)),$G$5)),'
        'IFERROR(INDEX($AR$2:$AR$200,MATCH(B7,$AM$2:$AM$200,0)),'
        'IFERROR(INDEX($AI$2:$AI$100,MATCH(B5&"|"&B6,$AJ$2:$AJ$100,0)),'
        'IFERROR(INDEX($AB$2:$AB$50,MATCH(B5,$U$2:$U$50,0)),$G$5))))))'
    )

    style_range(ws, "A4:B7", fill=SELECTION_FILL, bold=True)
    style_range(ws, "E4:H4", fill=IMPLANT_FILL, bold=True)
    style_range(ws, "E5:H5", fill=RESULT_FILL, bold=True)
    style_range(ws, "E6:F6", fill=RESULT_FILL, bold=True)
    for ref in ["B4", "B5", "B6", "B7"]:
        ws[ref].fill = INPUT_FILL
        ws[ref].border = SECTION_BORDER
    ws["F6"].fill = FORMULA_GREEN_FILL
    ws["F6"].border = SECTION_BORDER
    ws["C7"].fill = SELECTION_FILL
    ws["B4"].comment = Comment(
        "Default P50 keeps the typical implant estimate. Use Family, Brand, or Exact Item override only when you want to change the implant assumption.",
        "Codex",
    )
    ws["F6"].comment = Comment(
        "This is the implant amount currently feeding the main estimate.",
        "Codex",
    )

    mode_validation = DataValidation(type="list", formula1='"Default P50,Family Override,Brand Override,Exact Item Override"', allow_blank=False)
    family_validation = DataValidation(type="list", formula1=f"='{SHEET_IMPLANTS_SELECT}'!$R$2:$R${len(families) + 2}", allow_blank=False)
    brand_values = sorted({brand["brand"] for brand in brands})
    brand_validation = DataValidation(type="list", formula1=f"='{SHEET_IMPLANTS_SELECT}'!$S$2:$S${len(brand_values) + 2}", allow_blank=False)
    item_validation = DataValidation(type="list", formula1=f"='{SHEET_IMPLANTS_SELECT}'!$T$2:$T${len(items) + 2}", allow_blank=False)
    for dv, ref in [
        (mode_validation, "B4"),
        (family_validation, "B5"),
        (brand_validation, "B6"),
        (item_validation, "B7"),
    ]:
        ws.add_data_validation(dv)
        dv.add(ref)

    ws["A10"] = "Family Summary"
    family_headers = ["Family", "Presence Rate", "Qty P25", "Qty P50", "Qty P75", "Rate P25", "Rate P50", "Rate P75", "Amount P50"]
    for idx, header in enumerate(family_headers, start=1):
        ws.cell(row=11, column=idx, value=header)
    for row_idx, record in enumerate(families, start=12):
        values = [
            record["family"],
            record["presence_rate"],
            record["quantity_p25"],
            record["quantity_p50"],
            record["quantity_p75"],
            record["rate_p25"],
            record["rate_p50"],
            record["rate_p75"],
            record["amount_p50"],
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    family_end = 11 + len(families)
    style_range(ws, "A10:I10", fill=SUBHEADER_FILL, bold=True, align="left")
    style_range(ws, "A11:I11", fill=IMPLANT_FILL, bold=True, wrap=True)
    if family_end >= 12:
        style_range(ws, f"A12:I{family_end}", fill=SELECTION_FILL, wrap=True)
        set_number_format(ws, f"B12:I{family_end}", '#,##0.00')

    ws["K10"] = "Brand View"
    brand_headers = ["Family", "Brand", "Presence Rate", "Qty P50", "Rate P50", "Amount P50", "Matches Family"]
    for idx, header in enumerate(brand_headers, start=11):
        ws.cell(row=11, column=idx, value=header)
    for row_idx, record in enumerate(brands, start=12):
        values = [
            record["family"],
            record["brand"],
            record["presence_rate"],
            record["quantity_p50"],
            record["rate_p50"],
            record["amount_p50"],
        ]
        for col_idx, value in enumerate(values, start=11):
            ws.cell(row=row_idx, column=col_idx, value=value)
        ws.cell(row=row_idx, column=17, value=f'=IF($B$5="All","Yes",IF(K{row_idx}=$B$5,"Yes","No"))')
    brand_end = 11 + len(brands)
    style_range(ws, "K10:Q10", fill=SUBHEADER_FILL, bold=True, align="left")
    style_range(ws, "K11:Q11", fill=IMPLANT_FILL, bold=True, wrap=True)
    if brand_end >= 12:
        style_range(ws, f"K12:Q{brand_end}", fill=SELECTION_FILL, wrap=True)
        set_number_format(ws, f"M12:P{brand_end}", '#,##0.00')

    ws["K26"] = "Exact Item View"
    item_headers = ["Family", "Brand", "Item Code", "Item Name", "Presence Rate", "Qty P50", "Rate P50", "Amount P50", "Matches Selection"]
    for idx, header in enumerate(item_headers, start=11):
        ws.cell(row=27, column=idx, value=header)
    for row_idx, record in enumerate(items, start=28):
        values = [
            record["family"],
            record["brand"],
            record["item_code"],
            record["item_name"],
            record["presence_rate"],
            record["quantity_p50"],
            record["rate_p50"],
            record["amount_p50"],
        ]
        for col_idx, value in enumerate(values, start=11):
            ws.cell(row=row_idx, column=col_idx, value=value)
        ws.cell(
            row=row_idx,
            column=19,
            value=(
                f'=IF($B$5="All","Yes",'
                f'IF($B$6="All",IF(K{row_idx}=$B$5,"Yes","No"),'
                f'IF(AND(K{row_idx}=$B$5,L{row_idx}=$B$6),"Yes","No")))'
            ),
        )
    item_end = 27 + len(items)
    style_range(ws, "K26:S26", fill=SUBHEADER_FILL, bold=True, align="left")
    style_range(ws, "K27:S27", fill=IMPLANT_FILL, bold=True, wrap=True)
    if item_end >= 28:
        style_range(ws, f"K28:S{item_end}", fill=SELECTION_FILL, wrap=True)
        set_number_format(ws, f"O28:R{item_end}", '#,##0.00')

    ws["R1"] = "Family List"
    ws["S1"] = "Brand List"
    ws["T1"] = "Item List"
    ws["R2"] = "All"
    for idx, record in enumerate(families, start=3):
        ws[f"R{idx}"] = record["family"]
    ws["S2"] = "All"
    for idx, brand_name in enumerate(brand_values, start=3):
        ws[f"S{idx}"] = brand_name
    ws["T2"] = "None"
    for idx, record in enumerate(items, start=3):
        ws[f"T{idx}"] = record["item_code"]

    ws["U1"], ws["V1"], ws["W1"], ws["X1"], ws["Y1"], ws["Z1"], ws["AA1"], ws["AB1"] = (
        "Family", "Presence", "Qty P25", "Qty P50", "Qty P75", "Rate P25", "Rate P50", "Amount P50"
    )
    for idx, record in enumerate(families, start=2):
        values = [
            record["family"], record["presence_rate"], record["quantity_p25"], record["quantity_p50"],
            record["quantity_p75"], record["rate_p25"], record["rate_p50"], record["amount_p50"],
        ]
        for col_idx, value in enumerate(values, start=21):
            ws.cell(row=idx, column=col_idx, value=value)

    ws["AD1"], ws["AE1"], ws["AF1"], ws["AG1"], ws["AH1"], ws["AI1"], ws["AJ1"] = (
        "Family", "Brand", "Presence", "Qty P50", "Rate P50", "Amount P50", "Brand Key"
    )
    for idx, record in enumerate(brands, start=2):
        values = [
            record["family"], record["brand"], record["presence_rate"], record["quantity_p50"], record["rate_p50"], record["amount_p50"], f'{record["family"]}|{record["brand"]}',
        ]
        for col_idx, value in enumerate(values, start=30):
            ws.cell(row=idx, column=col_idx, value=value)

    ws["AK1"], ws["AL1"], ws["AM1"], ws["AN1"], ws["AO1"], ws["AP1"], ws["AQ1"], ws["AR1"] = (
        "Family", "Brand", "Item Code", "Item Name", "Presence", "Qty P50", "Rate P50", "Amount P50"
    )
    for idx, record in enumerate(items, start=2):
        values = [
            record["family"], record["brand"], record["item_code"], record["item_name"],
            record["presence_rate"], record["quantity_p50"], record["rate_p50"], record["amount_p50"],
        ]
        for col_idx, value in enumerate(values, start=37):
            ws.cell(row=idx, column=col_idx, value=value)

    for col in ["R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ","AR"]:
        ws.column_dimensions[col].hidden = True

    for col, width in {
        "A": 18, "B": 18, "C": 42, "D": 10, "E": 18, "F": 14, "G": 14, "H": 14,
        "I": 2, "K": 18, "L": 22, "M": 14, "N": 12, "O": 12, "P": 14, "Q": 16, "R": 14, "S": 18,
    }.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A11"
    return f"'{SHEET_IMPLANTS_SELECT}'!F6"


def write_builder_sheet(ws, quartiles_json: dict[str, Any], args: argparse.Namespace) -> None:
    supported_basis_options = supported_basis_options_from_resolution_rows(getattr(args, "payer_basis_resolution_rows", []))
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"
    ws["A1"] = "FC Estimate Builder"
    ws["A2"] = "Procedure"
    ws["B2"] = args.sheet1_template_name
    ws["A3"] = "Estimate Pricing"
    ws["B3"] = "=E2"
    ws["D2"] = "Pricing Mode"
    ws["E2"] = PRICING_MODE_OPTIONS[0]
    ws["D3"] = "Historical Payer Basis"
    ws["E3"] = AUTO_BASIS
    ws["F2"] = "Insurance Org Code"
    ws["G2"] = ""
    ws["F3"] = "Resolved Org Name"
    empty_text_lookup = '""'
    ws["G3"] = build_org_reference_lookup_formula("G2", ORG_TARIFF_REFERENCE_COLS["organization_name"])
    ws["D4"] = "Resolved Payor Bucket"
    ws["E4"] = f'=IF(E2="{PRICING_MODE_OPTIONS[0]}","Cash",{build_org_reference_lookup_formula("G2", ORG_TARIFF_REFERENCE_COLS["payor_bucket"], empty_text_lookup).lstrip("=")})'
    ws["F4"] = "Resolved Tariff Name"
    ws["G4"] = f'=IF(E2="{PRICING_MODE_OPTIONS[0]}","KIMS",{build_org_reference_lookup_formula("G2", ORG_TARIFF_REFERENCE_COLS["tariff_name"], empty_text_lookup).lstrip("=")})'
    ws["D5"] = "Resolved Tariff Code"
    ws["E5"] = f'=IF(E2="{PRICING_MODE_OPTIONS[0]}","TR1",{build_org_reference_lookup_formula("G2", ORG_TARIFF_REFERENCE_COLS["tariff_code"], empty_text_lookup).lstrip("=")})'
    ws["F5"] = "Resolved Pharmacy Basis"
    pharmacy_resolution_formula = build_resolution_lookup_formula(
        PAYER_BASIS_RESOLUTION_COLS["selected_basis"],
        '"pharmacy_basis"',
        "E4",
        '"Cash"',
    ).lstrip("=")
    service_resolution_formula = build_resolution_lookup_formula(
        PAYER_BASIS_RESOLUTION_COLS["selected_basis"],
        '"service_basis"',
        "E4",
        '"Cash"',
    ).lstrip("=")
    pf_resolution_formula = build_resolution_lookup_formula(
        PAYER_BASIS_RESOLUTION_COLS["selected_basis"],
        '"pf_basis"',
        "E4",
        '"Cash"',
    ).lstrip("=")
    service_reason_formula = build_resolution_lookup_formula(
        PAYER_BASIS_RESOLUTION_COLS["selection_reason"],
        '"service_basis"',
        "E4",
        '""',
    ).lstrip("=")
    ws["G5"] = (
        f'=IF(E3<>"{AUTO_BASIS}",E3,{pharmacy_resolution_formula})'
    )
    ws["F6"] = "Resolved Service Basis"
    ws["G6"] = f'=IF(E3<>"{AUTO_BASIS}",E3,{service_resolution_formula})'
    ws["F7"] = "Resolved PF Basis"
    ws["G7"] = f'=IF(E3<>"{AUTO_BASIS}",E3,{pf_resolution_formula})'
    ws["F8"] = "Basis Resolver"
    ws["G8"] = f'=IF(E3<>"{AUTO_BASIS}","Manual override applied",{service_reason_formula})'
    ws["A4"] = "Selected Room Type"
    ws["B4"] = "Single"
    ws["A5"] = "Selected Estimate Mode"
    ws["B5"] = "Typical"
    ws["A6"] = "Emergency OT?"
    ws["B6"] = "No"
    ws["D6"] = "MLC?"
    ws["E6"] = "No"
    ws["A8"] = "Robotic?"
    ws["B8"] = getattr(args, "robotic_default_selection", "")
    ws["D8"] = "Robotic Charge Presence Rate"
    ws["E8"] = as_float(getattr(args, "robotic_charge_presence_rate", 0.0)) / 100.0
    ws["A7"] = "Clinical Drivers"
    ws["A9"], ws["B9"], ws["C9"], ws["D9"], ws["E9"], ws["F9"], ws["G9"] = (
        "Driver",
        "P25",
        "P50",
        "P75",
        "Selection",
        "Manual Value\n(used only if Selection = Manual)",
        "Selected Value",
    )

    driver_rows = [
        ("ICU Days", 11, build_day_rounding_formula(build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["icu_p25"], "Builder!G6")), build_day_rounding_formula(build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["icu_p50"], "Builder!G6")), build_day_rounding_formula(build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["icu_p75"], "Builder!G6")), True),
        ("Ward Days", 12, build_day_rounding_formula(build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ward_p25"], "Builder!G6")), build_day_rounding_formula(build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ward_p50"], "Builder!G6")), build_day_rounding_formula(build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ward_p75"], "Builder!G6")), True),
        ("OT Hours", 13, build_nearest_supported_ot_slot_formula(build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ot_p25"], "Builder!G6")), build_nearest_supported_ot_slot_formula(build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ot_p50"], "Builder!G6")), build_nearest_supported_ot_slot_formula(build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ot_p75"], "Builder!G6")), False),
    ]
    ws["A10"] = "LOS"
    ws["B10"] = "=B11+B12"
    ws["C10"] = "=C11+C12"
    ws["D10"] = "=D11+D12"
    ws["E10"] = "Derived"
    ws["F10"] = ""
    ws["G10"] = "=G11+G12"
    for label, row, p25, p50, p75, whole_number_driver in driver_rows:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = p25
        ws[f"C{row}"] = p50
        ws[f"D{row}"] = p75
        ws[f"E{row}"] = "P50"
        ws[f"F{row}"] = ""
        selected_formula = build_builder_value_formula(f"E{row}", f"B{row}", f"C{row}", f"D{row}", f"F{row}")
        if whole_number_driver:
            ws[f"G{row}"] = build_day_rounding_formula(selected_formula)
        elif label == "OT Hours":
            ws[f"G{row}"] = build_nearest_supported_ot_slot_formula(selected_formula)
        else:
            ws[f"G{row}"] = selected_formula

    ws["A14"] = "Resolved OT Slot (Hours)"
    ws["B14"] = build_resolved_ot_slot_formula("G13")
    ws["A15"] = "Resolved OT Slot Code"
    ws["B15"] = build_ot_slot_lookup_from_resolved_formula("B14", "B6", "code")
    ws["A16"] = "Resolved OT Slot Label"
    ws["B16"] = build_ot_slot_lookup_from_resolved_formula("B14", "B6", "name")
    ws["A17"] = "Resolved OT Type"
    ws["B17"] = '=IF(B6="Yes","Emergency","Normal")'

    style_range(ws, "A1:G1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_range(ws, "A7:G7", fill=SUBHEADER_FILL, bold=True, align="left")
    style_range(ws, "A9:G9", fill=SELECTION_FILL, bold=True, wrap=True)
    style_range(ws, "A2:B8", fill=SELECTION_FILL)
    style_range(ws, "D2:G8", fill=SELECTION_FILL)
    style_range(ws, "A10:G13", fill=SELECTION_FILL)
    style_range(ws, "A14:B17", fill=SELECTION_FILL)
    set_number_format(ws, "B10:G12", '#,##0')
    set_number_format(ws, "B13:G13", '#,##0.0')
    set_number_format(ws, "B14:B14", '#,##0.0')
    set_number_format(ws, "E8:E8", '0.0%')
    for ref in ["E2", "E3", "G2", "B4", "B5", "B6", "E6", "B8"] + [f"E{row}" for row in range(11, 14)] + [f"F{row}" for row in range(11, 14)] + ["F13"]:
        ws[ref].fill = INPUT_FILL
        ws[ref].border = SECTION_BORDER
    for ref in ["B3", "E4", "E5", "E8", "G3", "G4", "G5", "G6", "G7", "G8"] + [f"G{row}" for row in range(10, 14)]:
        ws[ref].fill = FORMULA_GREEN_FILL
        ws[ref].border = SECTION_BORDER
        ws[ref].font = Font(bold=True)
    for ref in ["B14", "B15", "B16", "B17"]:
        ws[ref].fill = FORMULA_GREEN_FILL
        ws[ref].border = SECTION_BORDER
        ws[ref].font = Font(bold=True)
    ws["E3"].comment = Comment(
        "Auto resolves service, pharmacy, and PF historical bases separately. Any manual basis here overrides all resolved bases together.",
        "Codex",
    )
    ws["B5"].comment = Comment(
        "Estimate mode chooses the low, typical, or high amount band across line items. "
        "It does not change the selected driver value unless that driver row is set differently.",
        "Codex",
    )
    ws["B6"].comment = Comment(
        "Set to Yes to use the emergency OT tariff ladder for the resolved OT slab.",
        "Codex",
    )
    ws["E6"].comment = Comment(
        "Set to Yes to add one tariff-backed MLC charge.",
        "Codex",
    )
    ws["B8"].comment = Comment(
        "Robotic charge is controlled here, not from Service Add-Ons. Robotic-specific variants default to Yes; mixed variants stay blank unless robotic-charge presence is above the threshold.",
        "Codex",
    )
    ws["E8"].comment = Comment(
        "Historical presence rate for the treatment's mapped robotic-charge rows. If it is above the threshold, mixed variants default Robotic to Yes.",
        "Codex",
    )
    ws["E10"].comment = Comment(
        "LOS is always derived as the selected ICU Days plus selected Ward Days.",
        "Codex",
    )
    ws["G10"].comment = Comment(
        "This derived LOS value feeds all LOS-dependent line items so stay-day logic stays consistent with ICU + ward.",
        "Codex",
    )
    for row in range(11, 14):
        ws[f"E{row}"].comment = Comment(
            "Choose P25, P50, P75, or Manual for this driver.",
            "Codex",
        )
        ws[f"F{row}"].comment = Comment(
            "This manual number is used only when Selection is set to Manual for the same row.",
            "Codex",
        )
        ws[f"G{row}"].comment = Comment(
            "This is the driver value currently feeding the estimate.",
            "Codex",
        )

    room_validation = DataValidation(type="list", formula1='"General,Twin,Single"', allow_blank=False)
    mode_validation = DataValidation(type="list", formula1='"Low,Typical,High"', allow_blank=False)
    pricing_mode_validation = DataValidation(type="list", formula1=f'"{",".join(PRICING_MODE_OPTIONS)}"', allow_blank=False)
    basis_validation = DataValidation(
        type="list",
        formula1=f'"{AUTO_BASIS},{",".join(supported_basis_options)}"',
        allow_blank=False,
    )
    yes_no_validation = DataValidation(type="list", formula1='"No,Yes"', allow_blank=False)
    robotic_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    driver_validation = DataValidation(type="list", formula1='"P25,P50,P75,Manual"', allow_blank=False)
    org_validation = DataValidation(type="list", formula1=f"='{SHEET_REFERENCE}'!${ORG_TARIFF_REFERENCE_COLS['organization_cd']}${ORG_TARIFF_REFERENCE_START_ROW}:${ORG_TARIFF_REFERENCE_COLS['organization_cd']}$1000", allow_blank=True)
    ws.add_data_validation(room_validation)
    ws.add_data_validation(mode_validation)
    ws.add_data_validation(pricing_mode_validation)
    ws.add_data_validation(basis_validation)
    ws.add_data_validation(yes_no_validation)
    ws.add_data_validation(robotic_validation)
    ws.add_data_validation(driver_validation)
    ws.add_data_validation(org_validation)
    room_validation.add("B4")
    mode_validation.add("B5")
    pricing_mode_validation.add("E2")
    basis_validation.add("E3")
    org_validation.add("G2")
    yes_no_validation.add("B6")
    yes_no_validation.add("E6")
    robotic_validation.add("B8")
    driver_validation.add("E11:E13")

    notes_row = 19
    ws[f"A{notes_row}"] = "How To Use"
    ws[f"A{notes_row+1}"] = "1. Cash mode uses TR1 pricing. Insurance mode uses the tariff code resolved from the selected insurance organization."
    ws[f"A{notes_row+2}"] = "2. Historical Payer Basis can stay on Auto to resolve service, pharmacy, and PF reference cohorts separately, or be manually overridden for all three together."
    ws[f"A{notes_row+3}"] = "3. LOS is always derived as ICU Days + Ward Days. Manual Value is used only for ICU Days, Ward Days, or OT Hours when that row's Selection is set to Manual."
    ws[f"A{notes_row+4}"] = "4. Estimate Mode changes the amount band used downstream: Low = P25, Typical = P50, High = P75 for each eligible line item."
    ws[f"A{notes_row+5}"] = "5. OT pricing uses the nearest supported tariff OT slot from the selected OT duration and Emergency OT setting."
    ws[f"A{notes_row+6}"] = "6. Drug administration and insurance-excluded service items are automatically suppressed in insurance mode. Package details and patient-policy coverage are not applied in this workbook yet."
    ws[f"A{notes_row+7}"] = "7. Robotic charge is controlled from the Builder. When Robotic is Yes, all mapped robotic-charge rows are included; when No, they are fully excluded."
    ws[f"A{notes_row+8}"] = "8. MLC adds one tariff-backed MLC charge only when the MLC input is Yes. Use Advanced Controls, Implant Selection, Service Add-Ons, and Grouped Adjustments only when the default estimate needs refinement."
    ws[f"A{notes_row+9}"] = "9. Read the final estimate and bucket summary on Estimate Summary."
    style_range(ws, f"A{notes_row}:G{notes_row}", fill=SUBHEADER_FILL, bold=True, align="left")
    style_range(ws, f"A{notes_row+1}:G{notes_row+9}", fill=SELECTION_FILL, align="left", wrap=True)

    for col, width in {"A": 24, "B": 18, "C": 12, "D": 18, "E": 18, "F": 20, "G": 24}.items():
        ws.column_dimensions[col].width = width


def write_advanced_controls_sheet(
    ws,
    bucket_quartiles: dict[str, tuple[float, float, float]],
    pharmacy_rows: list[dict[str, str]],
    args: argparse.Namespace,
) -> dict[str, str]:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"
    ws["A1"] = "Advanced Controls"
    ws["A2"] = "Adjust only the OT consumables portion here. Implant overrides are handled on the Implant Selection sheet. Service add-ons are handled on the Service Add-Ons sheet."
    ws["A3"] = (
        "OT consumables: benchmark anchors stay fixed at cohort P25 / P50 / P75. "
        "The resolved applied value defaults to historical P50 when no shortlist items are selected, then moves by expected-contribution bands from the shortlist below: <=30% stays at P25, 30% to 50% uses P50, and >50% uses P75."
    )
    style_range(ws, "A1:H1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_range(ws, "A2:H2", fill=SELECTION_FILL, align="left", wrap=True)
    style_range(ws, "A3:H3", fill=SELECTION_FILL, align="left", wrap=True)

    ws["A4"], ws["B4"], ws["C4"], ws["D4"] = "OT Consumables Benchmark", "P25", "P50", "P75"
    ws["B5"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ot_consumables_p25"], "Builder!G5")
    ws["C5"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ot_consumables_p50"], "Builder!G5")
    ws["D5"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ot_consumables_p75"], "Builder!G5")
    style_range(ws, "A4:D4", fill=OT_FILL, bold=True)
    style_range(ws, "A5:D5", fill=RESULT_FILL, bold=True)
    ws["A6"] = "Resolved OT Consumables"
    ws["B6"] = "Applied Value"
    style_range(ws, "A6:B6", fill=OT_FILL, bold=True)

    shortlist_rows = build_ot_consumable_shortlist(pharmacy_rows, max_count=args.ot_consumable_shortlist_count)
    ws["A7"], ws["B7"], ws["C7"], ws["D7"], ws["E7"], ws["F7"], ws["G7"], ws["H7"] = "Item", "Typical Qty", "Typical Rate", "Typical Amount", "Presence Rate", "Expected Contribution", "Cumulative Share", "Selected"
    style_range(ws, "A7:H7", fill=OT_FILL, bold=True, wrap=True)
    ot_start = 8
    for idx, row in enumerate(shortlist_rows, start=ot_start):
        ws[f"A{idx}"] = normalize_text(row.get("item_name"))
        ws[f"B{idx}"] = build_reference_pharmacy_name_lookup_formula(f"A{idx}", PAYER_BASIS_PHARMACY_COLS["ot_quantity_typical_cleaned"], fallback=str(as_float(row.get("ot_quantity_typical_cleaned"))), basis_ref="Builder!G5")
        ws[f"D{idx}"] = build_reference_pharmacy_name_lookup_formula(f"A{idx}", PAYER_BASIS_PHARMACY_COLS["ot_amount_typical_cleaned"], fallback=str(as_float(row.get("ot_amount_typical_cleaned"))), basis_ref="Builder!G5")
        ws[f"C{idx}"] = f'=IF(B{idx}>0,D{idx}/B{idx},0)'
        ws[f"E{idx}"] = build_reference_pharmacy_name_lookup_formula(f"A{idx}", PAYER_BASIS_PHARMACY_COLS["case_presence_rate"], fallback=str(as_float(row.get("case_presence_rate"))), basis_ref="Builder!G5")
        ws[f"F{idx}"] = f"=E{idx}*B{idx}*C{idx}/100"
        if idx == ot_start:
            ws[f"G{idx}"] = f"=F{idx}/SUM($F${ot_start}:$F${ot_start + len(shortlist_rows)-1})"
        else:
            ws[f"G{idx}"] = f"=G{idx-1}+F{idx}/SUM($F${ot_start}:$F${ot_start + len(shortlist_rows)-1})"
        ws[f"H{idx}"] = SELECTION_EXCLUDE
    ot_resolved_cell = "C6"
    if shortlist_rows:
        ot_end = ot_start + len(shortlist_rows) - 1
        piecewise_formula = build_ot_consumables_piecewise_formula(
            p25_ref="B5",
            p50_ref="C5",
            p75_ref="D5",
            selected_flag_range=f"H{ot_start}:H{ot_end}",
            expected_contribution_range=f"F{ot_start}:F{ot_end}",
        ).lstrip("=")
        ws["C6"] = (
            f'=IF(COUNTIF(H{ot_start}:H{ot_end},"{SELECTION_INCLUDE}")=0,C5,{piecewise_formula})'
        )
        selection_validation = DataValidation(type="list", formula1=f'"{SELECTION_INCLUDE},{SELECTION_EXCLUDE}"', allow_blank=False)
        ws.add_data_validation(selection_validation)
        selection_validation.add(f"H{ot_start}:H{ot_end}")
        style_range(ws, f"A{ot_start}:H{ot_end}", fill=SELECTION_FILL, wrap=True)
        set_number_format(ws, f"B{ot_start}:F{ot_end}", '#,##0.00')
        set_number_format(ws, f"G{ot_start}:G{ot_end}", '0.0%')
        for row_num in range(ot_start, ot_end + 1):
            ws[f"A{row_num}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws[f"H{row_num}"].fill = INPUT_FILL
    else:
        ws["C6"] = "=C5"
    style_range(ws, "C6:D6", fill=FORMULA_GREEN_FILL, bold=True)
    set_number_format(ws, "B5:D6", '#,##0.00')

    for col, width in {
        "A": 40, "B": 12, "C": 12, "D": 14, "E": 14, "F": 18, "G": 14, "H": 14,
    }.items():
        ws.column_dimensions[col].width = width

    return {
        "ot_typical": f"'{SHEET_ADVANCED}'!C6",
        "ot_resolved": f"'{SHEET_ADVANCED}'!C6",
    }


def write_service_addons_sheet(
    ws,
    optional_service_rows: list[dict[str, str]],
    rate_lookup: dict[str, RateRow],
    service_line_count_metrics: dict[str, Any],
    args: argparse.Namespace,
) -> dict[str, str]:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"
    ws["A1"] = "Service Add-Ons"
    ws["A2"] = "Include only the service items that are clinically expected for this case. Rows are prioritized by expected historical contribution first, then presence rate, so the most likely and financially meaningful add-ons appear first."
    style_range(ws, "A1:M1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_range(ws, "A2:M2", fill=SELECTION_FILL, align="left", wrap=True)

    ws["O4"] = "Service Line Count Alert"
    ws["O5"] = "Historical P25"
    ws["O6"] = "Historical P50"
    ws["O7"] = "Historical P75"
    ws["O8"] = "Base Included Non-Pharmacy Count"
    ws["O9"] = "Selected Optional Count"
    ws["O10"] = "Current Included Non-Pharmacy Count"
    ws["O11"] = "Alert"
    ws["P5"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["service_line_p25"], "Builder!G6")
    ws["P6"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["service_line_p50"], "Builder!G6")
    ws["P7"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["service_line_p75"], "Builder!G6")
    core_rows = core_line_definitions(args)
    base_service_count = sum(
        1
        for row in core_rows
        if not str(row.get("kind", "")).startswith("pharmacy_")
        and row.get("robotic_controlled") != "true"
    )
    robotic_core_service_count = sum(
        1
        for row in core_rows
        if not str(row.get("kind", "")).startswith("pharmacy_")
        and row.get("robotic_controlled") == "true"
    )
    robotic_optional_service_count = int(getattr(args, "robotic_optional_service_count", 0) or 0)
    ws["P8"] = f'= {base_service_count} + IF(\'{SHEET_BUILDER}\'!B8="Yes",{robotic_core_service_count + robotic_optional_service_count},0)'
    style_range(ws, "O4:P4", fill=SUBHEADER_FILL, bold=True, align="left")
    style_range(ws, "O5:O11", fill=SELECTION_FILL, bold=True, align="left")
    style_range(ws, "P5:P11", fill=SELECTION_FILL, align="center")

    ws["A4"], ws["B4"], ws["C4"], ws["D4"] = "Optional Add-Ons", "Low", "Typical", "High"
    style_range(ws, "A4:D4", fill=SUBHEADER_FILL, bold=True)
    ws["A6"], ws["B6"], ws["C6"], ws["D6"], ws["E6"], ws["F6"], ws["G6"], ws["H6"], ws["I6"], ws["J6"], ws["K6"], ws["L6"], ws["M6"] = "Service Name", "Grouping", "Presence Rate", "Qty P25", "Qty P50", "Qty P75", "Selected Tariff Rate", "Typical Gross", "Selected", "Low Amt", "Typical Amt", "High Amt", "Code"
    style_range(ws, "A6:M6", fill=SUBHEADER_FILL, bold=True, wrap=True)
    add_start = 7
    for idx, row in enumerate(optional_service_rows, start=add_start):
        ws[f"A{idx}"] = normalize_text(row.get("item_name"))
        ws[f"B{idx}"] = normalize_text(row.get("grouping"))
        ws[f"C{idx}"] = build_reference_service_lookup_formula(f"M{idx}", PAYER_BASIS_SERVICE_COLS["case_presence_rate"], fallback=str(as_float(row.get("case_presence_rate"))))
        ws[f"D{idx}"] = build_reference_service_lookup_formula(f"M{idx}", PAYER_BASIS_SERVICE_COLS["quantity_p25"], fallback=str(as_float(row.get("quantity_p25"))))
        ws[f"E{idx}"] = build_reference_service_lookup_formula(f"M{idx}", PAYER_BASIS_SERVICE_COLS["quantity_p50"], fallback=str(as_float(row.get("quantity_p50"))))
        ws[f"F{idx}"] = build_reference_service_lookup_formula(f"M{idx}", PAYER_BASIS_SERVICE_COLS["quantity_p75"], fallback=str(as_float(row.get("quantity_p75"))))
        ws[f"G{idx}"] = build_room_pick_formula(
            f"'{SHEET_BUILDER}'!B4",
            build_tariff_rate_lookup_formula(f"M{idx}", TARIFF_RATE_MATRIX_COLS["general"], fallback="0", tariff_ref="Builder!E5"),
            build_tariff_rate_lookup_formula(f"M{idx}", TARIFF_RATE_MATRIX_COLS["twin"], fallback="0", tariff_ref="Builder!E5"),
            build_tariff_rate_lookup_formula(f"M{idx}", TARIFF_RATE_MATRIX_COLS["single"], fallback="0", tariff_ref="Builder!E5"),
        )
        ws[f"H{idx}"] = wrap_insurance_exclusion(f"M{idx}", f"=E{idx}*G{idx}")
        ws[f"I{idx}"] = SELECTION_EXCLUDE
        ws[f"J{idx}"] = wrap_insurance_exclusion(f"M{idx}", f"=D{idx}*G{idx}")
        ws[f"K{idx}"] = wrap_insurance_exclusion(f"M{idx}", f"=E{idx}*G{idx}")
        ws[f"L{idx}"] = wrap_insurance_exclusion(f"M{idx}", f"=F{idx}*G{idx}")
        ws[f"M{idx}"] = normalize_code(row.get("item_code"))
    add_end = add_start + len(optional_service_rows) - 1
    totals_row = 4
    ws["B5"] = f'=SUMIF(I{add_start}:I{add_end},"{SELECTION_INCLUDE}",J{add_start}:J{add_end})'
    ws["C5"] = f'=SUMIF(I{add_start}:I{add_end},"{SELECTION_INCLUDE}",K{add_start}:K{add_end})'
    ws["D5"] = f'=SUMIF(I{add_start}:I{add_end},"{SELECTION_INCLUDE}",L{add_start}:L{add_end})'
    ws["P9"] = f'=COUNTIF(I{add_start}:I{add_end},"{SELECTION_INCLUDE}")'
    ws["P10"] = "=P8+P9"
    ws["P11"] = '=IF(P10<P5,"Below historical P25",IF(P10>P7,"Above historical P75","Within historical range"))'
    style_range(ws, "A5:D5", fill=RESULT_FILL, bold=True)
    ws["P10"].fill = FORMULA_GREEN_FILL
    ws["P11"].fill = RESULT_FILL

    add_validation = DataValidation(type="list", formula1=f'"{SELECTION_INCLUDE},{SELECTION_EXCLUDE}"', allow_blank=False)
    ws.add_data_validation(add_validation)
    add_validation.add(f"I{add_start}:I{add_end}")
    style_range(ws, f"A{add_start}:M{add_end}", fill=SELECTION_FILL, wrap=True)
    set_number_format(ws, f"C{add_start}:L{add_end}", '#,##0.00')
    for row_num in range(add_start, add_end + 1):
        ws[f"A{row_num}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"B{row_num}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"I{row_num}"].fill = INPUT_FILL
        ws[f"H{row_num}"].fill = FORMULA_GREEN_FILL
        ws[f"J{row_num}"].fill = FORMULA_BLUE_FILL
        ws[f"K{row_num}"].fill = FORMULA_GREEN_FILL
        ws[f"L{row_num}"].fill = FORMULA_BLUE_FILL

    ws.column_dimensions["M"].hidden = True
    for col, width in {
        "A": 40, "B": 18, "C": 12, "D": 12, "E": 12, "F": 12, "G": 14, "H": 14, "I": 12,
        "J": 14, "K": 14, "L": 14, "M": 14, "O": 28, "P": 16,
    }.items():
        ws.column_dimensions[col].width = width

    return {
        "optional_low": f"'{SHEET_SERVICE_ADDONS}'!B5",
        "optional_typical": f"'{SHEET_SERVICE_ADDONS}'!C5",
        "optional_high": f"'{SHEET_SERVICE_ADDONS}'!D5",
        "service_line_count_current": f"'{SHEET_SERVICE_ADDONS}'!P10",
        "service_line_count_alert": f"'{SHEET_SERVICE_ADDONS}'!P11",
        "service_line_count_p25": f"'{SHEET_SERVICE_ADDONS}'!P5",
        "service_line_count_p50": f"'{SHEET_SERVICE_ADDONS}'!P6",
        "service_line_count_p75": f"'{SHEET_SERVICE_ADDONS}'!P7",
        "optional_start": str(add_start),
        "optional_end": str(add_end),
    }


def write_line_item_detail_sheet(
    ws,
    quartiles_json: dict[str, Any],
    bucket_quartiles: dict[str, tuple[float, float, float]],
    ip_pharmacy_per_day_metrics: dict[str, Any],
    service_lookup: dict[str, dict[str, str]],
    cleaned_service_rows: list[dict[str, str]],
    auto_included_service_rows: list[dict[str, str]],
    optional_service_rows: list[dict[str, str]],
    robotic_service_rows: list[dict[str, str]],
    grouped_adjustment_rows: list[dict[str, str]],
    rate_lookup: dict[str, RateRow],
    advanced_refs: dict[str, str],
    args: argparse.Namespace,
) -> dict[str, int]:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    headers = [
        "Line Item", "Parent Bucket", "Sub-Bucket", "Source", "How", "Item Code",
        "Selected Qty", "Qty Low", "Qty Typical", "Qty High",
        "Rate General", "Rate Twin", "Rate Single",
        "General Low", "General Typical", "General High",
        "Twin Low", "Twin Typical", "Twin High",
        "Single Low", "Single Typical", "Single High",
        "Selected Total General", "Selected Total Twin", "Selected Total Single",
    ]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        apply_cell_style(cell, fill=HEADER_FILL, bold=True, font_color="FFFFFF", wrap=True)

    rate_mode_cell = f"'{SHEET_BUILDER}'!B5"
    los_low, los_typ, los_high = "Builder!B10", "Builder!C10", "Builder!D10"
    icu_low, icu_typ, icu_high = "Builder!B11", "Builder!C11", "Builder!D11"
    ward_low, ward_typ, ward_high = "Builder!B12", "Builder!C12", "Builder!D12"
    ot_low, ot_typ, ot_high = "Builder!B13", "Builder!C13", "Builder!D13"
    selected_los, selected_icu, selected_ward, selected_ot = "Builder!G10", "Builder!G11", "Builder!G12", "Builder!G13"

    core_rows = core_line_definitions(args)
    all_rows = core_rows + [
        {
            "name": normalize_text(row.get("item_name")),
            "parent": "Optional Add-Ons",
            "sub": normalize_text(row.get("grouping")),
            "source": "Advanced",
            "how": "Include / Exclude selection",
            "code": normalize_code(row.get("item_code")),
            "kind": "optional_service",
            "service_row": row,
        }
        for row in optional_service_rows
    ] + [
        {
            "name": normalize_text(row.get("item_name")),
            "parent": "Procedure / OT Charges",
            "sub": "Robotic Charges",
            "source": "Builder Control",
            "how": "Included only when Builder Robotic is Yes",
            "code": normalize_code(row.get("item_code")),
            "kind": "robotic_service",
            "service_row": row,
        }
        for row in robotic_service_rows
    ] + [
        {
            "name": f"{normalize_text(row.get('grouping'))} Residual",
            "parent": grouped_residual_parent_bucket(row),
            "sub": normalize_text(row.get("grouping")),
            "source": "Grouped Residual",
            "how": "Mode-aware grouped residual net of selected child add-ons from same grouping",
            "code": "",
            "kind": "grouped_residual",
            "group_row": row,
        }
        for row in grouped_adjustment_rows
    ]

    bucket_refs: dict[str, int] = {}
    current_row = 2
    pf_multipliers = get_professional_fee_multipliers("Cash")
    insurance_mode_formula = build_is_insurance_mode_formula()

    for row_def in all_rows:
        ws[f"A{current_row}"] = row_def["name"]
        ws[f"B{current_row}"] = bucket_label(row_def["parent"])
        ws[f"C{current_row}"] = row_def["sub"]
        ws[f"D{current_row}"] = row_def["source"]
        ws[f"E{current_row}"] = row_def["how"]
        ws[f"F{current_row}"] = row_def["code"]
        kind = row_def["kind"]
        rate_general = rate_twin = rate_single = None

        if kind == "template":
            service_row = service_lookup.get(normalize_code(row_def["code"]), {})
            rate = rate_lookup.get(normalize_code(row_def["code"]))
            rate_general = rate.general if rate else None
            rate_twin = rate.twin if rate else None
            rate_single = rate.single if rate else None
            ws[f"H{current_row}"] = build_reference_service_lookup_formula(
                f"F{current_row}",
                PAYER_BASIS_SERVICE_COLS["quantity_p25"],
                fallback=str(maybe_float(service_row.get("quantity_p25")) or 1.0),
            )
            ws[f"I{current_row}"] = build_reference_service_lookup_formula(
                f"F{current_row}",
                PAYER_BASIS_SERVICE_COLS["quantity_p50"],
                fallback=str(maybe_float(service_row.get("quantity_p50")) or 1.0),
            )
            ws[f"J{current_row}"] = build_reference_service_lookup_formula(
                f"F{current_row}",
                PAYER_BASIS_SERVICE_COLS["quantity_p75"],
                fallback=str(maybe_float(service_row.get("quantity_p75")) or 1.0),
            )
            ws[f"G{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"H{current_row}", f"I{current_row}", f"J{current_row}")
            ws[f"K{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["general"], fallback=str(rate_general or 0), tariff_ref="Builder!E5")
            ws[f"L{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["twin"], fallback=str(rate_twin or 0), tariff_ref="Builder!E5")
            ws[f"M{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["single"], fallback=str(rate_single or 0), tariff_ref="Builder!E5")
            robotic_formula_prefix = 'IF(Builder!B8="Yes",' if row_def.get("robotic_controlled") == "true" else ""
            robotic_formula_suffix = ",0)" if robotic_formula_prefix else ""
            ws[f"N{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={robotic_formula_prefix}H{current_row}*K{current_row}{robotic_formula_suffix}")
            ws[f"O{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={robotic_formula_prefix}I{current_row}*K{current_row}{robotic_formula_suffix}")
            ws[f"P{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={robotic_formula_prefix}J{current_row}*K{current_row}{robotic_formula_suffix}")
            ws[f"Q{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={robotic_formula_prefix}H{current_row}*L{current_row}{robotic_formula_suffix}")
            ws[f"R{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={robotic_formula_prefix}I{current_row}*L{current_row}{robotic_formula_suffix}")
            ws[f"S{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={robotic_formula_prefix}J{current_row}*L{current_row}{robotic_formula_suffix}")
            ws[f"T{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={robotic_formula_prefix}H{current_row}*M{current_row}{robotic_formula_suffix}")
            ws[f"U{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={robotic_formula_prefix}I{current_row}*M{current_row}{robotic_formula_suffix}")
            ws[f"V{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={robotic_formula_prefix}J{current_row}*M{current_row}{robotic_formula_suffix}")
            ws[f"W{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"N{current_row}", f"O{current_row}", f"P{current_row}")
            ws[f"X{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"Q{current_row}", f"R{current_row}", f"S{current_row}")
            ws[f"Y{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"T{current_row}", f"U{current_row}", f"V{current_row}")
        elif kind == "driver":
            code = normalize_code(row_def["code"])
            rate = rate_lookup.get(code)
            if row_def.get("driver") == "los":
                low_ref, typ_ref, high_ref, selected_ref = los_low, los_typ, los_high, selected_los
            elif row_def.get("driver") == "ward":
                low_ref, typ_ref, high_ref, selected_ref = ward_low, ward_typ, ward_high, selected_ward
            else:
                low_ref, typ_ref, high_ref, selected_ref = icu_low, icu_typ, icu_high, selected_icu
            ws[f"G{current_row}"] = f"={selected_ref}"
            ws[f"H{current_row}"] = f"={low_ref}"
            ws[f"I{current_row}"] = f"={typ_ref}"
            ws[f"J{current_row}"] = f"={high_ref}"
            if row_def.get("icu_only") == "true":
                rate_general = rate_twin = rate_single = rate.icu if rate else None
            else:
                rate_general = rate.general if rate else None
                rate_twin = rate.twin if rate else None
                rate_single = rate.single if rate else None
            if row_def.get("icu_only") == "true":
                icu_formula = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["icu"], fallback=str(rate_general or 0), tariff_ref="Builder!E5")
                ws[f"K{current_row}"], ws[f"L{current_row}"], ws[f"M{current_row}"] = icu_formula, icu_formula, icu_formula
            else:
                ws[f"K{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["general"], fallback=str(rate_general or 0), tariff_ref="Builder!E5")
                ws[f"L{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["twin"], fallback=str(rate_twin or 0), tariff_ref="Builder!E5")
                ws[f"M{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["single"], fallback=str(rate_single or 0), tariff_ref="Builder!E5")
            for total_col, qty_col, rate_col in [("N", "H", "K"), ("O", "I", "K"), ("P", "J", "K"), ("Q", "H", "L"), ("R", "I", "L"), ("S", "J", "L"), ("T", "H", "M"), ("U", "I", "M"), ("V", "J", "M"), ("W", "G", "K"), ("X", "G", "L"), ("Y", "G", "M")]:
                ws[f"{total_col}{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={qty_col}{current_row}*{rate_col}{current_row}")
        elif kind == "ward_bed":
            ws[f"G{current_row}"] = f"={selected_ward}"
            ws[f"H{current_row}"] = f"={ward_low}"
            ws[f"I{current_row}"] = f"={ward_typ}"
            ws[f"J{current_row}"] = f"={ward_high}"
            g = rate_lookup.get("ROM0001")
            t = rate_lookup.get("ROM0024")
            s = rate_lookup.get("ROM0036")
            ws[f"K{current_row}"] = build_tariff_rate_lookup_formula('"ROM0001"', TARIFF_RATE_MATRIX_COLS["general"], fallback=str((g.general if g and g.general is not None else 0)), tariff_ref="Builder!E5")
            ws[f"L{current_row}"] = build_tariff_rate_lookup_formula('"ROM0024"', TARIFF_RATE_MATRIX_COLS["general"], fallback=str((t.general if t and t.general is not None else 0)), tariff_ref="Builder!E5")
            ws[f"M{current_row}"] = build_tariff_rate_lookup_formula('"ROM0036"', TARIFF_RATE_MATRIX_COLS["general"], fallback=str((s.general if s and s.general is not None else 0)), tariff_ref="Builder!E5")
            for total_col, qty_col, rate_col in [("N", "H", "K"), ("O", "I", "K"), ("P", "J", "K"), ("Q", "H", "L"), ("R", "I", "L"), ("S", "J", "L"), ("T", "H", "M"), ("U", "I", "M"), ("V", "J", "M"), ("W", "G", "K"), ("X", "G", "L"), ("Y", "G", "M")]:
                ws[f"{total_col}{current_row}"] = f"={qty_col}{current_row}*{rate_col}{current_row}"
        elif kind == "fixed_one":
            rate = rate_lookup.get(normalize_code(row_def["code"]))
            ws[f"G{current_row}"], ws[f"H{current_row}"], ws[f"I{current_row}"], ws[f"J{current_row}"] = 1, 1, 1, 1
            ws[f"K{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["general"], fallback=str((rate.general if rate and rate.general is not None else 0)), tariff_ref="Builder!E5")
            ws[f"L{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["twin"], fallback=str((rate.twin if rate and rate.twin is not None else 0)), tariff_ref="Builder!E5")
            ws[f"M{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["single"], fallback=str((rate.single if rate and rate.single is not None else 0)), tariff_ref="Builder!E5")
            for total_col, rate_col in [("N", "K"), ("O", "K"), ("P", "K"), ("Q", "L"), ("R", "L"), ("S", "L"), ("T", "M"), ("U", "M"), ("V", "M"), ("W", "K"), ("X", "L"), ("Y", "M")]:
                ws[f"{total_col}{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={rate_col}{current_row}")
        elif kind == "ot_hours":
            ws[f"E{current_row}"] = "Selected OT duration snapped to the nearest supported tariff OT slot using the normal or emergency ladder"
            ws[f"F{current_row}"] = "=Builder!B15"
            ws[f"G{current_row}"] = f"={selected_ot}"
            ws[f"H{current_row}"] = f"={selected_ot}"
            ws[f"I{current_row}"] = f"={selected_ot}"
            ws[f"J{current_row}"] = f"={selected_ot}"
            ws[f"K{current_row}"] = build_tariff_rate_lookup_formula("Builder!B15", TARIFF_RATE_MATRIX_COLS["general"], fallback="0", tariff_ref="Builder!E5")
            ws[f"L{current_row}"] = build_tariff_rate_lookup_formula("Builder!B15", TARIFF_RATE_MATRIX_COLS["twin"], fallback="0", tariff_ref="Builder!E5")
            ws[f"M{current_row}"] = build_tariff_rate_lookup_formula("Builder!B15", TARIFF_RATE_MATRIX_COLS["single"], fallback="0", tariff_ref="Builder!E5")
            for total_col, rate_col in [("N", "K"), ("O", "K"), ("P", "K"), ("Q", "L"), ("R", "L"), ("S", "L"), ("T", "M"), ("U", "M"), ("V", "M"), ("W", "K"), ("X", "L"), ("Y", "M")]:
                ws[f"{total_col}{current_row}"] = f"={rate_col}{current_row}"
        elif kind == "cath_lab_history":
            ws[f"G{current_row}"], ws[f"H{current_row}"], ws[f"I{current_row}"], ws[f"J{current_row}"] = 1, 1, 1, 1
            ws[f"N{current_row}"] = f"='{SHEET_REFERENCE}'!{CATH_LAB_REFERENCE_COLS['p25']}4"
            ws[f"O{current_row}"] = f"='{SHEET_REFERENCE}'!{CATH_LAB_REFERENCE_COLS['p50']}4"
            ws[f"P{current_row}"] = f"='{SHEET_REFERENCE}'!{CATH_LAB_REFERENCE_COLS['p75']}4"
            ws[f"Q{current_row}"] = f"=N{current_row}"
            ws[f"R{current_row}"] = f"=O{current_row}"
            ws[f"S{current_row}"] = f"=P{current_row}"
            ws[f"T{current_row}"] = f"=N{current_row}"
            ws[f"U{current_row}"] = f"=O{current_row}"
            ws[f"V{current_row}"] = f"=P{current_row}"
            ws[f"W{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"N{current_row}", f"O{current_row}", f"P{current_row}")
            ws[f"X{current_row}"] = f"=W{current_row}"
            ws[f"Y{current_row}"] = f"=W{current_row}"
        elif kind == "mlc_charge":
            rate = rate_lookup.get(normalize_code(row_def["code"]))
            toggle_ref = "Builder!E6"
            ws[f"G{current_row}"] = f'=IF({toggle_ref}="Yes",1,0)'
            ws[f"H{current_row}"] = f'=IF({toggle_ref}="Yes",1,0)'
            ws[f"I{current_row}"] = f'=IF({toggle_ref}="Yes",1,0)'
            ws[f"J{current_row}"] = f'=IF({toggle_ref}="Yes",1,0)'
            ws[f"K{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["general"], fallback=str((rate.general if rate and rate.general is not None else 0)), tariff_ref="Builder!E5")
            ws[f"L{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["twin"], fallback=str((rate.twin if rate and rate.twin is not None else 0)), tariff_ref="Builder!E5")
            ws[f"M{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["single"], fallback=str((rate.single if rate and rate.single is not None else 0)), tariff_ref="Builder!E5")
            for total_col, qty_col, rate_col in [("N", "H", "K"), ("O", "I", "K"), ("P", "J", "K"), ("Q", "H", "L"), ("R", "I", "L"), ("S", "J", "L"), ("T", "H", "M"), ("U", "I", "M"), ("V", "J", "M"), ("W", "G", "K"), ("X", "G", "L"), ("Y", "G", "M")]:
                ws[f"{total_col}{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f"={qty_col}{current_row}*{rate_col}{current_row}")
        elif kind == "drug_admin":
            ws[f"W{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"N{current_row}", f"O{current_row}", f"P{current_row}")
            ws[f"X{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"Q{current_row}", f"R{current_row}", f"S{current_row}")
            ws[f"Y{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"T{current_row}", f"U{current_row}", f"V{current_row}")
        elif kind == "pf_surgeon":
            pass
        elif kind == "pf_asst_surgeon":
            pass
        elif kind == "pf_anesthetist":
            pass
        elif kind == "pf_asst_anesthetist":
            pass
        elif kind == "pharmacy_ip_drugs":
            ws[f"E{current_row}"] = "Historic per-LOS-day percentile x selected LOS"
            ws[f"G{current_row}"] = f"={selected_los}"
            ws[f"H{current_row}"] = f"={los_low}"
            ws[f"I{current_row}"] = f"={los_typ}"
            ws[f"J{current_row}"] = f"={los_high}"
            ws[f"K{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ip_drugs_day_p50"], "Builder!G5")
            ws[f"L{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ip_drugs_day_p50"], "Builder!G5")
            ws[f"M{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ip_drugs_day_p50"], "Builder!G5")
            ws[f"N{current_row}"] = f"=G{current_row}*{build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS['ip_drugs_day_p25'], 'Builder!G5').lstrip('=')}"
            ws[f"O{current_row}"] = f"=G{current_row}*{build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS['ip_drugs_day_p50'], 'Builder!G5').lstrip('=')}"
            ws[f"P{current_row}"] = f"=G{current_row}*{build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS['ip_drugs_day_p75'], 'Builder!G5').lstrip('=')}"
            ws[f"Q{current_row}"] = f"=N{current_row}"
            ws[f"R{current_row}"] = f"=O{current_row}"
            ws[f"S{current_row}"] = f"=P{current_row}"
            ws[f"T{current_row}"] = f"=N{current_row}"
            ws[f"U{current_row}"] = f"=O{current_row}"
            ws[f"V{current_row}"] = f"=P{current_row}"
            ws[f"W{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"N{current_row}", f"O{current_row}", f"P{current_row}")
            ws[f"X{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"Q{current_row}", f"R{current_row}", f"S{current_row}")
            ws[f"Y{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"T{current_row}", f"U{current_row}", f"V{current_row}")
        elif kind == "pharmacy_ip_consumables":
            ws[f"E{current_row}"] = "Historic per-LOS-day percentile x selected LOS"
            ws[f"G{current_row}"] = f"={selected_los}"
            ws[f"H{current_row}"] = f"={los_low}"
            ws[f"I{current_row}"] = f"={los_typ}"
            ws[f"J{current_row}"] = f"={los_high}"
            ws[f"K{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ip_consumables_day_p50"], "Builder!G5")
            ws[f"L{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ip_consumables_day_p50"], "Builder!G5")
            ws[f"M{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ip_consumables_day_p50"], "Builder!G5")
            ws[f"N{current_row}"] = f"=G{current_row}*{build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS['ip_consumables_day_p25'], 'Builder!G5').lstrip('=')}"
            ws[f"O{current_row}"] = f"=G{current_row}*{build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS['ip_consumables_day_p50'], 'Builder!G5').lstrip('=')}"
            ws[f"P{current_row}"] = f"=G{current_row}*{build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS['ip_consumables_day_p75'], 'Builder!G5').lstrip('=')}"
            ws[f"Q{current_row}"] = f"=N{current_row}"
            ws[f"R{current_row}"] = f"=O{current_row}"
            ws[f"S{current_row}"] = f"=P{current_row}"
            ws[f"T{current_row}"] = f"=N{current_row}"
            ws[f"U{current_row}"] = f"=O{current_row}"
            ws[f"V{current_row}"] = f"=P{current_row}"
            ws[f"W{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"N{current_row}", f"O{current_row}", f"P{current_row}")
            ws[f"X{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"Q{current_row}", f"R{current_row}", f"S{current_row}")
            ws[f"Y{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"T{current_row}", f"U{current_row}", f"V{current_row}")
        elif kind == "pharmacy_ot_drugs":
            ws[f"N{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ot_drugs_p25"], "Builder!G5")
            ws[f"O{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ot_drugs_p50"], "Builder!G5")
            ws[f"P{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ot_drugs_p75"], "Builder!G5")
            ws[f"Q{current_row}"] = f"=N{current_row}"
            ws[f"R{current_row}"] = f"=O{current_row}"
            ws[f"S{current_row}"] = f"=P{current_row}"
            ws[f"T{current_row}"] = f"=N{current_row}"
            ws[f"U{current_row}"] = f"=O{current_row}"
            ws[f"V{current_row}"] = f"=P{current_row}"
            ws[f"W{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"N{current_row}", f"O{current_row}", f"P{current_row}")
            ws[f"X{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"Q{current_row}", f"R{current_row}", f"S{current_row}")
            ws[f"Y{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"T{current_row}", f"U{current_row}", f"V{current_row}")
        elif kind == "pharmacy_ot_consumables":
            ws[f"N{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ot_consumables_p25"], "Builder!G5")
            ws[f"P{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["ot_consumables_p75"], "Builder!G5")
            ws[f"Q{current_row}"] = f"=N{current_row}"
            ws[f"S{current_row}"] = f"=P{current_row}"
            ws[f"T{current_row}"] = f"=N{current_row}"
            ws[f"V{current_row}"] = f"=P{current_row}"
            ws[f"O{current_row}"] = f"={advanced_refs['ot_typical']}"
            ws[f"R{current_row}"] = f"={advanced_refs['ot_typical']}"
            ws[f"U{current_row}"] = f"={advanced_refs['ot_typical']}"
            ws[f"W{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"N{current_row}", f"O{current_row}", f"P{current_row}")
            ws[f"X{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"Q{current_row}", f"R{current_row}", f"S{current_row}")
            ws[f"Y{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"T{current_row}", f"U{current_row}", f"V{current_row}")
        elif kind == "pharmacy_implants":
            ws[f"N{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["implants_p25"], "Builder!G5")
            ws[f"P{current_row}"] = build_reference_basis_lookup_formula(PAYER_BASIS_SUMMARY_COLS["implants_p75"], "Builder!G5")
            ws[f"Q{current_row}"] = f"=N{current_row}"
            ws[f"S{current_row}"] = f"=P{current_row}"
            ws[f"T{current_row}"] = f"=N{current_row}"
            ws[f"V{current_row}"] = f"=P{current_row}"
            ws[f"O{current_row}"] = f"={advanced_refs['implant_typical']}"
            ws[f"R{current_row}"] = f"={advanced_refs['implant_typical']}"
            ws[f"U{current_row}"] = f"={advanced_refs['implant_typical']}"
            ws[f"W{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"N{current_row}", f"O{current_row}", f"P{current_row}")
            ws[f"X{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"Q{current_row}", f"R{current_row}", f"S{current_row}")
            ws[f"Y{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"T{current_row}", f"U{current_row}", f"V{current_row}")
        elif kind == "optional_service":
            idx = optional_service_rows.index(row_def["service_row"]) + int(advanced_refs["optional_start"])
            rate = tariff_rate_for_add_on(rate_lookup.get(normalize_code(row_def["code"])))
            ws[f"H{current_row}"] = build_reference_service_lookup_formula(
                f"F{current_row}",
                PAYER_BASIS_SERVICE_COLS["quantity_p25"],
                fallback=str(as_float(row_def["service_row"].get("quantity_p25"))),
            )
            ws[f"I{current_row}"] = build_reference_service_lookup_formula(
                f"F{current_row}",
                PAYER_BASIS_SERVICE_COLS["quantity_p50"],
                fallback=str(as_float(row_def["service_row"].get("quantity_p50"))),
            )
            ws[f"J{current_row}"] = build_reference_service_lookup_formula(
                f"F{current_row}",
                PAYER_BASIS_SERVICE_COLS["quantity_p75"],
                fallback=str(as_float(row_def["service_row"].get("quantity_p75"))),
            )
            rate_row = rate_lookup.get(normalize_code(row_def["code"]))
            ws[f"K{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["general"], fallback=str((rate_row.general if rate_row and rate_row.general is not None else 0)), tariff_ref="Builder!E5")
            ws[f"L{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["twin"], fallback=str((rate_row.twin if rate_row and rate_row.twin is not None else 0)), tariff_ref="Builder!E5")
            ws[f"M{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["single"], fallback=str((rate_row.single if rate_row and rate_row.single is not None else 0)), tariff_ref="Builder!E5")
            ws[f"N{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF(\'{SHEET_SERVICE_ADDONS}\'!I{idx}="{SELECTION_INCLUDE}",H{current_row}*K{current_row},0)')
            ws[f"O{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF(\'{SHEET_SERVICE_ADDONS}\'!I{idx}="{SELECTION_INCLUDE}",I{current_row}*K{current_row},0)')
            ws[f"P{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF(\'{SHEET_SERVICE_ADDONS}\'!I{idx}="{SELECTION_INCLUDE}",J{current_row}*K{current_row},0)')
            ws[f"Q{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF(\'{SHEET_SERVICE_ADDONS}\'!I{idx}="{SELECTION_INCLUDE}",H{current_row}*L{current_row},0)')
            ws[f"R{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF(\'{SHEET_SERVICE_ADDONS}\'!I{idx}="{SELECTION_INCLUDE}",I{current_row}*L{current_row},0)')
            ws[f"S{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF(\'{SHEET_SERVICE_ADDONS}\'!I{idx}="{SELECTION_INCLUDE}",J{current_row}*L{current_row},0)')
            ws[f"T{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF(\'{SHEET_SERVICE_ADDONS}\'!I{idx}="{SELECTION_INCLUDE}",H{current_row}*M{current_row},0)')
            ws[f"U{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF(\'{SHEET_SERVICE_ADDONS}\'!I{idx}="{SELECTION_INCLUDE}",I{current_row}*M{current_row},0)')
            ws[f"V{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF(\'{SHEET_SERVICE_ADDONS}\'!I{idx}="{SELECTION_INCLUDE}",J{current_row}*M{current_row},0)')
            ws[f"G{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"H{current_row}", f"I{current_row}", f"J{current_row}")
            ws[f"W{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"N{current_row}", f"O{current_row}", f"P{current_row}")
            ws[f"X{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"Q{current_row}", f"R{current_row}", f"S{current_row}")
            ws[f"Y{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"T{current_row}", f"U{current_row}", f"V{current_row}")
            ws[f"D{current_row}"] = "Advanced"
        elif kind == "robotic_service":
            service_row = row_def["service_row"]
            ws[f"H{current_row}"] = build_reference_service_lookup_formula(
                f"F{current_row}",
                PAYER_BASIS_SERVICE_COLS["quantity_p25"],
                fallback=str(as_float(service_row.get("quantity_p25"))),
            )
            ws[f"I{current_row}"] = build_reference_service_lookup_formula(
                f"F{current_row}",
                PAYER_BASIS_SERVICE_COLS["quantity_p50"],
                fallback=str(as_float(service_row.get("quantity_p50"))),
            )
            ws[f"J{current_row}"] = build_reference_service_lookup_formula(
                f"F{current_row}",
                PAYER_BASIS_SERVICE_COLS["quantity_p75"],
                fallback=str(as_float(service_row.get("quantity_p75"))),
            )
            rate_row = rate_lookup.get(normalize_code(row_def["code"]))
            ws[f"K{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["general"], fallback=str((rate_row.general if rate_row and rate_row.general is not None else 0)), tariff_ref="Builder!E5")
            ws[f"L{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["twin"], fallback=str((rate_row.twin if rate_row and rate_row.twin is not None else 0)), tariff_ref="Builder!E5")
            ws[f"M{current_row}"] = build_tariff_rate_lookup_formula(f"F{current_row}", TARIFF_RATE_MATRIX_COLS["single"], fallback=str((rate_row.single if rate_row and rate_row.single is not None else 0)), tariff_ref="Builder!E5")
            robotic_include = 'Builder!B8="Yes"'
            ws[f"G{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"H{current_row}", f"I{current_row}", f"J{current_row}")
            ws[f"N{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF({robotic_include},H{current_row}*K{current_row},0)')
            ws[f"O{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF({robotic_include},I{current_row}*K{current_row},0)')
            ws[f"P{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF({robotic_include},J{current_row}*K{current_row},0)')
            ws[f"Q{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF({robotic_include},H{current_row}*L{current_row},0)')
            ws[f"R{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF({robotic_include},I{current_row}*L{current_row},0)')
            ws[f"S{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF({robotic_include},J{current_row}*L{current_row},0)')
            ws[f"T{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF({robotic_include},H{current_row}*M{current_row},0)')
            ws[f"U{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF({robotic_include},I{current_row}*M{current_row},0)')
            ws[f"V{current_row}"] = wrap_insurance_exclusion(f"F{current_row}", f'=IF({robotic_include},J{current_row}*M{current_row},0)')
            ws[f"W{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"N{current_row}", f"O{current_row}", f"P{current_row}")
            ws[f"X{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"Q{current_row}", f"R{current_row}", f"S{current_row}")
            ws[f"Y{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"T{current_row}", f"U{current_row}", f"V{current_row}")
            ws[f"D{current_row}"] = "Builder Control"
        elif kind == "grouped_residual":
            idx = grouped_adjustment_rows.index(row_def["group_row"]) + int(advanced_refs["grouped_start"])
            group_include_guard = (
                f'IF(AND({insurance_mode_formula},'
                f'ISNUMBER(SEARCH("excluded for insurance",LOWER(\'{SHEET_GROUPED_ADJUSTMENTS}\'!N{idx})))),0,'
                f'IF(\'{SHEET_GROUPED_ADJUSTMENTS}\'!L{idx}="{SELECTION_INCLUDE}",'
            )
            ws[f"G{current_row}"] = ""
            ws[f"H{current_row}"] = ""
            ws[f"I{current_row}"] = ""
            ws[f"J{current_row}"] = ""
            ws[f"K{current_row}"] = f"='{SHEET_GROUPED_ADJUSTMENTS}'!M{idx}"
            ws[f"L{current_row}"] = f"='{SHEET_GROUPED_ADJUSTMENTS}'!M{idx}"
            ws[f"M{current_row}"] = f"='{SHEET_GROUPED_ADJUSTMENTS}'!M{idx}"
            ws[f"N{current_row}"] = f"={group_include_guard}'{SHEET_GROUPED_ADJUSTMENTS}'!I{idx},0))"
            ws[f"O{current_row}"] = f"={group_include_guard}'{SHEET_GROUPED_ADJUSTMENTS}'!J{idx},0))"
            ws[f"P{current_row}"] = f"={group_include_guard}'{SHEET_GROUPED_ADJUSTMENTS}'!K{idx},0))"
            ws[f"Q{current_row}"] = f"=N{current_row}"
            ws[f"R{current_row}"] = f"=O{current_row}"
            ws[f"S{current_row}"] = f"=P{current_row}"
            ws[f"T{current_row}"] = f"=N{current_row}"
            ws[f"U{current_row}"] = f"=O{current_row}"
            ws[f"V{current_row}"] = f"=P{current_row}"
            ws[f"W{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"N{current_row}", f"O{current_row}", f"P{current_row}")
            ws[f"X{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"Q{current_row}", f"R{current_row}", f"S{current_row}")
            ws[f"Y{current_row}"] = build_mode_pick_formula(rate_mode_cell, f"T{current_row}", f"U{current_row}", f"V{current_row}")
            ws[f"D{current_row}"] = "Grouped Residual"
        current_row += 1

    last_core_data_row = current_row - 1
    kind_to_row = {all_rows[idx]["kind"]: 2 + idx for idx in range(len(core_rows)) if idx < len(core_rows)}
    drug_admin_row = kind_to_row["drug_admin"]
    ip_drugs_row = kind_to_row["pharmacy_ip_drugs"]
    ip_cons_row = kind_to_row["pharmacy_ip_consumables"]
    ot_drugs_row = kind_to_row["pharmacy_ot_drugs"]
    ot_cons_row = kind_to_row["pharmacy_ot_consumables"]
    implants_row = kind_to_row["pharmacy_implants"]
    for target_col, source_cols in {
        "N": ["N", "N", "N", "N", "N"],
        "O": ["O", "O", "O", "O", "O"],
        "P": ["P", "P", "P", "P", "P"],
        "Q": ["Q", "Q", "Q", "Q", "Q"],
        "R": ["R", "R", "R", "R", "R"],
        "S": ["S", "S", "S", "S", "S"],
        "T": ["T", "T", "T", "T", "T"],
        "U": ["U", "U", "U", "U", "U"],
        "V": ["V", "V", "V", "V", "V"],
        "W": ["W", "W", "W", "W", "W"],
        "X": ["X", "X", "X", "X", "X"],
        "Y": ["Y", "Y", "Y", "Y", "Y"],
    }.items():
        refs = [f"{col}{row}" for col, row in zip(source_cols, [ip_drugs_row, ip_cons_row, ot_drugs_row, ot_cons_row, implants_row])]
        ws[f"{target_col}{drug_admin_row}"] = f'=IF({insurance_mode_formula},0,0.125*SUM({",".join(refs)}))'

    ws["N2"] = ws["N2"].value  # no-op to avoid linter-ish issues

    pf_rows = {
        "surgeon": kind_to_row["pf_surgeon"],
        "asst_surgeon": kind_to_row["pf_asst_surgeon"],
        "anesthetist": kind_to_row["pf_anesthetist"],
        "asst_anesthetist": kind_to_row["pf_asst_anesthetist"],
    }
    subtotal_row = current_row
    ws[f"A{subtotal_row}"] = "Subtotal Before Professional Fees"
    ws[f"B{subtotal_row}"] = "Grand Total"
    pf_detail_rows = [
        pf_rows["surgeon"],
        pf_rows["asst_surgeon"],
        pf_rows["anesthetist"],
        pf_rows["asst_anesthetist"],
    ]
    pf_start = min(pf_detail_rows)
    pf_end = max(pf_detail_rows)
    pre_pf_start = 2
    pre_pf_end = pf_start - 1
    post_pf_start = pf_end + 1
    post_pf_end = subtotal_row - 1
    for low_col in ["N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y"]:
        sum_parts: list[str] = []
        if pre_pf_end >= pre_pf_start:
            sum_parts.append(f"{low_col}{pre_pf_start}:{low_col}{pre_pf_end}")
        if post_pf_end >= post_pf_start:
            sum_parts.append(f"{low_col}{post_pf_start}:{low_col}{post_pf_end}")
        if sum_parts:
            ws[f"{low_col}{subtotal_row}"] = f"=SUM({','.join(sum_parts)})"
        else:
            ws[f"{low_col}{subtotal_row}"] = 0

    ws[f"W{pf_rows['surgeon']}"] = f'=IF({insurance_mode_formula},0,{pf_multipliers["surgeon"]}*W{subtotal_row})'
    ws[f"X{pf_rows['surgeon']}"] = f'=IF({insurance_mode_formula},0,{pf_multipliers["surgeon"]}*X{subtotal_row})'
    ws[f"Y{pf_rows['surgeon']}"] = f'=IF({insurance_mode_formula},0,{pf_multipliers["surgeon"]}*Y{subtotal_row})'
    for col_base, room_col in [("N", "K"), ("O", "K"), ("P", "K"), ("Q", "L"), ("R", "L"), ("S", "L"), ("T", "M"), ("U", "M"), ("V", "M")]:
        total_ref = {"N": "N", "O": "O", "P": "P", "Q": "Q", "R": "R", "S": "S", "T": "T", "U": "U", "V": "V"}[col_base]
        ws[f"{col_base}{pf_rows['surgeon']}"] = f'=IF({insurance_mode_formula},0,{pf_multipliers["surgeon"]}*{total_ref}{subtotal_row})'
    for room_col in ["W", "X", "Y"]:
        base = room_col + str(pf_rows["surgeon"])
        ws[f"{room_col}{pf_rows['asst_surgeon']}"] = f'=IF({insurance_mode_formula},0,{pf_multipliers["assistant_surgeon"]}*{base})'
        ws[f"{room_col}{pf_rows['anesthetist']}"] = f'=IF({insurance_mode_formula},0,{pf_multipliers["anesthetist"]}*{base})'
        if pf_multipliers["assistant_anesthetist"] > 0:
            an_base = room_col + str(pf_rows["anesthetist"])
            ws[f"{room_col}{pf_rows['asst_anesthetist']}"] = f'=IF({insurance_mode_formula},0,{pf_multipliers["assistant_anesthetist"]}*{an_base})'
        else:
            ws[f"{room_col}{pf_rows['asst_anesthetist']}"] = 0
    for low_col in ["N", "O", "P", "Q", "R", "S", "T", "U", "V"]:
        room_selected_col = {"N": "W", "O": "W", "P": "W", "Q": "X", "R": "X", "S": "X", "T": "Y", "U": "Y", "V": "Y"}[low_col]
        # mirror selected room formulas into low/typ/high for detail completeness
        if low_col in {"N", "O", "P"}:
            surgeon_base = {"N": "N", "O": "O", "P": "P"}[low_col]
        elif low_col in {"Q", "R", "S"}:
            surgeon_base = {"Q": "Q", "R": "R", "S": "S"}[low_col]
        else:
            surgeon_base = {"T": "T", "U": "U", "V": "V"}[low_col]
        ws[f"{low_col}{pf_rows['surgeon']}"] = f'=IF({insurance_mode_formula},0,{pf_multipliers["surgeon"]}*{surgeon_base}{subtotal_row})'
        ws[f"{low_col}{pf_rows['asst_surgeon']}"] = f'=IF({insurance_mode_formula},0,{pf_multipliers["assistant_surgeon"]}*{low_col}{pf_rows["surgeon"]})'
        ws[f"{low_col}{pf_rows['anesthetist']}"] = f'=IF({insurance_mode_formula},0,{pf_multipliers["anesthetist"]}*{low_col}{pf_rows["surgeon"]})'
        if pf_multipliers["assistant_anesthetist"] > 0:
            ws[f"{low_col}{pf_rows['asst_anesthetist']}"] = f'=IF({insurance_mode_formula},0,{pf_multipliers["assistant_anesthetist"]}*{low_col}{pf_rows["anesthetist"]})'
        else:
            ws[f"{low_col}{pf_rows['asst_anesthetist']}"] = 0
    grand_total_row = subtotal_row + 1
    ws[f"A{grand_total_row}"] = "Grand Total"
    ws[f"B{grand_total_row}"] = "Grand Total"
    for col in ["N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y"]:
        ws[f"{col}{grand_total_row}"] = f"={col}{subtotal_row}+SUM({col}{pf_start}:{col}{pf_end})"

    style_range(ws, f"A2:Y{grand_total_row}", fill=SELECTION_FILL, wrap=True)
    for row_num in [subtotal_row, grand_total_row]:
        style_row(ws, row_num, 1, 25, fill=RESULT_FILL, bold=True, wrap=True)
    for row_num in range(2, grand_total_row + 1):
        ws[f"A{row_num}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"C{row_num}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"D{row_num}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"E{row_num}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    set_number_format(ws, f"G2:Y{grand_total_row}", '#,##0.00')
    ws.column_dimensions["F"].hidden = True
    for col, width in {
        "A": 34, "B": 20, "C": 18, "D": 14, "E": 24, "F": 12, "G": 12, "H": 10, "I": 10, "J": 10,
        "K": 12, "L": 12, "M": 12, "N": 12, "O": 12, "P": 12, "Q": 12, "R": 12, "S": 12,
        "T": 12, "U": 12, "V": 12, "W": 14, "X": 14, "Y": 14,
    }.items():
        ws.column_dimensions[col].width = width
    return {"subtotal_row": subtotal_row, "grand_total_row": grand_total_row}


def write_estimate_summary_sheet(ws, detail_meta: dict[str, int], actual_rows: list[dict[str, Any]]) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Estimate Summary"
    ws["A2"] = "Selected Room"
    ws["B2"] = "=Builder!B4"
    ws["A3"] = "Selected Estimate Mode"
    ws["B3"] = "=Builder!B5"
    ws["A4"] = "Historical Payer Basis"
    ws["B4"] = "=Builder!E3"
    ws["A5"] = "Pricing Mode"
    ws["B5"] = "=Builder!E2"
    ws["A6"] = "Insurance Org Code"
    ws["B6"] = "=Builder!G2"
    ws["A7"] = "Resolved Payor Bucket"
    ws["B7"] = "=Builder!E4"
    ws["A8"] = "Resolved Tariff Code"
    ws["B8"] = "=Builder!E5"
    ws["A9"] = "Pharmacy Benchmark Basis"
    ws["B9"] = "=Builder!G5"
    ws["A10"] = "Resolved Service Basis"
    ws["B10"] = "=Builder!G6"
    ws["A11"] = "Resolved PF Basis"
    ws["B11"] = "=Builder!G7"
    ws["D3"] = "How estimate mode works"
    ws["E3"] = "Low uses P25 amounts, Typical uses P50, and High uses P75. Driver rows change only if their own Selection or Manual value changes. OT uses the resolved tariff OT slot from the selected OT duration and Emergency OT setting. Insurance mode prices services from the selected org tariff and suppresses insurance-excluded rows."
    style_range(ws, "A1:H1", fill=HEADER_FILL, bold=True, font_color="FFFFFF", align="left")
    style_range(ws, "A2:B11", fill=SELECTION_FILL, bold=True)
    style_range(ws, "D3:H3", fill=SELECTION_FILL, align="left", wrap=True)

    grand_total_row = detail_meta["grand_total_row"]
    last_detail_row = grand_total_row - 2
    room_mode_cells = {
        ("General", "Low"): f"'{SHEET_DETAIL}'!N{grand_total_row}",
        ("General", "Typical"): f"'{SHEET_DETAIL}'!O{grand_total_row}",
        ("General", "High"): f"'{SHEET_DETAIL}'!P{grand_total_row}",
        ("Twin", "Low"): f"'{SHEET_DETAIL}'!Q{grand_total_row}",
        ("Twin", "Typical"): f"'{SHEET_DETAIL}'!R{grand_total_row}",
        ("Twin", "High"): f"'{SHEET_DETAIL}'!S{grand_total_row}",
        ("Single", "Low"): f"'{SHEET_DETAIL}'!T{grand_total_row}",
        ("Single", "Typical"): f"'{SHEET_DETAIL}'!U{grand_total_row}",
        ("Single", "High"): f"'{SHEET_DETAIL}'!V{grand_total_row}",
    }
    ws["D2"] = "Final Estimate"
    ws["E2"] = build_room_pick_formula(
        "B2",
        f"'{SHEET_DETAIL}'!W{grand_total_row}",
        f"'{SHEET_DETAIL}'!X{grand_total_row}",
        f"'{SHEET_DETAIL}'!Y{grand_total_row}",
    )
    style_range(ws, "D2:E2", fill=RESULT_FILL, bold=True)

    ws["D6"], ws["E6"], ws["F6"], ws["G6"] = "Room", "Low", "Typical", "High"
    style_range(ws, "D6:G6", fill=SUBHEADER_FILL, bold=True)
    for idx, room in enumerate(["General", "Twin", "Single"], start=7):
        ws[f"D{idx}"] = room
        ws[f"E{idx}"] = f"={room_mode_cells[(room, 'Low')]}"
        ws[f"F{idx}"] = f"={room_mode_cells[(room, 'Typical')]}"
        ws[f"G{idx}"] = f"={room_mode_cells[(room, 'High')]}"
    style_range(ws, "D7:G9", fill=SELECTION_FILL)

    ws["I6"] = "Selected Drivers & Controls"
    style_range(ws, "I6:K6", fill=SUBHEADER_FILL, bold=True, align="left")
    summary_facts = [
        ("Selected LOS", "=Builder!G10"),
        ("Selected ICU Days", "=Builder!G11"),
        ("Selected Ward Days", "=Builder!G12"),
        ("Selected OT Hours", "=Builder!G13"),
        ("Emergency OT?", "=Builder!B6"),
        ("Resolved OT Slot Hours", "=Builder!B14"),
        ("Resolved OT Slot Code", "=Builder!B15"),
        ("Resolved OT Type", "=Builder!B17"),
        ("MLC?", "=Builder!E6"),
        ("OT Consumables Selected Typical", f"='{SHEET_ADVANCED}'!C6"),
        ("Implants Selected Typical", f"='{SHEET_IMPLANTS_SELECT}'!F6"),
        ("Optional Add-Ons Selected Typical", f"='{SHEET_SERVICE_ADDONS}'!C5"),
        ("Grouped Adjustments Selected Typical", f"='{SHEET_GROUPED_ADJUSTMENTS}'!C5"),
        ("Grouped Adjustments Included Count", f"='{SHEET_GROUPED_ADJUSTMENTS}'!E5"),
    ]
    for row_num, (label, formula) in enumerate(summary_facts, start=7):
        ws[f"I{row_num}"] = label
        ws[f"J{row_num}"] = formula
    style_range(ws, "I7:J20", fill=SELECTION_FILL)

    ws["I24"] = "Service Count Check"
    ws["I25"] = "Current Included"
    ws["J25"] = f"='{SHEET_SERVICE_ADDONS}'!P10"
    ws["I26"] = "Historical P25"
    ws["J26"] = f"='{SHEET_SERVICE_ADDONS}'!P5"
    ws["I27"] = "Historical P50"
    ws["J27"] = f"='{SHEET_SERVICE_ADDONS}'!P6"
    ws["I28"] = "Historical P75"
    ws["J28"] = f"='{SHEET_SERVICE_ADDONS}'!P7"
    ws["I29"] = "Alert"
    ws["J29"] = f"='{SHEET_SERVICE_ADDONS}'!P11"
    style_range(ws, "I24:J29", fill=SELECTION_FILL)

    ws["L6"] = "Cohort Basis Counts"
    style_range(ws, "L6:M6", fill=SUBHEADER_FILL, bold=True)
    for idx, (label, col) in enumerate(
        [
            ("Selected Cohort Size", PAYER_BASIS_SUMMARY_COLS["cohort_size"]),
            ("Cash Count", PAYER_BASIS_SUMMARY_COLS["cash_count"]),
            ("GIPSA Count", PAYER_BASIS_SUMMARY_COLS["gipsa_count"]),
            ("Non-GIPSA Count", PAYER_BASIS_SUMMARY_COLS["non_gipsa_count"]),
            ("Corporate Count", PAYER_BASIS_SUMMARY_COLS["corporate_count"]),
        ],
        start=7,
    ):
        ws[f"L{idx}"] = label
        ws[f"M{idx}"] = build_reference_basis_lookup_formula(col)
    style_range(ws, "L7:M11", fill=SELECTION_FILL)

    ws["L13"] = "Pharmacy P50 by Basis"
    style_range(ws, "L13:R13", fill=SUBHEADER_FILL, bold=True)
    headers = ["Bucket", "Cash", "GIPSA", "Non-GIPSA", "Corporate", "Insurance All", "All Payers"]
    for col_idx, value in enumerate(headers, start=12):
        ws.cell(row=14, column=col_idx, value=value)
    style_range(ws, "L14:R14", fill=SELECTION_FILL, bold=True)
    comparison_rows = [
        ("IP Drugs / Day", "ip_drugs_day_p50"),
        ("IP Consumables / Day", "ip_consumables_day_p50"),
        ("OT Drugs", "ot_drugs_p50"),
        ("OT Consumables", "ot_consumables_p50"),
        ("Implants", "implants_p50"),
    ]
    basis_labels = ["Cash", "GIPSA Insurance", "Non-GIPSA Insurance", "Corporate", "Insurance All", "All Payers"]
    for row_offset, (label, key) in enumerate(comparison_rows, start=15):
        ws[f"L{row_offset}"] = label
        for col_offset, basis_label in enumerate(basis_labels, start=13):
            lookup = (
                f'=IFERROR(INDEX(\'{SHEET_REFERENCE}\'!${PAYER_BASIS_SUMMARY_COLS[key]}${PAYER_BASIS_SUMMARY_START_ROW}:'
                f'${PAYER_BASIS_SUMMARY_COLS[key]}$500, MATCH("{basis_label}", '
                f'\'{SHEET_REFERENCE}\'!${PAYER_BASIS_SUMMARY_COLS["basis_label"]}${PAYER_BASIS_SUMMARY_START_ROW}:'
                f'${PAYER_BASIS_SUMMARY_COLS["basis_label"]}$500, 0)), 0)'
            )
            ws.cell(row=row_offset, column=col_offset, value=lookup)
    style_range(ws, "L15:R19", fill=SELECTION_FILL)

    ws["A25"] = "Professional Fees by Payer"
    style_range(ws, "A25:E25", fill=SUBHEADER_FILL, bold=True)
    for col_idx, value in enumerate(["Payer", "Cases", "P25", "P50", "P75"], start=1):
        ws.cell(row=26, column=col_idx, value=value)
    style_range(ws, "A26:E26", fill=SELECTION_FILL, bold=True)
    pf_payor_rows = [
        ("Cash", '"Cash"'),
        ("GIPSA Insurance", '"GIPSA Insurance"'),
        ("Non-GIPSA Insurance", '"Non-GIPSA Insurance"'),
        ("Corporate", '"Corporate"'),
        ("Insurance All", '"Insurance All"'),
        ("All Payers", '"All Payers"'),
    ]
    for row_idx, (label, lookup_key) in enumerate(pf_payor_rows, start=27):
        ws[f"A{row_idx}"] = label
        ws[f"B{row_idx}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["admission_count"], lookup_key)
        ws[f"C{row_idx}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p25"], lookup_key)
        ws[f"D{row_idx}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p50"], lookup_key)
        ws[f"E{row_idx}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p75"], lookup_key)
    style_range(ws, "A27:E32", fill=SELECTION_FILL)

    ws["L24"] = "Selected Basis PF Mix"
    style_range(ws, "L24:P24", fill=SUBHEADER_FILL, bold=True)
    for col_idx, value in enumerate(["PF Bucket", "P25", "P50", "P75", "Note"], start=7):
        ws.cell(row=25, column=col_idx + 5, value=value)
    style_range(ws, "L25:P25", fill=SELECTION_FILL, bold=True)
    selected_pf_mix_rows = [
        ("Collectible Historical PF", "pf_collectible_historical_total_p25", "pf_collectible_historical_total_p50", "pf_collectible_historical_total_p75", "Total collectible PF"),
        ("Named PF", "pf_named_total_p25", "pf_named_total_p50", "pf_named_total_p75", "Named doctor rows"),
        ("General Needed PF", "pf_general_needed_total_p25", "pf_general_needed_total_p50", "pf_general_needed_total_p75", "General-needed rows"),
        ("Surgeon Named", "surgeon_named_total_p25", "surgeon_named_total_p50", "surgeon_named_total_p75", "Named surgeon rows"),
        ("Assistant Surgeon Named", "assistant_surgeon_named_total_p25", "assistant_surgeon_named_total_p50", "assistant_surgeon_named_total_p75", "Named assistant surgeon rows"),
        ("Anesthetist Named", "anesthetist_named_total_p25", "anesthetist_named_total_p50", "anesthetist_named_total_p75", "Named anesthetist rows"),
        ("Assistant Anesthetist Named", "assistant_anesthetist_named_total_p25", "assistant_anesthetist_named_total_p50", "assistant_anesthetist_named_total_p75", "Named assistant anesthetist rows"),
        ("Consultant / Physician Named", "consultant_or_physician_named_total_p25", "consultant_or_physician_named_total_p50", "consultant_or_physician_named_total_p75", "Named consultant/physician rows"),
    ]
    for row_idx, (label, p25_key, p50_key, p75_key, note) in enumerate(selected_pf_mix_rows, start=26):
        ws[f"L{row_idx}"] = label
        ws[f"M{row_idx}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS[p25_key], "Builder!G7")
        ws[f"N{row_idx}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS[p50_key], "Builder!G7")
        ws[f"O{row_idx}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS[p75_key], "Builder!G7")
        ws[f"P{row_idx}"] = note
    style_range(ws, "L26:P33", fill=SELECTION_FILL)
    ws["L34"] = "Selected Basis PF Shape"
    ws["M34"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["dominant_pf_shape"], "Builder!G7", '""')
    style_range(ws, "L34:M34", fill=SELECTION_FILL, bold=True)

    ws["A12"], ws["B12"] = "Bucket", "Selected Estimate"
    style_range(ws, "A12:B12", fill=SUBHEADER_FILL, bold=True)
    bucket_names = [
        "Room Charges",
        "Investigations",
        "Procedure / OT Charges",
        "Bedside Services",
        "Pharmacy",
        "Drug Administration Charges",
        "Professional Fees",
        "Optional Add-Ons",
        "Grand Total",
        "Professional Fees (Historic Basis P50)",
        "Grand Total (Historic PF)",
    ]
    for row_num, bucket in enumerate(bucket_names, start=13):
        ws[f"A{row_num}"] = bucket
        if bucket == "Grand Total":
            ws[f"B{row_num}"] = "=E2"
        elif bucket == "Professional Fees (Historic Basis P50)":
            ws[f"B{row_num}"] = build_pf_payor_lookup_formula(PF_PAYOR_SUMMARY_COLS["pf_collectible_historical_total_p50"], "Builder!G7")
        elif bucket == "Grand Total (Historic PF)":
            ws[f"B{row_num}"] = "=B21-B19+B22"
        else:
            ws[f"B{row_num}"] = build_room_pick_formula(
                "B2",
                f"SUMIF('{SHEET_DETAIL}'!$B$2:$B${last_detail_row},A{row_num},'{SHEET_DETAIL}'!$W$2:$W${last_detail_row})",
                f"SUMIF('{SHEET_DETAIL}'!$B$2:$B${last_detail_row},A{row_num},'{SHEET_DETAIL}'!$X$2:$X${last_detail_row})",
                f"SUMIF('{SHEET_DETAIL}'!$B$2:$B${last_detail_row},A{row_num},'{SHEET_DETAIL}'!$Y$2:$Y${last_detail_row})",
            )
    style_range(ws, "A13:B23", fill=SELECTION_FILL)
    set_number_format(ws, "B7:E23", '#,##0.00')
    set_number_format(ws, "I7:N19", '#,##0.00')
    set_number_format(ws, "B27:E32", '#,##0.00')
    set_number_format(ws, "M26:O33", '#,##0.00')
    write_selected_basis_actuals_snapshot(
        ws,
        start_row=36,
        title="IP FC Actuals Selected-Basis Snapshot",
    )

    for col, width in {"A": 24, "B": 18, "C": 18, "D": 18, "E": 20, "F": 12, "G": 18, "H": 14, "I": 14, "J": 14, "K": 14, "L": 24, "M": 14, "N": 14, "O": 14, "P": 24, "Q": 16, "R": 14}.items():
        ws.column_dimensions[col].width = width


def build_guided_workbook(
    quartiles_json: dict[str, Any],
    bucket_quartiles: dict[str, tuple[float, float, float]],
    ip_pharmacy_per_day_metrics: dict[str, Any],
    cath_lab_metrics: dict[str, Any],
    service_line_count_metrics: dict[str, Any],
    rate_lookup: dict[str, RateRow],
    service_rows: list[dict[str, str]],
    service_lookup: dict[str, dict[str, str]],
    pharmacy_rows: list[dict[str, str]],
    per_ip_rows: list[dict[str, str]],
    ip_actual_rows: list[dict[str, Any]],
    implant_rows: list[dict[str, str]],
    implant_detail_rows: list[dict[str, str]],
    ot_slot_rows: list[OtSlotRateRow],
    org_tariff_reference_rows: list[dict[str, str]],
    tariff_rate_matrix_rows: list[dict[str, str]],
    tariff_ot_slot_matrix_rows: list[dict[str, str]],
    insurance_policy_rows: list[dict[str, str]],
    payer_basis_summary: dict[str, Any],
    payer_basis_service_rows: list[dict[str, str]],
    payer_basis_pharmacy_rows: list[dict[str, str]],
    payer_basis_resolution_rows: list[dict[str, str]],
    pf_payor_summary_rows: list[dict[str, str]],
    grouping_gap_summary_rows: list[dict[str, str]],
    grouping_gap_child_rows: list[dict[str, str]],
    args: argparse.Namespace,
) -> Workbook:
    cleaned_service_rows, auto_included_service_rows, optional_service_rows = clean_services_for_fc(service_rows)
    optional_service_rows = prioritize_optional_service_rows(optional_service_rows, rate_lookup)
    optional_service_rows, robotic_service_rows = split_robotic_optional_service_rows(
        optional_service_rows,
        procedure_code=args.procedure_code,
    )
    args.robotic_service_rows = collect_robotic_service_rows(
        cleaned_service_rows,
        procedure_code=args.procedure_code,
        include_procedure_row=args.include_procedure_row,
    )
    args.robotic_optional_service_count = len(robotic_service_rows)
    args.robotic_charge_presence_rate = compute_robotic_charge_presence_rate(
        collect_robotic_presence_signal_rows(
            service_rows,
            procedure_code=args.procedure_code,
        )
    )
    args.robotic_default_selection = resolve_robotic_default_selection(
        default_mode=args.robotic_default_mode,
        presence_rate=as_float(getattr(args, "robotic_charge_presence_rate", 0.0)),
        presence_threshold=args.robotic_presence_threshold,
    )
    grouped_adjustment_rows = build_grouped_residual_candidates(grouping_gap_summary_rows)
    insurance_excluded_groupings = build_insurance_excluded_groupings(
        grouping_gap_child_rows,
        insurance_policy_rows,
    )
    workbook = Workbook()
    apply_default_calc_settings(workbook)
    default_ws = workbook.active
    workbook.remove(default_ws)
    builder_ws = workbook.create_sheet(SHEET_BUILDER)
    summary_ws = workbook.create_sheet(SHEET_SUMMARY)
    estimate_vs_actual_ws = workbook.create_sheet(SHEET_ESTIMATE_VS_ACTUAL)
    advanced_ws = workbook.create_sheet(SHEET_ADVANCED)
    service_addons_ws = workbook.create_sheet(SHEET_SERVICE_ADDONS)
    grouped_adjustments_ws = workbook.create_sheet(SHEET_GROUPED_ADJUSTMENTS)
    grouping_review_ws = workbook.create_sheet(SHEET_GROUPING_REVIEW)
    implant_select_ws = workbook.create_sheet(SHEET_IMPLANTS_SELECT)
    breakdown_ws = workbook.create_sheet(SHEET_BREAKDOWN)
    detail_ws = workbook.create_sheet(SHEET_DETAIL)
    pharmacy_template_ws = workbook.create_sheet(SHEET_PHARMACY_TEMPLATE)
    service_template_ws = workbook.create_sheet(SHEET_SERVICE_TEMPLATE)
    pharmacy_metrics_ws = workbook.create_sheet(SHEET_PHARMACY_METRICS)
    ip_actuals_ws = workbook.create_sheet(SHEET_IP_ACTUALS)
    pf_review_ws = workbook.create_sheet(SHEET_PF_REVIEW)
    reference_ws = workbook.create_sheet(SHEET_REFERENCE)

    write_builder_sheet(builder_ws, quartiles_json, args)
    advanced_refs = write_advanced_controls_sheet(
        advanced_ws,
        bucket_quartiles,
        pharmacy_rows,
        args,
    )
    advanced_refs.update(
        write_service_addons_sheet(
            service_addons_ws,
            optional_service_rows,
            rate_lookup,
            service_line_count_metrics,
            args,
        )
    )
    advanced_refs.update(
        write_grouped_adjustments_sheet(
            grouped_adjustments_ws,
            grouped_adjustment_rows,
            insurance_excluded_groupings,
        )
    )
    implant_typical_ref = write_implant_selection_sheet(
        implant_select_ws,
        bucket_quartiles,
        implant_rows,
    )
    advanced_refs["implant_typical"] = implant_typical_ref
    detail_meta = write_line_item_detail_sheet(
        detail_ws,
        quartiles_json,
        bucket_quartiles,
        ip_pharmacy_per_day_metrics,
        service_lookup,
        cleaned_service_rows,
        auto_included_service_rows,
        optional_service_rows,
        robotic_service_rows,
        grouped_adjustment_rows,
        rate_lookup,
        advanced_refs,
        args,
    )
    write_estimate_summary_sheet(summary_ws, detail_meta, ip_actual_rows)
    write_estimate_vs_actual_sheet(estimate_vs_actual_ws, summary_ws)
    write_estimate_breakdown_sheet(breakdown_ws, detail_ws, detail_meta)
    write_pharmacy_template_sheet(pharmacy_template_ws, pharmacy_rows)
    write_service_template_sheet(service_template_ws, cleaned_service_rows)
    write_grouping_review_sheet(grouping_review_ws, grouping_gap_summary_rows, grouping_gap_child_rows)
    write_pharmacy_metrics_sheet(pharmacy_metrics_ws, per_ip_rows)
    write_ip_fc_actuals_sheet(ip_actuals_ws, ip_actual_rows)
    write_pf_review_sheet(pf_review_ws, args)
    write_reference_sheet(
        reference_ws,
        quartiles_json,
        bucket_quartiles,
        ip_pharmacy_per_day_metrics,
        cath_lab_metrics,
        service_line_count_metrics,
        cleaned_service_rows,
        optional_service_rows,
        rate_lookup,
        implant_rows,
        ot_slot_rows,
        org_tariff_reference_rows,
        tariff_rate_matrix_rows,
        tariff_ot_slot_matrix_rows,
        insurance_policy_rows,
        payer_basis_summary,
        payer_basis_service_rows,
        payer_basis_pharmacy_rows,
        payer_basis_resolution_rows,
        pf_payor_summary_rows,
        ip_actual_rows,
    )
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = None
    return workbook


def rebuild_workbook(args: argparse.Namespace) -> Path:
    runtime_inputs = load_builder_runtime_inputs(args)
    quartiles_json = runtime_inputs["quartiles_json"]
    bucket_quartiles = runtime_inputs["bucket_quartiles"]
    ip_pharmacy_per_day_metrics = runtime_inputs["ip_pharmacy_per_day_metrics"]
    cath_lab_metrics = load_json(args.cath_lab_metrics_json)
    service_line_count_metrics = load_json(args.service_line_count_metrics)
    payer_basis_summary = runtime_inputs["payer_basis_summary"]
    payer_basis_service_rows = runtime_inputs["payer_basis_service_rows"]
    payer_basis_pharmacy_rows = runtime_inputs["payer_basis_pharmacy_rows"]
    payer_basis_resolution_rows = load_csv_rows_if_exists(args.payer_basis_resolution_csv)
    args.payer_basis_resolution_rows = payer_basis_resolution_rows
    pf_payor_summary_rows = load_csv_rows_if_exists(args.pf_payor_summary_csv)
    grouping_gap_summary_rows = load_csv_rows_if_exists(args.grouping_gap_summary_csv)
    grouping_gap_child_rows = load_csv_rows_if_exists(args.grouping_gap_child_detail_csv)
    rate_lookup = load_rate_lookup(args.rate_csv)
    ot_slot_rows = load_ot_slot_rate_rows(args.ot_slot_rate_csv)
    org_tariff_reference_rows = load_csv_rows(args.org_tariff_reference_csv)
    tariff_rate_matrix_rows = load_csv_rows(args.tariff_rate_matrix_csv)
    tariff_ot_slot_matrix_rows = load_csv_rows(args.tariff_ot_slot_rate_matrix_csv)
    insurance_policy_rows = load_csv_rows(args.insurance_policy_csv)
    service_rows = runtime_inputs["service_rows"]
    service_lookup = load_service_lookup(args.services_template)
    cleaned_service_rows, _, _ = clean_services_for_fc(service_rows)
    cleaned_service_rows = filter_out_cath_lab_slot_rows(cleaned_service_rows)
    write_cleaned_services_csv(args.cleaned_services_output, cleaned_service_rows)
    pharmacy_rows = runtime_inputs["pharmacy_rows"]
    per_ip_rows = runtime_inputs["per_ip_rows"]
    ip_actual_rows = runtime_inputs["ip_actual_rows"]
    implant_rows = runtime_inputs["implant_rows"]
    implant_detail_rows = runtime_inputs["implant_detail_rows"]

    workbook = build_guided_workbook(
        quartiles_json,
        bucket_quartiles,
        ip_pharmacy_per_day_metrics,
        cath_lab_metrics,
        service_line_count_metrics,
        rate_lookup,
        service_rows,
        service_lookup,
        pharmacy_rows,
        per_ip_rows,
        ip_actual_rows,
        implant_rows,
        implant_detail_rows,
        ot_slot_rows,
        org_tariff_reference_rows,
        tariff_rate_matrix_rows,
        tariff_ot_slot_matrix_rows,
        insurance_policy_rows,
        payer_basis_summary,
        payer_basis_service_rows,
        payer_basis_pharmacy_rows,
        payer_basis_resolution_rows,
        pf_payor_summary_rows,
        grouping_gap_summary_rows,
        grouping_gap_child_rows,
        args,
    )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    apply_default_calc_settings(workbook)
    workbook.save(args.output)
    recalculate_workbook_with_soffice(args.output)
    validation_summary = validate_generated_surgical_workbook(args.output)
    if args.validation_output_json:
        args.validation_output_json.parent.mkdir(parents=True, exist_ok=True)
        args.validation_output_json.write_text(json.dumps(validation_summary, indent=2), encoding="utf-8")
    return args.output


def main() -> None:
    args = parse_args()
    output_path = rebuild_workbook(args)
    print(f"output={output_path}")


if __name__ == "__main__":
    main()
